"""Flask attendance viewer - MariaDB version (for NAS deployment)"""
import os
import re
import json
import time as _time
from dotenv import load_dotenv

load_dotenv()

# ── .env 파일 수동 로딩 (python-dotenv 없이) ──────────────
_env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
if os.path.exists(_env_path):
    with open(_env_path, encoding="utf-8") as _ef:
        for _line in _ef:
            _line = _line.strip()
            if _line and not _line.startswith("#") and "=" in _line:
                _k, _v = _line.split("=", 1)
                os.environ.setdefault(_k.strip(), _v.strip())

# ── 타임존을 한국(KST)으로 고정 (UTC 컨테이너 환경 대응) ──
os.environ["TZ"] = "Asia/Seoul"
_time.tzset()

from tuya_fire import start_tuya_poller
import ssl
import calendar
import threading
from collections import OrderedDict, defaultdict
from datetime import datetime, timedelta, date
from urllib.request import Request, urlopen
from urllib.parse import urlencode

import hashlib
import hmac as _hmac
import socket
from functools import wraps

import holidays
import pymysql
from flask import Flask, render_template, request, flash, jsonify, redirect, url_for, session, send_file

KR_HOLIDAYS = holidays.KR(years=range(2000, 2030), language="ko")
# 근로자의날 (5월 1일), 창립기념일 (7월 1일) — holidays.KR 미포함이므로 수동 추가
from datetime import date as _date
for _y in range(2000, 2030):
    KR_HOLIDAYS[_date(_y, 5, 1)] = "근로자의날"
    KR_HOLIDAYS[_date(_y, 7, 1)] = "창립기념일"
del _date, _y

app = Flask(__name__)
app.secret_key = os.environ["FLASK_SECRET_KEY"]

# 리버스 프록시(Synology DSM) 뒤에서 HTTPS 헤더 신뢰
from werkzeug.middleware.proxy_fix import ProxyFix
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

@app.before_request
def force_https():
    """HTTP로 들어오면 HTTPS로 강제 리다이렉트"""
    if request.headers.get('X-Forwarded-Proto', 'https') == 'http':
        return redirect(request.url.replace('http://', 'https://', 1), code=301)

# ── MES Blueprint ─────────────────────────────────────────
app.config["MARIA"]     = None   # _conn() 정의 후 아래에서 채움
app.config["MES_TOKEN"] = os.environ.get("MES_TOKEN", "")

@app.context_processor
def inject_lp_member():
    """사이드바용: lp_group 멤버/결재자 여부 + 오늘날짜 + admin 여부"""
    from datetime import datetime
    today_str = datetime.today().strftime("%m월 %d일 (%a)")
    day_map = {"Mon":"월","Tue":"화","Wed":"수","Thu":"목","Fri":"금","Sat":"토","Sun":"일"}
    for e, k in day_map.items():
        today_str = today_str.replace(e, k)
    uid = session.get("e_id")
    is_admin = session.get("role") == "admin"
    if not uid:
        return {"is_lp_member": False, "today_date_str": today_str, "is_admin": is_admin}
    try:
        conn = _conn(); cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM lp_group_members WHERE user_id=%s", (uid,))
        c1 = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM lp_group_reviewers WHERE user_id=%s", (uid,))
        c2 = cur.fetchone()[0]
        conn.close()
        return {"is_lp_member": (c1 + c2) > 0, "today_date_str": today_str, "is_admin": is_admin}
    except Exception:
        return {"is_lp_member": False, "today_date_str": today_str, "is_admin": is_admin}

TELEGRAM_TOKEN = os.environ["TELEGRAM_TOKEN"]
TELEGRAM_CHAT_ID = os.environ["TELEGRAM_CHAT_ID"]
FIRE_API_TOKEN = os.environ.get("FIRE_API_TOKEN", "")
SOLAPI_KEY    = os.environ.get("SOLAPI_KEY", "")
SOLAPI_SECRET = os.environ.get("SOLAPI_SECRET", "")
SOLAPI_SENDER = os.environ.get("SOLAPI_SENDER", "")
BASE_URL      = os.environ.get("BASE_URL", "http://192.168.100.11:5050").rstrip("/")
MAIL_HOST = os.environ.get("MAIL_HOST", "smtp.office365.com")
MAIL_PORT = int(os.environ.get("MAIL_PORT", "587"))
MAIL_USER = os.environ["MAIL_USER"]
MAIL_PASS = os.environ["MAIL_PASS"]

MARIA = {
    "host": os.environ.get("DB_HOST", "127.0.0.1"),
    "port": int(os.environ.get("DB_PORT", "3307")),
    "user": os.environ.get("DB_USER", "root"),
    "password": os.environ["DB_PASSWORD"],
    "db": os.environ.get("DB_NAME", "attendance"),
    "charset": "utf8mb4",
}

MODE_IN = "1"
MODE_OUT = "2"

DEPT_MAP = {
    "0001000000000000": "관리부",
    "0002000000000000": "전기사무",
    "0003000000000000": "전자사무",
    "0004000000000000": "전기생산",
    "0005000000000000": "전자생산",
    "0006000000000000": "임원",
    "0007000000000000": "경영기획팀",
}


def _conn():
    return pymysql.connect(**MARIA)

# MARIA 설정을 app.config에도 저장 (Blueprint에서 참조)
app.config["MARIA"] = MARIA


def _hash_pw(pw):
    return hashlib.sha256(pw.encode("utf-8")).hexdigest()


def _sync_leave_from_other(cur, e_id, e_date, value):
    """기타 컬럼의 연차/반차 입력 시 leave_records 자동 연동"""
    # 기존 해당 leave_records 삭제 (기타 컬럼 기반 연동분)
    cur.execute("DELETE FROM leave_records WHERE e_id=%s AND leave_date=%s", (e_id, e_date))
    # 값에 연차/반차 포함 시 새로 기록
    v = str(value).strip() if value else ""
    if "반차" in v:
        cur.execute("INSERT IGNORE INTO leave_records (e_id, leave_date, leave_type) VALUES (%s,%s,%s)",
                    (e_id, e_date, "반차"))
    elif "연차" in v:
        cur.execute("INSERT IGNORE INTO leave_records (e_id, leave_date, leave_type) VALUES (%s,%s,%s)",
                    (e_id, e_date, "연차"))
    # annual_leave.used ?ш퀎일
    year_val = int(e_date[:4])
    cur.execute("SELECT leave_type FROM leave_records WHERE e_id=%s AND leave_date LIKE %s",
                (e_id, f"{year_val}%"))
    total_used = 0
    for (lt,) in cur.fetchall():
        total_used += 0.5 if lt == "반차" else 1
    cur.execute("SELECT name FROM tuser WHERE id=%s", (e_id,))
    row = cur.fetchone()
    e_name = row[0] if row else ""
    cur.execute("""
        INSERT INTO annual_leave (e_id, e_name, year, used)
        VALUES (%s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE used=%s, e_name=%s, updated_at=NOW()
    """, (e_id, e_name, year_val, total_used, total_used, e_name))


def _send_telegram(msg, category=None):
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"

    def _post(chat_id):
        try:
            data = urlencode({"chat_id": chat_id, "text": msg}).encode()
            urlopen(Request(url, data), context=ctx, timeout=10)
        except Exception:
            pass

    try:
        conn2 = _conn(); cur2 = conn2.cursor()
        # 전체 발송
        global_on = True
        if category:
            cur2.execute("SELECT `enabled` FROM telegram_settings WHERE `key`=%s", (category,))
            row = cur2.fetchone()
            if row and not row[0]:
                global_on = False
        if global_on:
            _post(TELEGRAM_CHAT_ID)
        # 개인별 발송
        if category:
            cur2.execute("SELECT tg_chat_id, tg_alerts FROM app_users WHERE tg_chat_id != '' AND tg_chat_id IS NOT NULL")
            for chat_id, tg_alerts in cur2.fetchall():
                alerts = [a.strip() for a in (tg_alerts or "").split(",") if a.strip()]
                if category in alerts:
                    _post(chat_id)
        conn2.close()
    except Exception:
        pass


def _send_email(to_addrs, subject, body_html):
    """Office 365 SMTP 이메일 비동기 발송"""
    if not to_addrs or not MAIL_USER or not MAIL_PASS:
        return
    def _do():
        try:
            import smtplib
            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText
            msg = MIMEMultipart("alternative")
            msg["Subject"] = subject
            msg["From"] = f"근태관리시스템 <{MAIL_USER}>"
            msg["To"] = ", ".join(to_addrs)
            msg.attach(MIMEText(body_html, "html", "utf-8"))
            with smtplib.SMTP(MAIL_HOST, MAIL_PORT, timeout=15) as s:
                s.ehlo(); s.starttls(); s.ehlo()
                s.login(MAIL_USER, MAIL_PASS)
                s.sendmail(MAIL_USER, to_addrs, msg.as_string())
        except Exception:
            pass
    threading.Thread(target=_do, daemon=True).start()


_last_checked_sync_id = 0
_last_tg_update_id = 0

def _telegram_bot_polling():
    """백그라운드: 텔레그램 봇 메시지 폴링 — /start 사번 처리"""
    global _last_tg_update_id
    import time, json
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    base = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}"
    while True:
        try:
            time.sleep(3)
            url = f"{base}/getUpdates?offset={_last_tg_update_id + 1}&timeout=10"
            resp = urlopen(Request(url), context=ctx, timeout=15)
            data = json.loads(resp.read().decode())
            if not data.get("ok"):
                continue
            for upd in data.get("result", []):
                _last_tg_update_id = upd["update_id"]
                msg = upd.get("message")
                if not msg or not msg.get("text"):
                    continue
                text = msg["text"].strip()
                chat_id = str(msg["chat"]["id"])
                user_first = msg["from"].get("first_name", "")
                # /start 사번 처리
                if text.startswith("/start"):
                    parts = text.split()
                    if len(parts) < 2:
                        _tg_reply(chat_id, "사용법: /start 사번\n예) /start 12345")
                        continue
                    emp_id = parts[1].strip()
                    try:
                        conn2 = _conn(); cur2 = conn2.cursor()
                        # app_users에서 login_id로 먼저 찾기
                        cur2.execute("SELECT id, login_id, name, e_id FROM app_users WHERE login_id=%s", (emp_id,))
                        arow = cur2.fetchone()
                        if arow:
                            cur2.execute("UPDATE app_users SET tg_chat_id=%s WHERE id=%s", (chat_id, arow[0]))
                            conn2.commit(); conn2.close()
                            _tg_reply(chat_id, f"✅ 등록 완료!\n👤 {arow[2]} (사번: {emp_id})\n이제 알림을 받으실 수 있습니다.")
                            continue
                        # tuser.id로 찾기
                        cur2.execute("SELECT id, name FROM tuser WHERE id=%s", (emp_id,))
                        trow = cur2.fetchone()
                        if not trow:
                            _tg_reply(chat_id, f"❌ 사번 {emp_id}을(를) 찾을 수 없습니다.")
                            conn2.close()
                            continue
                        e_name = trow[1]
                        cur2.execute("SELECT id, login_id, name FROM app_users WHERE e_id=%s", (int(emp_id),))
                        arow = cur2.fetchone()
                        if not arow:
                            _tg_reply(chat_id, f"❌ 사번 {emp_id}({e_name})의 앱 계정이 없습니다.\n관리자에게 문의하세요.")
                            conn2.close()
                            continue
                        cur2.execute("UPDATE app_users SET tg_chat_id=%s WHERE id=%s", (chat_id, arow[0]))
                        conn2.commit(); conn2.close()
                        _tg_reply(chat_id, f"✅ 등록 완료!\n👤 {e_name} (사번: {emp_id})\n이제 알림을 받으실 수 있습니다.")
                    except Exception as ex:
                        _tg_reply(chat_id, f"⚠️ 오류가 발생했습니다: {str(ex)[:100]}")
                elif text == "/help":
                    _tg_reply(chat_id, "📋 사용 가능한 명령어:\n/start 사번 — 알림 등록\n/stop — 알림 해제\n/help — 도움말")
                elif text == "/stop":
                    try:
                        conn2 = _conn(); cur2 = conn2.cursor()
                        cur2.execute("UPDATE app_users SET tg_chat_id='' WHERE tg_chat_id=%s", (chat_id,))
                        conn2.commit(); conn2.close()
                        _tg_reply(chat_id, "🔕 알림이 해제되었습니다.")
                    except Exception:
                        _tg_reply(chat_id, "⚠️ 오류가 발생했습니다.")
        except Exception:
            import time; time.sleep(5)

def _tg_reply(chat_id, text):
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    try:
        data = urlencode({"chat_id": chat_id, "text": text}).encode()
        urlopen(Request(url, data), context=ctx, timeout=10)
    except Exception:
        pass

# 근무보고서 결재지연 알림: 발송한 날짜 기록 (하루 1회만)
_wr_overdue_alerted_date = None
# 결재지연 텔레그램/이메일 알림 시행 시작일 — 이 날짜 이후 보고서만 알림
# 7월분부터 발송 (그 이전 과거 밀린 건은 알림 제외)
WR_OVERDUE_START = "20260701"
# 화면 패널(관리자용) 표시 기준일 — 이 날짜 이전 보고서는 패널에서 제외 (과거 밀린 건 정리)
WR_OVERDUE_PANEL_START = "20260701"

def _wr_overdue_checker():
    """백그라운드: 매일 09:00에 결재 지연 근무보고서를 태인 알림방으로 통보.
    기준 = 보고일이 1일 이상 지났는데 status가 '결재'(완료)가 아닌 건."""
    global _wr_overdue_alerted_date
    while True:
        try:
            import time; time.sleep(300)  # 5분마다 시각 확인
            now = datetime.now()
            today = now.strftime("%Y-%m-%d")
            # 09시대에 하루 한 번만
            if now.hour != 9 or _wr_overdue_alerted_date == today:
                continue
            # 텔레그램 설정 off면 스킵
            conn = _conn(); cur = conn.cursor()
            cur.execute("SELECT `enabled` FROM telegram_settings WHERE `key`='wr_overdue'")
            row = cur.fetchone()
            if row and not row[0]:
                _wr_overdue_alerted_date = today
                conn.close()
                continue
            # 보고일 <= 어제 (=1일 이상 경과) & 실제 결재 대기('작성완료','검토')만
            # '대기'(작성자 미제출) 제외 / 기능 시행일 이전 과거분 제외 (알림 폭탄 방지)
            cutoff = (now.date() - timedelta(days=1)).strftime("%Y%m%d")
            cur.execute("""
                SELECT r.report_date, r.status, r.created_by_name, g.group_name, r.group_id
                FROM wr_reports r JOIN wr_groups g ON r.group_id = g.id
                WHERE r.report_date <= %s AND r.report_date >= %s
                  AND r.status IN ('작성완료', '검토')
                ORDER BY r.report_date DESC, g.group_name
            """, (cutoff, WR_OVERDUE_START))
            rows = cur.fetchall()
            if not rows:
                conn.close()
                _wr_overdue_alerted_date = today
                continue

            # ── 담당자별 이메일 취합 (현재 결재 차례인 사람에게만) ──
            # 상태 → 다음 결재 단계: 대기=1(작성자), 작성완료=2(검토), 검토=3(2차검토)
            step_map = {"대기": 1, "작성완료": 2, "검토": 3}
            # {이메일: {"name": 이름, "items": [지연내역 문자열]}}
            mail_map = {}
            lines = [f"⚠️ 근무보고서 결재 지연 {len(rows)}건", ""]
            for rd, st, writer, gname, gid in rows[:20]:
                rd = str(rd)
                disp = f"{rd[:4]}-{rd[4:6]}-{rd[6:8]}" if len(rd) == 8 else rd
                try:
                    late = (now.date() - datetime.strptime(rd, "%Y%m%d").date()).days
                    late_txt = f"{late}일 경과"
                except ValueError:
                    late_txt = ""
                lines.append(f"· {disp} ({late_txt}) {gname} / {writer or '-'} · 현재: {st}")
                # 이 건의 현재 담당자 찾기
                nstep = step_map.get(st)
                if not nstep:
                    continue
                try:
                    cur.execute(
                        "SELECT user_name FROM wr_group_reviewers WHERE group_id=%s AND step=%s",
                        (gid, nstep))
                    cand = [r[0] for r in cur.fetchall()]
                    # 1단계(작성)는 결재자 지정이 없을 수 있어 작성자에게
                    if not cand and nstep == 1 and writer:
                        cand = [writer]
                    if cand:
                        ph = ",".join(["%s"] * len(cand))
                        cur.execute(
                            f"SELECT name, email FROM employee_roster "
                            f"WHERE name IN ({ph}) AND email IS NOT NULL AND email <> ''", cand)
                        for nm, em in cur.fetchall():
                            mail_map.setdefault(em, {"name": nm, "items": []})["items"].append(
                                f"{disp} ({late_txt}) · {gname} · 현재: {st}")
                except Exception:
                    pass
            if len(rows) > 20:
                lines.append(f"... 외 {len(rows) - 20}건")
            lines += ["", "👉 https://app.taein.biz/work_report"]
            conn.close()
            _wr_overdue_alerted_date = today

            # 태인 알림방 (전체 요약)
            _send_telegram("\n".join(lines))

            # 담당자별 개별 이메일
            for em, info in mail_map.items():
                items_html = "".join(f"<li>{it}</li>" for it in info["items"])
                body = (
                    f"<p><b>{info['name']}</b>님, 결재 대기 중인 근무보고서가 있습니다.</p>"
                    f"<ul>{items_html}</ul>"
                    f"<p>근태관리시스템에 로그인하여 처리해 주세요.<br>"
                    f"<a href='https://app.taein.biz/work_report'>https://app.taein.biz/work_report</a></p>"
                )
                _send_email([em], f"[근태관리시스템] 근무보고서 결재 지연 {len(info['items'])}건", body)
        except Exception:
            import time; time.sleep(60)


# ── 근무보고서 '미작성일' 감지 ─────────────────────────────
_wr_missing_alerted_date = None
WR_MISSING_START = "20260701"   # 이 날짜 이후만 점검 (과거분 메일 폭탄 방지)
WR_MISSING_LOOKBACK = 30        # 최근 N일 점검
WR_MISSING_MAX_LINES = 12       # 메일에 나열할 최대 건수


def _wr_missing_scan(cur, now=None):
    """'보고서가 아예 작성되지 않은 날'을 찾는다.

    기존 결재지연 알림(_wr_overdue_checker)은 '만들어진 보고서'의 지연만 잡으므로
    보고서 자체가 없는 날은 감지되지 않았음 (2026-07 QA 7/9~7/12 누락 사례).

    판정: 그룹 구성원 중 1명이라도 그날 출입기록(tenter)이 있는데 보고서가 없으면 미작성.
          → 주말·공휴일·부서별 비근무일은 자동으로 제외된다.

    반환: (start, end, per_writer, per_group, total)
          per_writer = {작성자명: [(그룹, 'YYYY-MM-DD', 출근인원), ...]}
    """
    now = now or datetime.now()
    start = max(WR_MISSING_START,
                (now.date() - timedelta(days=WR_MISSING_LOOKBACK)).strftime("%Y%m%d"))
    end = (now.date() - timedelta(days=1)).strftime("%Y%m%d")   # 어제까지

    # alert_missing=0 인 그룹은 제외 (보고서를 쓰지 않는 그룹)
    try:
        cur.execute("""SELECT id, group_name FROM wr_groups
                       WHERE status='완료' AND alert_missing=1 ORDER BY id""")
        groups = cur.fetchall()
    except Exception:
        # 컬럼이 아직 없는 경우 (구버전 DB) — 전체 그룹 대상
        cur.execute("SELECT id, group_name FROM wr_groups WHERE status='완료' ORDER BY id")
        groups = cur.fetchall()

    per_writer, per_group, total = {}, [], 0
    for gid, gname in groups:
        cur.execute("""SELECT DISTINCT report_date FROM wr_reports
                       WHERE group_id=%s AND report_date BETWEEN %s AND %s""",
                    (gid, start, end))
        have = {r[0] for r in cur.fetchall()}
        # 구성원이 실제 출근한 날 + 출근 인원수
        cur.execute("""SELECT t.e_date, COUNT(DISTINCT t.e_name)
                       FROM tenter t
                       WHERE t.e_date BETWEEN %s AND %s AND t.e_id >= 0
                         AND t.e_name IN (SELECT user_name FROM wr_group_members
                                          WHERE group_id=%s)
                       GROUP BY t.e_date""", (start, end, gid))
        missing = sorted((d, c) for d, c in cur.fetchall() if d not in have)
        if not missing:
            continue
        # 그룹 작성자 (지정 없으면 최근 작성자로 대체)
        cur.execute("SELECT user_name FROM wr_group_writers WHERE group_id=%s", (gid,))
        writers = [r[0] for r in cur.fetchall()]
        if not writers:
            cur.execute("""SELECT created_by_name FROM wr_reports
                           WHERE group_id=%s AND created_by_name IS NOT NULL
                             AND created_by_name <> ''
                           ORDER BY id DESC LIMIT 1""", (gid,))
            r2 = cur.fetchone()
            writers = [r2[0]] if r2 else []
        per_group.append({"group": gname, "writers": writers,
                          "dates": [f"{d[:4]}-{d[4:6]}-{d[6:8]}" for d, _ in missing]})
        for d, c in missing:
            total += 1
            disp = f"{d[:4]}-{d[4:6]}-{d[6:8]}"
            for w in (writers or ["(작성자 미지정)"]):
                per_writer.setdefault(w, []).append((gname, disp, int(c or 0)))
    return start, end, per_writer, per_group, total


def _wr_missing_emails(cur, per_writer):
    names = [w for w in per_writer if w != "(작성자 미지정)"]
    if not names:
        return {}
    ph = ",".join(["%s"] * len(names))
    cur.execute(f"""SELECT name, email FROM employee_roster
                    WHERE name IN ({ph}) AND email IS NOT NULL AND email <> ''""", names)
    return dict(cur.fetchall())


def _wr_missing_send(per_writer, emails, start, end):
    """작성자별 미작성 안내 메일 발송 → 발송한 이메일 목록"""
    sent = []
    for w, items in per_writer.items():
        em = emails.get(w)
        if not em:
            continue
        shown = items[:WR_MISSING_MAX_LINES]
        rows_html = "".join(
            f"<tr><td style='padding:4px 10px;border-bottom:1px solid #eee;'>{g}</td>"
            f"<td style='padding:4px 10px;border-bottom:1px solid #eee;'>{d}</td>"
            f"<td style='padding:4px 10px;border-bottom:1px solid #eee;text-align:right;'>{c}명</td></tr>"
            for g, d, c in shown)
        more = (f"<p style='color:#888;'>… 외 {len(items) - len(shown)}건</p>"
                if len(items) > len(shown) else "")
        body = (
            f"<p><b>{w}</b>님, 근무보고서가 <b>작성되지 않은 날</b>이 있습니다.</p>"
            f"<p style='color:#666;font-size:0.9em;'>해당 날짜에 출근 기록이 있는데 "
            f"보고서가 없는 건입니다. (점검 기간 {start[:4]}-{start[4:6]}-{start[6:8]} ~ "
            f"{end[:4]}-{end[4:6]}-{end[6:8]})</p>"
            f"<table style='border-collapse:collapse;font-size:0.92em;'>"
            f"<tr><th style='padding:4px 10px;text-align:left;border-bottom:2px solid #ddd;'>그룹</th>"
            f"<th style='padding:4px 10px;text-align:left;border-bottom:2px solid #ddd;'>날짜</th>"
            f"<th style='padding:4px 10px;text-align:right;border-bottom:2px solid #ddd;'>출근</th></tr>"
            f"{rows_html}</table>{more}"
            f"<p>근태관리시스템에서 소급 작성해 주세요.<br>"
            f"<a href='https://app.taein.biz/work_report'>https://app.taein.biz/work_report</a></p>"
        )
        _send_email([em], f"[근태관리시스템] 근무보고서 미작성 {len(items)}건", body)
        sent.append(em)
    return sent


def _wr_missing_checker():
    """백그라운드: 매일 09시에 미작성일을 찾아 작성자에게 메일 발송"""
    global _wr_missing_alerted_date
    while True:
        try:
            import time; time.sleep(300)   # 5분마다 시각 확인
            now = datetime.now()
            today = now.strftime("%Y-%m-%d")
            if now.hour != 9 or _wr_missing_alerted_date == today:
                continue

            conn = _conn(); cur = conn.cursor()
            cur.execute("SELECT `enabled` FROM telegram_settings WHERE `key`='wr_missing'")
            row = cur.fetchone()
            if row and not row[0]:
                _wr_missing_alerted_date = today
                conn.close()
                continue

            start, end, per_writer, _pg, total = _wr_missing_scan(cur, now)
            emails = _wr_missing_emails(cur, per_writer) if total else {}
            conn.close()
            _wr_missing_alerted_date = today
            if total:
                _wr_missing_send(per_writer, emails, start, end)
        except Exception:
            import time; time.sleep(60)


# 출퇴근 싱크 알림 플래그 {날짜: {"morning_alerted": bool, "evening_alerted": bool}}
_sync_check_alerts = {}

def _sync_log_checker():
    """백그라운드: sync_log 테이블을 주기적으로 확인하여 텔레그램 알림 발송"""
    global _last_checked_sync_id
    while True:
        try:
            import time; time.sleep(60)
            conn = _conn(); cur = conn.cursor()
            # 최초 실행 시 마지막 ID 세팅
            if _last_checked_sync_id == 0:
                cur.execute("SELECT IFNULL(MAX(id),0) FROM sync_log")
                _last_checked_sync_id = cur.fetchone()[0]
                conn.close()
                continue
            # 새 로그 조회
            cur.execute("SELECT id, level, message FROM sync_log WHERE id > %s ORDER BY id",
                        (_last_checked_sync_id,))
            rows = cur.fetchall()
            for sid, level, message in rows:
                _last_checked_sync_id = sid
                msg = message or ""
                if level == "ERROR":
                    # ODBC 일시적 잠금 에러(-1905)는 무시
                    if "-1905" in msg or "SQLDriverConnect" in msg:
                        continue
                    _send_telegram(f"⚠️ 싱크 에러 발생!\n❌{msg[:200]}", "sync_error")

            # ── 출퇴근 싱크 건수 체크 ──
            _check_attendance_sync(cur)
            conn.close()
        except Exception:
            pass


def _is_alert_primary():
    """싱크 알림 발송 주체인지 판정.
    .5 대기(standby) 서버도 같은 코드를 돌리며 텔레그램 토큰을 갖고 있어
    중복/오탐 알림이 나간 사례가 있음 (2026-07-27 '퇴근 기록 0건' 오탐).
    운영 기준 서버(.11 = 마스터 DB 직결)에서만 알림을 보낸다."""
    db_host = (os.environ.get("DB_HOST") or "").strip()
    return db_host.startswith("192.168.100.11") or db_host in ("127.0.0.1", "localhost")


def _check_attendance_sync(cur):
    """출퇴근 싱크 데이터 건수 체크 → 금일 근태기록이 '0건'일 때만 텔레그램 알림"""
    from datetime import date, datetime

    # 대기서버에서는 알림 발송 안 함 (운영 서버만 발송)
    if not _is_alert_primary():
        return

    today = date.today()
    today_str = today.strftime("%Y%m%d")
    now = datetime.now()

    # 주말/공휴일은 체크 안 함
    if today.weekday() >= 5 or today in KR_HOLIDAYS:
        return
    # 회사 자체 휴무일(매년 7/1 창립기념일)도 체크 안 함
    if today.month == 7 and today.day == 1:
        return

    day_flags = _sync_check_alerts.setdefault(today_str, {
        "early_alerted": False, "morning_alerted": False, "evening_alerted": False
    })

    # ── 판정 기준: '근태기록' 화면의 금일 조회 건수 ──
    # ⚠️ 절대 원칙: 금일 근태기록이 1건이라도 조회되면 알림을 보내지 않는다.
    #   화면에 데이터가 보이는데 '싱크 이상' 알림이 가는 오탐은 금지.
    #   (2026-07-23 퇴근 133건인데 '퇴근 0건' 알림,
    #    2026-07-27 출근 100명인데 '퇴근 기록 0건(기준 5건)' 알림 — 모두 오탐 사례)
    #   → 건수 임계값(N건 미만) 방식 금지. 오직 '0건'일 때만 발송.
    def _today_record_cnt():
        cur.execute(
            "SELECT COUNT(*) FROM tenter WHERE e_date=%s AND e_id>=0",
            (today_str,)
        )
        return cur.fetchone()[0]

    # 데이터가 하나라도 있으면 어떤 알림도 보내지 않고 종료
    if _today_record_cnt() > 0:
        return

    # ── 조기 체크: 07:30 이후, 금일 근태기록 0건 ──
    if not day_flags.get("early_alerted") and (now.hour > 7 or (now.hour == 7 and now.minute >= 30)):
        _send_telegram(
            f"🚨 출근 데이터 이상 (조기 경보)!\n"
            f"📅 {today_str}\n"
            f"⏰ 07:30 기준 금일 근태기록: 0건\n"
            f"➡️ CAPS 서버/출입단말기 상태를 확인하세요.",
            "sync_check"
        )
        day_flags["early_alerted"] = True

    # ── 출근 체크: 09:00 이후, 금일 근태기록 0건 ──
    if not day_flags["morning_alerted"] and now.hour >= 9:
        _send_telegram(
            f"🚨 출근 싱크 이상!\n"
            f"📅 {today_str}\n"
            f"⏰ 금일 근태기록: 0건\n"
            f"➡️ CAPS 싱크 프로그램 상태를 확인하세요.",
            "sync_check"
        )
        day_flags["morning_alerted"] = True

    # ── 퇴근 체크: 19:00 이후, 금일 근태기록 0건 ──
    if not day_flags["evening_alerted"] and now.hour >= 19:
        _send_telegram(
            f"🚨 퇴근 싱크 이상!\n"
            f"📅 {today_str}\n"
            f"🏠 금일 근태기록: 0건\n"
            f"➡️ CAPS 싱크 프로그램 상태를 확인하세요.",
            "sync_check"
        )
        day_flags["evening_alerted"] = True

    # 오래된 날짜 플래그 정리 (3일 이전)
    old_keys = [k for k in _sync_check_alerts if k < (today - __import__('datetime').timedelta(days=3)).strftime("%Y%m%d")]
    for k in old_keys:
        del _sync_check_alerts[k]


# ── Docker 코드 싱크 체크 (1시간 주기) ──
_DOCKER_CHECK_FILES = [
    "/app/app_maria.py",
    "/app/tbm_app.py",
]

def _docker_sync_checker():
    """1시간마다 주요 파일 MD5를 DB에 기록 (컨테이너별)"""
    import time
    container_id = socket.gethostname()[:12]
    # 첫 체크는 30초 후
    time.sleep(30)
    while True:
        try:
            conn = _conn(); cur = conn.cursor()
            for fpath in _DOCKER_CHECK_FILES:
                try:
                    md5 = hashlib.md5(open(fpath, "rb").read()).hexdigest()
                except Exception:
                    md5 = "NOT_FOUND"
                cur.execute("""
                    INSERT INTO docker_code_hash (container_id, file_name, file_md5, checked_at)
                    VALUES (%s, %s, %s, NOW())
                    ON DUPLICATE KEY UPDATE file_md5=%s, checked_at=NOW()
                """, (container_id, fpath, md5, md5))
            conn.commit(); conn.close()
        except Exception:
            pass
        time.sleep(3600)


def _init_db():
    try:
        conn = _conn()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `work_override` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `e_id` INT NOT NULL,
                `e_date` VARCHAR(8) NOT NULL,
                `col_type` VARCHAR(10) NOT NULL,
                `value` VARCHAR(20) NOT NULL,
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_override` (`e_id`, `e_date`, `col_type`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # 기존 DECIMAL 컬럼이면 VARCHAR로 변경
        cur.execute("""
            SELECT DATA_TYPE FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA=DATABASE() AND TABLE_NAME='work_override' AND COLUMN_NAME='value'
        """)
        row = cur.fetchone()
        if row and row[0].upper() != 'VARCHAR':
            cur.execute("ALTER TABLE work_override MODIFY `value` VARCHAR(20) NOT NULL")
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `annual_leave` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `e_id` INT NOT NULL,
                `e_name` VARCHAR(30) NOT NULL DEFAULT '',
                `year` SMALLINT NOT NULL,
                `total` FLOAT NOT NULL DEFAULT 15,
                `used` FLOAT NOT NULL DEFAULT 0,
                `memo` VARCHAR(200) NOT NULL DEFAULT '',
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_annual` (`e_id`, `year`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `app_users` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `login_id` VARCHAR(30) NOT NULL,
                `password` VARCHAR(64) NOT NULL,
                `name` VARCHAR(30) NOT NULL DEFAULT '',
                `e_id` INT DEFAULT NULL,
                `role` VARCHAR(10) NOT NULL DEFAULT 'user',
                `permissions` VARCHAR(100) NOT NULL DEFAULT '',
                `perm_group_id` INT DEFAULT NULL,
                `must_change_pw` TINYINT NOT NULL DEFAULT 1,
                `tg_chat_id` VARCHAR(50) NOT NULL DEFAULT '',
                `tg_alerts` VARCHAR(100) NOT NULL DEFAULT '',
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_login` (`login_id`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # 기존 DB에 perm_group_id 컬럼이 없으면 추가
        try:
            cur.execute("SELECT perm_group_id FROM app_users LIMIT 1")
        except Exception:
            try:
                cur.execute("ALTER TABLE app_users ADD COLUMN perm_group_id INT DEFAULT NULL")
            except Exception:
                pass
        # 권한 그룹
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `perm_groups` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `name` VARCHAR(50) NOT NULL,
                `description` VARCHAR(200) NOT NULL DEFAULT '',
                `is_system` TINYINT NOT NULL DEFAULT 0,
                `sort_order` INT NOT NULL DEFAULT 0,
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_pg_name` (`name`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # 그룹별 메뉴 권한 매트릭스
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `perm_group_permissions` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `group_id` INT NOT NULL,
                `menu_key` VARCHAR(50) NOT NULL,
                `can_view` TINYINT NOT NULL DEFAULT 0,
                `can_create` TINYINT NOT NULL DEFAULT 0,
                `can_update` TINYINT NOT NULL DEFAULT 0,
                `can_delete` TINYINT NOT NULL DEFAULT 0,
                UNIQUE KEY `uq_gp` (`group_id`, `menu_key`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # 사용자 개별 오버라이드 (그룹 권한 + allow / - deny)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `user_perm_overrides` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `user_id` INT NOT NULL,
                `menu_key` VARCHAR(50) NOT NULL,
                `action` VARCHAR(10) NOT NULL,
                `grant_type` VARCHAR(10) NOT NULL DEFAULT 'allow',
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_upo` (`user_id`, `menu_key`, `action`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # ══════════════ 설문(Survey) 테이블 ══════════════
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `surveys` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `title` VARCHAR(200) NOT NULL,
                `description` TEXT,
                `start_at` DATETIME DEFAULT NULL,
                `end_at` DATETIME DEFAULT NULL,
                `status` VARCHAR(20) NOT NULL DEFAULT 'draft',
                `anonymous` TINYINT(1) NOT NULL DEFAULT 0,
                `allow_edit` TINYINT(1) NOT NULL DEFAULT 0,
                `one_per_user` TINYINT(1) NOT NULL DEFAULT 1,
                `remind_hours` INT NOT NULL DEFAULT 24,
                `send_email` TINYINT(1) NOT NULL DEFAULT 0,
                `send_sms` TINYINT(1) NOT NULL DEFAULT 0,
                `is_template` TINYINT(1) NOT NULL DEFAULT 0,
                `template_src` INT DEFAULT NULL,
                `created_by` INT NOT NULL,
                `created_by_name` VARCHAR(30) NOT NULL DEFAULT '',
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                INDEX `idx_sv_status` (`status`),
                INDEX `idx_sv_dates` (`start_at`, `end_at`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # 기존 DB에 send_email/send_sms 컬럼 없으면 추가
        for col in ("send_email", "send_sms"):
            try:
                cur.execute(f"SELECT {col} FROM surveys LIMIT 1")
            except Exception:
                try: cur.execute(f"ALTER TABLE surveys ADD COLUMN {col} TINYINT(1) NOT NULL DEFAULT 0")
                except Exception: pass
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `survey_sections` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `survey_id` INT NOT NULL,
                `title` VARCHAR(200) DEFAULT '',
                `description` TEXT,
                `sort_order` INT NOT NULL DEFAULT 0,
                INDEX `idx_svsec` (`survey_id`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `survey_questions` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `survey_id` INT NOT NULL,
                `section_id` INT DEFAULT NULL,
                `qtype` VARCHAR(20) NOT NULL,
                `title` VARCHAR(500) NOT NULL,
                `description` TEXT,
                `required` TINYINT(1) NOT NULL DEFAULT 0,
                `sort_order` INT NOT NULL DEFAULT 0,
                `branch_src_qid` INT DEFAULT NULL,
                `branch_src_value` VARCHAR(500) DEFAULT NULL,
                `meta` TEXT,
                INDEX `idx_svq_order` (`survey_id`, `sort_order`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `survey_question_options` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `question_id` INT NOT NULL,
                `label` VARCHAR(300) NOT NULL,
                `sort_order` INT NOT NULL DEFAULT 0,
                INDEX `idx_svqo` (`question_id`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `survey_targets` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `survey_id` INT NOT NULL,
                `target_type` VARCHAR(20) NOT NULL,
                `target_value` VARCHAR(100) DEFAULT NULL,
                INDEX `idx_svt` (`survey_id`),
                INDEX `idx_svt_tv` (`target_type`, `target_value`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `survey_responses` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `survey_id` INT NOT NULL,
                `respondent_id` INT DEFAULT NULL,
                `respondent_name` VARCHAR(30) DEFAULT NULL,
                `respondent_dept` VARCHAR(50) DEFAULT NULL,
                `respondent_position` VARCHAR(50) DEFAULT NULL,
                `submitted_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                `is_complete` TINYINT(1) NOT NULL DEFAULT 1,
                UNIQUE KEY `uq_svr_one` (`survey_id`, `respondent_id`),
                INDEX `idx_svr` (`survey_id`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `survey_answers` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `response_id` INT NOT NULL,
                `question_id` INT NOT NULL,
                `option_id` INT DEFAULT NULL,
                `value_text` TEXT,
                `value_number` DECIMAL(10,2) DEFAULT NULL,
                `file_path` VARCHAR(500) DEFAULT NULL,
                INDEX `idx_sva_resp` (`response_id`),
                INDEX `idx_sva_q` (`question_id`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `survey_notifications` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `survey_id` INT NOT NULL,
                `user_id` INT NOT NULL,
                `channel` VARCHAR(20) NOT NULL DEFAULT 'internal',
                `kind` VARCHAR(20) NOT NULL DEFAULT 'initial',
                `sent_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX `idx_svn` (`survey_id`, `user_id`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `leave_records` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `e_id` INT NOT NULL,
                `leave_date` VARCHAR(8) NOT NULL,
                `leave_type` VARCHAR(10) NOT NULL DEFAULT '연차',
                `status` VARCHAR(10) NOT NULL DEFAULT '승인',
                `memo` VARCHAR(100) NOT NULL DEFAULT '',
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_leave` (`e_id`, `leave_date`, `leave_type`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # 기존 테이블에 status 컬럼 추가 (없으면)
        try:
            cur.execute("ALTER TABLE leave_records ADD COLUMN `status` VARCHAR(10) NOT NULL DEFAULT '승인' AFTER `leave_type`")
        except Exception:
            pass
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `leave_approvers` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `dept` VARCHAR(20) NOT NULL DEFAULT '',
                `step` TINYINT NOT NULL,
                `user_id` INT NOT NULL,
                `user_name` VARCHAR(30) NOT NULL DEFAULT '',
                UNIQUE KEY `uq_dept_step_user` (`dept`, `step`, `user_id`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # 기존 테이블에 dept 컬럼 추가
        try:
            cur.execute("ALTER TABLE leave_approvers ADD COLUMN `dept` VARCHAR(20) NOT NULL DEFAULT '' FIRST")
            cur.execute("ALTER TABLE leave_approvers DROP INDEX `uq_step_user`")
            cur.execute("ALTER TABLE leave_approvers ADD UNIQUE KEY `uq_dept_step_user` (`dept`, `step`, `user_id`)")
        except Exception:
            pass
        # ── 그룹 기반 연차 검토 테이블 ──
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `leave_groups` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `group_name` VARCHAR(50) NOT NULL,
                `final_step` TINYINT NOT NULL DEFAULT 3,
                `status` VARCHAR(10) NOT NULL DEFAULT '설정중',
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_group_name` (`group_name`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        try:
            cur.execute("ALTER TABLE leave_groups ADD COLUMN `status` VARCHAR(10) NOT NULL DEFAULT '설정중' AFTER `final_step`")
        except Exception:
            pass
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `leave_group_reviewers` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `group_id` INT NOT NULL,
                `step` TINYINT NOT NULL,
                `user_id` INT NOT NULL,
                `user_name` VARCHAR(30) NOT NULL DEFAULT '',
                UNIQUE KEY `uq_gr_step_user` (`group_id`, `step`, `user_id`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `leave_group_members` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `group_id` INT NOT NULL,
                `user_id` INT NOT NULL,
                `user_name` VARCHAR(30) NOT NULL DEFAULT '',
                UNIQUE KEY `uq_gm_user` (`group_id`, `user_id`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `meal_count` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `year_month` VARCHAR(7) NOT NULL,
                `dept` VARCHAR(10) NOT NULL,
                `meal_type` VARCHAR(10) NOT NULL,
                `day` INT NOT NULL,
                `count` VARCHAR(20) NOT NULL DEFAULT '0',
                `memo` VARCHAR(200) NOT NULL DEFAULT '',
                `writer` VARCHAR(20) NOT NULL DEFAULT '',
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_meal` (`year_month`, `dept`, `meal_type`, `day`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # memo 컬럼 없으면 추가
        try:
            cur.execute("ALTER TABLE meal_count ADD COLUMN `memo` VARCHAR(200) NOT NULL DEFAULT '' AFTER `count`")
        except Exception:
            pass
        try:
            cur.execute("ALTER TABLE meal_count ADD COLUMN `writer` VARCHAR(20) NOT NULL DEFAULT '' AFTER `memo`")
        except Exception:
            pass
        try:
            cur.execute("ALTER TABLE meal_notice ADD COLUMN `writer` VARCHAR(20) NOT NULL DEFAULT '' AFTER `content`")
        except Exception:
            pass
        try:
            cur.execute("ALTER TABLE app_users ADD COLUMN `permissions` VARCHAR(100) NOT NULL DEFAULT '' AFTER `role`")
        except Exception:
            pass
        try:
            cur.execute("ALTER TABLE app_users ADD COLUMN `tg_chat_id` VARCHAR(50) NOT NULL DEFAULT '' AFTER `must_change_pw`")
        except Exception:
            pass
        try:
            cur.execute("ALTER TABLE app_users ADD COLUMN `tg_alerts` VARCHAR(100) NOT NULL DEFAULT '' AFTER `tg_chat_id`")
        except Exception:
            pass
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `meal_notice` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `year_month` VARCHAR(7) NOT NULL,
                `content` TEXT NOT NULL,
                `writer` VARCHAR(20) NOT NULL DEFAULT '',
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # leave_records에 reason 컬럼 추가
        try:
            cur.execute("ALTER TABLE leave_records ADD COLUMN `reason` VARCHAR(200) NOT NULL DEFAULT '' AFTER `memo`")
        except Exception:
            pass
        # 웹 알림 테이블
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `notifications` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `user_id` INT NOT NULL,
                `message` VARCHAR(300) NOT NULL,
                `link` VARCHAR(200) NOT NULL DEFAULT '',
                `is_read` TINYINT NOT NULL DEFAULT 0,
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX `idx_noti_user` (`user_id`, `is_read`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # 텔레그램 발송 설정 테이블
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `telegram_settings` (
                `key` VARCHAR(30) PRIMARY KEY,
                `enabled` TINYINT NOT NULL DEFAULT 1,
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        for k in ('sync_error', 'sync_check', 'leave', 'meal', 'notice',
                  'wr_overdue', 'wr_missing'):
            cur.execute("INSERT IGNORE INTO telegram_settings (`key`, `enabled`) VALUES (%s, 1)", (k,))

        # 근무보고서 미작성 알림 대상 여부 (0이면 해당 그룹은 알림 제외)
        # 보고서를 쓰지 않는 그룹(전기생기·전기품질 등)을 그룹 상태 변경 없이 제외하기 위함
        try:
            cur.execute("""SELECT COUNT(*) FROM information_schema.COLUMNS
                           WHERE TABLE_SCHEMA=DATABASE() AND TABLE_NAME='wr_groups'
                             AND COLUMN_NAME='alert_missing'""")
            if not cur.fetchone()[0]:
                cur.execute("ALTER TABLE wr_groups ADD COLUMN `alert_missing` "
                            "TINYINT NOT NULL DEFAULT 1")
        except Exception as _e:
            print(f"[init] wr_groups.alert_missing 추가 실패 (무시): {_e}")
        # 사원명부 테이블
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `employee_roster` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `name` VARCHAR(30) NOT NULL,
                `dept` VARCHAR(30) NOT NULL DEFAULT '',
                `emp_no` VARCHAR(20) NOT NULL DEFAULT '',
                `gender` VARCHAR(10) NOT NULL DEFAULT '',
                `birth_date` VARCHAR(20) NOT NULL DEFAULT '',
                `calendar_type` VARCHAR(10) NOT NULL DEFAULT '',
                `age` INT NOT NULL DEFAULT 0,
                `appoint_date` VARCHAR(20) NOT NULL DEFAULT '',
                `position` VARCHAR(20) NOT NULL DEFAULT '',
                `phone` VARCHAR(30) NOT NULL DEFAULT '',
                `email` VARCHAR(100) NOT NULL DEFAULT '',
                `address` VARCHAR(300) NOT NULL DEFAULT '',
                `hire_date` VARCHAR(20) NOT NULL DEFAULT '',
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_empno` (`emp_no`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # 근무기록 테이블
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `schedule_record` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `emp_name` VARCHAR(50) NOT NULL,
                `work_date` VARCHAR(8) NOT NULL,
                `basic_h` INT NOT NULL DEFAULT 0,
                `ot_h` INT NOT NULL DEFAULT 0,
                `night_h` INT NOT NULL DEFAULT 0,
                `etc` VARCHAR(20) DEFAULT NULL,
                `source_type` VARCHAR(10) NOT NULL DEFAULT '원본',
                `sheet_name` VARCHAR(30) NOT NULL DEFAULT '전자',
                `uploaded_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                `uploaded_by` VARCHAR(50) DEFAULT NULL,
                UNIQUE KEY `uq_sr` (`emp_name`, `work_date`, `source_type`, `sheet_name`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # basic_h 소수 허용 (휴일 조퇴/외출 시 8-N=6.5 등).
        # DOUBLE 사용 → pymysql이 python float 반환 → 기존 float/int 연산과 호환 (Decimal+float 에러 방지)
        try:
            cur.execute("""SELECT DATA_TYPE FROM information_schema.COLUMNS
                           WHERE TABLE_SCHEMA=DATABASE() AND TABLE_NAME='schedule_record' AND COLUMN_NAME='basic_h'""")
            _sr_bt = cur.fetchone()
            if _sr_bt and str(_sr_bt[0]).lower() not in ('double', 'float'):
                cur.execute("ALTER TABLE schedule_record MODIFY COLUMN `basic_h` DOUBLE NOT NULL DEFAULT 0")
        except Exception:
            pass
        # 공문서(문서요청) 테이블
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `doc_requests` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `requester_id` INT NOT NULL,
                `requester_name` VARCHAR(30) NOT NULL DEFAULT '',
                `doc_type` VARCHAR(30) NOT NULL DEFAULT '재직증명서',
                `purpose` VARCHAR(200) NOT NULL DEFAULT '',
                `memo` VARCHAR(500) NOT NULL DEFAULT '',
                `status` VARCHAR(10) NOT NULL DEFAULT '접수',
                `handler_name` VARCHAR(30) NOT NULL DEFAULT '',
                `reply_memo` VARCHAR(500) NOT NULL DEFAULT '',
                `file_name` VARCHAR(200) NOT NULL DEFAULT '',
                `file_path` VARCHAR(500) NOT NULL DEFAULT '',
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # ERP 동기화 로그
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `erp_sync_log` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `synced_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                `inserted` INT NOT NULL DEFAULT 0,
                `updated` INT NOT NULL DEFAULT 0,
                `detail` TEXT,
                INDEX `idx_synced_at` (`synced_at`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # ERP 출입기록 (CAPS 병행 검증용) — tenter와 동일 구조로 비교
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `tenter_erp` (
                `id` BIGINT AUTO_INCREMENT PRIMARY KEY,
                `e_date` VARCHAR(8) NOT NULL,
                `e_time` VARCHAR(6) NOT NULL,
                `e_card` VARCHAR(20) NOT NULL DEFAULT '',
                `e_mode` VARCHAR(1) NOT NULL DEFAULT '',
                `e_id` INT DEFAULT NULL,
                `e_name` VARCHAR(30) NOT NULL DEFAULT '',
                `e_idno` VARCHAR(32) NOT NULL DEFAULT '',
                `synced_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_erp_enter` (`e_date`, `e_time`, `e_card`),
                INDEX `idx_erp_date` (`e_date`),
                INDEX `idx_erp_eid` (`e_id`, `e_date`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # ERP 출입 대조 결과 로그
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `tenter_diff_log` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `check_date` VARCHAR(8) NOT NULL,
                `checked_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                `caps_cnt` INT NOT NULL DEFAULT 0,
                `erp_cnt` INT NOT NULL DEFAULT 0,
                `only_caps` INT NOT NULL DEFAULT 0,
                `only_erp` INT NOT NULL DEFAULT 0,
                `detail` TEXT,
                UNIQUE KEY `uq_diff_date` (`check_date`),
                INDEX `idx_diff_checked` (`checked_at`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # 근무보고서 그룹
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `wr_groups` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `group_name` VARCHAR(50) NOT NULL,
                `final_step` TINYINT NOT NULL DEFAULT 1,
                `status` VARCHAR(10) NOT NULL DEFAULT '설정중',
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_wrg_name` (`group_name`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `wr_group_reviewers` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `group_id` INT NOT NULL,
                `step` TINYINT NOT NULL,
                `user_id` INT NOT NULL,
                `user_name` VARCHAR(30) NOT NULL DEFAULT '',
                UNIQUE KEY `uq_wrr` (`group_id`, `step`, `user_id`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `wr_group_members` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `group_id` INT NOT NULL,
                `user_id` INT NOT NULL,
                `user_name` VARCHAR(30) NOT NULL DEFAULT '',
                UNIQUE KEY `uq_wrm` (`group_id`, `user_id`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # 작성자 테이블 (대상인원에 포함 안 되고 작성 권한만)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `wr_group_writers` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `group_id` INT NOT NULL,
                `user_id` INT NOT NULL,
                `user_name` VARCHAR(30) NOT NULL DEFAULT '',
                UNIQUE KEY `uq_wrw` (`group_id`, `user_id`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # 근무보고서 (일일 보고서 헤더)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `wr_reports` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `group_id` INT NOT NULL,
                `report_date` VARCHAR(8) NOT NULL,
                `status` VARCHAR(10) NOT NULL DEFAULT '대기',
                `created_by` INT NOT NULL,
                `created_by_name` VARCHAR(30) NOT NULL DEFAULT '',
                `memo` VARCHAR(1000) NOT NULL DEFAULT '',
                `reject_reason` VARCHAR(500) NOT NULL DEFAULT '',
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # uq_wr_gd 제약 제거 (같은 날 수정근무보고서 허용)
        try:
            cur.execute("ALTER TABLE wr_reports DROP INDEX uq_wr_gd")
        except:
            pass
        # 기존 상태값 마이그레이션 (승인→결재, 1차승인→작성완료, 2차승인→검토, 1차검토→검토, 확인→작성완료)
        try:
            cur.execute("UPDATE wr_reports SET status='결재' WHERE status='승인'")
            cur.execute("UPDATE wr_reports SET status='작성완료' WHERE status='1차승인'")
            cur.execute("UPDATE wr_reports SET status='검토' WHERE status='2차승인'")
            cur.execute("UPDATE wr_reports SET status='검토' WHERE status='1차검토'")
            cur.execute("UPDATE wr_reports SET status='작성완료' WHERE status='확인'")
        except:
            pass
        # 근무보고서 상세 (각 직원의 근무내역)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `wr_entries` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `report_id` INT NOT NULL,
                `user_id` INT NOT NULL,
                `user_name` VARCHAR(30) NOT NULL DEFAULT '',
                `category` VARCHAR(20) NOT NULL DEFAULT '주간기본',
                `deduction` INT NOT NULL DEFAULT 0,
                `etc_value` VARCHAR(30) NOT NULL DEFAULT '',
                UNIQUE KEY `uq_wre` (`report_id`, `user_id`, `category`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # wr_entries skipped 컬럼 추가
        try:
            cur.execute("ALTER TABLE wr_entries ADD COLUMN `skipped` TINYINT NOT NULL DEFAULT 0 AFTER `etc_value`")
        except:
            pass
        # wr_reports memo 컨럼 없으면 추가
        try:
            cur.execute("ALTER TABLE wr_reports ADD COLUMN `memo` VARCHAR(1000) NOT NULL DEFAULT '' AFTER `created_by_name`")
        except:
            pass
        # 소방관리 디바이스 상태
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `fire_devices` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `device_key` VARCHAR(100) NOT NULL,
                `device_name` VARCHAR(100) NOT NULL DEFAULT '',
                `location` VARCHAR(100) NOT NULL DEFAULT '',
                `status` VARCHAR(20) NOT NULL DEFAULT 'offline',
                `battery` INT DEFAULT NULL,
                `last_seen` DATETIME DEFAULT NULL,
                `last_payload` TEXT,
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_fire_device_key` (`device_key`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # 소방관리 이벤트 로그
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `fire_alert_logs` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `device_key` VARCHAR(100) NOT NULL,
                `event_type` VARCHAR(40) NOT NULL DEFAULT 'state',
                `severity` VARCHAR(10) NOT NULL DEFAULT 'info',
                `message` VARCHAR(255) NOT NULL DEFAULT '',
                `payload` TEXT,
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX `idx_fire_alert_created_at` (`created_at`),
                INDEX `idx_fire_alert_device_key` (`device_key`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # 근무보고서 결재이력
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `wr_approval_log` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `report_id` INT NOT NULL,
                `step` INT NOT NULL,
                `action` VARCHAR(10) NOT NULL DEFAULT '',
                `user_name` VARCHAR(30) NOT NULL DEFAULT '',
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # Docker 코드 싱크 체크 테이블
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `docker_code_hash` (
                `container_id` VARCHAR(64) NOT NULL,
                `file_name` VARCHAR(200) NOT NULL,
                `file_md5` VARCHAR(32) NOT NULL DEFAULT '',
                `checked_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (`container_id`, `file_name`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # ── 버스배차 테이블 ──
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `bus_members` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `user_id` INT DEFAULT NULL,
                `user_name` VARCHAR(30) NOT NULL,
                `dept` VARCHAR(20) NOT NULL DEFAULT '',
                `bus_no` TINYINT NOT NULL DEFAULT 1,
                `active` TINYINT NOT NULL DEFAULT 1,
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_bus_user` (`user_name`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `bus_dispatch_log` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `dispatch_date` VARCHAR(8) NOT NULL,
                `shift_type` VARCHAR(10) NOT NULL,
                `content` TEXT,
                `bus1_names` TEXT,
                `bus2_names` TEXT,
                `created_by` VARCHAR(30) DEFAULT '',
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

        # ── 통근버스 탑승인원 (경비실 운행일지 엑셀 취합) ──
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `bus_ridership` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `ride_date` DATE NOT NULL,
                `bus_no` TINYINT NOT NULL,
                `driver` VARCHAR(30) DEFAULT '',
                `on_0830` INT DEFAULT 0,
                `off_1730` INT DEFAULT 0,
                `off_2000` INT DEFAULT 0,
                `src_file` VARCHAR(120) DEFAULT '',
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_ride` (`ride_date`, `bus_no`),
                KEY `idx_ride_date` (`ride_date`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

        # ── 월별 일반관리비 (월결산자료 취합 — 식당/상하수도/전기 등) ──
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `mgmt_cost_monthly` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `year` SMALLINT NOT NULL,
                `month` TINYINT NOT NULL,
                `item` VARCHAR(20) NOT NULL,
                `amount` BIGINT NOT NULL DEFAULT 0,
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_cost` (`year`, `month`, `item`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

        # ── 월별 발주식수 (식수현황 엑셀 취합) ──
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `meal_order_monthly` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `year` SMALLINT NOT NULL,
                `month` TINYINT NOT NULL,
                `breakfast` INT DEFAULT 0,
                `lunch` INT DEFAULT 0,
                `dinner` INT DEFAULT 0,
                `night` INT DEFAULT 0,
                `days` SMALLINT DEFAULT 0,
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_meal_ym` (`year`, `month`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

        # ── 일별 발주식수 (식수현황 엑셀 취합) ──
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `meal_order_daily` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `order_date` DATE NOT NULL,
                `breakfast` INT DEFAULT 0,
                `lunch` INT DEFAULT 0,
                `dinner` INT DEFAULT 0,
                `night` INT DEFAULT 0,
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_meal_day` (`order_date`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

        # ── 수도 일일 검침 테이블 (통합관제) ──
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `water_meter` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `read_date` DATE NOT NULL,
                `reading` DECIMAL(14,2) NOT NULL,
                `memo` VARCHAR(200) DEFAULT '',
                `created_by` VARCHAR(30) DEFAULT '',
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_water_date` (`read_date`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # it_requests 사용장소(위치) 컬럼 추가 (없으면)
        try:
            cur.execute("ALTER TABLE it_requests ADD COLUMN `location` VARCHAR(100) NOT NULL DEFAULT '' AFTER `category`")
        except Exception:
            pass
        # admin 기본 계정
        cur.execute("""
            INSERT IGNORE INTO app_users (login_id, password, name, role, must_change_pw)
            VALUES ('admin', %s, '관리자', 'admin', 0)
        """, (_hash_pw("admin"),))
        # 기존 work_override 기타 컬럼의 연차/반차를 leave_records에 반영 시키기
        cur.execute("SELECT e_id, e_date, value FROM work_override WHERE col_type='other' AND (value LIKE '%%연차%%' OR value LIKE '%%반차%%')")
        for eid, edate, val in cur.fetchall():
            vstr = str(val).strip()
            if "반차" in vstr:
                cur.execute("INSERT IGNORE INTO leave_records (e_id, leave_date, leave_type) VALUES (%s,%s,'반차')", (eid, edate))
            if "연차" in vstr and "반차" not in vstr:
                cur.execute("INSERT IGNORE INTO leave_records (e_id, leave_date, leave_type) VALUES (%s,%s,'연차')", (eid, edate))
        # TBM 테이블 초기화 (통합 이후)
        try:
            from tbm_bp import init_tbm_db as _init_tbm_db
            _init_tbm_db(cur)
        except Exception as _e:
            print(f"[TBM init] {_e}")
        # 디지털 사이니지 테이블 초기화 (Phase 0)
        try:
            from signage_bp import init_signage_db as _init_signage_db
            _init_signage_db(cur)
        except Exception as _e:
            print(f"[Signage init] {_e}")
        conn.commit()
        conn.close()
    except Exception:
        pass


def _login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def _admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        if session.get("role") != "admin":
            flash("관리자 권한이 필요합니다.", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated


def _has_perm(perm):
    """admin은 모든 권한, 일반 사용자는 permissions 필드 확인"""
    if session.get("role") == "admin":
        return True
    perms = session.get("permissions", "")
    return perm in [p.strip() for p in perms.split(",") if p.strip()]


# ════════════════════════════════════════════════════════
# 권한 메뉴 레지스트리 (사이드바 메뉴 ↔ 권한키)
# ════════════════════════════════════════════════════════
# 각 항목: (key, parent, label, [available_actions])
#  - parent=None 이면 대분류(카테고리)
#  - actions: view / create / update / delete 중 해당 메뉴에 의미 있는 것만
MENU_STRUCTURE = [
    ("dashboard", None, "대시보드", ["view"]),

    ("cat_attendance", None, "근태관리", []),
    ("attendance_record", "cat_attendance", "근태기록", ["view"]),
    ("weekly52", "cat_attendance", "주52시간", ["view"]),
    ("work_schedule", "cat_attendance", "근무표", ["view", "update"]),
    ("schedule_record", "cat_attendance", "근무표기록관리", ["view", "create", "update", "delete"]),

    # 연차는 대시보드를 1차로 두고 그 아래 2차로 묶는다
    ("cat_leave", None, "연차관리 대시보드", ["view"]),
    ("annual_leave", "cat_leave", "연차관리", ["view", "create"]),
    ("leave_plan", "cat_leave", "연차사용계획", ["view", "create", "update", "delete"]),

    ("cat_hr", None, "인사관리", []),
    ("welfare", "cat_hr", "복지혜택", ["view"]),
    ("meal_mgmt", "cat_hr", "식수관리", ["view", "create", "update", "delete"]),
    ("work_report", "cat_hr", "근무보고서", ["view", "create", "update", "delete"]),
    ("document_mgmt", "cat_hr", "공문서관리", ["view", "create", "update", "delete"]),

    ("cat_bus", None, "배차관리", []),
    ("bus_members", "cat_bus", "탑승자명단", ["view", "create", "update", "delete"]),
    ("bus_dispatch", "cat_bus", "배차조회", ["view", "update"]),
    ("bus_sms", "cat_bus", "발송이력", ["view", "create"]),

    ("cat_mes", None, "MES관리", []),
    ("mes_realtime", "cat_mes", "실시간카운트", ["view"]),
    ("pacemaker", "cat_mes", "페이스메이커", ["view"]),
    ("mes_report", "cat_mes", "생산리포트", ["view"]),
    ("mes_devices", "cat_mes", "장치관리", ["view", "create", "update", "delete"]),

    ("cat_energy", None, "에너지", []),
    ("power_dashboard", "cat_energy", "전력사용현황", ["view"]),
    ("power_history", "cat_energy", "기간별조회", ["view"]),
    ("power_alerts", "cat_energy", "경보이력", ["view"]),
    ("power_settings", "cat_energy", "전력설정", ["view", "update"]),

    ("cat_it", None, "전산관리", []),
    ("it_request", "cat_it", "요청하기", ["view", "create"]),
    ("it_manage", "cat_it", "요청 관리", ["view", "update", "delete"]),

    ("cat_esg", None, "ESG경영", []),
    ("esg_environment", "cat_esg", "환경 (E)", ["view"]),
    ("esg_social", "cat_esg", "사회 (S)", ["view"]),
    ("esg_governance", "cat_esg", "지배구조 (G)", ["view"]),

    ("cat_survey", None, "설문관리", []),
    ("survey_admin", "cat_survey", "설문 관리", ["view", "create", "update", "delete"]),
    ("survey_my", "cat_survey", "내 설문", ["view", "create"]),

    ("cat_tbm", None, "TBM", []),
    ("tbm_sign", "cat_tbm", "서명 참여", ["view", "create"]),
    ("tbm_admin", "cat_tbm", "TBM 관리", ["view", "create", "update", "delete"]),

    ("cat_signage", None, "디지털 사이니지", []),
    ("signage", "cat_signage", "사이니지 관리", ["view", "create", "update", "delete", "publish"]),
]

# 기존 CSV 권한 → 새 메뉴키 매핑 (마이그레이션용)
_CSV_TO_MENUKEYS = {
    "meal": ["meal_mgmt"],
    "leave": ["leave_plan"],
    "schedule": ["work_schedule", "schedule_record"],
    "work_report": ["work_report"],
    "notice": [],  # 별도 공지 메뉴 없음 (대시보드 내부 기능)
    "bus": ["bus_members", "bus_dispatch", "bus_sms"],
    "document": ["document_mgmt"],
}


def _menu_leaves():
    """실제 권한 부여 대상인 리프 메뉴만 반환"""
    return [(k, p, label, acts) for k, p, label, acts in MENU_STRUCTURE if p is not None]


def _seed_perm_groups():
    """기본 권한 그룹 4개 + 각 그룹의 메뉴 권한 프리셋 시드.
    이미 그룹이 있으면 건너뜀 (최초 1회만 실행)."""
    conn = _conn(); cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM perm_groups")
    if cur.fetchone()[0] > 0:
        conn.close()
        return
    # 기본 4개 그룹
    seeds = [
        ("관리자", "모든 기능에 대한 전체 권한", 1, 10),
        ("팀장", "조회 전반 + 주요 등록/수정 권한 (삭제 제한)", 1, 20),
        ("일반직원", "본인 관련 조회 및 신청/작성 권한", 1, 30),
        ("조회전용", "모든 메뉴 조회만 가능", 1, 40),
    ]
    gid_by_name = {}
    for n, desc, sys_flag, order in seeds:
        cur.execute("INSERT INTO perm_groups (name, description, is_system, sort_order) VALUES (%s,%s,%s,%s)",
                    (n, desc, sys_flag, order))
        gid_by_name[n] = cur.lastrowid

    leaves = _menu_leaves()
    # 프리셋 정책 함수
    def preset(group_name, menu_key, actions):
        can = {"view": 0, "create": 0, "update": 0, "delete": 0}
        if group_name == "관리자":
            for a in actions:
                can[a] = 1
        elif group_name == "팀장":
            # 조회 모두 + 등록/수정 대부분, 삭제는 본인 작성 리소스 중심(여기선 쓰기 그룹 외 삭제 금지)
            for a in actions:
                if a == "view":
                    can[a] = 1
                elif a in ("create", "update"):
                    # 설정류·공문서는 관리자만
                    if menu_key in ("power_settings", "mes_devices", "document_mgmt"):
                        can[a] = 0
                    else:
                        can[a] = 1
                elif a == "delete":
                    can[a] = 0
        elif group_name == "일반직원":
            # 본인 관련 조회 + 연차/근무보고서 작성 정도
            if "view" in actions:
                # 설정·장치관리·공문서·근무표기록관리·발송이력·경보이력은 조회 불가
                if menu_key in ("power_settings", "mes_devices", "document_mgmt",
                                "schedule_record", "bus_sms", "power_alerts"):
                    can["view"] = 0
                else:
                    can["view"] = 1
            # 작성 가능 항목
            if menu_key == "annual_leave" and "create" in actions:
                can["create"] = 1
            if menu_key == "work_report" and "create" in actions:
                can["create"] = 1
        elif group_name == "조회전용":
            if "view" in actions:
                # 설정·장치관리·공문서관리는 제외
                if menu_key in ("power_settings", "mes_devices", "document_mgmt"):
                    can["view"] = 0
                else:
                    can["view"] = 1
        return can

    for gname, gid in gid_by_name.items():
        for mkey, _parent, _label, acts in leaves:
            c = preset(gname, mkey, acts)
            cur.execute("""INSERT INTO perm_group_permissions
                           (group_id, menu_key, can_view, can_create, can_update, can_delete)
                           VALUES (%s,%s,%s,%s,%s,%s)""",
                        (gid, mkey, c["view"], c["create"], c["update"], c["delete"]))
    conn.commit(); conn.close()


def _migrate_csv_to_overrides():
    """기존 CSV permissions 보유 사용자 → user_perm_overrides(allow)로 변환.
    이미 override 가진 사용자는 건너뜀. 최초 1회만 실행되는 idempotent 구조."""
    conn = _conn(); cur = conn.cursor()
    cur.execute("SELECT id, permissions FROM app_users WHERE permissions IS NOT NULL AND permissions<>''")
    rows = cur.fetchall()
    for uid, perms_csv in rows:
        # 이미 오버라이드가 있으면 생략
        cur.execute("SELECT COUNT(*) FROM user_perm_overrides WHERE user_id=%s", (uid,))
        if cur.fetchone()[0] > 0:
            continue
        keys = [p.strip() for p in perms_csv.split(",") if p.strip()]
        for csv_key in keys:
            menu_keys = _CSV_TO_MENUKEYS.get(csv_key, [])
            for mk in menu_keys:
                # 리프 메뉴의 가용 액션 전체에 allow
                acts = next((a for k, _p, _l, a in MENU_STRUCTURE if k == mk), [])
                for act in acts:
                    try:
                        cur.execute("""INSERT IGNORE INTO user_perm_overrides
                                       (user_id, menu_key, action, grant_type)
                                       VALUES (%s,%s,%s,'allow')""",
                                    (uid, mk, act))
                    except Exception:
                        pass
    conn.commit(); conn.close()


def _backfill_perm_group_menus():
    """MENU_STRUCTURE에 추가된 새 메뉴를 기존 그룹의 perm_group_permissions에 빈 행(0,0,0,0)으로 채움.
    관리자 그룹(is_system=1, name='관리자')에만 가용 액션 전부 allow로 자동 세팅."""
    conn = _conn(); cur = conn.cursor()
    cur.execute("SELECT id, name FROM perm_groups")
    groups = cur.fetchall()
    if not groups:
        conn.close()
        return
    leaves = _menu_leaves()
    for gid, gname in groups:
        for mk, _p, _l, acts in leaves:
            # 존재 여부 확인
            cur.execute("SELECT 1 FROM perm_group_permissions WHERE group_id=%s AND menu_key=%s", (gid, mk))
            if cur.fetchone():
                continue
            # 관리자 그룹은 가용 액션 전부 on
            if gname == "관리자":
                v = 1 if "view" in acts else 0
                c_ = 1 if "create" in acts else 0
                u_ = 1 if "update" in acts else 0
                d_ = 1 if "delete" in acts else 0
            # 조회전용 그룹은 view만 on (설정/삭제 제외)
            elif gname == "조회전용":
                v = 1 if "view" in acts else 0
                c_ = u_ = d_ = 0
            else:
                v = c_ = u_ = d_ = 0
            cur.execute("""INSERT INTO perm_group_permissions
                           (group_id, menu_key, can_view, can_create, can_update, can_delete)
                           VALUES (%s,%s,%s,%s,%s,%s)""", (gid, mk, v, c_, u_, d_))
    conn.commit(); conn.close()


_init_db()
try:
    _seed_perm_groups()
    _backfill_perm_group_menus()
    _migrate_csv_to_overrides()
except Exception as _e:
    print(f"[perm init] {_e}")


def _time_to_sec(t):
    if not t or len(t) < 6:
        return None
    return int(t[:2]) * 3600 + int(t[2:4]) * 60 + int(t[4:6])


def _sec_to_fmt(s):
    if s is None:
        return "-"
    h = int(s) // 3600
    m = (int(s) % 3600) // 60
    return f"{h:02d}:{m:02d}"


def _fmt_time(t):
    if not t or len(t) < 6:
        return ""
    return f"{t[:2]}:{t[2:4]}:{t[4:6]}"


def _fmt_date(d):
    if d and len(d) == 8:
        return f"{d[:4]}-{d[4:6]}-{d[6:8]}"
    return d or ""


def _merge_night_shifts(day_emp):
    """야간근무 대각선 패턴 병합
    패턴: 3/23 19:23출근 → 3/24 05:00퇴근 + 19:30출근 → 3/25 05:00퇴근 ...
    같은 날에 새벽퇴근(전날분) + 저녁출근(오늘분)이 섞인 경우를 분리 후 병합
    """
    emp_dates = defaultdict(list)
    for (e_date, eid) in list(day_emp.keys()):
        emp_dates[eid].append(e_date)
    remove_keys = []
    lost_morning_outs = {}
    for eid, dates in emp_dates.items():
        dates.sort()
        # 1단계: 시각선 분리 - 같은 날 새벽퇴근(전날분) + 정상출근(오늘분)
        for date in dates:
            key = (date, eid)
            rec = day_emp.get(key)
            if not rec:
                continue
            in_s = _time_to_sec(rec["in_time"])
            out_s = _time_to_sec(rec["out_time"])
            if (in_s is not None and in_s >= 64800 and
                    out_s is not None and out_s >= 43200 and out_s < in_s):
                # 야간 출근 직전 퇴근기록: 이전 근무의 퇴근이므로 제거
                rec["out_time"] = None
                if "out_gate" in rec:
                    rec["out_gate"] = ""
            elif (in_s is not None and in_s >= 64800 and
                    out_s is not None and out_s < 43200):
                # 새벽 퇴근을 전날 전날분으로 이동
                morning_out = rec["out_time"]
                morning_gate = rec.get("out_gate", "")
                rec["out_time"] = None
                if "out_gate" in rec:
                    rec["out_gate"] = ""
                try:
                    prev_date = (datetime.strptime(date, "%Y%m%d") - timedelta(days=1)).strftime("%Y%m%d")
                except ValueError:
                    continue
                prev_key = (prev_date, eid)
                prev = day_emp.get(prev_key)
                merged = False
                if prev:
                    prev_in_s = _time_to_sec(prev["in_time"])
                    if prev_in_s is not None and prev_in_s >= 64800 and prev["out_time"] is None:
                        prev["out_time"] = morning_out
                        if "out_gate" in prev:
                            prev["out_gate"] = morning_gate
                        prev["night_next_day"] = True
                        merged = True
                if not merged:
                    lost_morning_outs[prev_key] = {
                        "date": prev_date, "id": eid,
                        "name": rec["name"], "dept": rec.get("dept", ""),
                        "out_time": morning_out, "out_gate": morning_gate,
                    }
        # 2단계: 연속 미러링 - 전날 출근만+퇴근없음 / 다음날 출근없음+퇴근만
        for i in range(len(dates) - 1):
            curr_key = (dates[i], eid)
            next_key = (dates[i + 1], eid)
            curr = day_emp.get(curr_key)
            nxt = day_emp.get(next_key)
            if not curr or not nxt:
                continue
            try:
                curr_dt = datetime.strptime(dates[i], "%Y%m%d")
                next_dt = datetime.strptime(dates[i + 1], "%Y%m%d")
            except ValueError:
                continue
            if (next_dt - curr_dt).days != 1:
                continue
            curr_in_s = _time_to_sec(curr["in_time"])
            if curr_in_s is not None and curr_in_s >= 64800 and curr["out_time"] is None:
                nxt_out_s = _time_to_sec(nxt["out_time"])
                if nxt["in_time"] is None and nxt_out_s is not None and nxt_out_s < 43200:
                    curr["out_time"] = nxt["out_time"]
                    if "out_gate" in curr:
                        curr["out_gate"] = nxt.get("out_gate", "")
                    curr["night_next_day"] = True
                    remove_keys.append(next_key)
    for k in remove_keys:
        if k in day_emp:
            del day_emp[k]
    # 3단계: 유실된 새벽 퇴근 복원 (출근없음+퇴근만 전날분 생성)
    for key, info in lost_morning_outs.items():
        if key not in day_emp:
            day_emp[key] = {
                "date": info["date"], "id": info["id"],
                "name": info["name"], "dept": info.get("dept", ""),
                "in_time": None, "out_time": info["out_time"],
                "in_gate": "", "out_gate": info.get("out_gate", ""),
                "night_next_day": True,
            }
        elif day_emp[key]["out_time"] is None:
            day_emp[key]["out_time"] = info["out_time"]
            day_emp[key]["night_next_day"] = True
            if "out_gate" in day_emp[key]:
                day_emp[key]["out_gate"] = info.get("out_gate", "")
    # 4단계: 고아 새벽퇴근을 전날 야간근무로 이동
    orphan_moves = {}
    for key in list(day_emp.keys()):
        rec = day_emp[key]
        if rec["in_time"] is not None:
            continue
        out_s = _time_to_sec(rec["out_time"])
        if out_s is None or out_s >= 28800:
            continue
        if rec.get("night_next_day"):
            continue
        e_date, eid = key
        try:
            prev_date = (datetime.strptime(e_date, "%Y%m%d") - timedelta(days=1)).strftime("%Y%m%d")
        except ValueError:
            continue
        prev_key = (prev_date, eid)
        if prev_key not in day_emp:
            orphan_moves[key] = prev_key
    for old_key, new_key in orphan_moves.items():
        rec = day_emp.pop(old_key)
        rec["date"] = new_key[0]
        rec["night_next_day"] = True
        day_emp[new_key] = rec


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        login_id = request.form.get("login_id", "").strip()
        password = request.form.get("password", "").strip()
        if not login_id or not password:
            flash("아이디와 비밀번호를 입력하세요.", "danger")
            return render_template("login.html")
        conn = _conn(); cur = conn.cursor()
        cur.execute("SELECT id, login_id, password, name, e_id, role, must_change_pw, `permissions` FROM app_users WHERE login_id=%s", (login_id,))
        row = cur.fetchone()
        if not row:
            # idno(사번) 또는 id(DB PK)로 tuser 조회 후 자동 계정 생성
            cur.execute("SELECT id, name, idno FROM tuser WHERE (idno=%s OR CAST(id AS CHAR)=%s) AND name IS NOT NULL AND name <> ''", (login_id, login_id))
            tu = cur.fetchone()
            if tu:
                use_id = tu[2] if tu[2] else str(tu[0])  # idno 우선
                if password == login_id:
                    cur.execute("""
                        INSERT IGNORE INTO app_users (login_id, password, name, e_id, role, must_change_pw)
                        VALUES (%s, %s, %s, %s, 'user', 1)
                    """, (use_id, _hash_pw(use_id), tu[1], tu[0]))
                    conn.commit()
                    cur.execute("SELECT id, login_id, password, name, e_id, role, must_change_pw, `permissions` FROM app_users WHERE login_id=%s", (use_id,))
                    row = cur.fetchone()
        conn.close()
        if not row or row[2] != _hash_pw(password):
            flash("아이디 또는 비밀번호가 올바르지 않습니다.", "danger")
            return render_template("login.html")
        session["user_id"] = row[0]
        session["login_id"] = row[1]
        session["user_name"] = row[3]
        session["e_id"] = row[4]
        session["role"] = row[5]
        session["permissions"] = row[7] if len(row) > 7 else ""
        # 검토자/멤버 여부 세션 저장 (wr_groups, lp_groups)
        # wr_group_members/reviewers는 tuser.id (e_id) 사용
        try:
            conn2 = _conn(); cur2 = conn2.cursor()
            e_id = row[4]  # tuser.id
            cur2.execute("SELECT COUNT(*) FROM wr_group_reviewers WHERE user_id=%s", (e_id,))
            cnt1 = cur2.fetchone()[0]
            cur2.execute("SELECT COUNT(*) FROM wr_group_members WHERE user_id=%s", (e_id,))
            cnt1b = cur2.fetchone()[0]
            cur2.execute("SELECT COUNT(*) FROM wr_group_writers WHERE user_id=%s", (e_id,))
            cnt1c = cur2.fetchone()[0]
            session["is_wr_member"] = (cnt1 + cnt1b + cnt1c) > 0
            session["is_leave_approver"] = cnt1 > 0
            cur2.execute("SELECT COUNT(*) FROM lp_group_members WHERE user_id=%s", (e_id,))
            cnt2 = cur2.fetchone()[0]
            cur2.execute("SELECT COUNT(*) FROM lp_group_reviewers WHERE user_id=%s", (e_id,))
            cnt3 = cur2.fetchone()[0]
            session["is_lp_member"] = (cnt2 + cnt3) > 0
            # 최종 접속시간 업데이트
            cur2.execute("UPDATE app_users SET last_login=NOW() WHERE id=%s", (row[0],))
            conn2.commit()
            conn2.close()
        except Exception:
            session["is_leave_approver"] = False
            session["is_lp_member"] = False
            session["is_wr_member"] = False
        if row[6]:
            flash("초기 비밀번호입니다. 비밀번호를 변경해 주세요.", "warning")
            return redirect(url_for("change_password"))
        return redirect(url_for("dashboard"))
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/change_password", methods=["GET", "POST"])
@_login_required
def change_password():
    if request.method == "POST":
        cur_pw = request.form.get("current_password", "")
        new_pw = request.form.get("new_password", "")
        confirm_pw = request.form.get("confirm_password", "")
        if not new_pw or len(new_pw) < 4:
            flash("새 비밀번호는 4자리 이상이어야 합니다.", "danger")
            return render_template("change_password.html")
        if new_pw != confirm_pw:
            flash("새 비밀번호가 일치하지 않습니다.", "danger")
            return render_template("change_password.html")
        conn = _conn(); cur = conn.cursor()
        cur.execute("SELECT password FROM app_users WHERE id=%s", (session["user_id"],))
        row = cur.fetchone()
        if not row or row[0] != _hash_pw(cur_pw):
            flash("현재 비밀번호가 올바르지 않습니다.", "danger")
            conn.close()
            return render_template("change_password.html")
        cur.execute("UPDATE app_users SET password=%s, must_change_pw=0 WHERE id=%s",
                    (_hash_pw(new_pw), session["user_id"]))
        conn.commit(); conn.close()
        flash("비밀번호가 변경되었습니다.", "success")
        return redirect(url_for("dashboard"))
    return render_template("change_password.html")


@app.route("/user_management")
@_admin_required
def user_management():
    conn = _conn(); cur = conn.cursor()
    # company, last_login 컬럼 존재 여부 확인 및 추가
    cur.execute("SHOW COLUMNS FROM app_users LIKE 'company'")
    if not cur.fetchone():
        cur.execute("ALTER TABLE app_users ADD COLUMN company VARCHAR(50) DEFAULT NULL")
        conn.commit()
    cur.execute("SHOW COLUMNS FROM app_users LIKE 'last_login'")
    if not cur.fetchone():
        cur.execute("ALTER TABLE app_users ADD COLUMN last_login DATETIME DEFAULT NULL")
        conn.commit()
    # app_users와 tuser, employee_roster 조인해서 부서 정보 가져오기
    cur.execute("""
        SELECT a.id, a.login_id, a.name, a.role, a.`permissions`, a.must_change_pw,
               a.created_at, a.tg_chat_id, a.tg_alerts,
               COALESCE(er.dept, t.company, a.company) as company,
               a.last_login, a.perm_group_id, g.name as group_name
        FROM app_users a
        LEFT JOIN tuser t ON a.e_id = t.id
        LEFT JOIN employee_roster er ON a.name = er.name
        LEFT JOIN perm_groups g ON a.perm_group_id = g.id
        ORDER BY a.role DESC, a.name
    """)
    users = []
    dept_set = set()
    existing_names = set()
    # 오버라이드 개수 사전 집계
    cur2 = conn.cursor()
    cur2.execute("SELECT user_id, COUNT(*) FROM user_perm_overrides GROUP BY user_id")
    override_counts = {r[0]: r[1] for r in cur2.fetchall()}
    for uid, lid, name, role, perms, mcp, cat, tg_cid, tg_al, company, last_login, pgid, gname in cur.fetchall():
        comp_code = (company or "").strip()
        dept_name = DEPT_MAP.get(comp_code, comp_code)  # 매핑 없으면 원본 사용
        users.append({"id": uid, "login_id": lid, "name": name, "role": role,
                       "permissions": perms or "", "must_change_pw": mcp,
                       "created_at": cat.strftime("%Y-%m-%d") if cat else "",
                       "tg_chat_id": tg_cid or "", "tg_alerts": tg_al or "",
                       "company": dept_name,
                       "last_login": last_login.strftime("%m-%d %H:%M") if last_login else "",
                       "perm_group_id": pgid, "perm_group_name": gname or "",
                       "override_count": override_counts.get(uid, 0)})
        if dept_name:
            dept_set.add(dept_name)
        if name:
            existing_names.add(name)
    dept_list = sorted(dept_set)
    # 권한 그룹 목록 (드롭다운용)
    cur.execute("SELECT id, name FROM perm_groups ORDER BY sort_order, id")
    perm_groups_list = [{"id": r[0], "name": r[1]} for r in cur.fetchall()]
    # employee_roster에서 아직 등록 안된 사람 목록 + 세부 부서 수집
    cur.execute("SELECT name, emp_no, dept FROM employee_roster WHERE name IS NOT NULL AND name <> '' ORDER BY name")
    roster_list = []
    for rname, emp_no, rdept in cur.fetchall():
        if rdept:
            dept_set.add(rdept)  # 세부 부서도 필터에 추가
        if rname and rname not in existing_names:
            roster_list.append({"name": rname, "emp_no": emp_no or "", "dept": rdept or ""})
    # 세부 부서 포함한 전체 부서 목록
    dept_list = sorted(dept_set)
    conn.close()
    return render_template("user_management.html",
                           users=users,
                           dept_list=dept_list,
                           roster_list=roster_list,
                           perm_groups=perm_groups_list,
                           menu_structure=MENU_STRUCTURE)

@app.route("/sync_users", methods=["POST"])
@_admin_required
def sync_users():
    """연차관리 재직자를 app_users에 일괄 등록 (없는 사람만)"""
    year_int = datetime.now().year
    conn = _conn(); cur = conn.cursor()
    cur.execute("""SELECT t.id, t.name, t.idno FROM employee_roster er
                   INNER JOIN (
                       SELECT COALESCE(MAX(CASE WHEN company<>'' THEN id END), MAX(id)) AS id,
                              name, COALESCE(MAX(idno),'') AS idno
                       FROM tuser WHERE name IS NOT NULL AND name <> ''
                       GROUP BY name
                   ) t ON t.name = er.name""")
    employees = cur.fetchall()
    cur.execute("SELECT e_id FROM app_users WHERE e_id IS NOT NULL")
    existing = {r[0] for r in cur.fetchall()}
    added = 0
    for eid, name, idno in employees:
        if eid not in existing:
            login_id = idno if idno else str(eid)
            pw_hash = hashlib.sha256(login_id.encode()).hexdigest()
            cur.execute("""INSERT INTO app_users (login_id, password, name, e_id, role, must_change_pw, created_at)
                           VALUES (%s, %s, %s, %s, 'user', 1, NOW())""",
                        (login_id, pw_hash, name, eid))
            added += 1
    conn.commit(); conn.close()
    return jsonify(ok=True, added=added)


@app.route("/get_telegram_settings")
@_admin_required
def get_telegram_settings():
    conn = _conn(); cur = conn.cursor()
    cur.execute("SELECT `key`, `enabled` FROM telegram_settings")
    settings = {r[0]: bool(r[1]) for r in cur.fetchall()}
    conn.close()
    return jsonify(ok=True, settings=settings)


@app.route("/save_telegram_settings", methods=["POST"])
@_admin_required
def save_telegram_settings():
    data = request.get_json()
    conn = _conn(); cur = conn.cursor()
    for key in ("sync_error", "leave", "meal", "notice", "wr_overdue"):
        enabled = 1 if data.get(key) else 0
        cur.execute("UPDATE telegram_settings SET `enabled`=%s WHERE `key`=%s", (enabled, key))
    conn.commit(); conn.close()
    return jsonify(ok=True)


@app.route("/save_user_telegram", methods=["POST"])
@_admin_required
def save_user_telegram():
    try:
        data = request.get_json()
        uid = int(data["id"])
        tg_chat_id = str(data.get("tg_chat_id", "")).strip()
        tg_alerts = str(data.get("tg_alerts", "")).strip()
        conn = _conn(); cur = conn.cursor()
        cur.execute("UPDATE app_users SET tg_chat_id=%s, tg_alerts=%s WHERE id=%s", (tg_chat_id, tg_alerts, uid))
        conn.commit(); conn.close()
        return jsonify(ok=True)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/save_user_perm", methods=["POST"])
@_admin_required
def save_user_perm():
    try:
        data = request.get_json()
        uid = int(data["id"])
        perms = str(data.get("permissions", "")).strip()
        conn = _conn(); cur = conn.cursor()
        cur.execute("UPDATE app_users SET `permissions`=%s WHERE id=%s", (perms, uid))
        conn.commit(); conn.close()
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


# ════════════════════════════════════════════════════════
# 권한 그룹 관리 (UI + CRUD API)
# ════════════════════════════════════════════════════════

@app.route("/perm_groups")
@_admin_required
def perm_groups_page():
    """권한 그룹 관리 화면"""
    conn = _conn(); cur = conn.cursor()
    cur.execute("""SELECT id, name, description, is_system, sort_order
                   FROM perm_groups ORDER BY sort_order, id""")
    groups = [{"id": r[0], "name": r[1], "description": r[2],
               "is_system": r[3], "sort_order": r[4]} for r in cur.fetchall()]
    # 각 그룹의 권한 매트릭스
    group_perms = {}
    cur.execute("""SELECT group_id, menu_key, can_view, can_create, can_update, can_delete
                   FROM perm_group_permissions""")
    for gid, mk, v, c, u, d in cur.fetchall():
        group_perms.setdefault(gid, {})[mk] = {"view": v, "create": c, "update": u, "delete": d}
    # 각 그룹의 소속 사용자 수
    user_counts = {}
    cur.execute("SELECT perm_group_id, COUNT(*) FROM app_users WHERE perm_group_id IS NOT NULL GROUP BY perm_group_id")
    for gid, cnt in cur.fetchall():
        user_counts[gid] = cnt
    conn.close()
    return render_template("perm_groups.html",
                           groups=groups,
                           group_perms=group_perms,
                           user_counts=user_counts,
                           menu_structure=MENU_STRUCTURE)


@app.route("/api/perm_group", methods=["POST"])
@_admin_required
def api_perm_group_save():
    """그룹 생성/수정/삭제"""
    try:
        data = request.get_json()
        action = str(data.get("action", ""))
        conn = _conn(); cur = conn.cursor()
        if action == "create":
            name = str(data.get("name", "")).strip()
            desc = str(data.get("description", "")).strip()
            if not name:
                conn.close(); return jsonify(ok=False, error="이름 필수"), 400
            cur.execute("SELECT id FROM perm_groups WHERE name=%s", (name,))
            if cur.fetchone():
                conn.close(); return jsonify(ok=False, error="이미 존재하는 이름"), 400
            cur.execute("SELECT COALESCE(MAX(sort_order),0)+10 FROM perm_groups")
            order = cur.fetchone()[0]
            cur.execute("""INSERT INTO perm_groups (name, description, is_system, sort_order)
                           VALUES (%s,%s,0,%s)""", (name, desc, order))
            new_id = cur.lastrowid
            # 빈 권한 레코드 생성 (모든 리프 메뉴, 전부 0)
            for mk, _p, _l, _a in _menu_leaves():
                cur.execute("""INSERT INTO perm_group_permissions
                               (group_id, menu_key, can_view, can_create, can_update, can_delete)
                               VALUES (%s,%s,0,0,0,0)""", (new_id, mk))
            conn.commit(); conn.close()
            return jsonify(ok=True, id=new_id)
        elif action == "update":
            gid = int(data["id"])
            name = str(data.get("name", "")).strip()
            desc = str(data.get("description", "")).strip()
            cur.execute("SELECT is_system FROM perm_groups WHERE id=%s", (gid,))
            row = cur.fetchone()
            if not row:
                conn.close(); return jsonify(ok=False, error="그룹 없음"), 404
            # 시스템 그룹은 이름 변경 불가 (설명만)
            if row[0]:
                cur.execute("UPDATE perm_groups SET description=%s WHERE id=%s", (desc, gid))
            else:
                cur.execute("UPDATE perm_groups SET name=%s, description=%s WHERE id=%s", (name, desc, gid))
            conn.commit(); conn.close()
            return jsonify(ok=True)
        elif action == "delete":
            gid = int(data["id"])
            cur.execute("SELECT is_system FROM perm_groups WHERE id=%s", (gid,))
            row = cur.fetchone()
            if not row:
                conn.close(); return jsonify(ok=False, error="그룹 없음"), 404
            if row[0]:
                conn.close(); return jsonify(ok=False, error="시스템 기본 그룹은 삭제할 수 없습니다"), 400
            # 소속 사용자 확인
            cur.execute("SELECT COUNT(*) FROM app_users WHERE perm_group_id=%s", (gid,))
            if cur.fetchone()[0] > 0:
                conn.close(); return jsonify(ok=False, error="소속 사용자가 있어 삭제할 수 없습니다"), 400
            cur.execute("DELETE FROM perm_group_permissions WHERE group_id=%s", (gid,))
            cur.execute("DELETE FROM perm_groups WHERE id=%s", (gid,))
            conn.commit(); conn.close()
            return jsonify(ok=True)
        else:
            conn.close()
            return jsonify(ok=False, error="잘못된 action"), 400
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/api/perm_group_matrix_get")
@_admin_required
def api_perm_group_matrix_get():
    """단일 그룹의 권한 매트릭스 조회"""
    try:
        gid = int(request.args.get("gid", "0"))
        if not gid:
            return jsonify(ok=False, error="gid 필수"), 400
        conn = _conn(); cur = conn.cursor()
        cur.execute("""SELECT menu_key, can_view, can_create, can_update, can_delete
                       FROM perm_group_permissions WHERE group_id=%s""", (gid,))
        matrix = {}
        for mk, v, c, u, d in cur.fetchall():
            matrix[mk] = {"view": bool(v), "create": bool(c),
                          "update": bool(u), "delete": bool(d)}
        conn.close()
        return jsonify(ok=True, matrix=matrix)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/api/perm_group_matrix", methods=["POST"])
@_admin_required
def api_perm_group_matrix_save():
    """그룹의 메뉴×액션 매트릭스 일괄 저장"""
    try:
        data = request.get_json()
        gid = int(data["group_id"])
        matrix = data.get("matrix", {})  # {menu_key: {view, create, update, delete}}
        conn = _conn(); cur = conn.cursor()
        for mk, _p, _l, _a in _menu_leaves():
            c = matrix.get(mk, {})
            vals = (int(bool(c.get("view", 0))), int(bool(c.get("create", 0))),
                    int(bool(c.get("update", 0))), int(bool(c.get("delete", 0))))
            cur.execute("""INSERT INTO perm_group_permissions
                           (group_id, menu_key, can_view, can_create, can_update, can_delete)
                           VALUES (%s,%s,%s,%s,%s,%s)
                           ON DUPLICATE KEY UPDATE
                           can_view=VALUES(can_view), can_create=VALUES(can_create),
                           can_update=VALUES(can_update), can_delete=VALUES(can_delete)""",
                        (gid, mk) + vals)
        conn.commit(); conn.close()
        return jsonify(ok=True)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/api/user_perm_detail/<int:uid>")
@_admin_required
def api_user_perm_detail(uid):
    """특정 사용자의 그룹/오버라이드 상세 조회"""
    try:
        conn = _conn(); cur = conn.cursor()
        cur.execute("""SELECT u.perm_group_id, g.name, u.permissions
                       FROM app_users u
                       LEFT JOIN perm_groups g ON u.perm_group_id=g.id
                       WHERE u.id=%s""", (uid,))
        row = cur.fetchone()
        if not row:
            conn.close(); return jsonify(ok=False, error="사용자 없음"), 404
        gid, gname, old_csv = row
        # 그룹 권한
        group_matrix = {}
        if gid:
            cur.execute("""SELECT menu_key, can_view, can_create, can_update, can_delete
                           FROM perm_group_permissions WHERE group_id=%s""", (gid,))
            for mk, v, c, u, d in cur.fetchall():
                group_matrix[mk] = {"view": bool(v), "create": bool(c),
                                    "update": bool(u), "delete": bool(d)}
        # 오버라이드
        overrides = []
        cur.execute("""SELECT menu_key, action, grant_type FROM user_perm_overrides
                       WHERE user_id=%s""", (uid,))
        for mk, act, gt in cur.fetchall():
            overrides.append({"menu_key": mk, "action": act, "grant_type": gt})
        # 그룹 목록
        cur.execute("SELECT id, name FROM perm_groups ORDER BY sort_order, id")
        all_groups = [{"id": r[0], "name": r[1]} for r in cur.fetchall()]
        conn.close()
        return jsonify(ok=True,
                       group_id=gid, group_name=gname,
                       old_csv=old_csv or "",
                       group_matrix=group_matrix,
                       overrides=overrides,
                       all_groups=all_groups)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/api/user_perm", methods=["POST"])
@_admin_required
def api_user_perm_save():
    """사용자 그룹 + 오버라이드 저장"""
    try:
        data = request.get_json()
        uid = int(data["id"])
        group_id = data.get("group_id")
        overrides = data.get("overrides", [])  # [{menu_key, action, grant_type}, ...]
        conn = _conn(); cur = conn.cursor()
        # 그룹 설정
        if group_id in (None, "", 0, "0"):
            cur.execute("UPDATE app_users SET perm_group_id=NULL WHERE id=%s", (uid,))
        else:
            cur.execute("UPDATE app_users SET perm_group_id=%s WHERE id=%s", (int(group_id), uid))
        # 오버라이드 재작성 (전체 교체)
        cur.execute("DELETE FROM user_perm_overrides WHERE user_id=%s", (uid,))
        for ov in overrides:
            mk = str(ov.get("menu_key", "")).strip()
            act = str(ov.get("action", "")).strip()
            gt = str(ov.get("grant_type", "allow")).strip()
            if not mk or act not in ("view", "create", "update", "delete") or gt not in ("allow", "deny"):
                continue
            cur.execute("""INSERT IGNORE INTO user_perm_overrides
                           (user_id, menu_key, action, grant_type)
                           VALUES (%s,%s,%s,%s)""", (uid, mk, act, gt))
        conn.commit(); conn.close()
        return jsonify(ok=True)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


# ════════════════════════════════════════════════════════
# 설문(Survey) — Phase 1: 목록 페이지 + 스텁 라우트
# ════════════════════════════════════════════════════════

def _survey_can_admin():
    """관리자이거나 survey_admin 권한(view)이 있으면 True. Phase 6에서 세분화 예정."""
    if session.get("role") == "admin":
        return True
    # 기존 _has_perm 기반 임시 체크 (신 권한 시스템 연동은 권한 체크 전환 후)
    return False


@app.route("/surveys")
@_login_required
def surveys_admin():
    if not _survey_can_admin():
        flash("접근 권한이 없습니다.", "danger")
        return redirect("/")
    conn = _conn(); cur = conn.cursor()
    cur.execute("""SELECT id, title, status, start_at, end_at, anonymous,
                          created_by_name, created_at,
                          (SELECT COUNT(*) FROM survey_responses WHERE survey_id=surveys.id) AS resp_cnt
                   FROM surveys
                   WHERE is_template=0
                   ORDER BY FIELD(status,'published','draft','closed','archived'), id DESC""")
    rows = cur.fetchall()
    surveys_list = [{
        "id": r[0], "title": r[1], "status": r[2],
        "start_at": r[3].strftime("%Y-%m-%d %H:%M") if r[3] else "",
        "end_at": r[4].strftime("%Y-%m-%d %H:%M") if r[4] else "",
        "anonymous": bool(r[5]),
        "created_by_name": r[6],
        "created_at": r[7].strftime("%Y-%m-%d") if r[7] else "",
        "resp_cnt": r[8] or 0,
    } for r in rows]
    conn.close()
    return render_template("survey_list.html",
                           surveys=surveys_list,
                           active_page="survey_admin")


@app.route("/surveys/new")
@_login_required
def survey_new():
    if not _survey_can_admin():
        flash("접근 권한이 없습니다.", "danger"); return redirect("/")
    conn = _conn(); cur = conn.cursor()
    cur.execute("""INSERT INTO surveys (title, description, status, created_by, created_by_name)
                   VALUES (%s,%s,'draft',%s,%s)""",
                ("제목 없는 설문", "", session.get("user_id"), session.get("user_name","")))
    new_id = cur.lastrowid
    conn.commit(); conn.close()
    return redirect(f"/surveys/{new_id}/edit")


@app.route("/surveys/<int:sid>/edit")
@_login_required
def survey_edit(sid):
    if not _survey_can_admin():
        flash("접근 권한이 없습니다.", "danger"); return redirect("/")
    conn = _conn(); cur = conn.cursor()
    cur.execute("""SELECT id,title,description,start_at,end_at,status,anonymous,
                          allow_edit,one_per_user,remind_hours,send_email,send_sms
                   FROM surveys WHERE id=%s""", (sid,))
    r = cur.fetchone()
    if not r:
        conn.close()
        flash("설문을 찾을 수 없습니다.", "danger"); return redirect("/surveys")
    survey = {
        "id": r[0], "title": r[1], "description": r[2] or "",
        "start_at": r[3].strftime("%Y-%m-%dT%H:%M") if r[3] else "",
        "end_at": r[4].strftime("%Y-%m-%dT%H:%M") if r[4] else "",
        "status": r[5],
        "anonymous": bool(r[6]), "allow_edit": bool(r[7]),
        "one_per_user": bool(r[8]), "remind_hours": r[9],
        "send_email": bool(r[10]), "send_sms": bool(r[11]),
    }
    cur.execute("""SELECT id,qtype,title,description,required,sort_order,section_id,
                          branch_src_qid,branch_src_value
                   FROM survey_questions WHERE survey_id=%s ORDER BY sort_order, id""", (sid,))
    questions = []
    for qr in cur.fetchall():
        qid = qr[0]
        cur.execute("SELECT id,label,sort_order FROM survey_question_options WHERE question_id=%s ORDER BY sort_order, id", (qid,))
        options = [{"id":o[0],"label":o[1],"sort_order":o[2]} for o in cur.fetchall()]
        questions.append({
            "id": qid, "qtype": qr[1], "title": qr[2] or "",
            "description": qr[3] or "", "required": bool(qr[4]),
            "sort_order": qr[5], "section_id": qr[6],
            "branch_src_qid": qr[7], "branch_src_value": qr[8] or "",
            "options": options,
        })
    # 섹션 로드
    cur.execute("""SELECT id, title, description, sort_order FROM survey_sections
                   WHERE survey_id=%s ORDER BY sort_order, id""", (sid,))
    sections = [{"id": r[0], "title": r[1] or "", "description": r[2] or "", "sort_order": r[3]}
                for r in cur.fetchall()]
    conn.close()
    return render_template("survey_edit.html", survey=survey, questions=questions,
                           sections=sections, active_page="survey_admin")


@app.route("/surveys/<int:sid>/preview")
@_login_required
def survey_preview(sid):
    """관리자 미리보기 — 실제 응답자 화면과 동일하게 렌더, 제출 불가"""
    if not _survey_can_admin():
        flash("접근 권한이 없습니다.", "danger"); return redirect("/")
    conn = _conn(); cur = conn.cursor()
    cur.execute("""SELECT id,title,description,start_at,end_at,status,anonymous
                   FROM surveys WHERE id=%s""", (sid,))
    r = cur.fetchone()
    if not r:
        conn.close()
        flash("설문을 찾을 수 없습니다.", "danger"); return redirect("/surveys")
    survey = {"id": r[0], "title": r[1], "description": r[2] or "",
              "start_at": r[3].strftime("%Y-%m-%d %H:%M") if r[3] else "",
              "end_at": r[4].strftime("%Y-%m-%d %H:%M") if r[4] else "",
              "status": r[5], "anonymous": bool(r[6])}
    cur.execute("""SELECT id, title, description, sort_order FROM survey_sections
                   WHERE survey_id=%s ORDER BY sort_order, id""", (sid,))
    sections = [{"id": r2[0], "title": r2[1] or "", "description": r2[2] or "", "sort_order": r2[3]}
                for r2 in cur.fetchall()]
    cur.execute("""SELECT id,qtype,title,description,required,section_id,
                          branch_src_qid,branch_src_value,sort_order
                   FROM survey_questions WHERE survey_id=%s ORDER BY sort_order, id""", (sid,))
    questions = []
    for qr in cur.fetchall():
        qid = qr[0]
        cur.execute("SELECT id,label FROM survey_question_options WHERE question_id=%s ORDER BY sort_order, id", (qid,))
        options = [{"id":o[0],"label":o[1]} for o in cur.fetchall()]
        questions.append({
            "id": qid, "qtype": qr[1], "title": qr[2] or "",
            "description": qr[3] or "", "required": bool(qr[4]),
            "section_id": qr[5],
            "branch_src_qid": qr[6], "branch_src_value": qr[7] or "",
            "options": options,
        })
    conn.close()
    return render_template("survey_preview.html", survey=survey, questions=questions,
                           sections=sections, preview_mode=True, active_page="survey_admin")


@app.route("/surveys/<int:sid>/stats")
@_login_required
def survey_stats(sid):
    """설문 응답 통계 페이지"""
    if not _survey_can_admin():
        flash("접근 권한이 없습니다.", "danger"); return redirect("/")
    conn = _conn(); cur = conn.cursor()
    cur.execute("""SELECT id,title,description,start_at,end_at,status,anonymous
                   FROM surveys WHERE id=%s""", (sid,))
    r = cur.fetchone()
    if not r:
        conn.close()
        flash("설문을 찾을 수 없습니다.", "danger"); return redirect("/surveys")
    survey = {"id": r[0], "title": r[1], "description": r[2] or "",
              "start_at": r[3].strftime("%Y-%m-%d %H:%M") if r[3] else "",
              "end_at": r[4].strftime("%Y-%m-%d %H:%M") if r[4] else "",
              "status": r[5], "anonymous": bool(r[6])}
    # 응답자 목록
    cur.execute("""SELECT id, respondent_name, respondent_dept, respondent_position, submitted_at
                   FROM survey_responses WHERE survey_id=%s ORDER BY submitted_at""", (sid,))
    responses = [{"id": rr[0], "name": rr[1] or "익명", "dept": rr[2] or "",
                  "position": rr[3] or "",
                  "submitted_at": rr[4].strftime("%Y-%m-%d %H:%M") if rr[4] else ""}
                 for rr in cur.fetchall()]
    resp_ids = [rv["id"] for rv in responses]
    # 질문 목록
    cur.execute("""SELECT id,qtype,title,sort_order FROM survey_questions
                   WHERE survey_id=%s ORDER BY sort_order, id""", (sid,))
    questions = []
    for qr in cur.fetchall():
        qid = qr[0]
        cur.execute("SELECT id,label FROM survey_question_options WHERE question_id=%s ORDER BY sort_order,id", (qid,))
        options = [{"id": o[0], "label": o[1]} for o in cur.fetchall()]
        q = {"id": qid, "qtype": qr[1], "title": qr[2] or "", "sort_order": qr[3], "options": options}
        # 응답 집계
        if options:
            counts = {o["id"]: 0 for o in options}
            if resp_ids:
                fmt = ",".join(["%s"] * len(resp_ids))
                cur.execute(f"""SELECT option_id, COUNT(*) FROM survey_answers
                               WHERE question_id=%s AND response_id IN ({fmt})
                               AND option_id IS NOT NULL GROUP BY option_id""",
                            [qid] + resp_ids)
                for row in cur.fetchall():
                    if row[0] in counts:
                        counts[row[0]] = row[1]
            total = sum(counts.values()) or 1
            q["option_counts"] = [{"label": o["label"], "count": counts[o["id"]],
                                   "pct": round(counts[o["id"]] / total * 100)}
                                  for o in options]
        else:
            # 텍스트 응답 목록
            if resp_ids:
                fmt = ",".join(["%s"] * len(resp_ids))
                cur.execute(f"""SELECT value_text FROM survey_answers
                               WHERE question_id=%s AND response_id IN ({fmt})
                               AND value_text IS NOT NULL AND value_text != ''""",
                            [qid] + resp_ids)
                q["text_answers"] = [row[0] for row in cur.fetchall()]
            else:
                q["text_answers"] = []
        questions.append(q)
    conn.close()
    return render_template("survey_stats.html", survey=survey, responses=responses,
                           questions=questions, active_page="survey_admin")


# ── 설문 API ──
_SURVEY_QTYPES = {"short","long","radio","checkbox","rating","select","date","file"}
_SURVEY_TARGET_TYPES = {"user", "dept", "position", "all"}
_SURVEY_CHANNELS = {"internal", "telegram", "email", "sms"}


def _send_sms_solapi(phones, msg):
    """Solapi SMS 발송. 전화번호 리스트 + 메시지. return {ok, sent, error}"""
    from urllib.error import HTTPError
    if not SOLAPI_KEY or not SOLAPI_SECRET or not SOLAPI_SENDER:
        return {"ok": False, "error": "Solapi 설정 없음"}
    cleaned = []
    for p in phones:
        p = re.sub(r"[^0-9]", "", str(p))
        if p and p not in cleaned:
            cleaned.append(p)
    if not cleaned:
        return {"ok": False, "error": "유효 번호 없음"}
    dt = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.000Z")
    salt = os.urandom(16).hex()
    sig = _hmac.new(SOLAPI_SECRET.encode("utf-8"),
                    (dt + salt).encode("utf-8"), hashlib.md5).hexdigest()
    auth_header = f"HMAC-MD5 apiKey={SOLAPI_KEY}, date={dt}, salt={salt}, signature={sig}"
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    try:
        if len(cleaned) == 1:
            body = json.dumps({"message": {"to": cleaned[0], "from": SOLAPI_SENDER, "text": msg}}).encode("utf-8")
            url = "https://api.solapi.com/messages/v4/send"
        else:
            messages = [{"to": p, "from": SOLAPI_SENDER, "text": msg} for p in cleaned]
            body = json.dumps({"messages": messages}).encode("utf-8")
            url = "https://api.solapi.com/messages/v4/send-many"
        req = Request(url, data=body,
                      headers={"Content-Type": "application/json", "Authorization": auth_header},
                      method="POST")
        with urlopen(req, context=ctx, timeout=15) as resp:
            result = json.loads(resp.read().decode("utf-8"))
        if "groupId" in result:
            return {"ok": True, "sent": result.get("count", {}).get("total", len(cleaned))}
        return {"ok": False, "error": result.get("errorMessage", "전송 실패")}
    except HTTPError as e:
        err_body = e.read().decode("utf-8")
        try:
            err_json = json.loads(err_body)
            return {"ok": False, "error": err_json.get("errorMessage", err_body)}
        except Exception:
            return {"ok": False, "error": f"HTTP {e.code}: {err_body}"}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def _survey_resolve_targets(sid, cur):
    """설문 대상자 rows → app_users.id 집합. cur는 외부 cursor (caller가 관리)."""
    cur.execute("SELECT target_type, target_value FROM survey_targets WHERE survey_id=%s", (sid,))
    targets = cur.fetchall()
    if not targets:
        return set()
    if any(t[0] == 'all' for t in targets):
        cur.execute("SELECT id FROM app_users")
        return {r[0] for r in cur.fetchall()}
    user_ids = set()
    user_vals = [t[1] for t in targets if t[0] == 'user']
    dept_vals = [t[1] for t in targets if t[0] == 'dept']
    pos_vals = [t[1] for t in targets if t[0] == 'position']
    if user_vals:
        ph = ",".join(["%s"] * len(user_vals))
        cur.execute(f"SELECT id FROM app_users WHERE e_id IN ({ph})", user_vals)
        user_ids.update(r[0] for r in cur.fetchall())
    if dept_vals:
        ph = ",".join(["%s"] * len(dept_vals))
        cur.execute(f"""SELECT DISTINCT au.id FROM app_users au
                        LEFT JOIN employee_roster er ON au.name=er.name
                        LEFT JOIN tuser t ON au.e_id=t.id
                        WHERE er.dept IN ({ph}) OR t.company IN ({ph}) OR au.company IN ({ph})""",
                    dept_vals + dept_vals + dept_vals)
        user_ids.update(r[0] for r in cur.fetchall())
    if pos_vals:
        ph = ",".join(["%s"] * len(pos_vals))
        cur.execute(f"""SELECT DISTINCT au.id FROM app_users au
                        JOIN employee_roster er ON au.name=er.name
                        WHERE er.position IN ({ph})""", pos_vals)
        user_ids.update(r[0] for r in cur.fetchall())
    return user_ids


def _survey_send(sid, user_ids, channels, kind='initial'):
    """사용자 집합에 내부/텔레그램/이메일 발송 + 이력 기록. 이미 받은 사용자는 동일 채널 중복 제외."""
    if not user_ids:
        return {"internal": 0, "telegram": 0, "email": 0}
    conn = _conn(); cur = conn.cursor()
    cur.execute("SELECT title, end_at FROM surveys WHERE id=%s", (sid,))
    row = cur.fetchone()
    if not row:
        conn.close(); return {"internal": 0, "telegram": 0, "email": 0}
    title, end_at = row
    link = f"{BASE_URL}/s/{sid}"
    prefix = "재알림" if kind == 'reminder' else "설문 요청"
    end_str = end_at.strftime("%m/%d %H:%M") if end_at else "마감일 미정"
    full_msg = f"📝 [{prefix}] {title} (~ {end_str})"

    # 이미 같은 설문·같은 채널·같은 kind로 받은 사용자 제외 (중복 방지)
    ph = ",".join(["%s"] * len(user_ids))
    cur.execute(f"""SELECT user_id, channel FROM survey_notifications
                    WHERE survey_id=%s AND user_id IN ({ph}) AND kind=%s""",
                [sid] + list(user_ids) + [kind])
    already = set()
    for uid, ch in cur.fetchall():
        already.add((uid, ch))

    sent_cnt = {"internal": 0, "telegram": 0, "email": 0, "sms": 0}

    if 'internal' in channels:
        for uid in user_ids:
            if (uid, 'internal') in already: continue
            try:
                cur.execute("INSERT INTO notifications (user_id, message, link) VALUES (%s,%s,%s)",
                            (uid, full_msg, link))
                cur.execute("INSERT INTO survey_notifications (survey_id,user_id,channel,kind) VALUES (%s,%s,'internal',%s)",
                            (sid, uid, kind))
                sent_cnt['internal'] += 1
            except Exception:
                pass

    if 'telegram' in channels:
        cur.execute(f"SELECT id, tg_chat_id FROM app_users WHERE id IN ({ph}) AND tg_chat_id<>''",
                    list(user_ids))
        rows = cur.fetchall()
        if rows:
            ctx = ssl.create_default_context()
            ctx.check_hostname = False
            ctx.verify_mode = ssl.CERT_NONE
            tg_url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
            for uid, chat_id in rows:
                if (uid, 'telegram') in already: continue
                try:
                    data = urlencode({"chat_id": chat_id, "text": full_msg}).encode()
                    urlopen(Request(tg_url, data), context=ctx, timeout=5)
                    cur.execute("INSERT INTO survey_notifications (survey_id,user_id,channel,kind) VALUES (%s,%s,'telegram',%s)",
                                (sid, uid, kind))
                    sent_cnt['telegram'] += 1
                except Exception:
                    pass

    if 'email' in channels:
        cur.execute(f"""SELECT au.id, er.email FROM app_users au
                        JOIN employee_roster er ON au.name=er.name
                        WHERE au.id IN ({ph}) AND er.email IS NOT NULL AND er.email<>''""",
                    list(user_ids))
        rows = cur.fetchall()
        for uid, email in rows:
            if (uid, 'email') in already: continue
            try:
                body_html = (f"<p>{full_msg}</p>"
                             f"<p>아래 링크에서 설문에 응답해 주세요.</p>"
                             f"<p><a href='{link}' style='background:#1e40af;color:#fff;padding:8px 18px;text-decoration:none;border-radius:4px;'>응답하기</a></p>")
                _send_email([email], f"[태인관리시스템] 설문: {title}", body_html)
                cur.execute("INSERT INTO survey_notifications (survey_id,user_id,channel,kind) VALUES (%s,%s,'email',%s)",
                            (sid, uid, kind))
                sent_cnt['email'] += 1
            except Exception:
                pass

    if 'sms' in channels:
        cur.execute(f"""SELECT au.id, er.phone FROM app_users au
                        JOIN employee_roster er ON au.name=er.name
                        WHERE au.id IN ({ph}) AND er.phone IS NOT NULL AND er.phone<>''""",
                    list(user_ids))
        rows = cur.fetchall()
        # SMS는 한 번에 다수 발송 가능 — 미발송자만 수집
        sms_targets = []
        sms_uids = []
        for uid, phone in rows:
            if (uid, 'sms') in already: continue
            sms_targets.append(phone); sms_uids.append(uid)
        if sms_targets:
            # SMS 본문 간결하게 (80자 CBT 제한 고려, 링크 별도 안내)
            sms_body = f"[설문] {title}\n마감: {end_str}\n응답: {link}"
            if len(sms_body) > 150: sms_body = sms_body[:147] + "..."
            res = _send_sms_solapi(sms_targets, sms_body)
            if res.get("ok"):
                for uid in sms_uids:
                    try:
                        cur.execute("INSERT INTO survey_notifications (survey_id,user_id,channel,kind) VALUES (%s,%s,'sms',%s)",
                                    (sid, uid, kind))
                        sent_cnt['sms'] += 1
                    except Exception: pass
            else:
                sent_cnt['sms_error'] = res.get("error", "발송 실패")

    conn.commit(); conn.close()
    return sent_cnt


@app.route("/api/survey/save_meta", methods=["POST"])
@_login_required
def api_survey_save_meta():
    if not _survey_can_admin():
        return jsonify(ok=False, error="권한 없음"), 403
    try:
        data = request.get_json() or {}
        sid = int(data["id"])
        title = str(data.get("title","")).strip()[:200] or "제목 없는 설문"
        desc = str(data.get("description","")).strip()
        start_at = data.get("start_at") or None
        end_at = data.get("end_at") or None
        anon = 1 if data.get("anonymous") else 0
        aedit = 1 if data.get("allow_edit") else 0
        once = 1 if data.get("one_per_user") else 0
        remind = int(data.get("remind_hours") or 24)
        send_email = 1 if data.get("send_email") else 0
        send_sms = 1 if data.get("send_sms") else 0
        conn = _conn(); cur = conn.cursor()
        cur.execute("""UPDATE surveys SET title=%s,description=%s,start_at=%s,end_at=%s,
                       anonymous=%s,allow_edit=%s,one_per_user=%s,remind_hours=%s,
                       send_email=%s,send_sms=%s
                       WHERE id=%s""",
                    (title, desc, start_at, end_at, anon, aedit, once, remind,
                     send_email, send_sms, sid))
        conn.commit(); conn.close()
        return jsonify(ok=True, title=title)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/api/survey/add_question", methods=["POST"])
@_login_required
def api_survey_add_question():
    if not _survey_can_admin(): return jsonify(ok=False,error="권한 없음"),403
    try:
        data = request.get_json() or {}
        sid = int(data["survey_id"])
        qtype = str(data.get("qtype","short"))
        if qtype not in _SURVEY_QTYPES:
            return jsonify(ok=False, error="잘못된 유형"), 400
        conn = _conn(); cur = conn.cursor()
        cur.execute("SELECT COALESCE(MAX(sort_order),0)+10 FROM survey_questions WHERE survey_id=%s", (sid,))
        order = cur.fetchone()[0]
        cur.execute("""INSERT INTO survey_questions (survey_id,qtype,title,sort_order)
                       VALUES (%s,%s,'',%s)""", (sid, qtype, order))
        qid = cur.lastrowid
        options = []
        if qtype in ("radio","checkbox","select"):
            cur.execute("INSERT INTO survey_question_options (question_id,label,sort_order) VALUES (%s,%s,%s)", (qid,"선택지 1",10))
            o1 = cur.lastrowid
            cur.execute("INSERT INTO survey_question_options (question_id,label,sort_order) VALUES (%s,%s,%s)", (qid,"선택지 2",20))
            o2 = cur.lastrowid
            options = [{"id":o1,"label":"선택지 1","sort_order":10},
                       {"id":o2,"label":"선택지 2","sort_order":20}]
        conn.commit(); conn.close()
        return jsonify(ok=True, id=qid, qtype=qtype, sort_order=order, options=options)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/api/survey/save_question", methods=["POST"])
@_login_required
def api_survey_save_question():
    if not _survey_can_admin(): return jsonify(ok=False,error="권한 없음"),403
    try:
        data = request.get_json() or {}
        qid = int(data["id"])
        new_qtype = str(data.get("qtype","short"))
        if new_qtype not in _SURVEY_QTYPES:
            return jsonify(ok=False,error="잘못된 유형"),400
        # 조건부 표시 / 섹션
        section_id = data.get("section_id")
        section_id = int(section_id) if section_id else None
        branch_src_qid = data.get("branch_src_qid")
        branch_src_qid = int(branch_src_qid) if branch_src_qid else None
        branch_src_value = str(data.get("branch_src_value", ""))[:500] if branch_src_qid else None
        conn = _conn(); cur = conn.cursor()
        # 기존 qtype 확인
        cur.execute("SELECT qtype FROM survey_questions WHERE id=%s", (qid,))
        row = cur.fetchone()
        if not row:
            conn.close(); return jsonify(ok=False,error="질문 없음"),404
        old_qtype = row[0]
        cur.execute("""UPDATE survey_questions SET title=%s, description=%s, required=%s, qtype=%s,
                       section_id=%s, branch_src_qid=%s, branch_src_value=%s
                       WHERE id=%s""",
                    (str(data.get("title",""))[:500],
                     str(data.get("description","")),
                     1 if data.get("required") else 0,
                     new_qtype, section_id, branch_src_qid, branch_src_value, qid))
        # qtype 변경 시 선택지 정리
        added_options = []
        was_option_type = old_qtype in ("radio","checkbox","select")
        is_option_type = new_qtype in ("radio","checkbox","select")
        if was_option_type and not is_option_type:
            cur.execute("DELETE FROM survey_question_options WHERE question_id=%s", (qid,))
        elif is_option_type and not was_option_type:
            cur.execute("INSERT INTO survey_question_options (question_id,label,sort_order) VALUES (%s,%s,%s)", (qid,"선택지 1",10))
            o1 = cur.lastrowid
            cur.execute("INSERT INTO survey_question_options (question_id,label,sort_order) VALUES (%s,%s,%s)", (qid,"선택지 2",20))
            o2 = cur.lastrowid
            added_options = [{"id":o1,"label":"선택지 1","sort_order":10},
                             {"id":o2,"label":"선택지 2","sort_order":20}]
        conn.commit(); conn.close()
        return jsonify(ok=True, qtype_changed=(old_qtype != new_qtype), added_options=added_options)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/api/survey/delete_question", methods=["POST"])
@_login_required
def api_survey_delete_question():
    if not _survey_can_admin(): return jsonify(ok=False,error="권한 없음"),403
    try:
        data = request.get_json() or {}
        qid = int(data["id"])
        conn = _conn(); cur = conn.cursor()
        cur.execute("DELETE FROM survey_question_options WHERE question_id=%s", (qid,))
        cur.execute("DELETE FROM survey_answers WHERE question_id=%s", (qid,))
        cur.execute("DELETE FROM survey_questions WHERE id=%s", (qid,))
        conn.commit(); conn.close()
        return jsonify(ok=True)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/api/survey/reorder_questions", methods=["POST"])
@_login_required
def api_survey_reorder_questions():
    if not _survey_can_admin(): return jsonify(ok=False,error="권한 없음"),403
    try:
        data = request.get_json() or {}
        ids = data.get("ids", [])
        conn = _conn(); cur = conn.cursor()
        for i, qid in enumerate(ids):
            cur.execute("UPDATE survey_questions SET sort_order=%s WHERE id=%s", ((i+1)*10, int(qid)))
        conn.commit(); conn.close()
        return jsonify(ok=True)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/api/survey/add_section", methods=["POST"])
@_login_required
def api_survey_add_section():
    if not _survey_can_admin(): return jsonify(ok=False,error="권한 없음"),403
    try:
        data = request.get_json() or {}
        sid = int(data["survey_id"])
        conn = _conn(); cur = conn.cursor()
        cur.execute("SELECT COALESCE(MAX(sort_order),0)+10 FROM survey_sections WHERE survey_id=%s", (sid,))
        order = cur.fetchone()[0]
        title = str(data.get("title", "섹션"))[:200] or "섹션"
        cur.execute("INSERT INTO survey_sections (survey_id,title,description,sort_order) VALUES (%s,%s,%s,%s)",
                    (sid, title, "", order))
        sec_id = cur.lastrowid
        conn.commit(); conn.close()
        return jsonify(ok=True, id=sec_id, title=title, sort_order=order)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/api/survey/save_section", methods=["POST"])
@_login_required
def api_survey_save_section():
    if not _survey_can_admin(): return jsonify(ok=False,error="권한 없음"),403
    try:
        data = request.get_json() or {}
        sec_id = int(data["id"])
        conn = _conn(); cur = conn.cursor()
        cur.execute("UPDATE survey_sections SET title=%s, description=%s WHERE id=%s",
                    (str(data.get("title",""))[:200],
                     str(data.get("description","")),
                     sec_id))
        conn.commit(); conn.close()
        return jsonify(ok=True)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/api/survey/delete_section", methods=["POST"])
@_login_required
def api_survey_delete_section():
    """섹션 삭제 — 소속 질문은 섹션 NULL로 (질문 자체는 보존)"""
    if not _survey_can_admin(): return jsonify(ok=False,error="권한 없음"),403
    try:
        data = request.get_json() or {}
        sec_id = int(data["id"])
        conn = _conn(); cur = conn.cursor()
        cur.execute("UPDATE survey_questions SET section_id=NULL WHERE section_id=%s", (sec_id,))
        cur.execute("DELETE FROM survey_sections WHERE id=%s", (sec_id,))
        conn.commit(); conn.close()
        return jsonify(ok=True)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/api/survey/add_option", methods=["POST"])
@_login_required
def api_survey_add_option():
    if not _survey_can_admin(): return jsonify(ok=False,error="권한 없음"),403
    try:
        data = request.get_json() or {}
        qid = int(data["question_id"])
        label = str(data.get("label","선택지"))[:300] or "선택지"
        conn = _conn(); cur = conn.cursor()
        cur.execute("SELECT COALESCE(MAX(sort_order),0)+10 FROM survey_question_options WHERE question_id=%s", (qid,))
        order = cur.fetchone()[0]
        cur.execute("INSERT INTO survey_question_options (question_id,label,sort_order) VALUES (%s,%s,%s)",
                    (qid, label, order))
        oid = cur.lastrowid
        conn.commit(); conn.close()
        return jsonify(ok=True, id=oid, label=label, sort_order=order)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/api/survey/save_option", methods=["POST"])
@_login_required
def api_survey_save_option():
    if not _survey_can_admin(): return jsonify(ok=False,error="권한 없음"),403
    try:
        data = request.get_json() or {}
        oid = int(data["id"])
        label = str(data.get("label",""))[:300]
        conn = _conn(); cur = conn.cursor()
        cur.execute("UPDATE survey_question_options SET label=%s WHERE id=%s", (label, oid))
        conn.commit(); conn.close()
        return jsonify(ok=True)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/api/survey/delete_option", methods=["POST"])
@_login_required
def api_survey_delete_option():
    if not _survey_can_admin(): return jsonify(ok=False,error="권한 없음"),403
    try:
        data = request.get_json() or {}
        oid = int(data["id"])
        conn = _conn(); cur = conn.cursor()
        cur.execute("DELETE FROM survey_question_options WHERE id=%s", (oid,))
        conn.commit(); conn.close()
        return jsonify(ok=True)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/api/survey/publish/<int:sid>", methods=["POST"])
@_login_required
def api_survey_publish(sid):
    if not _survey_can_admin(): return jsonify(ok=False,error="권한 없음"),403
    try:
        data = request.get_json() or {}
        action = str(data.get("action","publish"))
        channels = [c for c in data.get("channels", ["internal"]) if c in _SURVEY_CHANNELS]
        status_map = {"publish":"published","unpublish":"draft","close":"closed"}
        new_status = status_map.get(action, "draft")
        conn = _conn(); cur = conn.cursor()
        if action == "publish":
            cur.execute("SELECT COUNT(*) FROM survey_questions WHERE survey_id=%s", (sid,))
            if cur.fetchone()[0] == 0:
                conn.close()
                return jsonify(ok=False, error="질문이 최소 1개 이상 있어야 발행할 수 있습니다"), 400
            # 대상자 확인
            user_ids = _survey_resolve_targets(sid, cur)
            if not user_ids:
                conn.close()
                return jsonify(ok=False, error="대상자가 지정되지 않았습니다. 대상자를 먼저 설정하세요."), 400
        cur.execute("UPDATE surveys SET status=%s WHERE id=%s", (new_status, sid))
        conn.commit(); conn.close()
        # 발행 시 자동 발송
        sent = None
        if action == "publish" and channels:
            sent = _survey_send(sid, user_ids, channels, kind='initial')
        return jsonify(ok=True, status=new_status, sent=sent)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/api/survey/target_options")
@_login_required
def api_survey_target_options():
    """대상자 선택 UI용 옵션 리스트 (부서/회사/직급/직원)"""
    if not _survey_can_admin(): return jsonify(ok=False,error="권한 없음"),403
    try:
        conn = _conn(); cur = conn.cursor()
        # 부서/직급 고유값
        cur.execute("SELECT DISTINCT dept FROM employee_roster WHERE dept IS NOT NULL AND dept<>'' ORDER BY dept")
        depts = [r[0] for r in cur.fetchall()]
        cur.execute("SELECT DISTINCT position FROM employee_roster WHERE position IS NOT NULL AND position<>'' ORDER BY position")
        positions = [r[0] for r in cur.fetchall()]
        # 직원 리스트 (e_id 기반) — 부서·직급도 포함
        cur.execute("""SELECT au.e_id, au.name,
                              COALESCE(er.dept, t.company, '') AS dept,
                              COALESCE(er.position, '') AS position
                       FROM app_users au
                       LEFT JOIN tuser t ON au.e_id=t.id
                       LEFT JOIN employee_roster er ON au.name=er.name
                       WHERE au.e_id IS NOT NULL
                       ORDER BY au.name""")
        employees = [{"e_id": r[0], "name": r[1], "dept": r[2] or "", "position": r[3] or ""}
                     for r in cur.fetchall()]
        conn.close()
        return jsonify(ok=True, depts=depts, positions=positions, employees=employees)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/api/survey/save_targets", methods=["POST"])
@_login_required
def api_survey_save_targets():
    """survey_targets 전면 교체"""
    if not _survey_can_admin(): return jsonify(ok=False,error="권한 없음"),403
    try:
        data = request.get_json() or {}
        sid = int(data["survey_id"])
        targets = data.get("targets", [])  # [{type,value}, ...]
        conn = _conn(); cur = conn.cursor()
        cur.execute("DELETE FROM survey_targets WHERE survey_id=%s", (sid,))
        for t in targets:
            tt = str(t.get("type",""))
            tv = str(t.get("value","")) if t.get("value") is not None else None
            if tt not in _SURVEY_TARGET_TYPES: continue
            if tt == 'all':
                cur.execute("INSERT INTO survey_targets (survey_id,target_type,target_value) VALUES (%s,'all',NULL)", (sid,))
            else:
                if not tv: continue
                cur.execute("INSERT INTO survey_targets (survey_id,target_type,target_value) VALUES (%s,%s,%s)",
                            (sid, tt, tv))
        # 바로 매칭 수 계산
        user_ids = _survey_resolve_targets(sid, cur)
        conn.commit(); conn.close()
        return jsonify(ok=True, target_count=len(user_ids))
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/api/survey/targets/<int:sid>")
@_login_required
def api_survey_get_targets(sid):
    if not _survey_can_admin(): return jsonify(ok=False,error="권한 없음"),403
    try:
        conn = _conn(); cur = conn.cursor()
        cur.execute("SELECT target_type, target_value FROM survey_targets WHERE survey_id=%s", (sid,))
        targets = [{"type": r[0], "value": r[1]} for r in cur.fetchall()]
        user_ids = _survey_resolve_targets(sid, cur)
        conn.close()
        return jsonify(ok=True, targets=targets, target_count=len(user_ids))
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/api/survey/send/<int:sid>", methods=["POST"])
@_login_required
def api_survey_send(sid):
    """수동 발송: 현재 대상자 전원에게 선택 채널로 전송."""
    if not _survey_can_admin(): return jsonify(ok=False,error="권한 없음"),403
    try:
        data = request.get_json() or {}
        channels = [c for c in data.get("channels", ["internal"]) if c in _SURVEY_CHANNELS]
        kind = str(data.get("kind", "initial"))
        if kind not in ("initial", "reminder"): kind = "initial"
        conn = _conn(); cur = conn.cursor()
        user_ids = _survey_resolve_targets(sid, cur)
        conn.close()
        if not user_ids:
            return jsonify(ok=False, error="대상자가 없습니다"), 400
        sent = _survey_send(sid, user_ids, channels, kind=kind)
        return jsonify(ok=True, sent=sent, target_count=len(user_ids))
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/api/survey/notifications/<int:sid>")
@_login_required
def api_survey_notifications(sid):
    """발송 이력 조회 (최근 200건)"""
    if not _survey_can_admin(): return jsonify(ok=False,error="권한 없음"),403
    try:
        conn = _conn(); cur = conn.cursor()
        cur.execute("""SELECT sn.id, sn.user_id, au.name, sn.channel, sn.kind, sn.sent_at,
                              (SELECT COUNT(*) FROM survey_responses sr WHERE sr.survey_id=sn.survey_id AND sr.respondent_id=sn.user_id) AS responded
                       FROM survey_notifications sn
                       LEFT JOIN app_users au ON sn.user_id=au.id
                       WHERE sn.survey_id=%s
                       ORDER BY sn.sent_at DESC, sn.id DESC
                       LIMIT 200""", (sid,))
        items = [{"id": r[0], "user_id": r[1], "user_name": r[2] or "",
                  "channel": r[3], "kind": r[4],
                  "sent_at": r[5].strftime("%Y-%m-%d %H:%M") if r[5] else "",
                  "responded": bool(r[6])} for r in cur.fetchall()]
        conn.close()
        return jsonify(ok=True, items=items)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/api/cron/survey_reminders", methods=["POST", "GET"])
def api_cron_survey_reminders():
    """cron 호출용 — 마감 remind_hours 이내의 진행중 설문에 미응답자에게 재알림.
    보안: CRON_TOKEN 환경변수 또는 localhost에서만 허용."""
    token = request.args.get("token") or (request.get_json(silent=True) or {}).get("token", "")
    expected = os.environ.get("CRON_TOKEN", "")
    remote_ip = request.remote_addr or ""
    if expected and token != expected and not remote_ip.startswith("127."):
        return jsonify(ok=False, error="unauthorized"), 403
    try:
        conn = _conn(); cur = conn.cursor()
        cur.execute("""SELECT id, title, end_at, remind_hours FROM surveys
                       WHERE status='published' AND end_at IS NOT NULL
                         AND end_at > NOW()
                         AND remind_hours > 0
                         AND TIMESTAMPDIFF(HOUR, NOW(), end_at) <= remind_hours""")
        surveys = cur.fetchall()
        totals = {"surveys": 0, "internal": 0, "telegram": 0, "email": 0, "recipients": 0}
        for sid, title, end_at, remind_hours in surveys:
            all_targets = _survey_resolve_targets(sid, cur)
            if not all_targets: continue
            # 미응답자 필터
            ph = ",".join(["%s"] * len(all_targets))
            cur.execute(f"""SELECT respondent_id FROM survey_responses
                            WHERE survey_id=%s AND respondent_id IN ({ph})""",
                        [sid] + list(all_targets))
            responded = {r[0] for r in cur.fetchall() if r[0]}
            pending = all_targets - responded
            if not pending: continue
            sent = _survey_send(sid, pending, ['internal', 'telegram', 'email'], kind='reminder')
            totals['surveys'] += 1
            totals['internal'] += sent['internal']
            totals['telegram'] += sent['telegram']
            totals['email'] += sent['email']
            totals['recipients'] += len(pending)
        conn.close()
        return jsonify(ok=True, result=totals)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/api/survey/delete/<int:sid>", methods=["POST"])
@_login_required
def api_survey_delete(sid):
    if not _survey_can_admin(): return jsonify(ok=False,error="권한 없음"),403
    try:
        conn = _conn(); cur = conn.cursor()
        cur.execute("DELETE FROM survey_answers WHERE response_id IN (SELECT id FROM survey_responses WHERE survey_id=%s)", (sid,))
        cur.execute("DELETE FROM survey_responses WHERE survey_id=%s", (sid,))
        cur.execute("DELETE FROM survey_question_options WHERE question_id IN (SELECT id FROM survey_questions WHERE survey_id=%s)", (sid,))
        cur.execute("DELETE FROM survey_questions WHERE survey_id=%s", (sid,))
        cur.execute("DELETE FROM survey_sections WHERE survey_id=%s", (sid,))
        cur.execute("DELETE FROM survey_targets WHERE survey_id=%s", (sid,))
        cur.execute("DELETE FROM survey_notifications WHERE survey_id=%s", (sid,))
        cur.execute("DELETE FROM surveys WHERE id=%s", (sid,))
        conn.commit(); conn.close()
        return jsonify(ok=True)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/s/<int:sid>")
def survey_short_link(sid):
    """SMS 단축 링크 → 로그인 후 해당 설문으로 이동"""
    next_url = f"/my/surveys/{sid}"
    if not session.get("user_id"):
        return redirect(f"/login?next={next_url}")
    return redirect(next_url)


@app.route("/my/surveys")
@_login_required
def surveys_my():
    """로그인 사용자가 응답 가능한 설문 목록 (대상자 매칭)."""
    uid = session.get("user_id")
    eid = session.get("e_id")
    conn = _conn(); cur = conn.cursor()
    # 사용자 부서/회사/직급 조회
    user_dept = user_company = user_position = ""
    cur.execute("""SELECT COALESCE(er.dept, t.company, a.company) AS dept, t.company AS comp, er.position
                   FROM app_users a
                   LEFT JOIN tuser t ON a.e_id = t.id
                   LEFT JOIN employee_roster er ON a.name = er.name
                   WHERE a.id=%s""", (uid,))
    ur = cur.fetchone()
    if ur:
        user_dept = (ur[0] or "").strip()
        user_company = (ur[1] or "").strip()
        user_position = (ur[2] or "").strip()
    # 진행 중 설문 조회 + 대상자 매칭
    cur.execute("""SELECT s.id, s.title, s.description, s.start_at, s.end_at, s.anonymous,
                          s.allow_edit, s.created_by_name
                   FROM surveys s
                   WHERE s.status='published'
                     AND (s.start_at IS NULL OR s.start_at <= NOW())
                     AND (s.end_at IS NULL OR s.end_at >= NOW())
                     AND s.is_template=0
                   ORDER BY s.end_at IS NULL, s.end_at, s.id DESC""")
    raw = cur.fetchall()
    my_list = []
    for r in raw:
        sid = r[0]
        cur.execute("SELECT target_type, target_value FROM survey_targets WHERE survey_id=%s", (sid,))
        targets = cur.fetchall()
        matched = False
        for tt, tv in targets:
            if tt == 'all':
                matched = True; break
            if tt == 'user' and str(tv) == str(eid):
                matched = True; break
            if tt == 'dept' and tv == user_dept:
                matched = True; break
            if tt == 'company' and tv == user_company:
                matched = True; break
            if tt == 'position' and tv == user_position:
                matched = True; break
        if not matched:
            continue
        # 이미 응답했는지
        cur.execute("SELECT id, is_complete FROM survey_responses WHERE survey_id=%s AND respondent_id=%s",
                    (sid, uid))
        rr = cur.fetchone()
        my_list.append({
            "id": sid, "title": r[1], "description": r[2] or "",
            "start_at": r[3].strftime("%Y-%m-%d %H:%M") if r[3] else "",
            "end_at": r[4].strftime("%Y-%m-%d %H:%M") if r[4] else "",
            "anonymous": bool(r[5]),
            "allow_edit": bool(r[6]),
            "created_by_name": r[7],
            "response_id": rr[0] if rr else None,
            "responded": bool(rr),
        })
    conn.close()
    return render_template("survey_my.html",
                           surveys=my_list,
                           active_page="survey_my")


@app.route("/my/surveys/<int:sid>")
@_login_required
def survey_respond(sid):
    """설문 응답 페이지 — 로그인 사용자가 직접 응답"""
    uid = session.get("user_id")
    eid = session.get("e_id")
    conn = _conn(); cur = conn.cursor()
    cur.execute("""SELECT id,title,description,start_at,end_at,status,anonymous,allow_edit
                   FROM surveys WHERE id=%s""", (sid,))
    r = cur.fetchone()
    if not r or r[5] != 'published':
        conn.close()
        flash("설문을 찾을 수 없거나 진행 중이 아닙니다.", "danger")
        return redirect("/my/surveys")
    survey = {"id": r[0], "title": r[1], "description": r[2] or "",
              "start_at": r[3].strftime("%Y-%m-%d %H:%M") if r[3] else "",
              "end_at": r[4].strftime("%Y-%m-%d %H:%M") if r[4] else "",
              "status": r[5], "anonymous": bool(r[6]), "allow_edit": bool(r[7])}
    # 기존 응답 확인
    cur.execute("SELECT id, is_complete FROM survey_responses WHERE survey_id=%s AND respondent_id=%s",
                (sid, uid))
    rr = cur.fetchone()
    if rr and not survey["allow_edit"]:
        conn.close()
        flash("이미 응답하셨습니다.", "info")
        return redirect("/my/surveys")
    response_id = rr[0] if rr else None
    # 기존 답변 로드
    existing_answers = {}
    if response_id:
        cur.execute("""SELECT question_id, option_id, value_text, value_number
                       FROM survey_answers WHERE response_id=%s""", (response_id,))
        for aw in cur.fetchall():
            qid = aw[0]
            if qid not in existing_answers:
                existing_answers[qid] = {"option_ids": [], "value_text": aw[2], "value_number": aw[3]}
            if aw[1]:
                existing_answers[qid]["option_ids"].append(aw[1])
    # 섹션 및 질문
    cur.execute("""SELECT id,title,description,sort_order FROM survey_sections
                   WHERE survey_id=%s ORDER BY sort_order, id""", (sid,))
    sections = [{"id": r2[0], "title": r2[1] or "", "description": r2[2] or "", "sort_order": r2[3]}
                for r2 in cur.fetchall()]
    cur.execute("""SELECT id,qtype,title,description,required,section_id,
                          branch_src_qid,branch_src_value,sort_order
                   FROM survey_questions WHERE survey_id=%s ORDER BY sort_order, id""", (sid,))
    questions = []
    for qr in cur.fetchall():
        qid2 = qr[0]
        cur.execute("SELECT id,label FROM survey_question_options WHERE question_id=%s ORDER BY sort_order, id", (qid2,))
        options = [{"id": o[0], "label": o[1]} for o in cur.fetchall()]
        q = {"id": qid2, "qtype": qr[1], "title": qr[2] or "",
             "description": qr[3] or "", "required": bool(qr[4]),
             "section_id": qr[5],
             "branch_src_qid": qr[6], "branch_src_value": qr[7] or "",
             "options": options}
        if qid2 in existing_answers:
            q["_existing"] = existing_answers[qid2]
        questions.append(q)
    conn.close()
    return render_template("survey_preview.html", survey=survey, questions=questions,
                           sections=sections, preview_mode=False,
                           response_id=response_id, active_page="survey_my")


@app.route("/api/survey/respond/<int:sid>", methods=["POST"])
@_login_required
def api_survey_respond(sid):
    """설문 응답 저장 API"""
    uid = session.get("user_id")
    eid = session.get("e_id")
    try:
        data = request.get_json(force=True)
        answers = data.get("answers", {})  # {question_id_str: {option_ids, value_text}}
        conn = _conn(); cur = conn.cursor()
        # 설문 확인
        cur.execute("SELECT id, status, allow_edit FROM surveys WHERE id=%s", (sid,))
        sr = cur.fetchone()
        if not sr or sr[1] != 'published':
            conn.close()
            return jsonify(ok=False, error="진행 중인 설문이 아닙니다"), 400
        # 사용자 정보
        name = dept = position = ""
        cur.execute("""SELECT a.name, COALESCE(er.dept,''), COALESCE(er.position,'')
                       FROM app_users a
                       LEFT JOIN employee_roster er ON a.name=er.name
                       WHERE a.id=%s""", (uid,))
        ur = cur.fetchone()
        if ur:
            name, dept, position = ur[0], ur[1], ur[2]
        # 기존 응답 확인
        cur.execute("SELECT id FROM survey_responses WHERE survey_id=%s AND respondent_id=%s", (sid, uid))
        rr = cur.fetchone()
        if rr:
            if not sr[2]:
                conn.close()
                return jsonify(ok=False, error="이미 응답하셨습니다"), 400
            resp_id = rr[0]
            cur.execute("DELETE FROM survey_answers WHERE response_id=%s", (resp_id,))
            cur.execute("UPDATE survey_responses SET submitted_at=NOW(), is_complete=1 WHERE id=%s", (resp_id,))
        else:
            cur.execute("""INSERT INTO survey_responses
                           (survey_id, respondent_id, respondent_name, respondent_dept, respondent_position, is_complete)
                           VALUES (%s,%s,%s,%s,%s,1)""", (sid, uid, name, dept, position))
            resp_id = cur.lastrowid
        # 답변 저장
        for qid_str, ans in answers.items():
            qid = int(qid_str)
            option_ids = ans.get("option_ids", [])
            value_text = ans.get("value_text", "")
            value_number = ans.get("value_number")
            if option_ids:
                for oid in option_ids:
                    cur.execute("""INSERT INTO survey_answers (response_id, question_id, option_id, value_text)
                                   VALUES (%s,%s,%s,%s)""", (resp_id, qid, oid, value_text))
            else:
                cur.execute("""INSERT INTO survey_answers (response_id, question_id, value_text, value_number)
                               VALUES (%s,%s,%s,%s)""", (resp_id, qid, value_text or "", value_number))
        conn.commit(); conn.close()
        return jsonify(ok=True)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/reset_user_pw", methods=["POST"])
@_admin_required
def reset_user_pw():
    try:
        data = request.get_json()
        uid = int(data["id"])
        conn = _conn(); cur = conn.cursor()
        cur.execute("SELECT login_id FROM app_users WHERE id=%s", (uid,))
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({"ok": False, "error": "사용자 없음"}), 404
        cur.execute("UPDATE app_users SET password=%s, must_change_pw=1 WHERE id=%s",
                    (_hash_pw(row[0]), uid))
        conn.commit(); conn.close()
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/api/add_user", methods=["POST"])
@_admin_required
def api_add_user():
    """신규 재직자 등록"""
    try:
        data = request.get_json()
        name = str(data.get("name", "")).strip()
        login_id = str(data.get("login_id", "")).strip()
        password = str(data.get("password", "")).strip()
        company = str(data.get("company", "")).strip()
        role = str(data.get("role", "user")).strip()
        if not name or not login_id or not password:
            return jsonify(ok=False, error="필수값 누락"), 400
        if role not in ("user", "admin"):
            role = "user"
        conn = _conn(); cur = conn.cursor()
        cur.execute("SELECT id FROM app_users WHERE login_id=%s", (login_id,))
        if cur.fetchone():
            conn.close()
            return jsonify(ok=False, error="이미 존재하는 로그인ID입니다"), 400
        hashed = _hash_pw(password)
        cur.execute("""INSERT INTO app_users (login_id, password, name, role, company, created_at)
                       VALUES (%s, %s, %s, %s, %s, NOW())""",
                    (login_id, hashed, name, role, company))
        conn.commit(); conn.close()
        return jsonify(ok=True)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/delete_user", methods=["POST"])
@_admin_required
def delete_user():
    try:
        data = request.get_json()
        uid = int(data["id"])
        if uid == session.get("user_id"):
            return jsonify({"ok": False, "error": "본인 계정은 삭제할 수 없습니다"}), 400
        conn = _conn(); cur = conn.cursor()
        cur.execute("DELETE FROM app_users WHERE id=%s AND role<>'admin'", (uid,))
        conn.commit(); conn.close()
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/")
@_login_required
def dashboard():
    """dashboard"""
    from datetime import date as _date
    today = _date.today()
    today_str = today.strftime("%Y%m%d")
    today_month = today.month
    today_day = today.day
    is_admin = session.get("role") == "admin"

    notices = []
    att_summary = {"total": 0, "checked_in": 0, "not_in": 0, "depts": {}}
    leave_today = []
    etc_today = []
    birthdays = []
    my_schedule = []
    wr_summary = None

    try:
        conn = _conn(); cur = conn.cursor()
        # 공지사항 (고정 우선, 최근 10건)
        cur.execute("SELECT id, title, content, is_pinned, created_by, created_at FROM board_notices ORDER BY is_pinned DESC, created_at DESC LIMIT 10")
        notices = [{"id": r[0], "title": r[1], "content": r[2], "is_pinned": r[3], "created_by": r[4], "created_at": r[5]} for r in cur.fetchall()]

        # 금일 출근현황 (근태기록기 출근 타각자만)
        cur.execute("""
            SELECT u.name, u.company,
                   MIN(t.e_time) in_time,
                   IFNULL(er.dept, '') as er_dept
            FROM tenter t
            JOIN tuser u ON t.e_id=u.id
            LEFT JOIN employee_roster er ON er.name=u.name
            WHERE t.e_date=%s AND t.e_mode=1 AND t.e_id>0 AND t.e_result='0'
            GROUP BY t.e_id, u.name, u.company, er.dept
        """, (today_str,))
        att_rows = cur.fetchall()
        # 임원 포함 (헤더 출근자 수와 동일 기준으로 맞춤)
        EXCLUDE_DEPTS = set()
        # 임원·경영기획팀은 별도 부서로 두지 않고 전기사무로 합산 (헤더 인원수와 일치)
        MERGE_DEPTS = {'임원': '전기사무', '경영기획팀': '전기사무',
                       '경영기획실': '전기사무', '비서팀': '전기사무'}
        dept_map = {}
        checked_in_count = 0
        for r in att_rows:
            er_dept = (r[3] or '').strip()
            dept = DEPT_MAP.get((r[1] or '').strip(), '')
            if not dept:
                dept = '기타'
            dept = MERGE_DEPTS.get(dept, dept)
            if dept in EXCLUDE_DEPTS or er_dept in EXCLUDE_DEPTS:
                continue
            checked_in_count += 1
            if dept not in dept_map:
                dept_map[dept] = {"in": 0}
            dept_map[dept]["in"] += 1
        att_summary["checked_in"] = checked_in_count
        DEPT_ORDER = ["관리부", "전기사무", "전기생산", "전자사무", "전자생산"]
        ordered_depts = {}
        for d in DEPT_ORDER:
            if d in dept_map:
                ordered_depts[d] = dept_map[d]
        for d in dept_map:
            if d not in ordered_depts:
                ordered_depts[d] = dept_map[d]
        att_summary["depts"] = ordered_depts

        # 금일 연차자
        cur.execute("""
            SELECT u.name, u.company, lr.leave_type
            FROM leave_records lr
            JOIN tuser u ON lr.e_id=u.id
            WHERE lr.leave_date=%s
            ORDER BY u.name
        """, (today_str,))
        leave_today = [{"name": r[0], "dept": DEPT_MAP.get((r[1] or '').strip(), '기타'), "type": r[2]} for r in cur.fetchall()]

        # 기타 근무신청자 (금일 보고서 기준) - 연차만 휴가자로 합산
        cur.execute("""
            SELECT e.user_name, e.etc_value
            FROM wr_entries e
            JOIN wr_reports r ON r.id=e.report_id
            WHERE e.category='기타' AND e.skipped=0
              AND r.report_date = %s
            ORDER BY e.user_name
        """, (today_str,))
        _leave_names = {lv["name"] for lv in leave_today}
        etc_today = []
        for row in cur.fetchall():
            name, etype = row[0], (row[1] or '')
            if etype == '연차':
                if name not in _leave_names:
                    leave_today.append({"name": name, "dept": '', "type": etype})
                    _leave_names.add(name)
            else:
                etc_today.append({"name": name, "type": etype})
        leave_today.sort(key=lambda x: x["name"])

        # 이번 달 생일자
        cur.execute("""
            SELECT er.name, er.dept, er.birth_date, IFNULL(er.position, '')
            FROM employee_roster er
            WHERE er.birth_date IS NOT NULL AND er.birth_date <> ''
              AND er.name IS NOT NULL AND er.name <> ''
        """)
        for r in cur.fetchall():
            bday = str(r[2]).replace('-', '')
            if len(bday) >= 4:
                try:
                    mm = int(bday[4:6]); dd = int(bday[6:8])
                    if mm == today_month:
                        pos = (r[3] or '').strip()
                        display_name = f"{r[0]} {pos}" if pos and pos != '사원' else f"{r[0]} 사원"
                        birthdays.append({"name": display_name, "dept": r[1] or '', "day": dd, "is_today": dd == today_day})
                except (ValueError, IndexError):
                    pass
        birthdays.sort(key=lambda x: x["day"])

        # 나의 근무일정 (이번 달) - 근무보고서 그룹 멤버만
        my_schedule = []
        login_name = session.get("user_name", "")
        login_uid = session.get("user_id")
        if login_name:
            cur.execute("SELECT COUNT(*) FROM wr_group_members WHERE user_id=%s", (login_uid,))
            is_member = cur.fetchone()[0] > 0
            if is_member:
                month_str_my = today.strftime("%Y%m")
                cur.execute("""
                    SELECT work_date, basic_h, ot_h, night_h, etc, source_type
                    FROM schedule_record
                    WHERE emp_name=%s AND work_date LIKE %s
                      AND source_type IN ('보고서','수정','수정근무보고서')
                    ORDER BY work_date, FIELD(source_type,'보고서','수정','수정근무보고서')
                """, (login_name, f"{month_str_my}%"))
                def _sr_cat(bh, oh, nh, etc):
                    if etc == "스킵": return "스킵"
                    if etc in ("연차", "무급"): return etc
                    if etc and etc not in ('0', ''):
                        try:
                            float(etc)
                            return "공제"
                        except (ValueError, TypeError):
                            return etc
                    if bh > 0 and oh > 0 and nh > 0: return "야간연장"
                    if bh > 0 and nh > 0: return "야간기본"
                    if bh > 0 and oh > 0: return "주간연장"
                    if bh > 0: return "주간기본"
                    return "근무"
                from collections import OrderedDict
                date_records = OrderedDict()
                for wd, bh, oh, nh, etc, src in cur.fetchall():
                    wd_s = str(wd)
                    disp = f"{wd_s[:4]}-{wd_s[4:6]}-{wd_s[6:8]}" if len(wd_s) == 8 else wd_s
                    cat = _sr_cat(bh, oh, nh, etc)
                    if disp not in date_records:
                        date_records[disp] = {}
                    date_records[disp][src] = cat
                for disp, recs in date_records.items():
                    orig = recs.get("보고서")
                    rev = recs.get("수정근무보고서")
                    if orig and rev:
                        my_schedule.append({"date": disp, "orig_cat": orig, "cat": rev, "is_revision": True})
                    elif rev:
                        my_schedule.append({"date": disp, "orig_cat": None, "cat": rev, "is_revision": True})
                    else:
                        my_schedule.append({"date": disp, "orig_cat": None, "cat": orig, "is_revision": False})

        # 근무보고서 현황 (관리자만)
        if is_admin:
            cur_month = today.strftime("%Y%m")
            # 보고서 상태별 건수
            cur.execute("""
                SELECT r.status, COUNT(*) FROM wr_reports r
                WHERE r.report_date LIKE %s
                GROUP BY r.status
            """, (f"{cur_month}%",))
            wr_status_counts = {}
            for st, cnt in cur.fetchall():
                wr_status_counts[st] = cnt
            # 그룹별 보고서 현황
            cur.execute("""
                SELECT g.group_name, r.status, COUNT(*) cnt
                FROM wr_reports r JOIN wr_groups g ON r.group_id = g.id
                WHERE r.report_date LIKE %s
                GROUP BY g.group_name, r.status
                ORDER BY g.group_name
            """, (f"{cur_month}%",))
            wr_group_detail = {}
            for gname, st, cnt in cur.fetchall():
                if gname not in wr_group_detail:
                    wr_group_detail[gname] = {}
                wr_group_detail[gname][st] = cnt
            # 이번달 전체 보고서의 엔트리 인원 현황
            cur.execute("""
                SELECT
                    COUNT(DISTINCT CASE WHEN e.skipped=1 THEN CONCAT(r.id,'-',e.user_id) END) skip_cnt,
                    COUNT(DISTINCT CASE WHEN e.skipped=0 AND e.category='휴무' AND e.etc_value='연차' THEN CONCAT(r.id,'-',e.user_id) END) leave_cnt,
                    COUNT(DISTINCT CASE WHEN e.skipped=0 AND e.category='기타' AND e.etc_value='무급' THEN CONCAT(r.id,'-',e.user_id) END) unpaid_cnt,
                    COUNT(DISTINCT CASE WHEN e.skipped=0 AND e.category IN ('주간기본','주간연장','야간기본','야간연장') THEN CONCAT(r.id,'-',e.user_id) END) work_cnt,
                    COUNT(DISTINCT CASE WHEN e.skipped=0 AND e.deduction>0 THEN CONCAT(r.id,'-',e.user_id) END) deduction_cnt
                FROM wr_entries e
                JOIN wr_reports r ON e.report_id = r.id
                WHERE r.report_date LIKE %s
            """, (f"{cur_month}%",))
            erow = cur.fetchone()
            wr_summary = {
                "status": wr_status_counts,
                "total": sum(wr_status_counts.values()),
                "groups": wr_group_detail,
                "work": erow[3] or 0,
                "leave": erow[1] or 0,
                "unpaid": erow[2] or 0,
                "skip": erow[0] or 0,
                "deduction": erow[4] or 0,
            }

        # 전력 현황 (smartview 실시간 + 오늘 사용량)
        power_info = None
        try:
            cur.execute("""
                SELECT rt_kwh, rt_bill, bill_start, bill_end, updated_at
                FROM power_smartview WHERE cust_no=%s
                ORDER BY updated_at DESC LIMIT 1
            """, (KEPCO_CUST,))
            sv = cur.fetchone()
            cur.execute("""
                SELECT COALESCE(SUM(kwh), 0) FROM power_usage_raw
                WHERE cust_no=%s AND DATE(measured_at) = CURDATE()
            """, (KEPCO_CUST,))
            td = cur.fetchone()
            cur.execute("""
                SELECT HOUR(measured_at) AS hr, ROUND(SUM(kwh), 2)
                FROM power_usage_raw
                WHERE cust_no=%s AND DATE(measured_at) = CURDATE()
                GROUP BY hr ORDER BY hr
            """, (KEPCO_CUST,))
            hr_map = {int(r[0]): float(r[1]) for r in cur.fetchall()}
            hourly_kwh = [hr_map.get(h, 0) for h in range(24)]
            if sv:
                def _fmt_bill_date(s):
                    # 'YYYY.MM.DD' → 'MM/DD'
                    if not s: return "-"
                    s = str(s)
                    return s[5:].replace('.', '/') if len(s) >= 7 else s
                power_info = {
                    "rt_kwh":     round(float(sv[0]), 1),
                    "rt_bill":    int(sv[1]),
                    "bill_start": _fmt_bill_date(sv[2]),
                    "bill_end":   _fmt_bill_date(sv[3]),
                    "updated_at": sv[4].strftime("%H:%M") if sv[4] else "-",
                    "today_kwh":  round(float(td[0]), 1) if td else 0.0,
                    "hourly_kwh": hourly_kwh,
                }
        except Exception:
            pass

        # 오늘 일정 건수
        today_events_count = 0
        try:
            today_str_dash = today.strftime("%Y-%m-%d")
            cur.execute("SELECT COUNT(*) FROM schedule_events WHERE event_date = %s", (today_str_dash,))
            r = cur.fetchone()
            today_events_count = r[0] if r else 0
        except Exception:
            pass

        conn.close()
    except Exception as e:
        flash(f"대시보드 오류: {e}", "danger")

    pending_reports = 0
    if wr_summary:
        s = wr_summary.get("status", {})
        pending_reports = s.get("대기", 0) + s.get("작성완료", 0) + s.get("검토", 0)

    # 이번 주 생일자 (월~일)
    from datetime import timedelta as _td
    week_days = set()
    _wd = today.weekday()  # 0=Mon
    for i in range(7):
        d = today + _td(days=(i - _wd))
        week_days.add(d.day) if d.month == today.month else None

    # 이번 주(월~일) 7일간 식수 데이터
    meal_week = []
    try:
        from datetime import timedelta as _td2
        conn2 = _conn(); cur2 = conn2.cursor()
        DOW_KR2 = ["월","화","수","목","금","토","일"]
        _monday = today - _td2(days=today.weekday())  # 이번 주 월요일
        for i in range(7):
            d = _monday + _td2(days=i)
            ym = d.strftime("%Y-%m")
            cur2.execute("SELECT dept, meal_type, `count` FROM meal_count WHERE `year_month`=%s AND `day`=%s", (ym, d.day))
            day_meals = {}
            for dept, mtype, cnt in cur2.fetchall():
                day_meals.setdefault(mtype, {})[dept] = cnt
            hol = KR_HOLIDAYS.get(d, "")
            is_hol = d.weekday() >= 5 or bool(hol)
            meal_week.append({"date": d.strftime("%m/%d"), "dow": DOW_KR2[d.weekday()],
                              "day": d.day, "is_today": d == today,
                              "is_holiday": is_hol, "meals": day_meals})
        conn2.close()
    except Exception as me:
        print(f"[dashboard] meal_week error: {me}")

    can_notice = is_admin or _has_perm("notice")
    can_schedule = is_admin or _has_perm("schedule")
    return render_template("dashboard.html", notices=notices, meal_week=meal_week,
                           att_summary=att_summary, leave_today=leave_today,
                           etc_today=etc_today,
                           birthdays=birthdays, today=today, can_notice=can_notice,
                           can_schedule=can_schedule,
                           my_schedule=my_schedule,
                           wr_summary=wr_summary if is_admin else None,
                           is_admin=is_admin,
                           power_info=power_info,
                           today_events_count=today_events_count,
                           pending_reports=pending_reports,
                           bday_week_days=week_days)


@app.route("/save_notice", methods=["POST"])
@_login_required
def save_notice():
    if session.get("role") != "admin" and not _has_perm("notice"):
        return jsonify(ok=False, msg="\uad8c\ud55c\uc774 \uc5c6\uc2b5\ub2c8\ub2e4")
    nid = request.form.get("id", "").strip()
    title = request.form.get("title", "").strip()
    content = request.form.get("content", "").strip()
    is_pinned = 1 if request.form.get("is_pinned") else 0
    if not title:
        return jsonify(ok=False, msg="제목을 입력하세요")
    conn = _conn(); cur = conn.cursor()
    if nid:
        cur.execute("UPDATE board_notices SET title=%s, content=%s, is_pinned=%s WHERE id=%s",
                     (title, content, is_pinned, nid))
    else:
        cur.execute("INSERT INTO board_notices (title, content, is_pinned, created_by) VALUES (%s,%s,%s,%s)",
                     (title, content, is_pinned, session.get("user_name", "")))
        writer = session.get("user_name", "")
        tg_msg = f"📢 공지사항 등록\n제목: {title}\n작성자: {writer}"
        if content:
            tg_msg += f"\n내용: {content[:200]}"
        _send_telegram(tg_msg, "notice")
    conn.commit(); conn.close()
    return jsonify(ok=True)


@app.route("/delete_notice", methods=["POST"])
@_login_required
def delete_notice():
    if session.get("role") != "admin" and not _has_perm("notice"):
        return jsonify(ok=False, msg="\uad8c\ud55c\uc774 \uc5c6\uc2b5\ub2c8\ub2e4")

    nid = request.form.get("id")
    conn = _conn(); cur = conn.cursor()
    cur.execute("DELETE FROM board_notices WHERE id=%s", (nid,))
    conn.commit(); conn.close()
    return jsonify(ok=True)


# 날씨 + 미세먼지 API (충북 청주)
@app.route("/api/weather")
@_login_required
def api_weather():
    """기상청 단기예보 + 에어코리아 미세먼지"""
    import json
    from datetime import datetime as dt
    
    # 청주 좌표 (기상청 격자)
    nx, ny = 69, 106  # 청주시
    
    now = dt.now()
    # 기상청 API는 매시 45분 이후 발표 -> 이전 시간 사용
    base_date = now.strftime("%Y%m%d")
    if now.minute < 45:
        base_hour = now.hour - 1
        if base_hour < 0:
            base_hour = 23
            base_date = (now - timedelta(days=1)).strftime("%Y%m%d")
    else:
        base_hour = now.hour
    base_time = f"{base_hour:02d}30"
    
    result = {"ok": True, "temp": "--", "sky": "정보없음", "humid": "--", "wind": "--", "pm10": "--"}
    
    # 기상청 초단기실황
    try:
        service_key = "Gex1GW5D4SpmXVWmqNEBlfLYpOp8nykpHBD8%2BXxvTmPAp1PzATJRSVWaAbqXWsJCc1HFPySmTvWKSl4%2BVdoFdQ%3D%3D"
        url = f"http://apis.data.go.kr/1360000/VilageFcstInfoService_2.0/getUltraSrtNcst?serviceKey={service_key}&numOfRows=10&pageNo=1&dataType=JSON&base_date={base_date}&base_time={base_time}&nx={nx}&ny={ny}"
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(req, timeout=5, context=ctx) as resp:
            data = json.loads(resp.read().decode())
            items = data.get("response", {}).get("body", {}).get("items", {}).get("item", [])
            for it in items:
                cat = it.get("category")
                val = it.get("obsrValue", "")
                if cat == "T1H": result["temp"] = val  # 기온
                elif cat == "REH": result["humid"] = val  # 습도
                elif cat == "WSD": result["wind"] = val  # 풍속
    except Exception as e:
        print(f"날씨API 오류: {e}")
    
    # 하늘상태 (초단기예보에서)
    try:
        fcst_time = f"{now.hour:02d}00"
        url2 = f"http://apis.data.go.kr/1360000/VilageFcstInfoService_2.0/getUltraSrtFcst?serviceKey={service_key}&numOfRows=60&pageNo=1&dataType=JSON&base_date={base_date}&base_time={base_time}&nx={nx}&ny={ny}"
        req2 = Request(url2, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(req2, timeout=5, context=ctx) as resp:
            data = json.loads(resp.read().decode())
            items = data.get("response", {}).get("body", {}).get("items", {}).get("item", [])
            for it in items:
                if it.get("category") == "SKY":
                    sky_code = it.get("fcstValue", "")
                    sky_map = {"1": "맑음", "3": "구름많음", "4": "흐림"}
                    result["sky"] = sky_map.get(sky_code, "맑음")
                    break
    except Exception as e:
        print(f"예보API 오류: {e}")
    
    # 에어코리아 미세먼지 (충북 청주)
    try:
        air_key = "Gex1GW5D4SpmXVWmqNEBlfLYpOp8nykpHBD8%2BXxvTmPAp1PzATJRSVWaAbqXWsJCc1HFPySmTvWKSl4%2BVdoFdQ%3D%3D"
        air_url = f"http://apis.data.go.kr/B552584/ArpltnInforInqireSvc/getCtprvnRltmMesureDnsty?serviceKey={air_key}&returnType=json&numOfRows=1&pageNo=1&sidoName=%EC%B6%A9%EB%B6%81&ver=1.0"
        req3 = Request(air_url, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(req3, timeout=5, context=ctx) as resp:
            data = json.loads(resp.read().decode())
            items = data.get("response", {}).get("body", {}).get("items", [])
            if items:
                result["pm10"] = items[0].get("pm10Value", "--")
    except Exception as e:
        print(f"미세먼지API 오류: {e}")
    
    return jsonify(result)


# 공휴일 여부 확인 API
@app.route("/api/check_holiday")
@_login_required
def api_check_holiday():
    """특정 날짜의 공휴일/주말 여부 확인"""
    date_str = request.args.get("date", "")
    if not date_str:
        return jsonify(ok=False, is_holiday=False)
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        is_weekend = dt.weekday() >= 5  # 토(5), 일(6)
        hol_name = KR_HOLIDAYS.get(dt.date(), "")
        is_holiday = is_weekend or bool(hol_name)
        return jsonify(ok=True, is_holiday=is_holiday, is_weekend=is_weekend, 
                       holiday_name=hol_name or ("토요일" if dt.weekday() == 5 else "일요일" if dt.weekday() == 6 else ""))
    except:
        return jsonify(ok=False, is_holiday=False)


# 상단바용 대시보드 통계 API
@app.route("/api/dashboard_stats")
@_login_required
def api_dashboard_stats():
    """상단바에 표시할 대시보드 통계"""
    from datetime import date as _date
    today = _date.today()
    today_str = today.strftime("%Y%m%d")
    today_month = today.month
    today_day = today.day
    
    result = {"ok": True, "checked_in": 0, "leave_count": 0, "report_count": 0, "birthday_count": 0}
    try:
        conn = _conn(); cur = conn.cursor()
        # 출근자 수 (인증 성공 + 실제 등록자만. e_id=-1은 미등록카드 인증실패 로그)
        cur.execute("SELECT COUNT(DISTINCT e_id) FROM tenter "
                    "WHERE e_date=%s AND e_mode=1 AND e_id>0 AND e_result='0'", (today_str,))
        result["checked_in"] = cur.fetchone()[0] or 0
        
        # 휴가자 수
        cur.execute("SELECT COUNT(*) FROM leave_records WHERE leave_date=%s", (today_str,))
        result["leave_count"] = cur.fetchone()[0] or 0
        
        # 이번달 보고서 수
        cur_month = today.strftime("%Y%m")
        cur.execute("SELECT COUNT(*) FROM wr_reports WHERE report_date LIKE %s", (f"{cur_month}%",))
        result["report_count"] = cur.fetchone()[0] or 0
        
        # 이번달 생일자 수
        cur.execute("""
            SELECT COUNT(*) FROM employee_roster 
            WHERE birth_date IS NOT NULL AND birth_date <> ''
              AND SUBSTRING(REPLACE(birth_date,'-',''),5,2) = %s
        """, (f"{today_month:02d}",))
        result["birthday_count"] = cur.fetchone()[0] or 0
        conn.close()
    except:
        pass
    return jsonify(result)


@app.route("/api/my_schedule")
@_login_required
def api_my_schedule():
    """나의 전체 근무일정 조회"""
    login_name = session.get("user_name", "")
    month = request.args.get("month", "")  # YYYYMM or empty for all
    if not login_name:
        return jsonify(ok=False, msg="로그인 필요")
    conn = _conn(); cur = conn.cursor()
    if month:
        cur.execute("""
            SELECT work_date, basic_h, ot_h, night_h, etc, source_type
            FROM schedule_record
            WHERE emp_name=%s AND work_date LIKE %s
              AND source_type IN ('보고서','수정','수정근무보고서')
            ORDER BY work_date, FIELD(source_type,'보고서','수정','수정근무보고서')
        """, (login_name, f"{month}%"))
    else:
        cur.execute("""
            SELECT work_date, basic_h, ot_h, night_h, etc, source_type
            FROM schedule_record
            WHERE emp_name=%s AND source_type IN ('보고서','수정','수정근무보고서')
            ORDER BY work_date, FIELD(source_type,'보고서','수정','수정근무보고서')
        """, (login_name,))
    from collections import OrderedDict
    date_records = OrderedDict()
    for wd, bh, oh, nh, etc, src in cur.fetchall():
        wd_s = str(wd)
        disp = f"{wd_s[:4]}-{wd_s[4:6]}-{wd_s[6:8]}" if len(wd_s) == 8 else wd_s
        cat = _sr_cat(bh, oh, nh, etc)
        if disp not in date_records:
            date_records[disp] = {"date": disp, "orig_cat": None, "cat": cat, "is_revision": src == "수정근무보고서"}
        else:
            rec = date_records[disp]
            if src == "수정근무보고서":
                rec["orig_cat"] = rec["cat"]
                rec["cat"] = cat
                rec["is_revision"] = True
            else:
                rec["orig_cat"] = cat
                rec["is_revision"] = True
    items = list(date_records.values())
    conn.close()
    # 사용 가능한 월 목록
    months_set = sorted(set(i["date"][:7] for i in items))
    return jsonify(ok=True, items=items, months=months_set)


@app.route("/api/schedule_bulk_etc", methods=["POST"])
@_login_required
def api_schedule_bulk_etc():
    """근무보고서 일괄 근무상태(etc) 업데이트.
    body: { work_date: "YYYYMMDD", emp_names: [...], etc: "연차" }
    """
    data = request.get_json(force=True, silent=True) or {}
    work_date = data.get("work_date", "").strip()
    emp_names = data.get("emp_names", [])
    etc_val   = data.get("etc", "").strip()

    if not work_date or not emp_names or etc_val is None:
        return jsonify(ok=False, error="필수 항목 누락"), 400
    if len(emp_names) > 200:
        return jsonify(ok=False, error="한번에 200명 이하로 처리하세요"), 400

    conn = _conn(); cur = conn.cursor()
    try:
        ph = ",".join(["%s"] * len(emp_names))
        params = [etc_val, work_date] + list(emp_names)
        cur.execute(
            f"UPDATE schedule_record SET etc=%s WHERE work_date=%s AND emp_name IN ({ph})",
            params,
        )
        conn.commit()
        return jsonify(ok=True, updated=cur.rowcount)
    except Exception as e:
        conn.rollback()
        return jsonify(ok=False, error=str(e)), 500
    finally:
        conn.close()


@app.route("/api/schedule_emp_list")
@_login_required
def api_schedule_emp_list():
    """근무보고서 사업부+날짜 기준 직원 목록 조회 (일괄처리 UI용).
    params: work_date=YYYYMMDD, sheet_name=... (복수 가능)
    """
    work_date   = request.args.get("work_date", "").strip()
    sheet_names = request.args.getlist("sheet_name")
    if not work_date or not sheet_names:
        return jsonify(ok=False, error="work_date, sheet_name 필수"), 400

    conn = _conn(); cur = conn.cursor()
    try:
        ph = ",".join(["%s"] * len(sheet_names))
        cur.execute(
            f"SELECT DISTINCT emp_name, etc FROM schedule_record "
            f"WHERE work_date=%s AND sheet_name IN ({ph}) ORDER BY emp_name",
            [work_date] + sheet_names,
        )
        rows = [{"name": r[0], "etc": r[1] or ""} for r in cur.fetchall()]
        return jsonify(ok=True, employees=rows)
    finally:
        conn.close()


@app.route("/api/schedule_events")
@_login_required
def api_schedule_events():
    """달력 일정 조회 (연차 자동 포함)"""
    year = request.args.get("year", "")
    month = request.args.get("month", "")
    if not year or not month:
        from datetime import date as _date
        t = _date.today()
        year, month = t.year, t.month
    else:
        year, month = int(year), int(month)

    events = []
    try:
        conn = _conn(); cur = conn.cursor()
        # 사용자 등록 일정
        ym_prefix = f"{year}-{month:02d}"
        cur.execute("SELECT id, title, event_date, color, created_by FROM schedule_events WHERE event_date LIKE %s ORDER BY event_date", (ym_prefix + '%',))
        for r in cur.fetchall():
            events.append({"id": r[0], "title": r[1], "date": r[2], "color": r[3], "created_by": r[4], "type": "custom"})

        # 보고서 결재분 근무일정 표시 (카테고리별 배지)
        login_name = session.get("user_name", "")
        month_str = f"{year}{month:02d}"
        if login_name:
            cur.execute("""
                SELECT work_date, basic_h, ot_h, night_h, etc, sheet_name, source_type
                FROM schedule_record
                WHERE emp_name=%s AND work_date LIKE %s AND source_type IN ('보고서','수정','수정근무보고서')
                ORDER BY work_date, FIELD(source_type,'보고서','수정','수정근무보고서')
            """, (login_name, f"{month_str}%"))
            from collections import OrderedDict
            sr_by_date = OrderedDict()
            for sr_date, sr_basic, sr_ot, sr_night, sr_etc, sr_sheet, sr_src in cur.fetchall():
                wd = str(sr_date)
                formatted = f"{wd[:4]}-{wd[4:6]}-{wd[6:8]}" if len(wd) == 8 else wd
                if formatted not in sr_by_date:
                    sr_by_date[formatted] = {"cats": [], "is_revision": False}
                if sr_src == "수정근무보고서":
                    sr_by_date[formatted]["is_revision"] = True
                # 카테고리 역추출
                if sr_etc == "스킵":
                    sr_by_date[formatted]["cats"].append("스킵")
                elif sr_etc in ("연차", "무급"):
                    sr_by_date[formatted]["cats"].append(sr_etc)
                elif sr_etc and sr_etc not in ('0', ''):
                    try:
                        float(sr_etc)
                        sr_by_date[formatted]["cats"].append("공제")
                    except (ValueError, TypeError):
                        sr_by_date[formatted]["cats"].append(sr_etc)
                elif sr_basic > 0 and sr_ot > 0 and sr_night > 0:
                    sr_by_date[formatted]["cats"].append("야간연장")
                elif sr_basic > 0 and sr_night > 0:
                    sr_by_date[formatted]["cats"].append("야간기본")
                elif sr_basic > 0 and sr_ot > 0:
                    sr_by_date[formatted]["cats"].append("주간연장")
                elif sr_basic > 0:
                    sr_by_date[formatted]["cats"].append("주간기본")
                else:
                    sr_by_date[formatted]["cats"].append("근무")
            _cat_colors = {
                "주간기본": "#6366f1", "주간연장": "#3b82f6",
                "야간기본": "#8b5cf6", "야간연장": "#7c3aed",
                "연차": "#f59e0b", "무급": "#f59e0b",
                "공제": "#ef4444", "스킵": "#dc2626",
            }
            _cat_icons = {
                "주간기본": "☀️", "주간연장": "☀️⏰",
                "야간기본": "🌙", "야간연장": "🌙⏰",
                "연차": "🏖️", "무급": "🚫",
                "공제": "➖", "스킵": "⛔",
            }
            _no_schedule_cats = {"연차", "무급", "스킵"}
            for formatted, info in sr_by_date.items():
                base_type = "work_report_revision" if info["is_revision"] else "work_report"
                for cat in info["cats"]:
                    icon = _cat_icons.get(cat, "📋")
                    color = _cat_colors.get(cat, "#6366f1")
                    title = f"{icon} {cat}"
                    ev_type = "work_report_leave" if cat in _no_schedule_cats else base_type
                    events.append({"id": None, "title": title, "date": formatted, "color": color, "created_by": "", "type": ev_type})

        conn.close()

        # 공휴일 자동 표시
        dim = calendar.monthrange(year, month)[1]
        for d in range(1, dim + 1):
            dt = date(year, month, d)
            hol_name = KR_HOLIDAYS.get(dt, "")
            if hol_name:
                formatted_date = f"{year}-{month:02d}-{d:02d}"
                events.append({"id": None, "title": f"🔴 {hol_name}", "date": formatted_date, "color": "#dc3545", "created_by": "", "type": "holiday"})

    except Exception as e:
        return jsonify(ok=False, msg=str(e))
    return jsonify(ok=True, events=events)


@app.route("/save_schedule_event", methods=["POST"])
@_login_required
def save_schedule_event():
    """일정 저장"""
    if session.get("role") != "admin" and not _has_perm("schedule"):
        return jsonify(ok=False, msg="권한이 없습니다")
    eid = request.form.get("id", "").strip()
    title = request.form.get("title", "").strip()
    event_date = request.form.get("event_date", "").strip()
    event_date_end = request.form.get("event_date_end", "").strip()
    color = request.form.get("color", "#0d6efd").strip()
    if not title or not event_date:
        return jsonify(ok=False, msg="제목과 날짜를 입력하세요")
    conn = _conn(); cur = conn.cursor()
    if eid:
        cur.execute("UPDATE schedule_events SET title=%s, event_date=%s, color=%s WHERE id=%s",
                     (title, event_date, color, eid))
    elif event_date_end and event_date_end > event_date:
        d = datetime.strptime(event_date, "%Y-%m-%d")
        d_end = datetime.strptime(event_date_end, "%Y-%m-%d")
        user = session.get("user_name", "")
        while d <= d_end:
            cur.execute("INSERT INTO schedule_events (title, event_date, color, created_by) VALUES (%s,%s,%s,%s)",
                         (title, d.strftime("%Y-%m-%d"), color, user))
            d += timedelta(days=1)
    else:
        cur.execute("INSERT INTO schedule_events (title, event_date, color, created_by) VALUES (%s,%s,%s,%s)",
                     (title, event_date, color, session.get("user_name", "")))
    conn.commit(); conn.close()
    return jsonify(ok=True)


@app.route("/delete_schedule_event", methods=["POST"])
@_login_required
def delete_schedule_event():
    """일정 삭제"""
    if session.get("role") != "admin" and not _has_perm("schedule"):
        return jsonify(ok=False, msg="권한이 없습니다")
    eid = request.form.get("id")
    if not eid:
        return jsonify(ok=False, msg="ID가 없습니다")
    conn = _conn(); cur = conn.cursor()
    cur.execute("DELETE FROM schedule_events WHERE id=%s", (eid,))
    conn.commit(); conn.close()
    return jsonify(ok=True)


@app.route("/attendance")
@_login_required
def attendance():
    search = request.args.get("search", "").strip()
    sel_date = request.args.get("date", "").strip()
    sel_dept = request.args.get("dept", "").strip()
    sel_month = request.args.get("month", "").strip()
    is_admin = session.get("role") == "admin"
    can_edit = is_admin or _has_perm("schedule")
    if not can_edit:
        search = session.get("user_name", "")
        sel_dept = ""
    sel_date_raw = sel_date.replace("-", "") if sel_date else ""
    records = []
    monthly = {}
    stats = {"total": 0, "on_time": 0, "late": 0, "no_in": 0, "no_out": 0}
    mode = ""
    ranking = OrderedDict()
    dept_list = sorted(set(DEPT_MAP.values()))  # fallback
    try:
        _conn2 = _conn()
        _cur2 = _conn2.cursor()
        _cur2.execute("SELECT DISTINCT dept FROM employee_roster WHERE dept IS NOT NULL AND dept <> '' ORDER BY dept")
        dept_list = [r[0] for r in _cur2.fetchall()]
        _conn2.close()
    except Exception:
        pass
    try:
        conn = _conn()
        cur = conn.cursor()
        if search:
            mode = "search"
            if not sel_month:
                sel_month = datetime.now().strftime("%Y-%m")
            # 데이터요약: 최근 1년 조회, 리스트 선택시
            now = datetime.now()
            m_from = (now - timedelta(days=365)).strftime("%Y%m") + "01"
            m_to = now.strftime("%Y%m") + "31"
            # 1단계: 커버링 인덱스로 e_id 먼저 조회 (중복 방지)
            cur.execute("""
                SELECT DISTINCT e_id FROM tenter
                WHERE e_name LIKE %s OR e_idno LIKE %s
                LIMIT 50
            """, (f"%{search}%", f"%{search}%"))
            eids = [r[0] for r in cur.fetchall()]
            if eids:
                ph = ",".join(["%s"] * len(eids))
                # 2단계: idx_eid 인덱스로 빠르게 조회
                cur.execute(f"""
                    SELECT t.e_date, t.e_time, t.e_id, t.e_idno, t.e_name,
                           t.e_mode, g.name AS gate_name, u.company,
                           IFNULL(er.dept, '') AS sub_dept
                    FROM tenter t
                         LEFT JOIN tgate g ON t.g_id = g.id
                         LEFT JOIN tuser u ON t.e_id = u.id
                         LEFT JOIN employee_roster er ON t.e_name = er.name
                    WHERE t.e_id IN ({ph}) AND t.e_date >= %s AND t.e_date <= %s
                    ORDER BY t.e_date, t.e_time
                """, eids + [m_from, m_to])
            else:
                cur.execute("SELECT 1 WHERE 1=0")  # 결과없음
        elif sel_date_raw or sel_dept or sel_month:
            mode = "date"
            if sel_month and not sel_date_raw:
                # 월 선택: 해당 월 전체 조회
                ym = sel_month.replace("-", "")
                ym_year, ym_mon = int(ym[:4]), int(ym[4:6])
                ym_dim = calendar.monthrange(ym_year, ym_mon)[1]
                date_start = f"{ym}01"
                date_end = f"{ym}{ym_dim:02d}"
                # 야간근무 병합을 위해 전후 2일 여유
                q_start = (datetime(ym_year, ym_mon, 1) - timedelta(days=2)).strftime("%Y%m%d")
                q_end = (datetime(ym_year, ym_mon, ym_dim) + timedelta(days=2)).strftime("%Y%m%d")
            else:
                if not sel_date_raw:
                    sel_date_raw = datetime.now().strftime("%Y%m%d")
                    sel_date = datetime.now().strftime("%Y-%m-%d")
                date_start = sel_date_raw
                date_end = sel_date_raw
                q_start = sel_date_raw
                q_end = sel_date_raw
            sql = """
                SELECT t.e_date, t.e_time, t.e_id, t.e_idno, t.e_name,
                       t.e_mode, g.name AS gate_name, u.company,
                       IFNULL(er.dept, '') AS sub_dept
                FROM tenter t
                     LEFT JOIN tgate g ON t.g_id = g.id
                     LEFT JOIN tuser u ON t.e_id = u.id
                     LEFT JOIN employee_roster er ON t.e_name = er.name
                WHERE t.e_id >= 0 AND t.e_date >= %s AND t.e_date <= %s
            """
            params = [q_start, q_end]
            if sel_dept:
                sql += " AND er.dept = %s"
                params += [sel_dept]
            sql += " ORDER BY t.e_name, t.e_time"
            cur.execute(sql, params)
        if mode:
            rows = cur.fetchall()
            day_emp = OrderedDict()
            for r in rows:
                e_date = r[0]
                e_time = r[1]
                eid = r[2]
                e_mode = (r[5] or "").strip()
                gate = (r[6] or "").strip()
                key = (e_date, eid)
                if key not in day_emp:
                    company = (r[7] or "").strip()
                    sub_dept = (r[8] or "").strip()
                    dept_display = sub_dept or DEPT_MAP.get(company, "")
                    day_emp[key] = {
                        "date": e_date, "date_fmt": _fmt_date(e_date),
                        "id": eid,
                        "er_dept": sub_dept,
                        "emp_no": (r[3] or "").strip() or str(eid),
                        "name": (r[4] or "").strip() or f"ID-{eid}",
                        "dept": dept_display,
                        "in_time": None, "out_time": None,
                        "in_gate": "", "out_gate": "",
                    }
                e = day_emp[key]
                if e_mode == MODE_IN:
                    if e["in_time"] is None or e_time < e["in_time"]:
                        e["in_time"] = e_time
                        e["in_gate"] = gate
                elif e_mode == MODE_OUT:
                    if e["out_time"] is None or e_time > e["out_time"]:
                        e["out_time"] = e_time
                        e["out_gate"] = gate
            # 야간근무 병합 (실제 평일 데이터)
            if mode == 'search' or (mode == 'date' and sel_month):
                _merge_night_shifts(day_emp)
            for e in day_emp.values():
                e["in_time_fmt"] = _fmt_time(e["in_time"])
                e["out_time_fmt"] = _fmt_time(e["out_time"]) + ("(+1일)" if e.get("night_next_day") else "")
                in_s = _time_to_sec(e["in_time"])
                out_s = _time_to_sec(e["out_time"])
                is_night = in_s is not None and in_s >= 64800  # 18:00:00
                # 평일/휴일 판단 (계산 전에 결정)
                dt = datetime.strptime(e["date"], "%Y%m%d")
                e["weekday"] = dt.weekday()
                e["holiday"] = KR_HOLIDAYS.get(dt.date(), "")
                is_hol = e["weekday"] >= 5 or bool(e["holiday"])
                e["basic_fmt"] = "-"
                e["night_fmt"] = "-"
                e["other_fmt"] = "-"
                e["ot_hours"] = 0
                if in_s is not None and out_s is not None and out_s > in_s and not is_night:
                    wh = (out_s - in_s) / 3600
                    e["basic_fmt"] = f"{min(wh, 8):.0f}" if wh >= 8 else f"{wh:.1f}"
                    e["work_hours"] = wh
                    e["overtime"] = "2" if out_s >= 68400 else ""  # 19:00 이후만 연장
                    e["ot_hours"] = 2 if out_s >= 68400 else 0
                elif is_night and out_s is not None and out_s < in_s:
                    wh = (86400 - in_s + out_s) / 3600  # 출근~자정 + 자정~퇴근
                    night_sec = max(0, 86400 - max(in_s, 79200)) + max(0, min(out_s, 21600))
                    ni_h = max(0, int(night_sec / 3600 - 1))  # 야간 22~06 - 1h 실제, 법령
                    e["basic_fmt"] = f"{min(wh, 8):.0f}" if wh >= 8 else f"{wh:.1f}"
                    e["night_fmt"] = str(ni_h) if ni_h > 0 else "-"
                    e["work_hours"] = wh
                    e["overtime"] = "2" if out_s >= 28200 else ""  # 07:50 이후 퇴근 시 연장
                    e["ot_hours"] = 2 if out_s >= 28200 else 0
                elif in_s is None and out_s is not None and out_s < 28800:  # 출근없음 + 퇴근 08:00 이전 = 야간근무
                    assumed_in = 68400  # 전날 19:00 출근 가정
                    wh = (86400 - assumed_in + out_s) / 3600
                    night_sec = max(0, 86400 - max(assumed_in, 79200)) + max(0, min(out_s, 21600))
                    ni_h = max(0, int(night_sec / 3600 - 1))
                    e["basic_fmt"] = f"{min(wh, 8):.0f}" if wh >= 8 else f"{wh:.1f}"
                    e["night_fmt"] = str(ni_h) if ni_h > 0 else "-"
                    e["work_hours"] = wh
                    e["overtime"] = "2" if out_s >= 28200 else ""
                    e["ot_hours"] = 2 if out_s >= 28200 else 0
                elif in_s is None and out_s is not None and out_s < 71400:  # 출근없음 + 퇴근 19:50 이전 (주간)
                    e["basic_fmt"] = "8"
                    e["work_hours"] = 8
                    e["overtime"] = ""
                elif in_s is None and out_s is not None and out_s >= 71400:  # 출근없음 + 퇴근 19:50 이후
                    e["basic_fmt"] = "8"
                    e["work_hours"] = 10
                    e["overtime"] = "2"
                    e["ot_hours"] = 2
                elif in_s is not None or out_s is not None:
                    e["basic_fmt"] = "8"
                    e["work_hours"] = 8
                    e["overtime"] = ""
                    if is_night and out_s is None:
                        assumed_out = 18000  # 다음날 05:00 퇴근 가정
                        night_sec = max(0, 86400 - max(in_s, 79200)) + max(0, min(assumed_out, 21600))
                        ni_h = max(0, int(night_sec / 3600 - 1))
                        e["night_fmt"] = str(ni_h)
                        e["work_hours"] = 8
                else:
                    e["work_hours"] = 0
                    e["overtime"] = ""
                e["month"] = e["date"][:6]
                day_names = ["월", "화", "수", "목", "금", "토", "일"]
                e["date_fmt"] = f"{_fmt_date(e['date'])}({day_names[dt.weekday()]})"
                if sel_dept and e.get("er_dept", "") != sel_dept:
                    continue
                if mode == 'date' and sel_month and e["date"] < date_start or \
                   mode == 'date' and sel_month and e["date"] > date_end:
                    continue
                records.append(e)
            # 월별 통계 (최근 1년)
            monthly = OrderedDict()
            all_in_secs_normal = []
            all_in_secs_late = []
            all_out_secs_normal = []
            all_out_secs_late = []
            month_cutoff = (datetime.now() - timedelta(days=365)).strftime("%Y%m")
            for r in records:
                m = r["month"]
                if m < month_cutoff:
                    continue
                if m not in monthly:
                    monthly[m] = {"label": f"{m[:4]}년 {int(m[4:6])}월",
                                  "total": 0, "on_time": 0, "late": 0, "no_in": 0, "no_out": 0,
                                  "total_work": 0, "total_overtime": 0,
                                  "sat_days": 0, "sat_work": 0, "sat_over": 0,
                                  "sun_days": 0, "sun_work": 0, "sun_over": 0,
                                  "hol_days": 0, "hol_work": 0, "hol_over": 0,
                                  "in_secs_normal": [], "in_secs_late": [],
                                  "out_secs_normal": [], "out_secs_late": []}
                ms = monthly[m]
                ms["total"] += 1
                ms["total_work"] += r["work_hours"]
                ms["total_overtime"] += r.get("ot_hours", 2 if r["overtime"] == "2" else 0)
                if r["weekday"] == 5:
                    ms["sat_days"] += 1
                    ms["sat_work"] += min(r["work_hours"], 8)
                    ms["sat_over"] += max(r["work_hours"] - 8, 0)
                elif r["weekday"] == 6:
                    ms["sun_days"] += 1
                    ms["sun_work"] += min(r["work_hours"], 8)
                    ms["sun_over"] += max(r["work_hours"] - 8, 0)
                if r.get("holiday"):
                    ms["hol_days"] += 1
                    ms["hol_work"] += min(r["work_hours"], 8)
                    ms["hol_over"] += max(r["work_hours"] - 8, 0)
                if r["in_time"] and r["in_time"] <= "115900":
                    ms["on_time"] += 1
                    stats["on_time"] += 1
                elif r["in_time"]:
                    ms["late"] += 1
                    stats["late"] += 1
                else:
                    ms["no_in"] += 1
                    stats["no_in"] += 1
                if r["in_time"] and not r["out_time"]:
                    ms["no_out"] += 1
                    stats["no_out"] += 1
                in_s = _time_to_sec(r["in_time"])
                out_s = _time_to_sec(r["out_time"])
                if r["in_time"] and r["in_time"] <= "115900":
                    if in_s is not None:
                        ms["in_secs_normal"].append(in_s)
                        all_in_secs_normal.append(in_s)
                    if out_s is not None:
                        ms["out_secs_normal"].append(out_s)
                        all_out_secs_normal.append(out_s)
                elif r["in_time"]:
                    if in_s is not None:
                        ms["in_secs_late"].append(in_s)
                        all_in_secs_late.append(in_s)
                    if out_s is not None:
                        ms["out_secs_late"].append(out_s)
                        all_out_secs_late.append(out_s)
            for ms in monthly.values():
                ms["avg_in_normal"] = _sec_to_fmt(sum(ms["in_secs_normal"]) / len(ms["in_secs_normal"])) if len(ms["in_secs_normal"]) >= 5 else "-"
                ms["avg_in_late"] = _sec_to_fmt(sum(ms["in_secs_late"]) / len(ms["in_secs_late"])) if len(ms["in_secs_late"]) >= 5 else "-"
                ms["avg_out_normal"] = _sec_to_fmt(sum(ms["out_secs_normal"]) / len(ms["out_secs_normal"])) if len(ms["out_secs_normal"]) >= 5 else "-"
                ms["avg_out_late"] = _sec_to_fmt(sum(ms["out_secs_late"]) / len(ms["out_secs_late"])) if len(ms["out_secs_late"]) >= 5 else "-"
            stats["total"] = sum(ms["total"] for ms in monthly.values())
            stats["total_work"] = sum(ms["total_work"] for ms in monthly.values())
            stats["total_overtime"] = sum(ms["total_overtime"] for ms in monthly.values())
            stats["has_both"] = stats["on_time"] > 0 and stats["late"] > 0
            stats["avg_in_normal"] = _sec_to_fmt(sum(all_in_secs_normal) / len(all_in_secs_normal)) if len(all_in_secs_normal) >= 5 else "-"
            stats["avg_in_late"] = _sec_to_fmt(sum(all_in_secs_late) / len(all_in_secs_late)) if len(all_in_secs_late) >= 5 else "-"
            stats["avg_out_normal"] = _sec_to_fmt(sum(all_out_secs_normal) / len(all_out_secs_normal)) if len(all_out_secs_normal) >= 5 else "-"
            stats["avg_out_late"] = _sec_to_fmt(sum(all_out_secs_late) / len(all_out_secs_late)) if len(all_out_secs_late) >= 5 else "-"
        else:
            # 금일 빨리 출근 TOP 20
            today_str_att = datetime.now().strftime("%Y%m%d")
            cur.execute("""
                SELECT t.e_id, t.e_name, u.company, MIN(t.e_time) AS in_time
                FROM tenter t LEFT JOIN tuser u ON t.e_id = u.id
                WHERE t.e_id >= 0 AND t.e_mode = %s AND t.e_date = %s
                GROUP BY t.e_id, t.e_name, u.company
                ORDER BY in_time
                LIMIT 20
            """, (MODE_IN, today_str_att))
            today_top = []
            for eid, ename, comp, itime in cur.fetchall():
                dept = DEPT_MAP.get((comp or "").strip(), "")
                if sel_dept and dept != sel_dept:
                    continue
                today_top.append({
                    "name": (ename or "").strip() or f"ID-{eid}",
                    "dept": dept,
                    "in_time": _fmt_time(itime),
                })
            # 최근 1개월 평균 출근 TOP 20 (롤링 30일)
            now = datetime.now()
            rank_start = (now - timedelta(days=30)).strftime("%Y%m%d")
            rank_end = now.strftime("%Y%m%d")
            cur.execute("""
                SELECT t.e_date, t.e_time, t.e_id, t.e_idno, t.e_name,
                       u.company
                FROM tenter t LEFT JOIN tuser u ON t.e_id = u.id
                WHERE t.e_id >= 0 AND t.e_mode = %s AND t.e_date >= %s AND t.e_date <= %s
                ORDER BY t.e_date, t.e_time
            """, (MODE_IN, rank_start, rank_end))
            all_rows = cur.fetchall()
            day_emp_in = {}
            emp_info = {}
            for r in all_rows:
                e_date, e_time, eid = r[0], r[1], r[2]
                key = (e_date, eid)
                if key not in day_emp_in or e_time < day_emp_in[key]:
                    day_emp_in[key] = e_time
                if eid not in emp_info:
                    company = (r[5] or "").strip()
                    emp_info[eid] = {
                        "name": (r[4] or "").strip() or f"ID-{eid}",
                        "dept": DEPT_MAP.get(company, "")
                    }
            emp_secs = defaultdict(list)
            for (e_date, eid), in_time in day_emp_in.items():
                # 주말(토·일) 출근기록은 평균 출근시간 산정에서 제외
                try:
                    if datetime.strptime(e_date, "%Y%m%d").weekday() >= 5:
                        continue
                except ValueError:
                    pass
                in_s = _time_to_sec(in_time)
                if in_s is not None:
                    emp_secs[eid].append(in_s)
            rank_list = []
            for eid, secs in emp_secs.items():
                if len(secs) < 5:
                    continue
                info = emp_info[eid]
                if sel_dept and info["dept"] != sel_dept:
                    continue
                avg = sum(secs) / len(secs)
                rank_list.append({
                    "name": info["name"],
                    "dept": info["dept"],
                    "avg_in": _sec_to_fmt(avg),
                    "avg_in_sec": avg,
                })
            rank_list.sort(key=lambda x: x["avg_in_sec"])
            d_from = (now - timedelta(days=30)).strftime("%m/%d")
            d_to = now.strftime("%m/%d")
            label = f"{d_from} ~ {d_to}"
            if rank_list:
                ranking[label] = rank_list[:20]

            # 전일 출근O 퇴근X 리스트
            yesterday_str = (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")
            cur.execute("""
                SELECT t.e_id, t.e_name, u.company,
                       MIN(CASE WHEN t.e_mode=%s THEN t.e_time END) AS in_time,
                       MAX(CASE WHEN t.e_mode=%s THEN t.e_time END) AS out_time
                FROM tenter t LEFT JOIN tuser u ON t.e_id = u.id
                WHERE t.e_id >= 0 AND t.e_date = %s
                GROUP BY t.e_id, t.e_name, u.company
            """, (MODE_IN, MODE_OUT, yesterday_str))
            prev2_str = (datetime.now() - timedelta(days=2)).strftime("%Y%m%d")
            # 타각 비대상: company 기준(임원·서울사무소)
            NO_TAG_DEPTS = {"임원", "경영기획팀"}
            # 타각 비대상: 명부 부서 기준(경비·비서 등 카드 타각 안 하는 직군)
            NO_TAG_ER_DEPTS = {"총무팀", "비서팀"}
            prev_no_out_list = []
            prev_no_in_list = []      # 전일 출근 미타각 (퇴근만 있거나 기록 없음)
            prev_seen_ids = set()
            for eid, ename, comp, in_t, out_t in cur.fetchall():
                dept = DEPT_MAP.get((comp or "").strip(), "")
                prev_seen_ids.add(eid)
                if sel_dept and dept != sel_dept:
                    continue
                nm = (ename or "").strip() or f"ID-{eid}"
                if in_t is not None and out_t is None:
                    prev_no_out_list.append({
                        "name": nm, "dept": dept,
                        "in_time": _fmt_time(in_t), "out_time": "-"
                    })
                elif in_t is None and out_t is not None:
                    # 퇴근만 찍힘 → 전전일 출근이 있으면 야간조 정상근무로 보고 제외
                    cur2 = conn.cursor()
                    cur2.execute("""SELECT COUNT(*) FROM tenter
                                    WHERE e_id=%s AND e_date=%s AND e_mode=%s""",
                                 (eid, prev2_str, MODE_IN))
                    had_prev_in = cur2.fetchone()[0] > 0
                    cur2.close()
                    if not had_prev_in:
                        prev_no_in_list.append({
                            "note": "", "name": nm, "dept": dept,
                            "in_time": "-", "out_time": _fmt_time(out_t)
                        })
            # 출입기록이 아예 없는 재직자 (휴가자 제외)
            cur.execute("SELECT e_id FROM leave_records WHERE leave_date=%s", (yesterday_str,))
            leave_ids = {r[0] for r in cur.fetchall()}
            # 카드 미발급자는 타각 자체가 불가하므로 제외 (근무표로 별도 관리)
            cur.execute("""
                SELECT u.id, u.name, u.company, IFNULL(er.dept,'') FROM tuser u
                JOIN employee_roster er ON er.emp_no = u.idno
                WHERE u.idno IS NOT NULL AND u.idno <> ''
                  AND (u.retire_date IS NULL OR u.retire_date = '')
                  AND u.cardnum IS NOT NULL AND u.cardnum <> ''
            """)
            no_in_cands = cur.fetchall()
            # 출입기록이 한 번도 없는 사용자 = 카드 미등록(장비 미반영)
            cur.execute("SELECT DISTINCT e_id FROM tenter WHERE e_id > 0")
            ever_tagged = {r[0] for r in cur.fetchall()}
            for uid, uname, comp, er_dept in no_in_cands:
                if uid in prev_seen_ids or uid in leave_ids:
                    continue
                dept = DEPT_MAP.get((comp or "").strip(), "")
                # 임원·서울사무소, 경비·비서 등 타각 비대상 제외
                if dept in NO_TAG_DEPTS or (er_dept or "").strip() in NO_TAG_ER_DEPTS:
                    continue
                if sel_dept and dept != sel_dept:
                    continue
                prev_no_in_list.append({
                    "note": "" if uid in ever_tagged else "카드 미발급",
                    "name": (uname or "").strip(), "dept": dept,
                    "in_time": "-", "out_time": "-"
                })
            prev_no_out_list.sort(key=lambda x: x["in_time"] or "")
            prev_no_in_list.sort(key=lambda x: (x["dept"] or "", x["name"] or ""))

            # ── 차트 데이터 ───────────────────────────────────────
            # ① 시간대별 출근 분포 (오늘)
            cur.execute("""
                SELECT CAST(SUBSTRING(e_time,1,2) AS UNSIGNED) as hr, COUNT(DISTINCT e_id) as cnt
                FROM tenter WHERE e_id>=0 AND e_date=%s AND e_mode=%s
                GROUP BY hr ORDER BY hr
            """, (today_str_att, MODE_IN))
            raw_hourly = {int(hr): int(cnt) for hr, cnt in cur.fetchall()}
            hourly_dist_labels = [f"{h:02d}시" for h in range(5, 10)] + ["10시+"]
            hourly_dist_data   = [raw_hourly.get(h, 0) for h in range(5, 10)] + \
                                  [sum(v for k, v in raw_hourly.items() if k >= 10)]

            # ② 최근 7일 퇴근 미처리 추이
            weekly_no_out_labels, weekly_no_out_data = [], []
            for d in range(6, -1, -1):
                day = datetime.now() - timedelta(days=d)
                day_str   = day.strftime("%Y%m%d")
                day_label = day.strftime("%m/%d") + ("(일)" if day.weekday()==6 else "(토)" if day.weekday()==5 else "")
                cur.execute("""
                    SELECT COUNT(*) FROM (
                      SELECT e_id,
                             MAX(CASE WHEN e_mode=%s THEN 1 ELSE 0 END) has_in,
                             MAX(CASE WHEN e_mode=%s THEN 1 ELSE 0 END) has_out
                      FROM tenter WHERE e_id>=0 AND e_date=%s GROUP BY e_id
                    ) sub WHERE has_in=1 AND has_out=0
                """, (MODE_IN, MODE_OUT, day_str))
                weekly_no_out_labels.append(day_label)
                weekly_no_out_data.append(int(cur.fetchone()[0]))

            # ③ 부서별 오늘 출근 현황
            cur.execute("""
                SELECT u.company, COUNT(DISTINCT t.e_id) as cnt
                FROM tenter t LEFT JOIN tuser u ON t.e_id=u.id
                WHERE t.e_id>=0 AND t.e_date=%s AND t.e_mode=%s
                GROUP BY u.company
            """, (today_str_att, MODE_IN))
            dept_raw = {(r[0] or "").strip(): int(r[1]) for r in cur.fetchall()}
            dept_chart_labels, dept_chart_data = [], []
            for code, name in DEPT_MAP.items():
                cnt = dept_raw.get(code, 0)
                if cnt > 0:
                    dept_chart_labels.append(name)
                    dept_chart_data.append(cnt)

            # ④ 오늘 출근 요약 (도넛) — 최근 30일 출근자 = 활성 직원
            cur.execute("""
                SELECT COUNT(DISTINCT e_id) FROM tenter
                WHERE e_id>=0 AND e_date>=%s
            """, ((datetime.now() - timedelta(days=30)).strftime("%Y%m%d"),))
            total_active = int(cur.fetchone()[0])
            today_in_count = len(today_top)  # already fetched, use len
            cur.execute("""
                SELECT COUNT(DISTINCT e_id) FROM tenter
                WHERE e_id>=0 AND e_date=%s AND e_mode=%s
            """, (today_str_att, MODE_IN))
            today_in_count = int(cur.fetchone()[0])
            today_absent   = max(0, total_active - today_in_count)
        conn.close()
    except Exception as e:
        flash(f"error: {e}", "danger")
    sel_date_holiday = ""
    if sel_date_raw and len(sel_date_raw) == 8:
        try:
            sel_date_holiday = KR_HOLIDAYS.get(datetime.strptime(sel_date_raw, "%Y%m%d").date(), "")
        except ValueError:
            pass
    return render_template("attendance.html", records=records,
                           stats=stats, monthly=monthly,
                           search=search, sel_date=sel_date, mode=mode,
                           ranking=ranking,
                           today_top=today_top if 'today_top' in dir() else [],
                           prev_no_out_list=prev_no_out_list if 'prev_no_out_list' in dir() else [],
                           prev_no_in_list=prev_no_in_list if 'prev_no_in_list' in dir() else [],
                           sel_date_holiday=sel_date_holiday,
                           sel_dept=sel_dept, dept_list=dept_list,
                           sel_month=sel_month, can_edit=can_edit,
                           hourly_dist_labels=hourly_dist_labels if 'hourly_dist_labels' in dir() else [],
                           hourly_dist_data=hourly_dist_data if 'hourly_dist_data' in dir() else [],
                           weekly_no_out_labels=weekly_no_out_labels if 'weekly_no_out_labels' in dir() else [],
                           weekly_no_out_data=weekly_no_out_data if 'weekly_no_out_data' in dir() else [],
                           dept_chart_labels=dept_chart_labels if 'dept_chart_labels' in dir() else [],
                           dept_chart_data=dept_chart_data if 'dept_chart_data' in dir() else [],
                           today_in_count=today_in_count if 'today_in_count' in dir() else 0,
                           today_absent=today_absent if 'today_absent' in dir() else 0)


@app.route("/weekly52")
@_login_required
def weekly52():
    search = request.args.get("search", "").strip()
    sel_month = request.args.get("month", "").strip()
    sel_dept = request.args.get("dept", "").strip()
    is_admin = session.get("role") == "admin"
    can_edit = is_admin or _has_perm("schedule")
    if not can_edit:
        search = session.get("user_name", "")
    weeks = []
    emp_name = ""
    dept_list = sorted(set(DEPT_MAP.values()))
    try:
        conn = _conn()
        cur = conn.cursor()
        if search:
            if not sel_month:
                sel_month = datetime.now().strftime("%Y-%m")
            # 선택월+7일(주 경계 포함)
            start_dt = datetime.strptime(sel_month + "-01", "%Y-%m-%d")
            q_start = (start_dt - timedelta(days=7)).strftime("%Y%m%d")
            dim52 = calendar.monthrange(start_dt.year, start_dt.month)[1]
            end_dt = datetime(start_dt.year, start_dt.month, dim52)
            q_end = (end_dt + timedelta(days=7)).strftime("%Y%m%d")
            # 1단계: 커버링 인덱스로 e_id 먼저 조회
            cur.execute("""
                SELECT DISTINCT e_id FROM tenter
                WHERE e_name LIKE %s OR e_idno LIKE %s LIMIT 50
            """, (f"%{search}%", f"%{search}%"))
            eids = [r[0] for r in cur.fetchall()]
            if eids:
                ph = ",".join(["%s"] * len(eids))
                # 2단계: 평일 범위 제한 조회
                cur.execute(f"""
                    SELECT t.e_date, t.e_time, t.e_id, t.e_idno, t.e_name, t.e_mode
                    FROM tenter t
                    WHERE t.e_id IN ({ph}) AND t.e_date >= %s AND t.e_date <= %s
                    ORDER BY t.e_date, t.e_time
                """, eids + [q_start, q_end])
            else:
                cur.execute("SELECT 1 WHERE 1=0")  # 결과없음
            rows = cur.fetchall()
            day_emp = OrderedDict()
            for r in rows:
                e_date, e_time, eid = r[0], r[1], r[2]
                e_mode = (r[5] or "").strip()
                key = (e_date, eid)
                if key not in day_emp:
                    day_emp[key] = {
                        "date": e_date, "id": eid,
                        "name": (r[4] or "").strip() or f"ID-{eid}",
                        "in_time": None, "out_time": None,
                    }
                e = day_emp[key]
                if e_mode == MODE_IN:
                    if e["in_time"] is None or e_time < e["in_time"]:
                        e["in_time"] = e_time
                elif e_mode == MODE_OUT:
                    if e["out_time"] is None or e_time > e["out_time"]:
                        e["out_time"] = e_time
            _merge_night_shifts(day_emp)
            records = []
            for e in day_emp.values():
                if not emp_name:
                    emp_name = e["name"]
                in_s = _time_to_sec(e["in_time"])
                out_s = _time_to_sec(e["out_time"])
                is_night = in_s is not None and in_s > 43140
                work_h = 0
                over_h = 0
                if in_s is not None and out_s is not None and out_s > in_s:
                    wh = (out_s - in_s) / 3600
                    work_h = min(wh, 8)
                    over_h = 2 if out_s >= 68400 else 0
                elif is_night and out_s is not None and out_s < in_s:
                    wh = (86400 - in_s + out_s) / 3600
                    work_h = min(wh, 8)
                    over_h = 2 if out_s >= 28200 else 0  # 07:50 이후
                elif in_s is None and out_s is not None and out_s < 71400:
                    work_h = 8
                elif in_s is None and out_s is not None and out_s >= 71400:
                    work_h, over_h = 8, 2
                elif in_s is not None or out_s is not None:
                    work_h = 8
                e["work_h"] = work_h
                e["over_h"] = over_h
                records.append(e)
            # 주간별 집계
            week_map = OrderedDict()
            for rec in records:
                dt = datetime.strptime(rec["date"], "%Y%m%d")
                if sel_month and dt.strftime("%Y-%m") != sel_month:
                    continue
                iso = dt.isocalendar()
                wk_key = f"{iso[0]}-W{iso[1]:02d}"
                if wk_key not in week_map:
                    mon = dt - timedelta(days=dt.weekday())
                    sun = mon + timedelta(days=6)
                    week_map[wk_key] = {
                        "week_no": iso[1],
                        "start": mon.strftime("%Y/%m/%d"),
                        "end": sun.strftime("%Y/%m/%d"),
                        "days": 0, "work": 0, "overtime": 0,
                    }
                w = week_map[wk_key]
                w["days"] += 1
                w["work"] += rec["work_h"]
                w["overtime"] += rec["over_h"]
            for w in week_map.values():
                w["work"] = int(w["work"])
                w["overtime"] = int(w["overtime"])
                w["total"] = w["work"] + w["overtime"]
            weeks = list(week_map.values())
        # 경고: 1주 직원당 주52시간 초과일
        over52_list = []
        over52_label = ""
        if not search:
            if sel_month:
                start_dt = datetime.strptime(sel_month + "-01", "%Y-%m-%d")
                dim_ov = calendar.monthrange(start_dt.year, start_dt.month)[1]
                d_from = (start_dt - timedelta(days=7)).strftime("%Y%m%d")
                d_to = (datetime(start_dt.year, start_dt.month, dim_ov) + timedelta(days=7)).strftime("%Y%m%d")
                over52_label = f"{start_dt.year}년 {start_dt.month}월 {start_dt.day}일주"
            else:
                today = datetime.now()
                four_weeks_ago = today - timedelta(weeks=4)
                d_from = four_weeks_ago.strftime("%Y%m%d")
                d_to = today.strftime("%Y%m%d")
                over52_label = "최근 4주"
            cur.execute("""
                SELECT t.e_date, t.e_time, t.e_id, t.e_idno, t.e_name,
                       t.e_mode, u.company
                FROM tenter t
                LEFT JOIN tgate g ON t.g_id = g.id
                LEFT JOIN tuser u ON t.e_id = u.id
                WHERE t.e_id >= 0 AND t.e_date >= %s AND t.e_date <= %s
                ORDER BY t.e_date, t.e_time
            """, (d_from, d_to))
            all52 = cur.fetchall()
            emp52 = {}
            for r in all52:
                e_date, e_time, eid = r[0], r[1], r[2]
                e_mode = (r[5] or "").strip()
                key = (e_date, eid)
                if key not in emp52:
                    company = (r[6] or "").strip()
                    emp52[key] = {
                        "date": e_date, "id": eid,
                        "name": (r[4] or "").strip() or f"ID-{eid}",
                        "emp_no": (r[3] or "").strip() or str(eid),
                        "dept": DEPT_MAP.get(company, ""),
                        "in_time": None, "out_time": None,
                    }
                e = emp52[key]
                if e_mode == MODE_IN:
                    if e["in_time"] is None or e_time < e["in_time"]:
                        e["in_time"] = e_time
                elif e_mode == MODE_OUT:
                    if e["out_time"] is None or e_time > e["out_time"]:
                        e["out_time"] = e_time
            _merge_night_shifts(emp52)
            for e in emp52.values():
                in_s = _time_to_sec(e["in_time"])
                out_s = _time_to_sec(e["out_time"])
                is_n = in_s is not None and in_s > 43140
                w, o = 0, 0
                if in_s is not None and out_s is not None and out_s > in_s:
                    wh = (out_s - in_s) / 3600
                    w = min(wh, 8)
                    o = 2 if out_s >= 68400 else 0
                elif is_n and out_s is not None and out_s < in_s:
                    wh = (86400 - in_s + out_s) / 3600
                    w, o = min(wh, 8), (2 if out_s >= 28200 else 0)  # 07:50 이후
                elif in_s is None and out_s is not None and out_s < 71400:
                    w = 8
                elif in_s is None and out_s is not None and out_s >= 71400:
                    w, o = 8, 2
                elif in_s is not None or out_s is not None:
                    w = 8
                e["work_h"] = w
                e["over_h"] = o
            week_emp = {}
            for e in emp52.values():
                dt = datetime.strptime(e["date"], "%Y%m%d")
                iso = dt.isocalendar()
                wk_key = f"{iso[0]}-W{iso[1]:02d}"
                eid = e["id"]
                k = (wk_key, eid)
                if k not in week_emp:
                    mon = dt - timedelta(days=dt.weekday())
                    sun = mon + timedelta(days=6)
                    week_emp[k] = {
                        "week_no": iso[1], "name": e["name"],
                        "emp_no": e["emp_no"], "dept": e["dept"],
                        "start": mon.strftime("%m/%d"),
                        "end": sun.strftime("%m/%d"),
                        "work": 0, "overtime": 0,
                    }
                we = week_emp[k]
                we["work"] += e["work_h"]
                we["overtime"] += e["over_h"]
            for we in week_emp.values():
                total = int(we["work"]) + int(we["overtime"])
                if total >= 52:
                    if sel_dept and we["dept"] != sel_dept:
                        continue
                    we["work"] = int(we["work"])
                    we["overtime"] = int(we["overtime"])
                    we["total"] = total
                    over52_list.append(we)
            over52_list.sort(key=lambda x: -x["total"])
        conn.close()
    except Exception as e:
        flash(f"error: {e}", "danger")
    return render_template("weekly52.html", weeks=weeks,
                           search=search, sel_month=sel_month,
                           emp_name=emp_name,
                           over52_list=over52_list if 'over52_list' in dir() else [],
                           over52_label=over52_label if 'over52_label' in dir() else "",
                           sel_dept=sel_dept, dept_list=dept_list,
                           can_edit=can_edit)


@app.route("/work_schedule")
@_login_required
def work_schedule():
    search = request.args.get("search", "").strip()
    sel_month = request.args.get("month", "").strip()
    sel_dept = request.args.get("dept", "").strip()
    is_admin = session.get("role") == "admin"
    can_edit = is_admin or _has_perm("schedule")
    own_dept = ""
    initial_dept_view = (
        not is_admin
        and "dept" not in request.args
        and "search" not in request.args
        and "page" not in request.args
    )
    if not is_admin and (initial_dept_view or not sel_dept):
        try:
            conn_user = _conn()
            cur_user = conn_user.cursor()
            cur_user.execute(
                "SELECT dept FROM employee_roster WHERE name = %s AND dept IS NOT NULL AND dept <> '' LIMIT 1",
                (session.get("user_name", ""),),
            )
            row_user = cur_user.fetchone()
            if row_user and row_user[0]:
                own_dept = (row_user[0] or "").strip()
            conn_user.close()
        except Exception:
            own_dept = ""
    if initial_dept_view and own_dept:
        sel_dept = own_dept
    elif not can_edit and not sel_dept:
        # 부서 필터가 없으면 본인만
        search = session.get("user_name", "")
    page = max(1, int(request.args.get("page", 1) or 1))
    export_salary = request.args.get("export") == "salary"
    per_page_val = request.args.get("per_page", "20")
    PER_PAGE = int(per_page_val) if per_page_val.isdigit() and int(per_page_val) > 0 else 0
    if not sel_month:
        sel_month = datetime.now().strftime("%Y-%m")
    year = int(sel_month[:4])
    mon = int(sel_month[5:7])
    dim = calendar.monthrange(year, mon)[1]
    month_label = f"{year}년 {mon}월"
    day_names = ["월", "화", "수", "목", "금", "토", "일"]
    day_info = []
    for d in range(1, dim + 1):
        dt = datetime(year, mon, d)
        wd = dt.weekday()
        hol = KR_HOLIDAYS.get(dt.date(), "")
        day_info.append({"day": d, "weekday": wd, "wd_name": day_names[wd],
                         "is_sat": wd == 5, "is_sun": wd == 6,
                         "is_holiday": bool(hol), "holiday_name": hol})
    employees = []
    total_pages = 1
    total_emp = 0
    override_log = []
    try:
        first_dt = datetime(year, mon, 1)
        last_dt = datetime(year, mon, dim)
        q_start = (first_dt - timedelta(days=2)).strftime("%Y%m%d")
        q_end = (last_dt + timedelta(days=2)).strftime("%Y%m%d")
        conn = _conn()
        cur = conn.cursor()
        # 1단계: e_id 조회 (부서 필터를 SQL에서 처리)
        params = []
        sql = "SELECT DISTINCT t.e_id FROM tenter t"
        if sel_dept:
            sql += " LEFT JOIN employee_roster er ON t.e_name = er.name"
        sql += " WHERE t.e_date >= %s AND t.e_date <= %s AND t.e_id <> -1"
        params += [q_start, q_end]
        if search:
            sql += " AND (t.e_name LIKE %s OR t.e_idno LIKE %s)"
            params += [f"%{search}%", f"%{search}%"]
        if sel_dept:
            sql += " AND er.dept = %s"
            params += [sel_dept]
        cur.execute(sql, params)
        all_eids = [r[0] for r in cur.fetchall()]
        total_emp = len(all_eids)
        if PER_PAGE > 0:
            total_pages = max(1, (total_emp + PER_PAGE - 1) // PER_PAGE)
            page = min(page, total_pages)
            eids = all_eids if export_salary else all_eids[(page - 1) * PER_PAGE: page * PER_PAGE]
        else:
            total_pages = 1
            page = 1
            eids = all_eids
        if eids:
            ph = ",".join(["%s"] * len(eids))
            # 2단계: idx_eid + 날짜 범위로 빠르게 조회
            cur.execute(f"""
                SELECT t.e_date, t.e_time, t.e_id, t.e_idno, t.e_name,
                                             t.e_mode, u.company, er.dept
                FROM tenter t
                     LEFT JOIN tgate g ON t.g_id = g.id
                     LEFT JOIN tuser u ON t.e_id = u.id
                                         LEFT JOIN employee_roster er ON t.e_name = er.name
                WHERE t.e_id IN ({ph})
                  AND t.e_date >= %s AND t.e_date <= %s
                ORDER BY t.e_date, t.e_time
            """, eids + [q_start, q_end])
        else:
            cur.execute("SELECT 1 WHERE 1=0")  # 결과없음
        rows = cur.fetchall()
        day_emp = OrderedDict()
        emp_info = {}
        for r in rows:
            e_date, e_time, eid = r[0], r[1], r[2]
            e_mode = (r[5] or "").strip()
            key = (e_date, eid)
            if key not in day_emp:
                company = (r[6] or "").strip()
                roster_dept = (r[7] or "").strip()
                day_emp[key] = {
                    "date": e_date, "id": eid,
                    "name": (r[4] or "").strip() or f"ID-{eid}",
                    "emp_no": (r[3] or "").strip() or str(eid),
                    "dept": roster_dept or DEPT_MAP.get(company, ""),
                    "in_time": None, "out_time": None,
                }
                if eid not in emp_info:
                    emp_info[eid] = {
                        "name": day_emp[key]["name"],
                        "emp_no": day_emp[key]["emp_no"],
                        "dept": day_emp[key]["dept"],
                    }
            e = day_emp[key]
            if e_mode == MODE_IN:
                if e["in_time"] is None or e_time < e["in_time"]:
                    e["in_time"] = e_time
            elif e_mode == MODE_OUT:
                if e["out_time"] is None or e_time > e["out_time"]:
                    e["out_time"] = e_time
        _merge_night_shifts(day_emp)
        month_prefix = f"{year}{mon:02d}"
        emp_days = defaultdict(dict)
        for e in day_emp.values():
            if not e["date"].startswith(month_prefix):
                continue
            eid = e["id"]
            day_num = int(e["date"][6:8])
            in_s = _time_to_sec(e["in_time"])
            out_s = _time_to_sec(e["out_time"])
            is_night = in_s is not None and in_s > 43140
            dt2 = datetime.strptime(e["date"], "%Y%m%d")
            wd = dt2.weekday()
            is_hol = wd >= 5 or bool(KR_HOLIDAYS.get(dt2.date(), ""))
            basic, ot_h, ni_h, etc_h = 0, 0, 0, 0
            if is_night:
                if out_s is not None and out_s < in_s:
                    wh = (86400 - in_s + out_s) / 3600
                    basic = int(min(wh, 8))
                    ot_h = 2 if out_s >= 28200 else 0
                    night_sec = max(0, 86400 - max(in_s, 79200)) + max(0, min(out_s, 21600))
                    ni_h = max(0, round(night_sec / 3600 - 1))
                elif out_s is not None and out_s > in_s:
                    pass
                elif in_s is not None:
                    basic = 8
                    assumed_out = 18000
                    night_sec = max(0, 86400 - max(in_s, 79200)) + max(0, min(assumed_out, 21600))
                    ni_h = max(0, round(night_sec / 3600 - 1))
            else:
                if in_s is not None and out_s is not None and out_s > in_s:
                    wh = (out_s - in_s) / 3600
                    basic = int(min(wh, 8))
                    ot_h = 2 if out_s >= 68400 else 0
                elif in_s is None and out_s is not None and out_s < 28800:
                    assumed_in = 68400
                    wh = (86400 - assumed_in + out_s) / 3600
                    basic = int(min(wh, 8))
                    ot_h = 2 if out_s >= 28200 else 0
                    night_sec = max(0, 86400 - max(assumed_in, 79200)) + max(0, min(out_s, 21600))
                    ni_h = max(0, round(night_sec / 3600 - 1))
                elif in_s is None and out_s is not None and out_s >= 71400:
                    basic = 8
                    ot_h = 2
                elif in_s is None and out_s is not None:
                    basic = 8
                elif in_s is not None or out_s is not None:
                    basic = 8
                    if is_night and out_s is None:
                        assumed_out = 18000
                        night_sec = max(0, 86400 - max(in_s, 79200)) + max(0, min(assumed_out, 21600))
                        ni_h = max(0, round(night_sec / 3600 - 1))
            emp_days[eid][day_num] = {
                "basic": basic, "overtime": ot_h,
                "night": ni_h, "other": etc_h,
                "is_holiday": is_hol,
            }
        # ── schedule_record(보고서 결재분) 병합 ──
        month_str_sr = f"{year}{mon:02d}"
        # tuser name→id 매핑
        cur.execute("SELECT id, name, idno, company FROM tuser WHERE name IS NOT NULL AND name <> ''")
        name_to_eid = {}
        for uid, uname, uidno, ucomp in cur.fetchall():
            n = (uname or "").strip()
            if n and n not in name_to_eid:
                name_to_eid[n] = uid
        # sel_dept 필터용 허용 이름 집합 (부서 필터 선택 시 다른 부서 결재분 제외)
        dept_allowed_names = None
        if sel_dept:
            cur.execute("SELECT name FROM employee_roster WHERE dept = %s AND name IS NOT NULL AND name <> ''", (sel_dept,))
            dept_allowed_names = {(r[0] or "").strip() for r in cur.fetchall()}
        # 현재 페이지 eids 집합 (페이지네이션 정합성: 이 집합 외 직원은 emp_info에 추가하지 않음)
        eids_set = set(eids)
        cur.execute("""
            SELECT emp_name, work_date, basic_h, ot_h, night_h, etc
            FROM schedule_record
            WHERE work_date LIKE %s AND source_type IN ('보고서','수정','수정근무보고서')
            ORDER BY FIELD(source_type,'보고서','수정','수정근무보고서')
        """, (f"{month_str_sr}%",))
        for sr_name, sr_date, sr_basic, sr_ot, sr_night, sr_etc in cur.fetchall():
            sr_name = (sr_name or "").strip()
            if dept_allowed_names is not None and sr_name not in dept_allowed_names:
                continue
            sr_eid = name_to_eid.get(sr_name)
            if not sr_eid:
                continue
            day_num_sr = int(str(sr_date)[6:8])
            dt_sr = datetime(year, mon, day_num_sr)
            wd_sr = dt_sr.weekday()
            is_hol_sr = wd_sr >= 5 or bool(KR_HOLIDAYS.get(dt_sr.date(), ""))
            # 보고서 결재 데이터가 출입기록보다 우선 적용
            # (현재 페이지 eids에 없는 직원은 emp_info 추가 생략 → 페이지네이션 정합성 유지)
            if sr_eid not in emp_info and sr_eid in eids_set:
                cur.execute(
                    """
                    SELECT tu.id, tu.name, tu.idno, tu.company, er.dept
                    FROM tuser tu
                    LEFT JOIN employee_roster er ON er.name = tu.name
                    WHERE tu.id=%s
                    """,
                    (sr_eid,),
                )
                tu = cur.fetchone()
                if tu:
                    emp_info[sr_eid] = {
                        "name": (tu[1] or "").strip(),
                        "emp_no": (tu[2] or "").strip() or str(tu[0]),
                        "dept": (tu[4] or "").strip() or DEPT_MAP.get((tu[3] or "").strip(), ""),
                    }
            # emp_days 업데이트는 emp_info에 있는 직원(현재 페이지)만 처리
            if sr_eid in emp_info:
                other_val = sr_etc if sr_etc else 0
                emp_days[sr_eid][day_num_sr] = {
                    "basic": sr_basic or 0, "overtime": sr_ot or 0,
                    "night": sr_night or 0, "other": other_val,
                    "is_holiday": is_hol_sr,
                }
        # 오버라이드 로딩
        # schedule_record에서 추가된 인원도 eids에 포함
        sr_extra_eids = [eid for eid in emp_days if eid not in eids]
        all_eids_for_ov = list(eids) + sr_extra_eids
        overrides = {}
        override_memos = {}
        override_updates = {}
        if all_eids_for_ov:
            ph2 = ",".join(["%s"] * len(all_eids_for_ov))
            month_str = f"{year}{mon:02d}"
            cur.execute(f"""
                SELECT e_id, e_date, col_type, value, memo, updated_at FROM work_override
                WHERE e_id IN ({ph2}) AND e_date LIKE %s
            """, all_eids_for_ov + [f"{month_str}%"])
            for row in cur.fetchall():
                v = row[3]
                try:
                    v = float(v)
                except (ValueError, TypeError):
                    pass
                key3 = (row[0], row[1], row[2])
                overrides[key3] = v
                if row[4]:
                    override_memos[key3] = row[4]
                if row[5]:
                    override_updates[key3] = row[5].strftime("%m/%d %H:%M") if hasattr(row[5], 'strftime') else str(row[5])
        col_label = {"basic": "기본", "overtime": "연장", "night": "야간", "other": "기타"}
        override_log = []
        for eid, info in emp_info.items():
            dd = emp_days.get(eid, {})
            if not dd:
                continue
            b_r, o_r, n_r, e_r = [], [], [], []
            b_ov, o_ov, n_ov, e_ov = [], [], [], []
            b_memo, o_memo, n_memo, e_memo = [], [], [], []
            s_ot = s_ni = wd_ot = nw = hw = ho = 0
            unpaid = 0
            hol_basic = hol_ot_h = hol_night = 0
            hol_time = 0
            for d in range(1, dim + 1):
                rec = dd.get(d)
                date_str = f"{year}{mon:02d}{d:02d}"
                is_hol = rec["is_holiday"] if rec else day_info[d-1]["is_sat"] or day_info[d-1]["is_sun"] or day_info[d-1]["is_holiday"]
                bv = overrides.get((eid, date_str, "basic"),   rec["basic"]   if rec else None)
                ov = overrides.get((eid, date_str, "overtime"),rec["overtime"]if rec else None)
                nv = overrides.get((eid, date_str, "night"),   rec["night"]   if rec else None)
                ev = overrides.get((eid, date_str, "other"),   rec["other"]   if rec else None)
                b_r.append(bv); o_r.append(ov); n_r.append(nv); e_r.append(ev)
                b_ov.append((eid, date_str, "basic")    in overrides)
                o_ov.append((eid, date_str, "overtime") in overrides)
                n_ov.append((eid, date_str, "night")    in overrides)
                e_ov.append((eid, date_str, "other")    in overrides)
                b_memo.append(override_memos.get((eid, date_str, "basic"), ""))
                o_memo.append(override_memos.get((eid, date_str, "overtime"), ""))
                n_memo.append(override_memos.get((eid, date_str, "night"), ""))
                e_memo.append(override_memos.get((eid, date_str, "other"), ""))
                for ct in ["basic", "overtime", "night", "other"]:
                    k3 = (eid, date_str, ct)
                    if k3 in overrides and k3 in override_memos:
                        override_log.append({
                            "name": info["name"], "date": f"{d}일",
                            "col": col_label[ct], "value": overrides[k3],
                            "memo": override_memos[k3],
                            "updated": override_updates.get(k3, ""),
                        })
                if isinstance(ev, str) and "무급" in ev:
                    unpaid += 1
                    hol_time += 8
                elif isinstance(ev, str) and "연차" in ev:
                    pass
                elif isinstance(ev, (int, float)) and ev > 0:
                    hol_time += ev
                bv = bv or 0; ov = ov or 0; nv = nv or 0
                s_ot += ov; s_ni += nv; nw += nv
                # 휴일근로: 토요일+일요일+공휴일의 기본시간만
                is_hol_strict = day_info[d-1]["is_sat"] or day_info[d-1]["is_sun"] or day_info[d-1]["is_holiday"]
                if is_hol_strict:
                    hw += bv
                    ho += ov
                if is_hol:
                    hol_basic += bv; hol_ot_h += ov; hol_night += nv
                if not is_hol:
                    wd_ot += ov
            # 합산 override 적용 (e_date = YYYYMM00)
            sum_date = f"{year}{mon:02d}00"
            final_sum_ot    = overrides.get((eid, sum_date, "sum_ot"),    0)
            final_sum_ni    = overrides.get((eid, sum_date, "sum_night"), 0)
            final_hol_work  = overrides.get((eid, sum_date, "hol_work"), 0)
            final_hol_ot    = overrides.get((eid, sum_date, "hol_ot"),   0)
            # 숫자 변환
            try: final_sum_ot   = float(final_sum_ot)
            except: final_sum_ot = 0
            try: final_sum_ni   = float(final_sum_ni)
            except: final_sum_ni = 0
            try: final_hol_work = float(final_hol_work)
            except: final_hol_work = 0
            try: final_hol_ot   = float(final_hol_ot)
            except: final_hol_ot = 0
            # 합산 override 메모
            sum_ov = {}
            sum_memo_map = {}
            for sc in ["sum_ot", "sum_night", "hol_work", "hol_ot"]:
                sk = (eid, sum_date, sc)
                sum_ov[sc] = sk in overrides
                sum_memo_map[sc] = override_memos.get(sk, "")
                if sk in overrides and sk in override_memos:
                    col_kr = {"sum_ot":"연장","sum_night":"야간","hol_work":"휴일","hol_ot":"휴연"}
                    override_log.append({
                        "name": info["name"], "date": "합산",
                        "col": col_kr[sc], "value": overrides[sk],
                        "memo": override_memos[sk],
                        "updated": override_updates.get(sk, ""),
                    })
            employees.append({
                "id": eid,
                "emp_no": info["emp_no"], "dept": info["dept"], "name": info["name"],
                "basic": b_r, "overtime": o_r, "night": n_r, "other": e_r,
                "basic_ov": b_ov, "overtime_ov": o_ov, "night_ov": n_ov, "other_ov": e_ov,
                "basic_memo": b_memo, "overtime_memo": o_memo, "night_memo": n_memo, "other_memo": e_memo,
                "sum_ot": int(final_sum_ot), "sum_night": int(final_sum_ni),
                "hol_work": int(final_hol_work), "hol_ot": int(final_hol_ot),
                "wd_ot": int(wd_ot + final_sum_ot), "night_work": int(nw + final_sum_ni),
                "calc_hw": int(hw + final_hol_work), "calc_ho": int(ho + final_hol_ot),
                "unpaid": unpaid, "hol_time": int(hol_time),
                "hol_basic": int(hol_basic), "hol_ot_h": int(hol_ot_h),
                "hol_night": int(hol_night), "hol_total": int(hol_basic + hol_ot_h + hol_night),
                "sum_ov": sum_ov, "sum_memo": sum_memo_map,
            })
        conn.close()
    except Exception as ex:
        flash(f"error: {ex}", "danger")
    try:
        conn2 = _conn()
        cur2 = conn2.cursor()
        cur2.execute("SELECT DISTINCT dept FROM employee_roster WHERE dept IS NOT NULL AND dept <> '' ORDER BY dept")
        dept_list = [r[0] for r in cur2.fetchall()]
        conn2.close()
    except Exception:
        dept_list = sorted(set(DEPT_MAP.values()))
    # 급여출력폼 Excel 내보내기
    if export_salary and employees:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        wbx = Workbook()
        wsx = wbx.active
        wsx.title = "급여출력폼"
        # R1: 타이틀 (병합)
        wsx.merge_cells("A1:F1")
        wsx["A1"].value = f"{month_label} 급여출력폼"
        wsx["A1"].font = Font(bold=True, size=14)
        wsx["A1"].alignment = Alignment(horizontal="center")
        # R2: 헤더
        headers = ["사원", "연장근로시간", "야간근로시간", "휴일근로시간", "휴일연장근로시간", "인정교통시간"]
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for ci, h in enumerate(headers, 1):
            c = wsx.cell(2, ci, h)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")
            c.border = border
        # R3~: 데이터
        for ri, emp in enumerate(employees, 3):
            vals = [emp["name"], emp["wd_ot"], emp["night_work"],
                    emp["calc_hw"], emp["calc_ho"], emp["hol_time"]]
            for ci, v in enumerate(vals, 1):
                c = wsx.cell(ri, ci, v)
                c.border = border
                if ci >= 2:
                    c.alignment = Alignment(horizontal="center")
        # 열너비
        wsx.column_dimensions["A"].width = 12
        for col in ["B", "C", "D", "E", "F"]:
            wsx.column_dimensions[col].width = 16
        buf = io.BytesIO()
        wbx.save(buf)
        buf.seek(0)
        fname = f"급여출력폼_{year}{mon:02d}.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return render_template("work_schedule.html",
                           search=search, sel_month=sel_month,
                           sel_dept=sel_dept, dept_list=dept_list,
                           month_label=month_label,
                           days_in_month=dim, day_info=day_info,
                           employees=employees,
                           page=page, total_pages=total_pages,
                           total_emp=total_emp,
                           per_page=PER_PAGE,
                           override_log=override_log,
                           can_edit=_has_perm("schedule"))


@app.route("/save_override", methods=["POST"])
@_login_required
def save_override():
    if not _has_perm("schedule"):
        return jsonify({"ok": False, "error": "권한이 없습니다"}), 403
    try:
        data = request.get_json()
        e_id     = int(data["e_id"])
        e_date   = str(data["e_date"])
        col_type = str(data["col_type"])
        value    = data.get("value", "")
        memo     = data.get("memo", "").strip()
        conn = _conn(); cur = conn.cursor()
        if value == "" or value is None:
            cur.execute("DELETE FROM work_override WHERE e_id=%s AND e_date=%s AND col_type=%s",
                        (e_id, e_date, col_type))
        else:
            if not memo:
                return jsonify({"ok": False, "error": "메모를 입력하세요."}), 400
            v = value
            if col_type != 'other':
                v = str(float(value))
            cur.execute("""
                INSERT INTO work_override (e_id, e_date, col_type, value, memo)
                VALUES (%s,%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE value=%s, memo=%s, updated_at=NOW()
            """, (e_id, e_date, col_type, v, memo, v, memo))
        # 기타 컬럼 연차/반차를 leave_records 자동 연동
        if col_type == "other":
            _sync_leave_from_other(cur, e_id, e_date, value)
        conn.commit(); conn.close()
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/schedule_record")
@app.route("/schedule_record/<fname>")  # URL 끝 파일명 → 다운로드 저장명 (프록시가 헤더 제거해도 유지)
@_login_required
def schedule_record(fname=None):
    is_admin = session.get("role") == "admin"
    can_edit = is_admin or _has_perm("schedule")
    if not can_edit:
        flash("접근 권한이 없습니다.", "danger")
        return redirect(url_for("dashboard"))
    sel_month = request.args.get("month", "").strip()
    search = request.args.get("search", "").strip()
    dept_search = request.args.get("dept", "").strip()
    if not can_edit:
        search = session.get("user_name", "")
        dept_search = ""
    if not sel_month:
        sel_month = datetime.now().strftime("%Y-%m")
    year = int(sel_month[:4])
    mon = int(sel_month[5:7])
    dim = calendar.monthrange(year, mon)[1]
    month_label = f"{year}년 {mon}월"
    day_names = ["월", "화", "수", "목", "금", "토", "일"]
    day_info = []
    for d in range(1, dim + 1):
        dt = datetime(year, mon, d)
        wd = dt.weekday()
        hol = KR_HOLIDAYS.get(dt.date(), "")
        day_info.append({"day": d, "weekday": wd, "wd_name": day_names[wd],
                         "is_sat": wd == 5, "is_sun": wd == 6,
                         "is_holiday": bool(hol), "holiday_name": hol})
    employees = []
    try:
        conn = _conn(); cur = conn.cursor()
        month_str = f"{year}{mon:02d}"
        # 수정 우선, 원본 fallback
        cur.execute("""
            SELECT emp_name, work_date, basic_h, ot_h, night_h, etc, source_type
            FROM schedule_record
            WHERE work_date LIKE %s
            ORDER BY emp_name, work_date
        """, (f"{month_str}%",))
        raw = cur.fetchall()
        # tuser 매칭 (이름과 사번 부분)
        cur.execute("SELECT id, name, idno, company FROM tuser")
        tuser_map = {}
        for uid, uname, uidno, ucomp in cur.fetchall():
            n = (uname or "").strip()
            if not n:
                continue
            eno = (uidno or "").strip()
            if n not in tuser_map or (eno and not tuser_map[n]["emp_no_raw"]):
                tuser_map[n] = {"emp_no": eno or str(uid), "emp_no_raw": eno, "dept": DEPT_MAP.get((ucomp or "").strip(), "")}
        # employee_roster 부서 우선 적용 (근무표와 동일 기준)
        cur.execute("SELECT name, dept FROM employee_roster WHERE dept IS NOT NULL AND dept <> ''")
        for rname, rdept in cur.fetchall():
            n = (rname or "").strip()
            if n in tuser_map:
                tuser_map[n]["dept"] = rdept
        # 퇴사일 조회 (중도 퇴사자 근무일 = 퇴사일)
        roster_retire = {}
        cur.execute("SELECT name, retire_date FROM employee_roster")
        for rname, rret in cur.fetchall():
            n = (rname or "").strip()
            if n:
                roster_retire[n] = str(rret or "").strip()
        conn.close()
        # ERP 근무일: 해당월 마지막날 (다음달 입력 기준). 중도 퇴사자는 퇴사일.
        month_end = f"{year}-{mon:02d}-{dim:02d}"
        _dow_kr = ["월요일", "화요일", "수요일", "목요일", "금요일", "토요일", "일요일"]
        def _erp_wk_date(emp_name):
            ret = roster_retire.get(emp_name, "")
            if ret:
                digits = ret.replace("-", "").replace("/", "")[:8]
                if len(digits) == 8 and digits[:4] == str(year) and digits[4:6] == f"{mon:02d}":
                    return f"{digits[:4]}-{digits[4:6]}-{digits[6:8]}"
            return month_end
        def _erp_wk_dayname(date_str):
            try:
                dt = datetime.strptime(date_str, "%Y-%m-%d")
                return _dow_kr[dt.weekday()]
            except Exception:
                return ""
        # 수정 우선 병합
        merged = {}
        for name, wdate, bh, oh, nh, etc, src in raw:
            key = (name, wdate)
            if key not in merged or src == '수정':
                merged[key] = {"basic": bh, "ot": oh, "night": nh, "etc": etc, "src": src}
        # 직원별 그룹핑
        emp_data = OrderedDict()
        for (name, wdate), vals in merged.items():
            if search and search not in name:
                continue
            info = tuser_map.get(name, {"emp_no": "", "dept": ""})
            if dept_search and info["dept"] != dept_search:
                continue
            if name not in emp_data:
                emp_data[name] = {"name": name, "emp_no": info["emp_no"], "dept": info["dept"], "days": {}}
            day_num = int(wdate[6:8])
            emp_data[name]["days"][day_num] = vals
        for idx, (name, ed) in enumerate(emp_data.items()):
            dd = ed["days"]
            b_r, o_r, n_r, e_r = [], [], [], []
            s_ot = s_ni = wd_ot = nw = hw = ho = 0
            unpaid = 0
            hol_basic = hol_ot_h = hol_night = hol_time = 0
            etc_sum = 0.0   # 기타 행 숫자(조퇴/외출/공제) 합계 → "조퇴외출" 칸
            work_days = 0     # 근무일수 (기본근무 있는 날)
            basic_total = 0   # 기본근무시간 총합
            annual_days = 0.0 # 연차일수 (기타에 '연차')
            for d in range(1, dim + 1):
                rec = dd.get(d)
                is_hol = day_info[d-1]["is_sat"] or day_info[d-1]["is_sun"] or day_info[d-1]["is_holiday"]
                if rec:
                    bv, ov, nv, ev = rec["basic"], rec["ot"], rec["night"], rec["etc"]
                else:
                    bv, ov, nv, ev = None, None, None, None
                b_r.append(bv); o_r.append(ov); n_r.append(nv); e_r.append(ev)
                bval = bv or 0; oval = ov or 0; nval = nv or 0
                if bval > 0:
                    work_days += 1; basic_total += bval
                s_ot += oval; s_ni += nval; nw += nval
                is_hol_strict = day_info[d-1]["is_sat"] or day_info[d-1]["is_sun"] or day_info[d-1]["is_holiday"]
                if is_hol_strict:
                    hw += bval; ho += oval
                if is_hol:
                    hol_basic += bval; hol_ot_h += oval; hol_night += nval
                if not is_hol:
                    wd_ot += oval
                if isinstance(ev, str) and "무급" in ev:
                    unpaid += 1; hol_time += 8
                elif isinstance(ev, str) and "연차" in ev:
                    annual_days += 0.5 if "반차" in ev else 1
                elif ev is not None:
                    try:
                        fv = float(str(ev))
                        if fv > 0:
                            hol_time += fv     # 소수점 유지 (3.5 등)
                            etc_sum += fv      # 소수점 유지 합산 (조퇴/외출/공제)
                    except (ValueError, TypeError):
                        pass
            employees.append({
                "emp_no": ed["emp_no"], "dept": ed["dept"], "name": name,
                "basic": b_r, "overtime": o_r, "night": n_r, "other": e_r,
                "wd_ot": int(wd_ot), "night_work": int(nw),
                "calc_hw": int(hw), "calc_ho": int(ho),
                "unpaid": unpaid, "hol_time": ("%g" % hol_time) if hol_time else 0,
                "etc_sum": ("%g" % etc_sum) if etc_sum else 0,
                "hol_basic": int(hol_basic), "hol_ot_h": int(hol_ot_h),
                "hol_night": int(hol_night), "hol_total": int(hol_basic + hol_ot_h + hol_night),
                # ERP 근무내역등록 매핑용
                "wk_date": _erp_wk_date(name),
                "wk_dayname": _erp_wk_dayname(_erp_wk_date(name)),
                "work_days": work_days,
                "basic_total": int(basic_total),
                "annual_days": ("%g" % annual_days) if annual_days else 0,
            })
    except Exception as ex:
        flash(f"오류: {ex}", "danger")
    try:
        conn2 = _conn()
        cur2 = conn2.cursor()
        cur2.execute("SELECT DISTINCT dept FROM employee_roster WHERE dept IS NOT NULL AND dept <> '' ORDER BY dept")
        dept_list = [r[0] for r in cur2.fetchall()]
        conn2.close()
    except Exception:
        dept_list = sorted(set(DEPT_MAP.values()))

    # ERP 근무내역등록 붙여넣기용 엑셀 내보내기
    if request.args.get("export") == "erp_xlsx" and employees:
        import io
        from openpyxl import Workbook
        wbx = Workbook()
        wsx = wbx.active
        wsx.title = "ERP근무내역"
        headers = ["근무일", "요일", "사원", "영문명", "사번", "사원구분", "부서",
                   "급여형태", "근무유형", "일구분", "입사일", "퇴직일", "재직/퇴직",
                   "파견기관", "시작익일", "시작시간", "종료시간",
                   "근무일수", "기본근무시간", "연차지급일수", "연장근로시간",
                   "야간근로시간", "휴일근로시간", "휴일연장근로시간", "외출조퇴시간", "시간외수당"]
        for ci, h in enumerate(headers, 1):
            wsx.cell(1, ci, h)
        def _num(v):
            try:
                return float(str(v))
            except (ValueError, TypeError):
                return 0
        for ri, emp in enumerate(employees, 2):
            row = [emp["wk_date"], emp.get("wk_dayname", ""), emp["name"], "", emp["emp_no"],
                   "", emp["dept"], "", "", "", "", "", "", "", "", "", "",
                   dim, 209, 0,
                   _num(emp["wd_ot"]), _num(emp["night_work"]), _num(emp["calc_hw"]),
                   _num(emp["calc_ho"]), _num(emp["hol_time"]), ""]
            for ci, v in enumerate(row, 1):
                wsx.cell(ri, ci, v)
        buf = io.BytesIO()
        wbx.save(buf)
        # Content-Disposition 없이 xlsx 바이트만 반환 → <a download> 속성이 파일명 결정
        # (프록시가 Content-Disposition을 제거해 UUID로 저장되는 문제 회피)
        from flask import Response
        return Response(buf.getvalue(),
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    return render_template("schedule_record.html",
                           search=search, dept_search=dept_search,
                           dept_list=dept_list, sel_month=sel_month,
                           month_label=month_label,
                           days_in_month=dim, day_info=day_info,
                           employees=employees, can_edit=can_edit)


@app.route("/upload_schedule_record", methods=["POST"])
@_login_required
def upload_schedule_record():
    if session.get("role") != "admin" and not _has_perm("schedule"):
        flash("접근 권한이 없습니다.", "danger")
        return redirect(url_for("dashboard"))
    file = request.files.get("file")
    year = request.form.get("year", "")
    if not file or not file.filename:
        flash("파일을 선택하세요.", "warning")
        return redirect(url_for("schedule_record"))
    if not year:
        year = datetime.now().year
    year = int(year)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file, data_only=True)

        # 전기/전자 시트 자동 감지
        elec_sheet = None  # 전자
        power_sheet = None  # 전기
        for sn in wb.sheetnames:
            if "전자" in sn:
                elec_sheet = sn
            if "전기" in sn:
                power_sheet = sn

        records = []

        # ── 전자사업부 파싱 ──
        if elec_sheet:
            ws = wb[elec_sheet]
            blocks = []
            r1_left = ws.cell(1, 1).value or ""
            r1_right = ws.cell(1, 8).value or ""
            for title, cols in [(r1_left, "left"), (r1_right, "right")]:
                if not title:
                    continue
                if "(수정)" not in str(title):
                    continue
                m = re.search(r'(\d{2})/(\d{2})', str(title))
                if not m:
                    continue
                mm, dd = int(m.group(1)), int(m.group(2))
                wdate = f"{year}{mm:02d}{dd:02d}"
                src = "수정" if "(수정)" in str(title) else "원본"
                blocks.append({"wdate": wdate, "src": src, "side": cols})
            for blk in blocks:
                wdate, src, side = blk["wdate"], blk["src"], blk["side"]
                if side == "left":
                    col_map = {1: "주간기본", 3: "주간연장", 4: "야간기본", 5: "야간연장", 6: "기타"}
                    ded_col = 2
                    leave_col = 7
                else:
                    col_map = {8: "주간기본", 9: "주간연장", 10: "야간기본", 11: "야간연장", 12: "기타"}
                    ded_col = None
                    leave_col = None
                for r in range(4, ws.max_row + 1):
                    for col_idx, cat in col_map.items():
                        name = ws.cell(r, col_idx).value
                        if not name or not isinstance(name, str):
                            continue
                        name = name.strip()
                        if not name:
                            continue
                        basic, ot, night, etc = 0, 0, 0, None
                        if cat == "주간기본":
                            basic = 8
                            if ded_col:
                                ded = ws.cell(r, ded_col).value
                                if ded is not None:
                                    try:
                                        dv = int(float(str(ded)))
                                        if dv > 0:
                                            etc = str(dv)
                                    except (ValueError, TypeError):
                                        pass
                        elif cat == "주간연장":
                            basic, ot = 8, 2
                        elif cat == "야간기본":
                            basic, night = 8, 6
                        elif cat == "야간연장":
                            basic, ot, night = 8, 2, 7
                        elif cat == "기타":
                            if leave_col:
                                lt = ws.cell(r, leave_col).value
                                if lt and isinstance(lt, str) and lt.strip():
                                    etc = lt.strip()
                        records.append((name, wdate, basic, ot, night, etc, src, "전자",
                                        session.get("user_name", "")))

        # ── 전기사업부 파싱 ──
        if power_sheet:
            ws = wb[power_sheet]
            # R1 C2에서 날짜 파싱: "04/02 (목) 근무 보고서"
            r1_title = str(ws.cell(1, 2).value or "")
            m = re.search(r'(\d{2})/(\d{2})', r1_title)
            if m:
                mm, dd = int(m.group(1)), int(m.group(2))
                wdate = f"{year}{mm:02d}{dd:02d}"
                # R3 헤더: C1=현인원, C2=주간연장, C3=연차, C4=무급, C5=조퇴, C6=조퇴시간
                # C1: 현인원(주간기본) → basic=8
                # C2: 주간연장 → basic=8, ot=2
                # C3: 연차 → etc="연차"
                # C4: 무급 → etc="무급"
                # C5: 조퇴 → etc="조퇴", C6=조퇴시간
                # R3 헤더명 제외 목록
                header_names = {'현인원', '주간연장', '연차', '무급', '조퇴', '조퇴시간'}
                col_cats = {1: "현인원", 2: "주간연장", 3: "연차", 4: "무급", 5: "조퇴"}
                # 같은 사람이 여러 열에 있을 수 있으므로 dict로 병합
                power_merged = {}
                for r in range(4, ws.max_row + 1):
                    for col_idx, cat in col_cats.items():
                        name = ws.cell(r, col_idx).value
                        if not name or not isinstance(name, str):
                            continue
                        name = name.strip()
                        if not name or name in header_names:
                            continue
                        basic, ot, night, etc = 0, 0, 0, None
                        if cat == "현인원":
                            basic = 8
                        elif cat == "주간연장":
                            basic, ot = 8, 2
                        elif cat == "연차":
                            etc = "연차"
                        elif cat == "무급":
                            etc = "무급"
                        elif cat == "조퇴":
                            basic = 8
                            ct = ws.cell(r, 6).value
                            if ct is not None:
                                try:
                                    etc = str(int(float(str(ct))))
                                except:
                                    etc = "조퇴"
                            else:
                                etc = "조퇴"
                        key = name
                        if key in power_merged:
                            prev = power_merged[key]
                            prev["basic"] = max(prev["basic"], basic)
                            prev["ot"] = max(prev["ot"], ot)
                            prev["night"] = max(prev["night"], night)
                            if etc:
                                prev["etc"] = etc
                                if etc in ("무급", "연차"):
                                    prev["basic"] = 0
                                    prev["ot"] = 0
                                    prev["night"] = 0
                        else:
                            power_merged[key] = {"basic": basic, "ot": ot, "night": night, "etc": etc}
                for name, vals in power_merged.items():
                    # 최종 검증: 무급/연차면 basic/ot/night 강제 0
                    if vals["etc"] in ("무급", "연차"):
                        vals["basic"] = 0
                        vals["ot"] = 0
                        vals["night"] = 0
                    records.append((name, wdate, vals["basic"], vals["ot"], vals["night"], vals["etc"],
                                    "원본", "전기", session.get("user_name", "")))

        if not elec_sheet and not power_sheet:
            flash("'전자' 또는 '전기' 시트를 찾을 수 없습니다.", "danger")
            return redirect(url_for("schedule_record"))

        if not records:
            flash("파싱된 데이터가 없습니다.", "warning")
            return redirect(url_for("schedule_record"))
        conn = _conn(); cur = conn.cursor()
        inserted = 0
        for rec in records:
            cur.execute("""
                INSERT INTO schedule_record (emp_name, work_date, basic_h, ot_h, night_h, etc,
                                             source_type, sheet_name, uploaded_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE basic_h=%s, ot_h=%s, night_h=%s, etc=%s,
                                        uploaded_at=NOW(), uploaded_by=%s
            """, (rec[0], rec[1], rec[2], rec[3], rec[4], rec[5], rec[6], rec[7], rec[8],
                  rec[2], rec[3], rec[4], rec[5], rec[8]))
            inserted += 1
        conn.commit(); conn.close()
        dates_str = ", ".join(sorted(set(r[1] for r in records)))
        flash(f"일{inserted}건 업로드 완료 (날짜: {dates_str})", "success")
        # 업로드한 달로 이동
        first_date = records[0][1]
        redir_month = f"{first_date[:4]}-{first_date[4:6]}"
        return redirect(url_for("schedule_record", month=redir_month))
    except Exception as ex:
        flash(f"업로드 오류: {ex}", "danger")
        return redirect(url_for("schedule_record"))


@app.route("/welfare")
@_login_required
def welfare():
    return render_template("welfare.html")


@app.route("/leave_dashboard")
@_login_required
def leave_dashboard():
    """연차관리 대시보드 — annual_leave(ERP 동기화 결과)를 집계해 보여준다.

    숫자는 전부 annual_leave 한 곳에서 나온다. 화면에서 다시 계산하지 않는다.
    """
    try:
        year = int(request.args.get("year", datetime.now().year))
    except ValueError:
        year = datetime.now().year
    conn = _conn(); cur = conn.cursor()

    cur.execute("SELECT DISTINCT year FROM annual_leave ORDER BY year DESC")
    year_list = [r[0] for r in cur.fetchall()] or [year]

    cur.execute("""SELECT COUNT(*), IFNULL(SUM(generated),0), IFNULL(SUM(total),0),
                          IFNULL(SUM(used),0)
                     FROM annual_leave WHERE year=%s""", (year,))
    cnt, gen, tot, used = cur.fetchone()
    s = {"people": cnt, "generated": float(gen), "total": float(tot), "used": float(used),
         "rest": float(tot) - float(used),
         "rate": (float(used) / float(tot) * 100) if float(tot) else 0.0}

    # 부서는 화면 다른 곳과 같게 tuser.company → DEPT_MAP 으로 묶는다
    cur.execute("""SELECT t.company, COUNT(*), IFNULL(SUM(a.total),0), IFNULL(SUM(a.used),0)
                     FROM annual_leave a JOIN tuser t ON t.id = a.e_id
                    WHERE a.year=%s GROUP BY t.company""", (year,))
    agg = {}
    for comp, n, t2, u2 in cur.fetchall():
        d = agg.setdefault(DEPT_MAP.get(comp, "기타"),
                           {"people": 0, "total": 0.0, "used": 0.0})
        d["people"] += n; d["total"] += float(t2); d["used"] += float(u2)
    depts = []
    for name, d in agg.items():
        d["name"] = name
        d["rest"] = d["total"] - d["used"]
        d["rate"] = (d["used"] / d["total"] * 100) if d["total"] else 0.0
        depts.append(d)
    depts.sort(key=lambda x: -x["total"])

    cur.execute("SELECT %s FROM annual_leave WHERE year=%%s"
                % ", ".join("IFNULL(SUM(m%d),0)" % i for i in range(1, 13)), (year,))
    mvals = [float(x or 0) for x in cur.fetchone()]
    mmax = max(mvals) if mvals else 0
    months = [{"v": v, "h": int(v / mmax * 110) if mmax else 4} for v in mvals]

    cur.execute("""SELECT a.e_name, t.company, a.total, a.used, (a.total - a.used) AS rest
                     FROM annual_leave a LEFT JOIN tuser t ON t.id = a.e_id
                    WHERE a.year=%s ORDER BY rest DESC""", (year,))
    rows = []
    for nm, comp, t2, u2, rest in cur.fetchall():
        t2, u2, rest = float(t2), float(u2), float(rest)
        rows.append({"name": nm or "", "dept": DEPT_MAP.get(comp, "기타"),
                     "total": t2, "used": u2, "rest": rest,
                     "rate": (u2 / t2 * 100) if t2 else 0.0})

    negatives = sorted([r for r in rows if r["rest"] < 0], key=lambda r: r["rest"])
    unused = [r for r in rows if r["used"] == 0 and r["total"] > 0]
    cur.execute("""SELECT e_name FROM annual_leave
                    WHERE year=%s AND IFNULL(generated,0)=0 AND used>0
                    ORDER BY e_name""", (year,))
    no_grant = [{"name": r[0]} for r in cur.fetchall()]

    cur.execute("""SELECT DATE_FORMAT(MAX(updated_at), '%%Y-%%m-%%d %%H:%%i')
                     FROM annual_leave WHERE year=%s""", (year,))
    last_sync = cur.fetchone()[0]
    conn.close()

    return render_template("leave_dashboard.html",
                           year=year, year_list=year_list, s=s, depts=depts,
                           months=months, month_max=mmax,
                           negatives=negatives, unused=unused, no_grant=no_grant,
                           top_rest=rows[:15], last_sync=last_sync,
                           active_page="leave_dashboard")


@app.route("/annual_leave")
@_login_required
def annual_leave():
    search = request.args.get("search", "").strip()
    sel_month = request.args.get("month", "").strip()
    sel_dept = request.args.get("dept", "").strip()
    is_admin = session.get("role") == "admin"
    if not sel_month:
        sel_month = datetime.now().strftime("%Y-%m")
    try:
        ym = sel_month.split("-")
        year_int, month_int = int(ym[0]), int(ym[1])
    except (ValueError, IndexError):
        year_int, month_int = datetime.now().year, datetime.now().month
        sel_month = f"{year_int:04d}-{month_int:02d}"
    dept_list = sorted(set(DEPT_MAP.values()))
    # 달력 정보
    _, dim = calendar.monthrange(year_int, month_int)
    first_wd = calendar.weekday(year_int, month_int, 1)  # 0=Mon
    day_info = []
    for d in range(1, dim + 1):
        dt = datetime(year_int, month_int, d)
        ds = dt.strftime("%Y%m%d")
        wd = dt.weekday()
        hol_name = KR_HOLIDAYS.get(dt.date(), "")
        day_info.append({
            "day": d, "ds": ds, "wd": wd,
            "is_sat": wd == 5, "is_sun": wd == 6,
            "is_holiday": bool(hol_name), "hol_name": hol_name,
        })
    conn = _conn(); cur = conn.cursor()
    # 직원 목록 (annual_leave에 기록된 직원만)
    sql_u = """SELECT t.id, t.name, t.company FROM employee_roster er
               INNER JOIN (
                   SELECT COALESCE(MAX(CASE WHEN company<>'' THEN id END), MAX(id)) AS id,
                          name, COALESCE(MAX(company),'') AS company,
                          COALESCE(MAX(idno),'') AS idno
                   FROM tuser WHERE name IS NOT NULL AND name <> ''
                   GROUP BY name
               ) t ON t.name = er.name
               WHERE 1=1"""
    params_u = []
    if sel_dept:
        dept_codes = [k for k, v in DEPT_MAP.items() if v == sel_dept]
        if dept_codes:
            ph = ",".join(["%s"] * len(dept_codes))
            sql_u += f" AND t.company IN ({ph})"
            params_u += dept_codes
    if search:
        sql_u += " AND (t.name LIKE %s OR CAST(t.id AS CHAR) LIKE %s OR t.idno LIKE %s)"
        params_u += [f"%{search}%", f"%{search}%", f"%{search}%"]
    # 관리자/leave권한 없으면 자기 것만
    if not is_admin and not _has_perm("leave"):
        my_eid = session.get("e_id")
        if my_eid:
            sql_u += " AND t.id = %s"
            params_u.append(my_eid)
    sql_u += " ORDER BY t.company, t.name"
    cur.execute(sql_u, params_u)
    all_users = cur.fetchall()
    total_emp = len(all_users)
    users = all_users
    user_ids = [u[0] for u in users]
    # 연차 요약
    cur.execute("SELECT e_id, total, used, memo, deduct_prev, generated, m1,m2,m3,m4,m5,m6,m7,m8,m9,m10,m11,m12 FROM annual_leave WHERE year=%s", (year_int,))
    summary_map = {}
    for r in cur.fetchall():
        summary_map[r[0]] = {"total": r[1], "used": r[2], "memo": r[3],
                             "deduct_prev": r[4], "generated": r[5],
                             "months": list(r[6:18])}
    # 해당월 개별 기록
    m_start = f"{year_int:04d}{month_int:02d}01"
    m_end = f"{year_int:04d}{month_int:02d}{dim:02d}"
    leave_map = {}  # e_id -> {date_str: [(type, status)]}
    year_leave_map = {}  # e_id -> total used (연간, 승인건만)
    if user_ids:
        ph = ",".join(["%s"] * len(user_ids))
        cur.execute(f"SELECT e_id, leave_date, leave_type, status FROM leave_records WHERE e_id IN ({ph}) AND leave_date BETWEEN %s AND %s",
                    user_ids + [m_start, m_end])
        for eid, ld, lt, st in cur.fetchall():
            leave_map.setdefault(eid, {}).setdefault(ld, []).append((lt, st))
        # 연간 전체 사용량 계산 (승인건만)
        y_start = f"{year_int:04d}0101"
        y_end = f"{year_int:04d}1231"
        cur.execute(f"SELECT e_id, leave_type FROM leave_records WHERE e_id IN ({ph}) AND leave_date BETWEEN %s AND %s AND status='승인'",
                    user_ids + [y_start, y_end])
        for eid, lt in cur.fetchall():
            year_leave_map[eid] = year_leave_map.get(eid, 0) + (0.5 if lt == "반차" else 1)
    # 수정보고서: 기존 계획서 날짜 로드
    amend_id = request.args.get("amend", "").strip()
    amend_dates = []
    amend_type = "연차"
    amend_reason = ""
    if amend_id.isdigit():
        cur.execute("SELECT leave_date, leave_type FROM leave_plan_dates WHERE plan_id=%s ORDER BY leave_date", (int(amend_id),))
        for ld, lt in cur.fetchall():
            amend_dates.append(str(ld))
            amend_type = lt
        cur.execute("SELECT reason FROM leave_plans WHERE id=%s", (int(amend_id),))
        rrow = cur.fetchone()
        if rrow and rrow[0]:
            amend_reason = rrow[0]
        # 수정보고서 모드: 본인의 모든 연차 표시 제거 (자유롭게 날짜 선택 가능하게)
        my_eid = session.get("e_id")
        if my_eid and int(my_eid) in leave_map:
            leave_map[int(my_eid)] = {}
    conn.close()
    employees = []
    for uid, uname, company in users:
        if not uname:
            continue
        dept = DEPT_MAP.get((company or "").strip(), "")
        sm = summary_map.get(uid, {})
        total = sm.get("total", 15)
        used_calc = year_leave_map.get(uid, 0)
        used_db = sm.get("used", 0)
        used_final = max(used_calc, used_db)
        memo = sm.get("memo", "")
        deduct_prev = sm.get("deduct_prev", 0)
        generated = sm.get("generated", 0)
        remain = total - used_final
        emp_leaves = leave_map.get(uid, {})
        days = []
        for di in day_info:
            dl = emp_leaves.get(di["ds"], [])  # [(type, status), ...]
            days.append(dl)
        employees.append({
            "id": uid, "name": uname, "dept": dept,
            "total": total, "used": used_final, "remain": remain, "memo": memo,
            "deduct_prev": deduct_prev, "generated": generated,
            "months": sm.get("months", [0]*12),
            "days": days,
        })
    # 대기 건수 (관리자용)
    pending_count = 0
    if _has_perm("leave"):
        conn2 = _conn(); cur2 = conn2.cursor()
        cur2.execute("SELECT COUNT(*) FROM leave_records WHERE status='대기'")
        pending_count = cur2.fetchone()[0]
        conn2.close()
    # lp_group 멤버 여부 (leave 권한자도 본인 계획서 신청 가능 여부 판단)
    _cur_eid = session.get("e_id")
    conn3 = _conn(); cur3 = conn3.cursor()
    cur3.execute("SELECT COUNT(*) FROM lp_group_members WHERE user_id=%s", (_cur_eid,))
    is_lp_member = cur3.fetchone()[0] > 0
    conn3.close()
    return render_template("annual_leave.html",
                           employees=employees, sel_month=sel_month,
                           sel_dept=sel_dept, search=search,
                           dept_list=dept_list, day_info=day_info,
                           year=year_int, month=month_int, dim=dim,
                           first_wd=first_wd,
                           total_emp=total_emp,
                           can_edit=_has_perm("leave"),
                           user_eid=session.get("e_id"),
                           is_lp_member=is_lp_member,
                           pending_count=pending_count,
                           amend_dates=amend_dates,
                           amend_type=amend_type,
                           amend_reason=amend_reason)


@app.route("/save_leave_record", methods=["POST"])
@_login_required
def save_leave_record():
    if not _has_perm("leave"):
        return jsonify({"ok": False, "error": "권한이 없습니다"}), 403
    try:
        data = request.get_json()
        e_id = int(data["e_id"])
        leave_date = str(data["date"])
        leave_type = str(data.get("type", "연차"))
        action = str(data.get("action", "add"))
        conn = _conn(); cur = conn.cursor()
        if action == "delete":
            cur.execute("DELETE FROM leave_records WHERE e_id=%s AND leave_date=%s AND leave_type=%s",
                        (e_id, leave_date, leave_type))
        else:
            cur.execute("""
                INSERT INTO leave_records (e_id, leave_date, leave_type, status)
                VALUES (%s, %s, %s, '승인')
                ON DUPLICATE KEY UPDATE status='승인'
            """, (e_id, leave_date, leave_type))
        # 근무표 기타 컬럼에도 자동 연동 (승인된 것만)
        cur.execute("SELECT leave_type FROM leave_records WHERE e_id=%s AND leave_date=%s AND status='승인'", (e_id, leave_date))
        day_leaves = [r[0] for r in cur.fetchall()]
        if day_leaves:
            other_val = "/".join(day_leaves)
            cur.execute("""
                INSERT INTO work_override (e_id, e_date, col_type, value)
                VALUES (%s,%s,'other',%s)
                ON DUPLICATE KEY UPDATE value=%s, updated_at=NOW()
            """, (e_id, leave_date, other_val, other_val))
        else:
            cur.execute("SELECT value FROM work_override WHERE e_id=%s AND e_date=%s AND col_type='other'", (e_id, leave_date))
            ov_row = cur.fetchone()
            if ov_row and ov_row[0] in ("연차", "반차", "연차/반차", "반차/연차"):
                cur.execute("DELETE FROM work_override WHERE e_id=%s AND e_date=%s AND col_type='other'", (e_id, leave_date))
        # used 자동 갱신 + 월별(m1~m12) 갱신 (승인건만)
        year_val = int(leave_date[:4])
        cur.execute("SELECT leave_type, leave_date FROM leave_records WHERE e_id=%s AND leave_date LIKE %s AND status='승인'",
                    (e_id, f"{year_val}%"))
        total_used = 0
        month_used = [0] * 12
        for lt, ld in cur.fetchall():
            v = 0.5 if lt == "반차" else 1
            total_used += v
            m_idx = int(ld[4:6]) - 1
            month_used[m_idx] += v
        cur.execute("SELECT name FROM tuser WHERE id=%s", (e_id,))
        row = cur.fetchone()
        e_name = row[0] if row else ""
        cur.execute("""
            INSERT INTO annual_leave (e_id, e_name, year, used, m1,m2,m3,m4,m5,m6,m7,m8,m9,m10,m11,m12)
            VALUES (%s, %s, %s, %s, %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE used=%s, e_name=%s,
                m1=%s,m2=%s,m3=%s,m4=%s,m5=%s,m6=%s,m7=%s,m8=%s,m9=%s,m10=%s,m11=%s,m12=%s,
                updated_at=NOW()
        """, (e_id, e_name, year_val, total_used,
              *month_used,
              total_used, e_name,
              *month_used))
        conn.commit(); conn.close()
        d = leave_date
        disp_date = f"{d[:4]}-{d[4:6]}-{d[6:8]}" if len(d) == 8 else d
        if action == "delete":
            _send_telegram(f"🗑️ [연차삭제] {e_name} {disp_date} {leave_type} 삭제", "leave")
        else:
            _send_telegram(f"🌴 [연차등록] {e_name} {disp_date} {leave_type}", "leave")
        return jsonify({"ok": True, "used": total_used})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/request_leave", methods=["POST"])
@_login_required
def request_leave():
    """일반 직원 연차 신청 (단일 또는 복수 날짜, 대기 상태로 저장)"""
    try:
        data = request.get_json()
        e_id = session.get("e_id")
        if not e_id:
            return jsonify({"ok": False, "error": "직원 정보가 없습니다"}), 400
        e_id = int(e_id)
        # 복수 날짜 지원: dates 배열 또는 단일 date
        dates = data.get("dates", [])
        if not dates:
            single = data.get("date")
            if single:
                dates = [str(single)]
        if not dates:
            return jsonify({"ok": False, "error": "날짜를 선택해 주세요"}), 400
        leave_type = str(data.get("type", "연차"))
        reason = str(data.get("reason", "")).strip()
        conn = _conn(); cur = conn.cursor()
        cur.execute("SELECT name FROM tuser WHERE id=%s", (e_id,))
        row = cur.fetchone()
        e_name = row[0] if row else ""
        created_dates = []
        for leave_date in dates:
            leave_date = str(leave_date).replace("-", "")
            cur.execute("""
                INSERT INTO leave_records (e_id, leave_date, leave_type, status, reason)
                VALUES (%s, %s, %s, '대기', %s)
                ON DUPLICATE KEY UPDATE status='대기', reason=%s
            """, (e_id, leave_date, leave_type, reason, reason))
            d = leave_date
            created_dates.append(f"{d[:4]}-{d[4:6]}-{d[6:8]}" if len(d) == 8 else d)
        conn.commit(); conn.close()
        if len(created_dates) == 1:
            _send_telegram(f"📋 [연차신청] {e_name} {created_dates[0]} {leave_type} (승인대기)\n📝 사유: {reason}", "leave")
        else:
            date_list = ", ".join(created_dates)
            _send_telegram(f"📋 [연차일괄신청] {e_name} {leave_type} {len(created_dates)}건 (승인대기)\n📅 {date_list}\n📝 사유: {reason}", "leave")
        return jsonify({"ok": True, "count": len(created_dates)})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/cancel_leave_request", methods=["POST"])
@_login_required
def cancel_leave_request():
    """일반 직원이 본인의 대기 중 신청 취소"""
    try:
        data = request.get_json()
        e_id = session.get("e_id")
        if not e_id:
            return jsonify({"ok": False, "error": "직원 정보가 없습니다"}), 400
        e_id = int(e_id)
        leave_date = str(data["date"])
        leave_type = str(data.get("type", "연차"))
        conn = _conn(); cur = conn.cursor()
        cur.execute("DELETE FROM leave_records WHERE e_id=%s AND leave_date=%s AND leave_type=%s AND status='대기'",
                    (e_id, leave_date, leave_type))
        conn.commit(); conn.close()
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/leave_approval")
@_login_required
def leave_approval():
    is_admin = session.get("role") == "admin"
    login_uid = session.get("user_id")
    login_eid = session.get("e_id")  # tuser.id (lp_group용)
    # 검토자/멤버 여부 확인 (lp_group_reviewers + lp_group_members)
    conn_chk = _conn(); cur_chk = conn_chk.cursor()
    cur_chk.execute("SELECT group_id, step FROM lp_group_reviewers WHERE user_id=%s", (login_eid,))
    user_reviewer_rows = cur_chk.fetchall()
    cur_chk.execute("SELECT COUNT(*) FROM lp_group_members WHERE user_id=%s", (login_eid,))
    is_lp_member = cur_chk.fetchone()[0] > 0
    conn_chk.close()
    if not is_admin and not _has_perm("leave") and not user_reviewer_rows and not is_lp_member:
        flash("접근 권한이 없습니다.", "danger")
        return redirect(url_for("dashboard"))
    conn = _conn(); cur = conn.cursor()
    # ── 그룹 목록 (lp_groups) ──
    cur.execute("SELECT id, group_name, IFNULL(status,'설정중') FROM lp_groups ORDER BY group_name")
    groups = [{"id": r[0], "name": r[1], "status": r[2]} for r in cur.fetchall()]
    # 선택된 그룹
    sel_group = request.args.get("gid", "").strip()
    sel_group = int(sel_group) if sel_group.isdigit() else 0
    reviewers = {"1": [], "2": []}
    members = []
    group_status = "설정중"
    if sel_group:
        cur.execute("SELECT IFNULL(status,'설정중') FROM lp_groups WHERE id=%s", (sel_group,))
        fs = cur.fetchone()
        group_status = fs[0] if fs else "설정중"
        cur.execute("SELECT step, user_id, user_name FROM lp_group_reviewers WHERE group_id=%s ORDER BY step, user_name", (sel_group,))
        for step, uid, uname in cur.fetchall():
            reviewers[str(step)].append({"user_id": uid, "user_name": uname})
        cur.execute("SELECT user_id, user_name FROM lp_group_members WHERE group_id=%s ORDER BY user_name", (sel_group,))
        members = [{"user_id": r[0], "user_name": r[1]} for r in cur.fetchall()]
    # 로그인 사용자의 그룹별 step 확인
    cur.execute("SELECT group_id, step FROM lp_group_reviewers WHERE user_id=%s", (login_eid,))
    user_group_steps = {}
    for gid, step in cur.fetchall():
        user_group_steps.setdefault(gid, []).append(step)
    # 전체 그룹 현황 (요약)
    all_group_info = {}
    for g in groups:
        cur.execute("SELECT step, user_name FROM lp_group_reviewers WHERE group_id=%s ORDER BY step", (g["id"],))
        revs = {"1": [], "2": []}
        for step, uname in cur.fetchall():
            revs[str(step)].append(uname)
        cur.execute("SELECT COUNT(*) FROM lp_group_members WHERE group_id=%s", (g["id"],))
        mcnt = cur.fetchone()[0]
        cur.execute("SELECT user_name FROM lp_group_members WHERE group_id=%s ORDER BY user_name", (g["id"],))
        mnames = [r[0] for r in cur.fetchall()]
        all_group_info[g["id"]] = {"revs": revs, "member_count": mcnt, "member_names": mnames, "status": g.get("status", "설정중")}
    # 직원 목록 (사원 제외)
    cur_year = datetime.now().year
    cur.execute("""
        SELECT t.id, t.name, IFNULL(t.company,'') as company,
               IFNULL(er.dept,'') as sub_dept,
               IFNULL(er.position,'') as position
        FROM employee_roster er
        INNER JOIN (
            SELECT COALESCE(MAX(CASE WHEN company<>'' THEN id END), MAX(id)) AS id,
                   name, COALESCE(MAX(company),'') AS company
            FROM tuser WHERE name IS NOT NULL AND name <> ''
            GROUP BY name
        ) t ON t.name = er.name
        ORDER BY t.company, t.name
    """)
    all_users = []
    reviewer_users = []
    dept_set = set(); sub_dept_set = set()
    exclude_positions = ('사원', '')
    include_override_ids = (307,)  # 이민영(총무)
    for r in cur.fetchall():
        dept_name = DEPT_MAP.get((r[2] or "").strip(), "")
        sub_dept_name = (r[3] or "").strip()
        pos = (r[4] or "").strip()
        u = {"id": r[0], "name": r[1], "dept": dept_name, "sub_dept": sub_dept_name, "position": pos}
        if pos not in exclude_positions or r[0] in include_override_ids:
            all_users.append(u)
        reviewer_users.append(u)
        if dept_name: dept_set.add(dept_name)
        if sub_dept_name: sub_dept_set.add(sub_dept_name)
    dept_list = sorted(dept_set)
    sub_dept_list = sorted(sub_dept_set)
    # 연차사용계획서 목록
    is_reviewer = is_admin or _has_perm("leave")
    plan_status = request.args.get("plan_status", "전체").strip()
    if plan_status not in ("전체", "대기", "승인", "반려"):
        plan_status = "대기"
    member_ids = set()
    if is_admin or _has_perm("leave"):
        cur.execute("SELECT DISTINCT user_id FROM lp_group_members")
        member_ids = set(r[0] for r in cur.fetchall())
    else:
        # 개인: 자기 계획서만 (검토자 step이 있어도 본인만)
        member_ids.add(int(login_eid))
    # 개인은 전체 상태, 관리자/검토자는 필터 적용
    if not is_reviewer:
        status_filter_sql = "1=1"
    elif plan_status == "대기":
        status_filter_sql = "p.status IN ('대기','1차승인')"
    else:
        status_filter_sql = "p.status = %s"
    plans_list = []
    if member_ids:
        ph2 = ','.join(['%s'] * len(member_ids))
        cur.execute(f"""
            SELECT p.id, p.user_id, p.user_name, p.group_id, p.plan_month, p.status, p.reason, p.created_at,
                   g.group_name,
                   (SELECT COUNT(*) FROM leave_plan_dates WHERE plan_id=p.id) as date_cnt,
                   (SELECT SUM(CASE WHEN leave_type='반차' THEN 0.5 ELSE 1 END) FROM leave_plan_dates WHERE plan_id=p.id) as total_days,
                   p.parent_id
            FROM leave_plans p
            LEFT JOIN lp_groups g ON p.group_id = g.id
            WHERE p.user_id IN ({ph2})
            ORDER BY COALESCE(p.parent_id, p.id), p.parent_id IS NOT NULL, p.id
        """, tuple(member_ids))
        all_plans = []
        for r in cur.fetchall():
            pm = str(r[4])
            disp_month = f"{pm[:4]}-{pm[4:6]}" if len(pm) >= 6 else pm
            all_plans.append({
                "id": r[0], "user_id": r[1], "user_name": r[2], "group_id": r[3],
                "plan_month": pm, "disp_month": disp_month,
                "status": r[5], "reason": r[6], "created_at": r[7],
                "group_name": r[8] or "", "date_count": r[9] or 0,
                "total_days": float(r[10] or 0),
                "parent_id": r[11],
                "is_amendment": bool(r[11]),
            })
        # 필터 적용: 원본이 필터에 맞으면 하위 수정보고서도 함께 표시
        if not is_reviewer or plan_status == "전체":
            plans_list = all_plans
        else:
            # 필터에 맞는 ID 수집
            match_ids = set()
            for p in all_plans:
                st = p["status"]
                if plan_status == "대기" and st in ("대기", "1차승인"):
                    match_ids.add(p["id"])
                elif plan_status != "대기" and st == plan_status:
                    match_ids.add(p["id"])
            # 매칭된 원본의 수정보고서도 포함, 매칭된 수정보고서의 원본도 포함
            include_ids = set()
            for p in all_plans:
                pid = p["id"]
                parent = p["parent_id"]
                if pid in match_ids:
                    include_ids.add(pid)
                    if parent:
                        include_ids.add(parent)
                if parent and parent in match_ids:
                    include_ids.add(pid)
                    include_ids.add(parent)
            plans_list = [p for p in all_plans if p["id"] in include_ids]
    conn.close()
    # 개인용: 월별 그룹핑 (1~12월)
    cur_year = datetime.now().year
    plans_by_month = {}
    for m in range(1, 13):
        plans_by_month[m] = []
    for p in plans_list:
        pm = p.get("plan_month", "")
        try:
            pm_month = int(pm[4:6]) if len(pm) >= 6 else 0
        except (ValueError, TypeError):
            pm_month = 0
        if 1 <= pm_month <= 12:
            plans_by_month[pm_month].append(p)
    return render_template("leave_approval.html",
                           plans=plans_list,
                           plans_by_month=plans_by_month,
                           cur_year=cur_year,
                           plan_status=plan_status,
                           groups=groups,
                           sel_group=sel_group,
                           reviewers=reviewers, members=members,
                           group_status=group_status,
                           all_group_info=all_group_info,
                           all_users=all_users,
                           reviewer_users=reviewer_users,
                           dept_list=dept_list, sub_dept_list=sub_dept_list,
                           user_group_steps=user_group_steps,
                           is_admin=is_admin,
                           is_reviewer=is_reviewer)


@app.route("/process_leave_request", methods=["POST"])
@_login_required
def process_leave_request():
    try:
        data = request.get_json()
        record_id = int(data["id"])
        new_status = str(data["status"])  # '승인' or '반려'
        reject_reason = str(data.get("reject_reason", "")).strip()
        if new_status not in ("승인", "반려"):
            return jsonify({"ok": False, "error": "잘못된 상태"}), 400
        if new_status == "반려" and not reject_reason:
            return jsonify({"ok": False, "error": "반려 사유를 입력해 주세요"}), 400
        conn = _conn(); cur = conn.cursor()
        cur.execute("SELECT e_id, leave_date, leave_type, status FROM leave_records WHERE id=%s", (record_id,))
        rec = cur.fetchone()
        if not rec:
            conn.close()
            return jsonify({"ok": False, "error": "기록을 찾을 수 없습니다"}), 404
        e_id, leave_date, leave_type, cur_status = rec
        # 신청자가 속한 그룹과 final_step 확인 (wr_groups 공유)
        cur.execute("""
            SELECT g.id, g.final_step FROM wr_groups g
            JOIN wr_group_members m ON g.id = m.group_id
            WHERE m.user_id = %s LIMIT 1
        """, (e_id,))
        grow = cur.fetchone()
        req_group_id = grow[0] if grow else 0
        final_step = grow[1] if grow else 3
        # 검토자 권한 확인
        login_uid = session.get("user_id")
        is_admin = session.get("role") == "admin"
        user_steps = []
        if req_group_id:
            cur.execute("SELECT step FROM wr_group_reviewers WHERE group_id=%s AND user_id=%s", (req_group_id, login_uid))
            user_steps = [r[0] for r in cur.fetchall()]
        # 승인 로직 (final_step에 따라 동적)
        if new_status == "반려":
            if not is_admin and not user_steps:
                conn.close()
                return jsonify({"ok": False, "error": "해당 그룹의 검토자 권한이 없습니다"}), 403
            actual_status = "반려"
        else:  # 승인
            if cur_status == "대기":
                if not is_admin and 1 not in user_steps:
                    conn.close()
                    return jsonify({"ok": False, "error": "1차 검토자만 처리 가능합니다"}), 403
                if final_step == 1:
                    actual_status = "승인"
                else:
                    actual_status = "1차승인"
            elif cur_status == "1차승인":
                if not is_admin and 2 not in user_steps:
                    conn.close()
                    return jsonify({"ok": False, "error": "2차 검토자만 처리 가능합니다"}), 403
                if final_step == 2:
                    actual_status = "승인"
                else:
                    actual_status = "2차승인"
            elif cur_status == "2차승인":
                if not is_admin and 3 not in user_steps:
                    conn.close()
                    return jsonify({"ok": False, "error": "3차 검토자(최종)만 처리 가능합니다"}), 403
                actual_status = "승인"
            else:
                conn.close()
                return jsonify({"ok": False, "error": f"현재 상태({cur_status})에서 승인할 수 없습니다"}), 400
        cur.execute("UPDATE leave_records SET status=%s WHERE id=%s", (actual_status, record_id))
        if actual_status == "반려" and reject_reason:
            cur.execute("UPDATE leave_records SET memo=%s WHERE id=%s", (reject_reason, record_id))
        cur.execute("SELECT name FROM tuser WHERE id=%s", (e_id,))
        row = cur.fetchone()
        e_name = row[0] if row else ""
        if actual_status == "승인":
            # 최종 승인 시 근무표 기타 컬럼 자동 연동
            cur.execute("SELECT leave_type FROM leave_records WHERE e_id=%s AND leave_date=%s AND status='승인'", (e_id, leave_date))
            day_leaves = [r[0] for r in cur.fetchall()]
            if day_leaves:
                other_val = "/".join(day_leaves)
                cur.execute("""
                    INSERT INTO work_override (e_id, e_date, col_type, value)
                    VALUES (%s,%s,'other',%s)
                    ON DUPLICATE KEY UPDATE value=%s, updated_at=NOW()
                """, (e_id, leave_date, other_val, other_val))
            # 최종 승인 시 근무보고서(wr_entries) 자동 반영
            if req_group_id:
                cur.execute("""
                    SELECT r.id, r.status FROM wr_reports r
                    WHERE r.group_id=%s AND r.report_date=%s
                    ORDER BY r.id DESC LIMIT 1
                """, (req_group_id, leave_date))
                rpt = cur.fetchone()
                if rpt:
                    rpt_id, rpt_status = rpt
                    leave_cat = leave_type if leave_type in ('연차', '무급') else '연차'
                    cur.execute("""
                        INSERT INTO wr_entries (report_id, user_id, user_name, category, deduction, etc_value, skipped)
                        VALUES (%s,%s,%s,'기타',0,%s,0)
                        ON DUPLICATE KEY UPDATE category='기타', etc_value=%s, deduction=0, skipped=0
                    """, (rpt_id, e_id, e_name, leave_cat, leave_cat))
                    # 이미 결재된 보고서면 schedule_record에도 반영
                    if rpt_status == '결재':
                        cur.execute("SELECT group_name FROM wr_groups WHERE id=%s", (req_group_id,))
                        gn = cur.fetchone()
                        sname = gn[0] if gn else '기타'
                        cur.execute("""
                            INSERT INTO schedule_record (emp_name, work_date, basic_h, ot_h, night_h, etc,
                                                         source_type, sheet_name, uploaded_by)
                            VALUES (%s,%s,0,0,0,%s,'보고서',%s,%s)
                            ON DUPLICATE KEY UPDATE basic_h=0, ot_h=0, night_h=0, etc=%s
                        """, (e_name, leave_date, leave_cat, sname,
                              session.get("user_name", ""), leave_cat))
                else:
                    # 보고서가 없으면 자동 생성 (그룹 멤버 전원 포함)
                    approver_id = session.get("user_id")
                    cur.execute("""INSERT INTO wr_reports (group_id, report_date, status, created_by, created_by_name, memo)
                                   VALUES (%s,%s,'대기',%s,%s,'연차승인 자동생성')""",
                                (req_group_id, leave_date, approver_id, approver_name))
                    new_rpt_id = cur.lastrowid
                    # 그룹 멤버 전원 로드
                    cur.execute("""SELECT m.user_id, t.name FROM wr_group_members m
                                   JOIN tuser t ON m.user_id = t.id
                                   WHERE m.group_id=%s""", (req_group_id,))
                    members = cur.fetchall()
                    for m_uid, m_name in members:
                        if m_uid == e_id:
                            m_leave_cat = leave_type if leave_type in ('연차', '무급') else '연차'
                            cur.execute("""INSERT INTO wr_entries (report_id, user_id, user_name, category, deduction, etc_value, skipped)
                                           VALUES (%s,%s,%s,'기타',0,%s,0)""",
                                        (new_rpt_id, m_uid, m_name, m_leave_cat))
                        else:
                            # 다른 멤버도 해당 날짜에 승인된 연차가 있으면 기타 처리
                            cur.execute("SELECT leave_type FROM leave_records WHERE e_id=%s AND leave_date=%s AND status='승인'",
                                        (m_uid, leave_date))
                            m_leave = cur.fetchone()
                            if m_leave:
                                m_lc = m_leave[0] if m_leave[0] in ('연차', '무급') else '연차'
                                cur.execute("""INSERT INTO wr_entries (report_id, user_id, user_name, category, deduction, etc_value, skipped)
                                               VALUES (%s,%s,%s,'기타',0,%s,0)""",
                                            (new_rpt_id, m_uid, m_name, m_lc))
                            else:
                                cur.execute("""INSERT INTO wr_entries (report_id, user_id, user_name, category, deduction, etc_value, skipped)
                                               VALUES (%s,%s,%s,'주간기본',0,'',0)""",
                                            (new_rpt_id, m_uid, m_name))
        # used 자동 갱신 (최종 승인건만)
        year_val = int(leave_date[:4])
        cur.execute("SELECT leave_type, leave_date FROM leave_records WHERE e_id=%s AND leave_date LIKE %s AND status='승인'",
                    (e_id, f"{year_val}%"))
        total_used = 0
        month_used = [0] * 12
        for lt, ld in cur.fetchall():
            v = 0.5 if lt == "반차" else 1
            total_used += v
            m_idx = int(ld[4:6]) - 1
            month_used[m_idx] += v
        cur.execute("""
            INSERT INTO annual_leave (e_id, e_name, year, used, m1,m2,m3,m4,m5,m6,m7,m8,m9,m10,m11,m12)
            VALUES (%s, %s, %s, %s, %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE used=%s, e_name=%s,
                m1=%s,m2=%s,m3=%s,m4=%s,m5=%s,m6=%s,m7=%s,m8=%s,m9=%s,m10=%s,m11=%s,m12=%s,
                updated_at=NOW()
        """, (e_id, e_name, year_val, total_used,
              *month_used,
              total_used, e_name,
              *month_used))
        # 웹 알림 생성 (신청자에게)
        d = leave_date
        disp_date = f"{d[:4]}-{d[4:6]}-{d[6:8]}" if len(d) == 8 else d
        approver_name = session.get("user_name", "")
        # 신청자의 app_users.id 찾기
        cur.execute("SELECT id FROM app_users WHERE e_id=%s LIMIT 1", (e_id,))
        noti_row = cur.fetchone()
        noti_uid = noti_row[0] if noti_row else None
        if noti_uid:
            if actual_status == "승인":
                noti_msg = f"✅ {disp_date} {leave_type}가 최종승인 되었습니다 (승인자: {approver_name})"
            elif actual_status == "1차승인":
                noti_msg = f"📋 {disp_date} {leave_type}가 1차승인 되었습니다 (승인자: {approver_name})"
            elif actual_status == "2차승인":
                noti_msg = f"📋 {disp_date} {leave_type}가 2차승인 되었습니다 (승인자: {approver_name})"
            else:
                noti_msg = f"❌ {disp_date} {leave_type}가 반려되었습니다 (처리자: {approver_name})\n📝 사유: {reject_reason}"
            cur.execute("INSERT INTO notifications (user_id, message, link) VALUES (%s, %s, %s)",
                        (noti_uid, noti_msg, "/annual_leave"))
        conn.commit(); conn.close()
        if actual_status == "승인":
            _send_telegram(f"✅ [연차최종승인] {e_name} {disp_date} {leave_type} 최종승인", "leave")
        elif actual_status == "1차승인":
            _send_telegram(f"📝 [연차1차승인] {e_name} {disp_date} {leave_type} 1차승인", "leave")
        elif actual_status == "2차승인":
            _send_telegram(f"📝 [연차2차승인] {e_name} {disp_date} {leave_type} 2차승인", "leave")
        else:
            _send_telegram(f"❌ [연차반려] {e_name} {disp_date} {leave_type} 반려\n📝 사유: {reject_reason}", "leave")
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/api/notifications")
@_login_required
def api_notifications():
    uid = session.get("user_id")
    conn = _conn(); cur = conn.cursor()
    cur.execute("SELECT id, message, link, is_read, created_at FROM notifications WHERE user_id=%s ORDER BY created_at DESC LIMIT 20", (uid,))
    items = [{"id": r[0], "message": r[1], "link": r[2], "is_read": r[3], "created_at": str(r[4])} for r in cur.fetchall()]
    cur.execute("SELECT COUNT(*) FROM notifications WHERE user_id=%s AND is_read=0", (uid,))
    unread = cur.fetchone()[0]
    conn.close()
    return jsonify({"ok": True, "items": items, "unread": unread})


@app.route("/api/notifications/read", methods=["POST"])
@_login_required
def api_notifications_read():
    uid = session.get("user_id")
    data = request.get_json()
    nid = data.get("id")
    conn = _conn(); cur = conn.cursor()
    if nid == "all":
        cur.execute("UPDATE notifications SET is_read=1 WHERE user_id=%s AND is_read=0", (uid,))
    else:
        cur.execute("UPDATE notifications SET is_read=1 WHERE id=%s AND user_id=%s", (int(nid), uid))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/save_annual_leave", methods=["POST"])
@_login_required
def save_annual_leave():
    if not _has_perm("leave"):
        return jsonify({"ok": False, "error": "권한이 없습니다"}), 403
    try:
        data = request.get_json()
        e_id = int(data["e_id"])
        year = int(data["year"])
        field = data["field"]
        value = data["value"]
        conn = _conn(); cur = conn.cursor()
        # 이름 조회
        cur.execute("SELECT name FROM tuser WHERE id=%s", (e_id,))
        row = cur.fetchone()
        e_name = row[0] if row else ""
        if field == "total":
            cur.execute("""
                INSERT INTO annual_leave (e_id, e_name, year, total)
                VALUES (%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE total=%s, e_name=%s, updated_at=NOW()
            """, (e_id, e_name, year, float(value), float(value), e_name))
        elif field == "used":
            cur.execute("""
                INSERT INTO annual_leave (e_id, e_name, year, used)
                VALUES (%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE used=%s, e_name=%s, updated_at=NOW()
            """, (e_id, e_name, year, float(value), float(value), e_name))
        elif field == "memo":
            cur.execute("""
                INSERT INTO annual_leave (e_id, e_name, year, memo)
                VALUES (%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE memo=%s, e_name=%s, updated_at=NOW()
            """, (e_id, e_name, year, str(value), str(value), e_name))
        else:
            return jsonify({"ok": False, "error": "invalid field"}), 400
        conn.commit(); conn.close()
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/meal_management")
@_login_required
def meal_management():
    if session.get("role") != "admin" and not _has_perm("meal"):
        flash("접근 권한이 없습니다.", "danger")
        return redirect(url_for("dashboard"))
    # 모바일 자동 전환 (view=pc 파라미터 없으면)
    if request.args.get("view") != "pc":
        ua = request.headers.get("User-Agent", "").lower()
        if any(k in ua for k in ["iphone", "android", "mobile", "ipod"]):
            return redirect(url_for("meal_mobile", month=request.args.get("month", datetime.now().strftime("%Y-%m"))))
    conn = _conn(); cur = conn.cursor()

    month_str = request.args.get("month", datetime.now().strftime("%Y-%m"))
    year, mon = int(month_str[:4]), int(month_str[5:7])
    days_in_month = calendar.monthrange(year, mon)[1]

    MEAL_DEPTS = ["전기", "전자", "관리"]
    MEAL_TYPES = ["조식", "중식", "석식", "야식"]

    # DB에서 해당 월 데이터 조회
    cur.execute("SELECT dept, meal_type, `day`, `count`, `memo`, `writer`, updated_at FROM meal_count WHERE `year_month`=%s",
                (month_str,))
    meal_data = {}
    meal_memo_data = {}
    meal_time_data = {}
    meal_writer_data = {}
    meal_memos = []
    for dept, mtype, day, cnt, memo, writer, updated_at in cur.fetchall():
        meal_data[(dept, mtype, day)] = cnt
        time_str = updated_at.strftime("%m/%d %H:%M") if updated_at else ""
        if memo:
            meal_memo_data[(dept, mtype, day)] = memo
            meal_time_data[(dept, mtype, day)] = time_str
            meal_writer_data[(dept, mtype, day)] = writer or ""
            meal_memos.append({"dept": dept, "type": mtype, "day": day, "count": cnt, "memo": memo, "writer": writer or "", "time": time_str})
    # 공지 조회
    cur.execute("SELECT id, content, `writer`, created_at FROM meal_notice WHERE `year_month`=%s ORDER BY created_at DESC", (month_str,))
    meal_notices = []
    for nid, content, writer, created_at in cur.fetchall():
        t = created_at.strftime("%m/%d %H:%M") if created_at else ""
        meal_notices.append({"id": nid, "content": content, "writer": writer or "", "time": t})
    conn.close()

    # 요일 / 공휴일 정보
    DOW_KR = ["월", "화", "수", "목", "금", "토", "일"]
    days_info = []
    for d in range(1, days_in_month + 1):
        dt = datetime(year, mon, d)
        dow = DOW_KR[dt.weekday()]
        hol = KR_HOLIDAYS.get(dt.date(), "")
        is_holiday = (dt.weekday() >= 5) or bool(hol)
        days_info.append({"day": d, "dow": dow, "is_holiday": is_holiday, "hol": hol})

    return render_template("meal_management.html",
                           month=month_str, year=year, mon=mon,
                           days_in_month=days_in_month,
                           days_info=days_info,
                           meal_depts=MEAL_DEPTS,
                           meal_types=MEAL_TYPES,
                           meal_data=meal_data,
                           meal_memo_data=meal_memo_data,
                           meal_time_data=meal_time_data,
                           meal_writer_data=meal_writer_data,
                           meal_memos=meal_memos,
                           meal_notices=meal_notices,
                           can_edit=_has_perm("meal"))


@app.route("/meal_mobile")
@_login_required
def meal_mobile():
    if session.get("role") != "admin" and not _has_perm("meal"):
        flash("접근 권한이 없습니다.", "danger")
        return redirect(url_for("dashboard"))
    conn = _conn(); cur = conn.cursor()

    month_str = request.args.get("month", datetime.now().strftime("%Y-%m"))
    year, mon = int(month_str[:4]), int(month_str[5:7])
    days_in_month = calendar.monthrange(year, mon)[1]

    MEAL_DEPTS = ["전기", "전자", "관리"]
    MEAL_TYPES = ["조식", "중식", "석식", "야식"]

    cur.execute("SELECT dept, meal_type, `day`, `count`, `memo`, `writer`, updated_at FROM meal_count WHERE `year_month`=%s",
                (month_str,))
    meal_data = {}
    meal_memo_data = {}
    meal_time_data = {}
    meal_writer_data = {}
    for dept, mtype, day, cnt, memo, writer, updated_at in cur.fetchall():
        meal_data[(dept, mtype, day)] = cnt
        time_str = updated_at.strftime("%m/%d %H:%M") if updated_at else ""
        if memo:
            meal_memo_data[(dept, mtype, day)] = memo
            meal_time_data[(dept, mtype, day)] = time_str
            meal_writer_data[(dept, mtype, day)] = writer or ""

    cur.execute("SELECT id, content, `writer`, created_at FROM meal_notice WHERE `year_month`=%s ORDER BY created_at DESC", (month_str,))
    meal_notices = []
    for nid, content, writer, created_at in cur.fetchall():
        t = created_at.strftime("%m/%d %H:%M") if created_at else ""
        meal_notices.append({"id": nid, "content": content, "writer": writer or "", "time": t})
    conn.close()

    DOW_KR = ["월", "화", "수", "목", "금", "토", "일"]
    days_info = []
    for d in range(1, days_in_month + 1):
        dt = datetime(year, mon, d)
        dow = DOW_KR[dt.weekday()]
        hol = KR_HOLIDAYS.get(dt.date(), "")
        is_holiday = (dt.weekday() >= 5) or bool(hol)
        days_info.append({"day": d, "dow": dow, "is_holiday": is_holiday, "hol": hol})

    today_day = datetime.now().day if month_str == datetime.now().strftime("%Y-%m") else 1

    return render_template("meal_mobile.html",
                           month=month_str, year=year, mon=mon,
                           days_in_month=days_in_month,
                           days_info=days_info,
                           meal_depts=MEAL_DEPTS,
                           meal_types=MEAL_TYPES,
                           meal_data=meal_data,
                           meal_memo_data=meal_memo_data,
                           meal_time_data=meal_time_data,
                           meal_writer_data=meal_writer_data,
                           meal_notices=meal_notices,
                           today_day=today_day,
                           can_edit=_has_perm("meal"))


@app.route("/save_meal", methods=["POST"])
@_login_required
def save_meal():
    if not _has_perm("meal"):
        return jsonify({"ok": False, "error": "권한이 없습니다"}), 403
    try:
        data = request.get_json()
        year_month = str(data["year_month"])
        dept = str(data["dept"])
        meal_type = str(data["meal_type"])
        day = int(data["day"])
        count = str(data.get("count", "")).strip()
        memo = str(data.get("memo", "")).strip()
        conn = _conn(); cur = conn.cursor()
        if count == "":
            cur.execute("DELETE FROM meal_count WHERE `year_month`=%s AND dept=%s AND meal_type=%s AND `day`=%s",
                        (year_month, dept, meal_type, day))
            _send_telegram(f"🗑️ [식수삭제] {dept} {meal_type} {year_month}-{day:02d} (by {session.get('user_name','')})", "meal")
        else:
            writer = session.get("user_name", "")
            cur.execute("""
                INSERT INTO meal_count (`year_month`, dept, meal_type, `day`, `count`, `memo`, `writer`)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE `count`=%s, `memo`=%s, `writer`=%s, updated_at=NOW()
            """, (year_month, dept, meal_type, day, count, memo, writer, count, memo, writer))
        conn.commit(); conn.close()
        from datetime import datetime as _dt
        if count:
            memo_txt = f" 메모:{memo}" if memo else ""
            _send_telegram(f"🍽️ [식수등록] {dept} {meal_type} {year_month}-{day:02d} {count}명{memo_txt} (by {session.get('user_name','')})", "meal")
        return jsonify({"ok": True, "time": _dt.now().strftime("%m/%d %H:%M"), "writer": session.get("user_name", "")})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/save_meal_notice", methods=["POST"])
@_login_required
def save_meal_notice():
    if not _has_perm("meal"):
        return jsonify({"ok": False, "error": "권한이 없습니다"}), 403
    try:
        data = request.get_json()
        year_month = str(data["year_month"])
        content = str(data.get("content", "")).strip()
        if not content:
            return jsonify({"ok": False, "error": "내용을 입력하세요."}), 400
        writer = session.get("user_name", "")
        conn = _conn(); cur = conn.cursor()
        cur.execute("INSERT INTO meal_notice (`year_month`, content, `writer`) VALUES (%s, %s, %s)", (year_month, content, writer))
        nid = cur.lastrowid
        conn.commit(); conn.close()
        from datetime import datetime as _dt
        _send_telegram(f"📢 [식수공지] {content} (by {writer})", "notice")
        return jsonify({"ok": True, "id": nid, "time": _dt.now().strftime("%m/%d %H:%M"), "writer": writer})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/delete_meal_notice", methods=["POST"])
@_login_required
def delete_meal_notice():
    if not _has_perm("meal"):
        return jsonify({"ok": False, "error": "권한이 없습니다"}), 403
    try:
        data = request.get_json()
        nid = int(data["id"])
        conn = _conn(); cur = conn.cursor()
        cur.execute("SELECT content FROM meal_notice WHERE id=%s", (nid,))
        row = cur.fetchone()
        cur.execute("DELETE FROM meal_notice WHERE id=%s", (nid,))
        conn.commit(); conn.close()
        if row:
            _send_telegram(f"🗑️ [공지삭제] {row[0][:50]} (by {session.get('user_name','')})", "notice")
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/dev_history")
@_login_required
def dev_history():
    if session.get("role") != "admin":
        flash("접근 권한이 없습니다.", "danger"); return redirect("/")
    import json as _json
    commits = []
    try:
        json_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "dev_history.json")
        with open(json_path, encoding="utf-8") as f:
            raw = _json.load(f)
        commits = []
        for c in raw:
            msg = c.get("msg", "")
            if msg.startswith("[") and "]" in msg:
                idx = msg.index("]")
                cat = msg[1:idx]
                body = msg[idx+2:]
            else:
                cat = ""
                body = msg
            commits.append({"sha": c.get("sha",""), "date": c.get("date",""), "msg": msg, "cat": cat, "body": body})
    except Exception as e:
        commits = [{"sha": "-", "date": "-", "msg": f"이력 파일 없음: {e}", "cat": "", "body": f"이력 파일 없음: {e}"}]
    return render_template("dev_history.html", commits=commits, active_page="dev_history")


@app.route("/log_management")
@_login_required
def log_management():
    if session.get("role") != "admin":
        flash("접근 권한이 없습니다.", "danger")
        return redirect(url_for("dashboard"))
    conn = _conn(); cur = conn.cursor()

    # --- 필터 파라미터 ---
    search   = request.args.get("search", "").strip()
    log_type = request.args.get("type", "all")          # all / override / leave / sync
    date_from = request.args.get("from", "")
    date_to   = request.args.get("to", "")

    # 기간 미입력시 최근 7일 기본값
    if not date_from:
        date_from = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
    if not date_to:
        date_to = datetime.now().strftime("%Y-%m-%d")

    logs = []

    # --- work_override 로그 ---
    if log_type in ("all", "override"):
        sql = """
            SELECT wo.e_id, u.name, wo.e_date, wo.col_type, wo.value,
                   wo.memo, wo.updated_at
            FROM work_override wo
            LEFT JOIN tuser u ON wo.e_id = u.id
            WHERE 1=1
        """
        params = []
        if search:
            sql += " AND (u.name LIKE %s OR CAST(wo.e_id AS CHAR) LIKE %s)"
            params += [f"%{search}%", f"%{search}%"]
        if date_from:
            sql += " AND wo.updated_at >= %s"
            params.append(date_from + " 00:00:00")
        if date_to:
            sql += " AND wo.updated_at <= %s"
            params.append(date_to + " 23:59:59")
        sql += " ORDER BY wo.updated_at DESC LIMIT 500"
        cur.execute(sql, params)
        col_label = {
            "basic": "기본", "overtime": "연장", "night": "야간", "other": "기타",
            "sum_ot": "전월연장", "sum_night": "전월야간",
            "hol_work": "전월휴일", "hol_ot": "전월휴연",
        }
        for eid, name, edate, ctype, val, memo, upd in cur.fetchall():
            # e_date ?щ㎎: YYYYMMDD or YYYYMM00
            dt = str(edate)
            if dt.endswith("00"):
                disp_date = f"{dt[:4]}-{dt[4:6]} 합산"
            else:
                disp_date = f"{dt[:4]}-{dt[4:6]}-{dt[6:8]}"
            logs.append({
                "type": "수정",
                "name": name or str(eid),
                "date": disp_date,
                "detail": col_label.get(ctype, ctype),
                "value": val,
                "memo": memo or "",
                "updated_at": upd,
            })

    # --- leave_records 로그 ---
    if log_type in ("all", "leave"):
        sql = """
            SELECT lr.e_id, u.name, lr.leave_date, lr.leave_type,
                   lr.memo, lr.created_at
            FROM leave_records lr
            LEFT JOIN tuser u ON lr.e_id = u.id
            WHERE 1=1
        """
        params = []
        if search:
            sql += " AND (u.name LIKE %s OR CAST(lr.e_id AS CHAR) LIKE %s)"
            params += [f"%{search}%", f"%{search}%"]
        if date_from:
            sql += " AND lr.created_at >= %s"
            params.append(date_from + " 00:00:00")
        if date_to:
            sql += " AND lr.created_at <= %s"
            params.append(date_to + " 23:59:59")
        sql += " ORDER BY lr.created_at DESC LIMIT 500"
        cur.execute(sql, params)
        for eid, name, ldate, ltype, memo, crt in cur.fetchall():
            dt = str(ldate)
            disp_date = f"{dt[:4]}-{dt[4:6]}-{dt[6:8]}" if len(dt) == 8 else dt
            logs.append({
                "type": ltype,
                "name": name or str(eid),
                "date": disp_date,
                "detail": "연차관리",
                "value": ltype,
                "memo": memo or "",
                "updated_at": crt,
            })

    # --- sync_log 로그 ---
    if log_type in ("all", "sync"):
        sql = "SELECT log_time, host, level, message FROM sync_log WHERE 1=1"
        params = []
        if date_from:
            sql += " AND log_time >= %s"
            params.append(date_from + " 00:00:00")
        if date_to:
            sql += " AND log_time <= %s"
            params.append(date_to + " 23:59:59")
        if search:
            sql += " AND message LIKE %s"
            params.append(f"%{search}%")
        sql += " ORDER BY id DESC LIMIT 500"
        cur.execute(sql, params)
        for lt, host, level, msg in cur.fetchall():
            logs.append({
                "type": "싱크",
                "name": host or "-",
                "date": "-",
                "detail": level,
                "value": (msg or "")[:120],
                "memo": "",
                "updated_at": lt,
            })

    # --- MDB 싱크 현황 (최근 조회) ---
    cur.execute("SELECT COUNT(*) FROM tenter")
    sync_total = cur.fetchone()[0]
    cur.execute("SELECT MAX(e_date), MAX(e_uptime) FROM tenter")
    row = cur.fetchone()
    sync_max_date = row[0] if row else "-"
    sync_max_uptime = row[1] if row else "-"

    d7 = (datetime.now() - timedelta(days=7)).strftime("%Y%m%d")
    cur.execute("SELECT e_date, COUNT(*) FROM tenter WHERE e_date >= %s GROUP BY e_date ORDER BY e_date", (d7,))
    sync_daily = cur.fetchall()

    cur.execute("SELECT log_time, host, level, message FROM sync_log ORDER BY id DESC LIMIT 10")
    sync_recent = cur.fetchall()

    # 留덉?留일깃났 싱크
    cur.execute("SELECT log_time, message FROM sync_log WHERE level='OK' OR (level='INFO' AND message LIKE '%%sync done%%') ORDER BY id DESC LIMIT 1")
    last_ok = cur.fetchone()
    sync_last_ok_time = last_ok[0] if last_ok else None
    sync_last_ok_msg = last_ok[1] if last_ok else "-"

    # 留덉?留일먮윭
    cur.execute("SELECT log_time, message FROM sync_log WHERE level='ERROR' ORDER BY id DESC LIMIT 1")
    last_err = cur.fetchone()
    sync_last_err_time = last_err[0] if last_err else None
    sync_last_err_msg = last_err[1] if last_err else None

    # --- 오늘 출퇴근 싱크 현황 ---
    today_str = datetime.now().strftime("%Y%m%d")
    cur.execute("SELECT COUNT(*) FROM tenter WHERE e_date=%s AND e_mode='1' AND e_time BETWEEN '060000' AND '083000'", (today_str,))
    today_in_cnt = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM tenter WHERE e_date=%s AND e_mode='2'", (today_str,))
    today_out_cnt = cur.fetchone()[0]

    # --- Docker 코드 싱크 현황 ---
    cur.execute("SELECT container_id, file_name, file_md5, checked_at FROM docker_code_hash ORDER BY file_name, container_id")
    dch_rows = cur.fetchall()
    # 파일별 컨테이너 해시 비교
    from collections import defaultdict as _dd
    dch_by_file = _dd(list)
    for cid, fname, md5, cat in dch_rows:
        dch_by_file[fname].append({"cid": cid, "md5": md5, "at": cat})
    docker_sync_ok = True
    docker_sync_files = []
    for fname, entries in dch_by_file.items():
        md5s = set(e["md5"] for e in entries)
        ok = len(md5s) == 1 and "NOT_FOUND" not in md5s
        if not ok:
            docker_sync_ok = False
        short = fname.replace("/app/", "")
        docker_sync_files.append({"name": short, "ok": ok, "entries": entries, "md5s": list(md5s)})
    docker_sync_cnt = len(set(r[0] for r in dch_rows)) if dch_rows else 0
    docker_sync_last = max((r[3] for r in dch_rows if r[3]), default=None) if dch_rows else None

    conn.close()

    # 통합 정렬 (최신순)
    logs.sort(key=lambda x: x["updated_at"] or datetime.min, reverse=True)
    logs = logs[:500]

    return render_template("log_management.html",
                           logs=logs, search=search, log_type=log_type,
                           date_from=date_from, date_to=date_to,
                           sync_total=sync_total,
                           sync_max_date=sync_max_date,
                           sync_max_uptime=sync_max_uptime,
                           sync_daily=sync_daily,
                           sync_recent=sync_recent,
                           sync_last_ok_time=sync_last_ok_time,
                           sync_last_ok_msg=sync_last_ok_msg,
                           sync_last_err_time=sync_last_err_time,
                           sync_last_err_msg=sync_last_err_msg,
                           today_in_cnt=today_in_cnt,
                           today_out_cnt=today_out_cnt,
                           docker_sync_ok=docker_sync_ok,
                           docker_sync_files=docker_sync_files,
                           docker_sync_cnt=docker_sync_cnt,
                           docker_sync_last=docker_sync_last)


@app.route("/fire_management")
@_login_required
def fire_management():
    if session.get("role") != "admin":
        flash("접근 권한이 없습니다.", "danger")
        return redirect(url_for("dashboard"))
    return render_template("fire_management.html")


@app.route("/api/fire/status")
@_login_required
def api_fire_status():
    if session.get("role") != "admin":
        return jsonify({"ok": False, "error": "권한이 없습니다."}), 403
    try:
        conn = _conn(); cur = conn.cursor()
        cur.execute("""
            SELECT device_key, device_name, location, status, battery, last_seen, updated_at
            FROM fire_devices
            ORDER BY location, device_name, device_key
        """)
        devices = []
        for dk, dn, loc, st, bat, ls, upd in cur.fetchall():
            devices.append({
                "device_key": dk,
                "device_name": dn or dk,
                "location": loc or "-",
                "status": st or "unknown",
                "battery": bat,
                "last_seen": ls.strftime("%Y-%m-%d %H:%M:%S") if ls else None,
                "updated_at": upd.strftime("%Y-%m-%d %H:%M:%S") if upd else None,
            })

        cur.execute("""
            SELECT device_key, event_type, severity, message, created_at
            FROM fire_alert_logs
            ORDER BY id DESC
            LIMIT 50
        """)
        logs = []
        for dk, et, sv, msg, cat in cur.fetchall():
            logs.append({
                "device_key": dk,
                "event_type": et,
                "severity": sv,
                "message": msg,
                "created_at": cat.strftime("%Y-%m-%d %H:%M:%S") if cat else "",
            })
        conn.close()

        summary = {"normal": 0, "alarm": 0, "offline": 0, "other": 0}
        for d in devices:
            st = (d.get("status") or "").lower()
            if st in ("normal", "ok", "safe"):
                summary["normal"] += 1
            elif st in ("alarm", "smoke", "fire", "fault"):
                summary["alarm"] += 1
            elif st in ("offline", "disconnected"):
                summary["offline"] += 1
            else:
                summary["other"] += 1

        return jsonify({"ok": True, "summary": summary, "devices": devices, "logs": logs})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/api/fire/event", methods=["POST"])
def api_fire_event():
    try:
        if FIRE_API_TOKEN:
            token = request.headers.get("X-Fire-Token", "")
            if token != FIRE_API_TOKEN:
                return jsonify({"ok": False, "error": "invalid token"}), 401

        data = request.get_json(silent=True) or {}
        device_key = str(data.get("device_key", "")).strip()
        if not device_key:
            return jsonify({"ok": False, "error": "device_key is required"}), 400

        device_name = str(data.get("device_name", "")).strip()
        location = str(data.get("location", "")).strip()
        status = str(data.get("status", "normal")).strip().lower() or "normal"
        event_type = str(data.get("event_type", "state")).strip().lower() or "state"
        message = str(data.get("message", "")).strip()

        battery = data.get("battery")
        try:
            battery = int(battery) if battery is not None and str(battery).strip() != "" else None
        except (TypeError, ValueError):
            battery = None

        if status in ("alarm", "smoke", "fire"):
            severity = "critical"
        elif status in ("fault", "offline", "disconnected"):
            severity = "warning"
        else:
            severity = "info"

        payload_text = json.dumps(data, ensure_ascii=False)
        conn = _conn(); cur = conn.cursor()
        cur.execute("""
            INSERT INTO fire_devices (device_key, device_name, location, status, battery, last_seen, last_payload)
            VALUES (%s, %s, %s, %s, %s, NOW(), %s)
            ON DUPLICATE KEY UPDATE
                device_name=VALUES(device_name),
                location=VALUES(location),
                status=VALUES(status),
                battery=VALUES(battery),
                last_seen=NOW(),
                last_payload=VALUES(last_payload)
        """, (device_key, device_name, location, status, battery, payload_text))

        if event_type != "heartbeat" or severity in ("critical", "warning"):
            cur.execute("""
                INSERT INTO fire_alert_logs (device_key, event_type, severity, message, payload)
                VALUES (%s, %s, %s, %s, %s)
            """, (device_key, event_type, severity, message, payload_text))

        conn.commit(); conn.close()

        # 위험물 안전 알림 훅
        if severity == "critical":
            try:
                from hazmat_bp import on_fire_event
                on_fire_event(device_key, device_name, location, event_type, status)
            except Exception:
                pass

        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/api/fire/mock_alarm", methods=["POST"])
@_login_required
def api_fire_mock_alarm():
    if session.get("role") != "admin":
        return jsonify({"ok": False, "error": "권한이 없습니다."}), 403
    try:
        payload = {
            "device_key": "mock-smoke-001",
            "device_name": "테스트 감지기",
            "location": "관리부",
            "status": "alarm",
            "event_type": "smoke_detected",
            "battery": 92,
            "message": "테스트 경보 이벤트",
        }

        conn = _conn(); cur = conn.cursor()
        payload_text = json.dumps(payload, ensure_ascii=False)
        cur.execute("""
            INSERT INTO fire_devices (device_key, device_name, location, status, battery, last_seen, last_payload)
            VALUES (%s, %s, %s, %s, %s, NOW(), %s)
            ON DUPLICATE KEY UPDATE
                device_name=VALUES(device_name),
                location=VALUES(location),
                status=VALUES(status),
                battery=VALUES(battery),
                last_seen=NOW(),
                last_payload=VALUES(last_payload)
        """, (payload["device_key"], payload["device_name"], payload["location"], payload["status"], payload["battery"], payload_text))
        cur.execute("""
            INSERT INTO fire_alert_logs (device_key, event_type, severity, message, payload)
            VALUES (%s, %s, 'critical', %s, %s)
        """, (payload["device_key"], payload["event_type"], payload["message"], payload_text))
        conn.commit(); conn.close()
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/api/fire/clear_mock", methods=["POST"])
@_login_required
def api_fire_clear_mock():
    """이벤트 로그 및 mock 장치 삭제"""
    if session.get("role") != "admin":
        return jsonify({"ok": False, "error": "권한이 없습니다."}), 403
    try:
        conn = _conn(); cur = conn.cursor()
        cur.execute("DELETE FROM fire_alert_logs")
        cur.execute("DELETE FROM fire_devices WHERE device_key LIKE 'mock-%'")
        conn.commit(); conn.close()
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/api/fire/heartbeat", methods=["POST"])
def api_fire_heartbeat():
    """Tuya 폴러에서 호출 - 장치 상태 업데이트 + 1시간 1회 로그"""
    data = request.get_json(silent=True) or {}
    device_key = str(data.get("device_key", "")).strip()
    if not device_key:
        return jsonify({"ok": False, "error": "device_key required"}), 400
    try:
        conn = _conn(); cur = conn.cursor()
        cur.execute("""
            INSERT INTO fire_devices (device_key, device_name, location, status, battery, last_seen)
            VALUES (%s, %s, %s, %s, %s, NOW())
            ON DUPLICATE KEY UPDATE
                device_name=VALUES(device_name),
                location=VALUES(location),
                status=VALUES(status),
                battery=VALUES(battery),
                last_seen=NOW()
        """, (device_key, data.get("device_name", device_key),
              data.get("location", ""), data.get("status", "normal"),
              data.get("battery")))
        # 1시간에 1번 heartbeat 로그 기록
        cur.execute("""
            SELECT COUNT(*) FROM fire_alert_logs
            WHERE device_key=%s AND event_type='heartbeat'
              AND created_at >= NOW() - INTERVAL 1 HOUR
        """, (device_key,))
        if cur.fetchone()[0] == 0:
            status = data.get("status", "normal")
            cur.execute("""
                INSERT INTO fire_alert_logs (device_key, event_type, severity, message, payload)
                VALUES (%s, 'heartbeat', 'info', %s, %s)
            """, (device_key, f"정상 감시 중 ({status})", json.dumps(data, ensure_ascii=False)))
        conn.commit(); conn.close()
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


# ── 사원명부 ──────────────────────────
@app.route("/roster")
@_login_required
def roster():
    if session.get("role") != "admin":
        flash("접근 권한이 없습니다.", "danger")
        return redirect(url_for("dashboard"))
    search = request.args.get("search", "").strip()
    sel_dept = request.args.get("dept", "").strip()
    sel_pos = request.args.get("pos", "").strip()
    conn = _conn(); cur = conn.cursor()
    sql = "SELECT id, name, dept, emp_no, gender, birth_date, calendar_type, age, appoint_date, position, phone, email, address, hire_date FROM employee_roster WHERE 1=1"
    params = []
    if sel_dept:
        sql += " AND dept=%s"; params.append(sel_dept)
    if sel_pos:
        sql += " AND position=%s"; params.append(sel_pos)
    if search:
        sql += " AND (name LIKE %s OR emp_no LIKE %s OR phone LIKE %s OR position LIKE %s)"
        params += [f"%{search}%"] * 4
    sql += " ORDER BY dept, name"
    cur.execute(sql, params)
    rows = []
    for r in cur.fetchall():
        hire_date_str = r[13] or ""
        years_worked = ""
        if hire_date_str and len(hire_date_str) >= 10:
            try:
                hd = datetime.strptime(hire_date_str[:10], "%Y-%m-%d")
                diff = (datetime.now() - hd).days
                y = diff // 365
                years_worked = f"{y}년차"
            except:
                pass
        rows.append({"id": r[0], "name": r[1], "dept": r[2], "emp_no": r[3],
                      "gender": r[4], "birth_date": r[5], "calendar_type": r[6],
                      "age": r[7], "appoint_date": r[8], "position": r[9],
                      "phone": r[10], "email": r[11], "address": r[12], "hire_date": r[13],
                      "years_worked": years_worked})
    cur.execute("SELECT DISTINCT dept FROM employee_roster ORDER BY dept")
    dept_list = [r[0] for r in cur.fetchall()]
    cur.execute("SELECT DISTINCT position FROM employee_roster WHERE position!='' ORDER BY position")
    pos_list = [r[0] for r in cur.fetchall()]
    # 이번달 생일자
    from datetime import datetime as _dt
    cur_month = _dt.now().strftime("%m")
    cur.execute("""SELECT name, dept, position, birth_date, calendar_type, age, phone
                   FROM employee_roster
                   WHERE SUBSTRING(birth_date,6,2)=%s
                   ORDER BY SUBSTRING(birth_date,9,2)""", (cur_month,))
    birthday_list = [{"name":r[0],"dept":r[1],"position":r[2],"birth_date":r[3],
                      "calendar_type":r[4],"age":r[5],"phone":r[6],
                      "day":r[3][8:10] if r[3] and len(r[3])>=10 else ""} for r in cur.fetchall()]
    # 정년퇴직 대상자 (만56세이상, 이사/회장/사장 제외)
    cur_year = _dt.now().year
    cur.execute("""SELECT name, dept, position, birth_date, age, hire_date, phone
                   FROM employee_roster
                   WHERE age >= 56
                     AND position NOT LIKE '%%이사%%'
                     AND position NOT LIKE '%%회장%%'
                     AND position NOT LIKE '%%사장%%'
                   ORDER BY age DESC, name""")
    peak_list = [{"name":r[0],"dept":r[1],"position":r[2],"birth_date":r[3],
                  "age":r[4],"hire_date":r[5],"phone":r[6]} for r in cur.fetchall()]
    # ERP 동기화 내역 (최근 10건)
    cur.execute("""SELECT synced_at, inserted, updated, detail
                   FROM erp_sync_log ORDER BY synced_at DESC LIMIT 10""")
    sync_logs = [{"synced_at": r[0], "inserted": r[1], "updated": r[2], "detail": r[3]}
                 for r in cur.fetchall()]
    conn.close()
    is_admin = session.get("role") == "admin"
    return render_template("roster.html", rows=rows, dept_list=dept_list, pos_list=pos_list,
                           search=search, sel_dept=sel_dept, sel_pos=sel_pos, is_admin=is_admin,
                           birthday_list=birthday_list, peak_list=peak_list, cur_month=cur_month,
                           sync_logs=sync_logs)


@app.route("/upload_roster", methods=["POST"])
@_admin_required
def upload_roster():
    f = request.files.get("file")
    if not f or not f.filename.endswith((".xlsx", ".xls")):
        return jsonify(ok=False, error="xlsx 파일을 선택하세요."), 400
    import openpyxl, io
    wb = openpyxl.load_workbook(io.BytesIO(f.read()))
    ws = wb.active
    # 헤더 행 찾기 - 헤더 이름으로 컬럼 인덱스 동적 매칭
    header_map = {"사원": "name", "부서": "dept", "사번": "emp_no",
                  "성별": "gender", "생년월일": "birth_date", "양력": "calendar_type",
                  "나이": "age", "발령일": "appoint_date", "직위": "position",
                  "휴대전화": "phone", "이메일": "email", "현주소": "address", "입사일": "hire_date"}
    col_map = {}
    header_row = None
    for r in range(1, min(10, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and str(v).strip() in header_map:
                col_map[header_map[str(v).strip()]] = c
        if "name" in col_map and "emp_no" in col_map:
            header_row = r
            break
        col_map.clear()
    if header_row is None or "emp_no" not in col_map:
        return jsonify(ok=False, error="헤더 행(사원,부서,사번,...)을 찾을 수 없습니다"), 400
    def _cell(r, key):
        ci = col_map.get(key)
        return ws.cell(r, ci).value if ci else None
    def _str(r, key):
        return str(_cell(r, key) or "").strip()
    def _date(r, key):
        v = _cell(r, key)
        return v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else str(v or "")
    conn = _conn(); cur = conn.cursor()
    cur.execute("DELETE FROM employee_roster")
    added = 0
    for r in range(header_row + 1, ws.max_row + 1):
        name = _str(r, "name")
        if not name:
            continue
        emp_no = _str(r, "emp_no")
        if not emp_no:
            continue
        try:
            age = int(_cell(r, "age"))
        except (ValueError, TypeError):
            age = 0
        cur.execute("""INSERT INTO employee_roster (name, dept, emp_no, gender, birth_date,
                       calendar_type, age, appoint_date, position, phone, email, address, hire_date)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (name, _str(r, "dept"), emp_no, _str(r, "gender"),
                     _date(r, "birth_date"), _str(r, "calendar_type"), age,
                     _date(r, "appoint_date"), _str(r, "position"), _str(r, "phone"),
                     _str(r, "email"), _str(r, "address"), _date(r, "hire_date")))
        added += 1
    # 명부에 있으나 tuser(직원 마스터)에 없는 사람 → tuser 자동 생성
    # (근무보고서 대상인원 등은 employee_roster ∩ tuser 로 후보를 만들기 때문)
    # tuser.id는 AUTO_INCREMENT가 아니라 수동 할당이며, 출입통제 장비(CAPS)가 쓰는
    # 순차 id(현재 ~332)와 충돌 방지를 위해 앱이 만드는 행은 90000+ 대역을 사용한다.
    # idno=emp_no(사번)로 넣어 기존 매핑 규칙과 일치시킨다.
    # 부서(roster.dept) → tuser.company 코드 매핑: 기존 직원(id<90000)의 부서별 최빈 company.
    # 근무보고서 후보 화면의 "부서" 필터(DEPT_MAP 기반)가 신규 인원에도 동작하도록 company를 채운다.
    cur.execute("""SELECT er.dept, t.company, COUNT(*) c
                   FROM employee_roster er JOIN tuser t ON t.name = er.name
                   WHERE IFNULL(t.company,'') <> '' AND t.id < 90000
                   GROUP BY er.dept, t.company""")
    _best = {}
    for d, comp, c in cur.fetchall():
        if d not in _best or c > _best[d][1]:
            _best[d] = (comp, c)
    dept_company = {d: v[0] for d, v in _best.items()}

    # 명부에 있으나 tuser에 없는 사람 → tuser 자동 생성
    # (근무보고서 대상인원 등은 employee_roster ∩ tuser 로 후보를 만들기 때문)
    # tuser.id는 AUTO_INCREMENT가 아니라 수동 할당이며, 출입통제 장비(CAPS)가 쓰는
    # 순차 id(현재 ~332)와 충돌 방지를 위해 앱이 만드는 행은 90000+ 대역을 사용한다.
    # idno=emp_no(사번), company=부서매핑 으로 넣어 기존 규칙/필터와 일치시킨다.
    tuser_added = 0
    cur.execute("""SELECT DISTINCT er.name, er.emp_no, er.dept
                   FROM employee_roster er
                   WHERE er.name IS NOT NULL AND er.name <> ''
                     AND NOT EXISTS (SELECT 1 FROM tuser t WHERE t.name = er.name)""")
    missing = cur.fetchall()
    if missing:
        cur.execute("SELECT IFNULL(MAX(id), 0) FROM tuser WHERE id >= 90000")
        next_id = max(cur.fetchone()[0], 89999) + 1
        today_ymd = datetime.now().strftime("%Y%m%d")
        for mname, memp, mdept in missing:
            cur.execute("""INSERT INTO tuser (id, name, idno, dept, company, reg_date)
                           VALUES (%s, %s, %s, %s, %s, %s)""",
                        (next_id, mname, (memp or None), (mdept or None),
                         dept_company.get(mdept), today_ymd))
            next_id += 1
            tuser_added += 1

    # 과거에 company 없이 만들어진 앱 생성분(id>=90000) 백필
    tuser_fixed = 0
    cur.execute("""SELECT t.id, er.dept FROM tuser t JOIN employee_roster er ON er.name = t.name
                   WHERE t.id >= 90000 AND IFNULL(t.company,'') = ''""")
    for tid, dept in cur.fetchall():
        comp = dept_company.get(dept)
        if comp:
            cur.execute("UPDATE tuser SET company=%s WHERE id=%s", (comp, tid))
            tuser_fixed += 1

    conn.commit(); conn.close()
    return jsonify(ok=True, added=added, updated=0, tuser_added=tuser_added, tuser_fixed=tuser_fixed)


@app.route("/api/roster_detail/<emp_no>")
@_login_required
def api_roster_detail(emp_no):
    if session.get("role") != "admin":
        return jsonify(ok=False, error="권한 없음"), 403
    conn = _conn(); cur = conn.cursor()
    cur.execute("""SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS
                   WHERE TABLE_SCHEMA=%s AND TABLE_NAME='employee_roster'
                   ORDER BY ORDINAL_POSITION""", (MARIA['db'],))
    cols = [r[0] for r in cur.fetchall()]
    cur.execute(f"SELECT {','.join(cols)} FROM employee_roster WHERE emp_no=%s", (emp_no,))
    row = cur.fetchone()
    conn.close()
    if not row:
        return jsonify(ok=False, error="not found"), 404
    data = {}
    for c, v in zip(cols, row):
        if hasattr(v, 'strftime'):
            data[c] = v.strftime('%Y-%m-%d %H:%M:%S')
        else:
            data[c] = v
    return jsonify(ok=True, data=data)


@app.route("/export_roster")
@_login_required
def export_roster():
    import openpyxl, io
    conn = _conn(); cur = conn.cursor()
    cur.execute("SELECT name, dept, emp_no, gender, birth_date, calendar_type, age, appoint_date, position, phone, email, address, hire_date FROM employee_roster ORDER BY dept, name")
    rows = cur.fetchall()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "사원명부"
    headers = ["사원", "부서", "사번", "성별", "생년월일", "양력", "나이", "발령일", "직위", "휴대전화", "이메일", "현주소", "입사일"]
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, download_name="사원명부.xlsx", as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ─── 공문서관리 (문서요청 워크플로우) ───
DOC_UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads", "documents")
os.makedirs(DOC_UPLOAD_DIR, exist_ok=True)

DOC_TYPES = ["재직증명서", "경력증명서", "급여명세서", "원천징수영수증", "퇴직금 관련", "기타"]


@app.route("/document_management")
@_login_required
def document_management():
    conn = _conn(); cur = conn.cursor()
    is_admin = _has_perm("document")
    user_id = session.get("user_id")
    status_filter = request.args.get("status", "all")
    doc_type_filter = request.args.get("doc_type", "all")

    if is_admin:
        sql = """SELECT id, requester_name, doc_type, purpose, memo, status,
                        handler_name, reply_memo, file_name, created_at, updated_at
                 FROM doc_requests WHERE 1=1"""
        params = []
    else:
        cur.execute("SELECT e_id FROM app_users WHERE id=%s", (user_id,))
        row = cur.fetchone()
        req_id = row[0] if row and row[0] else user_id
        sql = """SELECT id, requester_name, doc_type, purpose, memo, status,
                        handler_name, reply_memo, file_name, created_at, updated_at
                 FROM doc_requests WHERE requester_id=%s"""
        params = [req_id]

    if status_filter and status_filter != "all":
        sql += " AND status=%s"
        params.append(status_filter)
    if doc_type_filter and doc_type_filter != "all":
        sql += " AND doc_type=%s"
        params.append(doc_type_filter)
    sql += " ORDER BY created_at DESC LIMIT 500"
    cur.execute(sql, params)
    reqs = []
    for r in cur.fetchall():
        reqs.append({
            "id": r[0], "requester_name": r[1], "doc_type": r[2],
            "purpose": r[3], "memo": r[4], "status": r[5],
            "handler_name": r[6], "reply_memo": r[7], "file_name": r[8],
            "created_at": r[9], "updated_at": r[10],
        })
    # 상태별 카운트 (관리자용)
    counts = {"all": 0, "접수": 0, "처리중": 0, "완료": 0, "반려": 0}
    if is_admin:
        cur.execute("SELECT status, COUNT(*) FROM doc_requests GROUP BY status")
    else:
        cur.execute("SELECT status, COUNT(*) FROM doc_requests WHERE requester_id=%s GROUP BY status",
                    (params[0] if not is_admin else None,))
    for st, cnt in cur.fetchall():
        counts[st] = cnt
        counts["all"] += cnt
    conn.close()
    return render_template("document_management.html", reqs=reqs, is_handler=is_admin,
                           doc_types=DOC_TYPES, counts=counts,
                           status_filter=status_filter, doc_type_filter=doc_type_filter)


@app.route("/submit_doc_request", methods=["POST"])
@_login_required
def submit_doc_request():
    doc_type = request.json.get("doc_type", "").strip()
    purpose = request.json.get("purpose", "").strip()
    memo = request.json.get("memo", "").strip()
    if not doc_type:
        return jsonify(ok=False, msg="문서 종류를 선택하세요.")
    user_id = session.get("user_id")
    conn = _conn(); cur = conn.cursor()
    cur.execute("SELECT e_id, name FROM app_users WHERE id=%s", (user_id,))
    row = cur.fetchone()
    req_id = row[0] if row and row[0] else user_id
    req_name = row[1] if row else session.get("user_name", "")
    cur.execute("""INSERT INTO doc_requests (requester_id, requester_name, doc_type, purpose, memo)
                   VALUES (%s,%s,%s,%s,%s)""", (req_id, req_name, doc_type, purpose, memo))
    conn.commit(); conn.close()
    return jsonify(ok=True)


@app.route("/update_doc_request", methods=["POST"])
@_login_required
def update_doc_request():
    if not _has_perm("document"):
        return jsonify(ok=False, msg="권한 없음"), 403
    req_id = request.form.get("req_id")
    status = request.form.get("status", "").strip()
    reply_memo = request.form.get("reply_memo", "").strip()
    handler = session.get("user_name", "")

    file = request.files.get("file")
    file_name = ""
    file_path = ""
    if file and file.filename:
        from werkzeug.utils import secure_filename
        orig = file.filename
        safe = secure_filename(orig) or orig
        ts = datetime.now().strftime("%Y%m%d%H%M%S")
        save_name = f"{ts}_{safe}"
        dest = os.path.join(DOC_UPLOAD_DIR, save_name)
        file.save(dest)
        file_name = orig
        file_path = save_name

    conn = _conn(); cur = conn.cursor()
    if file_name:
        cur.execute("""UPDATE doc_requests SET status=%s, reply_memo=%s, handler_name=%s,
                       file_name=%s, file_path=%s WHERE id=%s""",
                    (status, reply_memo, handler, file_name, file_path, req_id))
    else:
        cur.execute("""UPDATE doc_requests SET status=%s, reply_memo=%s, handler_name=%s
                       WHERE id=%s""", (status, reply_memo, handler, req_id))
    conn.commit(); conn.close()
    return jsonify(ok=True)


@app.route("/get_doc_request/<int:req_id>")
@_login_required
def get_doc_request(req_id):
    conn = _conn(); cur = conn.cursor()
    cur.execute("""SELECT id, requester_id, requester_name, doc_type, purpose, memo, status,
                   handler_name, reply_memo, file_name, file_path, created_at, updated_at
                   FROM doc_requests WHERE id=%s""", (req_id,))
    row = cur.fetchone()
    conn.close()
    if not row:
        return jsonify(ok=False, msg="요청을 찾을 수 없습니다.")
    is_admin = _has_perm("document")
    user_id = session.get("user_id")
    if not is_admin:
        conn2 = _conn(); cur2 = conn2.cursor()
        cur2.execute("SELECT e_id FROM app_users WHERE id=%s", (user_id,))
        u = cur2.fetchone()
        conn2.close()
        my_eid = u[0] if u and u[0] else user_id
        if row[1] != my_eid:
            return jsonify(ok=False, msg="권한이 없습니다.")
    return jsonify(ok=True, req={
        "id": row[0], "requester_id": row[1], "requester_name": row[2],
        "doc_type": row[3], "purpose": row[4], "memo": row[5], "status": row[6],
        "handler_name": row[7], "reply_memo": row[8], "file_name": row[9],
        "file_path": row[10], "created_at": str(row[11]), "updated_at": str(row[12]),
    })


@app.route("/delete_doc_request", methods=["POST"])
@_login_required
def delete_doc_request():
    req_id = request.json.get("id")
    conn = _conn(); cur = conn.cursor()
    cur.execute("SELECT requester_id, file_path, status FROM doc_requests WHERE id=%s", (req_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return jsonify(ok=False, msg="요청을 찾을 수 없습니다.")
    is_admin = _has_perm("document")
    if not is_admin:
        cur.execute("SELECT e_id FROM app_users WHERE id=%s", (session.get("user_id"),))
        u = cur.fetchone()
        my_eid = u[0] if u and u[0] else session.get("user_id")
        if row[0] != my_eid:
            conn.close()
            return jsonify(ok=False, msg="본인 요청만 삭제 가능합니다.")
        if row[2] not in ("접수",):
            conn.close()
            return jsonify(ok=False, msg="접수 상태에서만 삭제 가능합니다.")
    if row[1]:
        fpath = os.path.join(DOC_UPLOAD_DIR, row[1])
        if os.path.exists(fpath):
            os.remove(fpath)
    cur.execute("DELETE FROM doc_requests WHERE id=%s", (req_id,))
    conn.commit(); conn.close()
    return jsonify(ok=True)


@app.route("/download_doc_file/<int:req_id>")
@_login_required
def download_doc_file(req_id):
    conn = _conn(); cur = conn.cursor()
    cur.execute("SELECT requester_id, file_name, file_path FROM doc_requests WHERE id=%s", (req_id,))
    row = cur.fetchone()
    conn.close()
    if not row or not row[2]:
        flash("파일이 없습니다.", "warning")
        return redirect(url_for("document_management"))
    is_admin = _has_perm("document")
    if not is_admin:
        conn2 = _conn(); cur2 = conn2.cursor()
        cur2.execute("SELECT e_id FROM app_users WHERE id=%s", (session.get("user_id"),))
        u = cur2.fetchone(); conn2.close()
        my_eid = u[0] if u and u[0] else session.get("user_id")
        if row[0] != my_eid:
            flash("권한이 없습니다.", "danger")
            return redirect(url_for("document_management"))
    fpath = os.path.join(DOC_UPLOAD_DIR, row[2])
    if not os.path.realpath(fpath).startswith(os.path.realpath(DOC_UPLOAD_DIR)):
        flash("잘못된 파일 경로입니다.", "danger")
        return redirect(url_for("document_management"))
    if not os.path.exists(fpath):
        flash("파일을 찾을 수 없습니다.", "warning")
        return redirect(url_for("document_management"))
    return send_file(fpath, download_name=row[1], as_attachment=True)



# ──────────────────────────────────────
#  근무보고서 관리
# ──────────────────────────────────────
WR_CATEGORIES = ["주간기본", "주간연장", "야간기본", "야간연장", "기타"]

@app.route("/work_report")
@_login_required
def work_report():
    is_admin = session.get("role") == "admin"
    login_uid = session.get("user_id")
    login_eid = session.get("e_id")  # tuser.id (wr_group용)
    conn = _conn(); cur = conn.cursor()
    # 검토자 여부 (e_id로 조회)
    cur.execute("SELECT group_id, step FROM wr_group_reviewers WHERE user_id=%s", (login_eid,))
    user_reviewer_rows = cur.fetchall()
    # 멤버 여부도 체크
    cur.execute("SELECT group_id FROM wr_group_members WHERE user_id=%s", (login_eid,))
    user_member_rows = cur.fetchall()
    # 작성자 여부도 체크
    cur.execute("SELECT group_id FROM wr_group_writers WHERE user_id=%s", (login_eid,))
    user_writer_rows = cur.fetchall()
    has_wr_perm = is_admin or _has_perm("schedule") or _has_perm("work_report") or bool(user_writer_rows) or bool(user_reviewer_rows)
    if not has_wr_perm and not user_reviewer_rows and not user_member_rows and not user_writer_rows:
        flash("접근 권한이 없습니다.", "danger")
        conn.close()
        return redirect(url_for("dashboard"))
    status_filter = request.args.get("status", "all").strip()
    mine_pending = request.args.get("mine_pending", "").strip() == "1"
    recent_mode = "month" not in request.args
    sel_month = request.args.get("month", datetime.now().strftime("%Y-%m"))
    wr_page = max(1, int(request.args.get("wr_page", 1) or 1))
    # 목록 필터: 그룹 / 작성자 / 페이지당 건수
    filter_gid = (request.args.get("fgid") or "").strip()
    filter_writer = (request.args.get("writer") or "").strip()
    try:
        WR_PER_PAGE = int(request.args.get("per_page", 20) or 20)
    except ValueError:
        WR_PER_PAGE = 20
    if WR_PER_PAGE not in (10, 20, 50, 100):
        WR_PER_PAGE = 20
    # ── 작성자 목록 (필터 드롭다운용) ──
    cur.execute("""SELECT created_by_name, COUNT(*) FROM wr_reports
                   WHERE created_by_name IS NOT NULL AND created_by_name <> ''
                   GROUP BY created_by_name ORDER BY COUNT(*) DESC""")
    writer_list = [{"name": r[0], "cnt": int(r[1])} for r in cur.fetchall()]
    # ── 그룹 목록 ──
    cur.execute("SELECT id, group_name, final_step, IFNULL(status,'설정중') FROM wr_groups ORDER BY group_name")
    all_groups = [{"id": r[0], "name": r[1], "final_step": r[2], "status": r[3]} for r in cur.fetchall()]
    # 검토자인 그룹 ID 집합
    reviewer_group_ids = set(r[0] for r in user_reviewer_rows)
    # 멤버인 그룹 ID 집합
    member_group_ids = set(r[0] for r in user_member_rows)
    # 작성자인 그룹 ID 집합
    writer_group_ids = set(r[0] for r in user_writer_rows)
    user_related_group_ids = reviewer_group_ids | member_group_ids | writer_group_ids
    # 관리자/schedule 권한자는 전체 그룹, 그 외는 검토자/멤버/작성자인 그룹
    if is_admin or _has_perm("schedule") or _has_perm("work_report"):
        groups = all_groups
    else:
        groups = [g for g in all_groups if g["id"] in user_related_group_ids]
    sel_group = request.args.get("gid", "").strip()
    sel_group = int(sel_group) if sel_group.isdigit() else 0
    # gid 미지정 시 검토자 그룹 자동 선택
    if not sel_group and not is_admin and groups:
        sel_group = groups[0]["id"]
    reviewers = {"1": [], "2": [], "3": []}
    members = []
    writers = []
    group_final_step = 1
    group_status = "설정중"
    if sel_group:
        cur.execute("SELECT final_step, IFNULL(status,'설정중') FROM wr_groups WHERE id=%s", (sel_group,))
        fs = cur.fetchone()
        group_final_step = fs[0] if fs else 1
        group_status = fs[1] if fs else "설정중"
        cur.execute("SELECT step, user_id, user_name FROM wr_group_reviewers WHERE group_id=%s ORDER BY step, user_name", (sel_group,))
        for step, uid, uname in cur.fetchall():
            reviewers[str(step)].append({"user_id": uid, "user_name": uname})
        cur.execute("SELECT user_id, user_name FROM wr_group_members WHERE group_id=%s ORDER BY user_name", (sel_group,))
        members = [{"user_id": r[0], "user_name": r[1]} for r in cur.fetchall()]
        cur.execute("SELECT user_id, user_name FROM wr_group_writers WHERE group_id=%s ORDER BY user_name", (sel_group,))
        writers = [{"user_id": r[0], "user_name": r[1]} for r in cur.fetchall()]
    # 그룹 요약 정보
    all_group_info = {}
    for g in groups:
        cur.execute("SELECT step, user_name FROM wr_group_reviewers WHERE group_id=%s ORDER BY step", (g["id"],))
        revs = {"1": [], "2": [], "3": []}
        for step, uname in cur.fetchall():
            revs[str(step)].append(uname)
        cur.execute("SELECT COUNT(*) FROM wr_group_members WHERE group_id=%s", (g["id"],))
        mcnt = cur.fetchone()[0]
        cur.execute("SELECT user_name FROM wr_group_members WHERE group_id=%s ORDER BY user_name", (g["id"],))
        mnames = [r[0] for r in cur.fetchall()]
        all_group_info[g["id"]] = {"revs": revs, "member_count": mcnt, "member_names": mnames,
                                    "final_step": g["final_step"], "status": g.get("status", "설정중"),
                                    # name: 미작성 경고 배너에서 그룹명으로 그룹을 찾기 위해 사용
                                    "name": g.get("group_name", "")}
    # 로그인 사용자별 그룹-step (reviewer + member + writer) - e_id 사용
    cur.execute("SELECT group_id, step FROM wr_group_reviewers WHERE user_id=%s", (login_eid,))
    user_group_steps = {}
    for gid, step in cur.fetchall():
        user_group_steps.setdefault(gid, []).append(step)
    # 멤버도 step=1 권한 부여
    cur.execute("SELECT group_id FROM wr_group_members WHERE user_id=%s", (login_eid,))
    for (gid,) in cur.fetchall():
        if gid not in user_group_steps:
            user_group_steps[gid] = [1]
        elif 1 not in user_group_steps[gid]:
            user_group_steps[gid].append(1)
    # 작성자도 step=1 권한 부여 (대상인원에 안 포함되지만 작성 가능)
    cur.execute("SELECT group_id FROM wr_group_writers WHERE user_id=%s", (login_eid,))
    for (gid,) in cur.fetchall():
        if gid not in user_group_steps:
            user_group_steps[gid] = [1]
        elif 1 not in user_group_steps[gid]:
            user_group_steps[gid].append(1)
    # 직원 목록
    cur_year = datetime.now().year
    cur.execute("""
        SELECT t.id, t.name, IFNULL(t.company,'') as company,
               IFNULL(er.dept,'') as sub_dept,
               IFNULL(er.position,'') as position
        FROM employee_roster er
        INNER JOIN (
            SELECT COALESCE(MAX(CASE WHEN company<>'' THEN id END), MAX(id)) AS id,
                   name, COALESCE(MAX(company),'') AS company
            FROM tuser WHERE name IS NOT NULL AND name <> ''
            GROUP BY name
        ) t ON t.name = er.name
        ORDER BY t.company, t.name
    """)
    all_users = []
    reviewer_users = []
    dept_set = set(); sub_dept_set = set()
    non_reviewer_positions = ('사원', '')
    for r in cur.fetchall():
        dept_name = DEPT_MAP.get((r[2] or "").strip(), "")
        sub_dept_name = (r[3] or "").strip()
        pos = (r[4] or "").strip()
        u = {"id": r[0], "name": r[1], "dept": dept_name, "sub_dept": sub_dept_name, "position": pos}
        all_users.append(u)
        if pos not in non_reviewer_positions:
            reviewer_users.append(u)
        if dept_name: dept_set.add(dept_name)
        if sub_dept_name: sub_dept_set.add(sub_dept_name)
    dept_list = sorted(dept_set)
    sub_dept_list = sorted(sub_dept_set)
    # ── 보고서 목록 ──
    member_group_ids = set()
    if has_wr_perm:
        member_group_ids = set(g["id"] for g in groups)
    else:
        for gid in user_group_steps:
            member_group_ids.add(gid)
    reports_list = []
    wr_total = 0
    wr_total_pages = 1
    # ── 관리자: 해당일 경과 + 결재 미완료 보고서 목록 ──
    overdue_reports = []
    if is_admin:
        # 1일 이상 경과 & 실제 결재 대기('작성완료','검토')만. '대기'(작성자 미제출)·미래건 제외.
        # 최근 60일로 제한 + WR_OVERDUE_PANEL_START(2026-07-01) 이전 과거분 제외.
        cutoff = (datetime.now().date() - timedelta(days=1)).strftime("%Y%m%d")
        since = (datetime.now().date() - timedelta(days=60)).strftime("%Y%m%d")
        since = max(since, WR_OVERDUE_PANEL_START)
        cur.execute("""
            SELECT r.id, r.group_id, r.report_date, r.status, r.created_by_name,
                   g.group_name
            FROM wr_reports r JOIN wr_groups g ON r.group_id = g.id
            WHERE r.report_date <= %s AND r.report_date >= %s
              AND r.status IN ('작성완료', '검토')
            ORDER BY r.report_date DESC, g.group_name
            LIMIT 100
        """, (cutoff, since))
        for oid, ogid, ord_, ost, owriter, ogname in cur.fetchall():
            rd = str(ord_)
            disp = f"{rd[:4]}-{rd[4:6]}-{rd[6:8]}" if len(rd) == 8 else rd
            try:
                days_late = (datetime.now().date() - datetime.strptime(rd, "%Y%m%d").date()).days
            except ValueError:
                days_late = 0
            overdue_reports.append({
                "id": oid, "gid": ogid, "date": disp, "status": ost,
                "writer": owriter or "", "group_name": ogname, "days_late": days_late,
            })
    if member_group_ids:
        ph = ','.join(['%s'] * len(member_group_ids))
        sql = f"""SELECT r.id, r.group_id, r.report_date, r.status, r.created_by_name,
                         r.created_at, g.group_name, r.reject_reason
                  FROM wr_reports r JOIN wr_groups g ON r.group_id = g.id
                  WHERE r.group_id IN ({ph})"""
        params = list(member_group_ids)
        if not recent_mode:
            sql += " AND DATE_FORMAT(STR_TO_DATE(r.report_date, '%%Y%%m%%d'), '%%Y-%%m') = %s"
            params.append(sel_month)
        if status_filter and status_filter != "all":
            sql += " AND r.status = %s"
            params.append(status_filter)
        # 그룹 필터
        if filter_gid:
            try:
                sql += " AND r.group_id = %s"
                params.append(int(filter_gid))
            except ValueError:
                pass
        # 작성자 필터 (드롭다운에서 선택 → 정확히 일치)
        if filter_writer:
            sql += " AND r.created_by_name = %s"
            params.append(filter_writer)
        # ── 내 미결재만: 사용자가 결재자(reviewer)인 보고서 중, 사용자의 결재 이력이 없는 것 ──
        if mine_pending and user_reviewer_rows:
            # 사용자가 결재자인 그룹 ID
            my_rev_gids = list(set(r[0] for r in user_reviewer_rows))
            ph_rev = ','.join(['%s'] * len(my_rev_gids))
            sql += f""" AND r.group_id IN ({ph_rev})
                       AND r.status NOT IN ('결재', '반려')
                       AND NOT EXISTS (
                           SELECT 1 FROM wr_approval_log al
                           WHERE al.report_id = r.id
                             AND al.user_id = %s
                             AND al.action != '반려'
                       )"""
            params.extend(my_rev_gids)
            params.append(login_eid)
        sql += " ORDER BY r.report_date DESC, g.group_name, r.id"
        # 전체 건수 조회
        cur.execute("SELECT COUNT(*) FROM (" + sql + ") _cnt", params)
        wr_total = cur.fetchone()[0]
        wr_total_pages = max(1, (wr_total + WR_PER_PAGE - 1) // WR_PER_PAGE)
        wr_page = min(wr_page, wr_total_pages)
        sql += " LIMIT %s OFFSET %s"
        cur.execute(sql, params + [WR_PER_PAGE, (wr_page - 1) * WR_PER_PAGE])
        # 같은 날짜+그룹 회차 계산용
        date_group_cnt = {}
        raw_rows = cur.fetchall()
        for row in raw_rows:
            key = (row[1], str(row[2]))
            date_group_cnt[key] = date_group_cnt.get(key, 0) + 1
        date_group_seq = {}
        for row in raw_rows:
            rd = str(row[2])
            is_holiday = False
            is_weekend = False
            if len(rd) == 8:
                disp_date = f"{rd[:4]}-{rd[4:6]}-{rd[6:8]}"
                dt_obj = datetime.strptime(rd, "%Y%m%d").date()
                wd_idx = dt_obj.weekday()
                disp_weekday = ["월","화","수","목","금","토","일"][wd_idx]
                is_weekend = wd_idx >= 5
                is_holiday = dt_obj in KR_HOLIDAYS
            else:
                disp_date = rd
                disp_weekday = ""
            key = (row[1], rd)
            date_group_seq[key] = date_group_seq.get(key, 0) + 1
            seq = date_group_seq[key]
            total = date_group_cnt[key]
            # 엔트리 수
            cur.execute("SELECT COUNT(*) FROM wr_entries WHERE report_id=%s", (row[0],))
            ecnt = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM wr_entries WHERE report_id=%s AND skipped=1", (row[0],))
            skip_cnt = cur.fetchone()[0]
            label = disp_date
            if total > 1:
                label += f" (수정{seq-1})" if seq > 1 else " (원본)"
            # 결재이력 + 그룹 결재자 정보
            group_id = row[1]
            cur.execute("SELECT final_step FROM wr_groups WHERE id=%s", (group_id,))
            fs_row = cur.fetchone()
            gfinal = fs_row[0] if fs_row else 1
            # 각 단계별 지정 결재자
            step_reviewers = {}
            cur.execute("SELECT step, user_name FROM wr_group_reviewers WHERE group_id=%s ORDER BY step, user_name", (group_id,))
            for sr_step, sr_name in cur.fetchall():
                step_reviewers.setdefault(sr_step, []).append(sr_name)
            # 실제 결재이력
            cur.execute("SELECT step, action, user_name FROM wr_approval_log WHERE report_id=%s ORDER BY id", (row[0],))
            done_steps = {}
            for astep, aaction, aname in cur.fetchall():
                if aaction != '반려':
                    done_steps[astep] = {"name": aname, "action": aaction}
            # 전체 흐름 조합
            step_labels = {1: "작성", 2: "검토", 3: "결재"}
            aflow = []
            for s in range(1, gfinal + 1):
                lbl = step_labels.get(s, f"{s}차")
                if s in done_steps:
                    aflow.append({"step": s, "label": lbl, "name": done_steps[s]["name"], "done": True})
                else:
                    names = step_reviewers.get(s, [])
                    if names:
                        fallback = ", ".join(names)
                    elif s == 1 and row[4]:
                        fallback = row[4]
                    else:
                        fallback = "?"
                    aflow.append({"step": s, "label": lbl, "name": fallback, "done": False})
            reports_list.append({
                "id": row[0], "group_id": group_id, "report_date": rd,
                "disp_date": label, "disp_weekday": disp_weekday,
                "is_holiday": is_holiday, "is_weekend": is_weekend,
                "status": row[3],
                "created_by_name": row[4], "created_at": row[5],
                "group_name": row[6], "entry_count": ecnt,
                "reject_reason": row[7] or "",
                "approval_flow": aflow,
                "skip_count": skip_cnt,
                "seq": seq, "total": total,
            })
    # 같은 날짜+그룹 보고서 그룹핑
    from collections import OrderedDict
    report_groups = OrderedDict()
    for r in reports_list:
        key = (r["report_date"], r["group_id"])
        if key not in report_groups:
            report_groups[key] = []
        report_groups[key].append(r)
    conn.close()
    return render_template("work_report.html",
                           reports=reports_list, sel_month=sel_month,
                           recent_mode=recent_mode,
                           wr_page=wr_page, wr_total_pages=wr_total_pages, wr_total=wr_total,
                           filter_gid=filter_gid, filter_writer=filter_writer,
                           per_page=WR_PER_PAGE, writer_list=writer_list,
                           report_groups=list(report_groups.values()),
                           status_filter=status_filter,
                           mine_pending=mine_pending,
                           is_reviewer=bool(user_reviewer_rows),
                           groups=groups, sel_group=sel_group,
                           reviewers=reviewers, members=members, writers=writers,
                           group_final_step=group_final_step,
                           group_status=group_status,
                           all_group_info=all_group_info,
                           all_users=all_users,
                           reviewer_users=reviewer_users,
                           dept_list=dept_list, sub_dept_list=sub_dept_list,
                           user_group_steps=user_group_steps,
                           user_default_group=list(user_group_steps.keys())[0] if user_group_steps else None,
                           login_user_name=session.get("user_name", ""),
                           is_admin=is_admin, has_wr_perm=has_wr_perm,
                           overdue_reports=overdue_reports,
                           categories=WR_CATEGORIES)


# ═══ 연차계획서 그룹 관리 API ═══
@app.route("/save_lp_group", methods=["POST"])
@_login_required
def save_lp_group():
    if session.get("role") != "admin":
        return jsonify(ok=False, error="관리자만 가능"), 403
    data = request.get_json()
    action = str(data.get("action", ""))
    conn = _conn(); cur = conn.cursor()
    try:
        if action == "create":
            name = str(data["name"]).strip()
            if not name:
                return jsonify(ok=False, error="그룹명을 입력하세요")
            cur.execute("INSERT INTO lp_groups (group_name) VALUES (%s)", (name,))
        elif action == "rename":
            cur.execute("UPDATE lp_groups SET group_name=%s WHERE id=%s", (str(data["name"]).strip(), int(data["id"])))
        elif action == "delete":
            gid = int(data["id"])
            cur.execute("DELETE FROM lp_group_reviewers WHERE group_id=%s", (gid,))
            cur.execute("DELETE FROM lp_group_members WHERE group_id=%s", (gid,))
            cur.execute("DELETE FROM lp_groups WHERE id=%s", (gid,))
        elif action == "confirm":
            gid = int(data["id"])
            cur.execute("UPDATE lp_groups SET status='완료' WHERE id=%s", (gid,))
        elif action == "reopen":
            cur.execute("UPDATE lp_groups SET status='설정중' WHERE id=%s", (int(data["id"]),))
        else:
            return jsonify(ok=False, error="알 수 없는 action")
        conn.commit()
    except Exception as ex:
        conn.close()
        return jsonify(ok=False, error=str(ex)), 500
    conn.close()
    return jsonify(ok=True)


@app.route("/save_lp_group_reviewer", methods=["POST"])
@_login_required
def save_lp_group_reviewer():
    if session.get("role") != "admin":
        return jsonify(ok=False, error="관리자만 가능"), 403
    data = request.get_json()
    gid, step, user_id = int(data["group_id"]), int(data["step"]), int(data["user_id"])
    action = str(data.get("action", "add"))
    conn = _conn(); cur = conn.cursor()
    if action == "delete":
        cur.execute("DELETE FROM lp_group_reviewers WHERE group_id=%s AND step=%s AND user_id=%s", (gid, step, user_id))
    else:
        cur.execute("SELECT name FROM tuser WHERE id=%s", (user_id,))
        row = cur.fetchone()
        if not row:
            cur.execute("SELECT name FROM app_users WHERE id=%s", (user_id,))
            row = cur.fetchone()
        uname = row[0] if row else ""
        cur.execute("""INSERT INTO lp_group_reviewers (group_id, step, user_id, user_name) VALUES (%s,%s,%s,%s)
                       ON DUPLICATE KEY UPDATE user_name=%s""", (gid, step, user_id, uname, uname))
    conn.commit(); conn.close()
    return jsonify(ok=True)


@app.route("/save_lp_group_member", methods=["POST"])
@_login_required
def save_lp_group_member():
    if session.get("role") != "admin":
        return jsonify(ok=False, error="관리자만 가능"), 403
    data = request.get_json()
    gid = int(data["group_id"])
    action = str(data.get("action", "add"))
    conn = _conn(); cur = conn.cursor()
    if action == "delete":
        cur.execute("DELETE FROM lp_group_members WHERE group_id=%s AND user_id=%s", (gid, int(data["user_id"])))
    elif action == "add_batch":
        for uid in data.get("user_ids", []):
            uid = int(uid)
            cur.execute("SELECT name FROM tuser WHERE id=%s", (uid,))
            row = cur.fetchone()
            if not row:
                cur.execute("SELECT name FROM app_users WHERE id=%s", (uid,))
                row = cur.fetchone()
            uname = row[0] if row else ""
            cur.execute("""INSERT INTO lp_group_members (group_id, user_id, user_name) VALUES (%s,%s,%s)
                           ON DUPLICATE KEY UPDATE user_name=%s""", (gid, uid, uname, uname))
    else:
        uid = int(data["user_id"])
        cur.execute("SELECT name FROM tuser WHERE id=%s", (uid,))
        row = cur.fetchone()
        if not row:
            cur.execute("SELECT name FROM app_users WHERE id=%s", (uid,))
            row = cur.fetchone()
        uname = row[0] if row else ""
        cur.execute("""INSERT INTO lp_group_members (group_id, user_id, user_name) VALUES (%s,%s,%s)
                       ON DUPLICATE KEY UPDATE user_name=%s""", (gid, uid, uname, uname))
    conn.commit(); conn.close()
    return jsonify(ok=True)


@app.route("/save_wr_group", methods=["POST"])
@_login_required
def save_wr_group():
    if session.get("role") != "admin":
        return jsonify(ok=False, error="관리자만 가능"), 403
    data = request.get_json()
    action = str(data.get("action", ""))
    conn = _conn(); cur = conn.cursor()
    try:
        if action == "create":
            name = str(data["name"]).strip()
            if not name: return jsonify(ok=False, error="그룹명을 입력하세요")
            final_step = 2  # 검토 + 결재 고정
            cur.execute("INSERT INTO wr_groups (group_name, final_step) VALUES (%s,%s)", (name, final_step))
        elif action == "rename":
            cur.execute("UPDATE wr_groups SET group_name=%s WHERE id=%s", (str(data["name"]).strip(), int(data["id"])))
        elif action == "set_final_step":
            gid = int(data["id"])
            new_step = int(data["final_step"])
            cur.execute("UPDATE wr_groups SET final_step=%s WHERE id=%s", (new_step, gid))
            # 새 final_step 초과 단계의 검토자 삭제
            cur.execute("DELETE FROM wr_group_reviewers WHERE group_id=%s AND step > %s", (gid, new_step))
        elif action == "delete":
            gid = int(data["id"])
            cur.execute("DELETE FROM wr_entries WHERE report_id IN (SELECT id FROM wr_reports WHERE group_id=%s)", (gid,))
            cur.execute("DELETE FROM wr_reports WHERE group_id=%s", (gid,))
            cur.execute("DELETE FROM wr_group_reviewers WHERE group_id=%s", (gid,))
            cur.execute("DELETE FROM wr_group_members WHERE group_id=%s", (gid,))
            cur.execute("DELETE FROM wr_groups WHERE id=%s", (gid,))
        elif action == "confirm":
            gid = int(data["id"])
            section = int(data.get("section", 0))
            if section == 3:
                cur.execute("UPDATE wr_groups SET status='완료' WHERE id=%s", (gid,))
        elif action == "reopen":
            cur.execute("UPDATE wr_groups SET status='설정중' WHERE id=%s", (int(data["id"]),))
        else:
            return jsonify(ok=False, error="알 수 없는 action")
        conn.commit()
    except Exception as ex:
        conn.close()
        return jsonify(ok=False, error=str(ex)), 500
    conn.close()
    return jsonify(ok=True)


@app.route("/save_wr_group_reviewer", methods=["POST"])
@_login_required
def save_wr_group_reviewer():
    if session.get("role") != "admin":
        return jsonify(ok=False, error="관리자만 가능"), 403
    data = request.get_json()
    gid, step, user_id = int(data["group_id"]), int(data["step"]), int(data["user_id"])
    action = str(data.get("action", "add"))
    conn = _conn(); cur = conn.cursor()
    if action == "delete":
        cur.execute("DELETE FROM wr_group_reviewers WHERE group_id=%s AND step=%s AND user_id=%s", (gid, step, user_id))
    else:
        cur.execute("SELECT name FROM tuser WHERE id=%s", (user_id,))
        row = cur.fetchone()
        if not row:
            cur.execute("SELECT name FROM app_users WHERE id=%s", (user_id,))
            row = cur.fetchone()
        uname = row[0] if row else ""
        cur.execute("""INSERT INTO wr_group_reviewers (group_id, step, user_id, user_name) VALUES (%s,%s,%s,%s)
                       ON DUPLICATE KEY UPDATE user_name=%s""", (gid, step, user_id, uname, uname))
    conn.commit(); conn.close()
    return jsonify(ok=True)


@app.route("/save_wr_group_member", methods=["POST"])
@_login_required
def save_wr_group_member():
    if session.get("role") != "admin":
        return jsonify(ok=False, error="관리자만 가능"), 403
    data = request.get_json()
    gid = int(data["group_id"])
    action = str(data.get("action", "add"))
    conn = _conn(); cur = conn.cursor()
    if action == "delete":
        cur.execute("DELETE FROM wr_group_members WHERE group_id=%s AND user_id=%s", (gid, int(data["user_id"])))
    elif action == "add_batch":
        for uid in data.get("user_ids", []):
            uid = int(uid)
            cur.execute("SELECT name FROM tuser WHERE id=%s", (uid,))
            row = cur.fetchone()
            if not row:
                cur.execute("SELECT name FROM app_users WHERE id=%s", (uid,))
                row = cur.fetchone()
            uname = row[0] if row else ""
            cur.execute("""INSERT INTO wr_group_members (group_id, user_id, user_name) VALUES (%s,%s,%s)
                           ON DUPLICATE KEY UPDATE user_name=%s""", (gid, uid, uname, uname))
    else:
        uid = int(data["user_id"])
        cur.execute("SELECT name FROM tuser WHERE id=%s", (uid,))
        row = cur.fetchone()
        if not row:
            cur.execute("SELECT name FROM app_users WHERE id=%s", (uid,))
            row = cur.fetchone()
        uname = row[0] if row else ""
        cur.execute("""INSERT INTO wr_group_members (group_id, user_id, user_name) VALUES (%s,%s,%s)
                       ON DUPLICATE KEY UPDATE user_name=%s""", (gid, uid, uname, uname))
    conn.commit(); conn.close()
    return jsonify(ok=True)


@app.route("/save_wr_group_writer", methods=["POST"])
@_login_required
def save_wr_group_writer():
    """작성자 추가/삭제 (대상인원에 포함 안 됨)"""
    if session.get("role") != "admin":
        return jsonify(ok=False, error="관리자만 가능"), 403
    data = request.get_json()
    gid = int(data["group_id"])
    action = str(data.get("action", "add"))
    conn = _conn(); cur = conn.cursor()
    if action == "delete":
        cur.execute("DELETE FROM wr_group_writers WHERE group_id=%s AND user_id=%s", (gid, int(data["user_id"])))
    elif action == "add_batch":
        for uid in data.get("user_ids", []):
            uid = int(uid)
            cur.execute("SELECT name FROM tuser WHERE id=%s", (uid,))
            row = cur.fetchone()
            if not row:
                cur.execute("SELECT name FROM app_users WHERE id=%s", (uid,))
                row = cur.fetchone()
            uname = row[0] if row else ""
            cur.execute("""INSERT INTO wr_group_writers (group_id, user_id, user_name) VALUES (%s,%s,%s)
                           ON DUPLICATE KEY UPDATE user_name=%s""", (gid, uid, uname, uname))
    else:
        uid = int(data["user_id"])
        cur.execute("SELECT name FROM tuser WHERE id=%s", (uid,))
        row = cur.fetchone()
        if not row:
            cur.execute("SELECT name FROM app_users WHERE id=%s", (uid,))
            row = cur.fetchone()
        uname = row[0] if row else ""
        cur.execute("""INSERT INTO wr_group_writers (group_id, user_id, user_name) VALUES (%s,%s,%s)
                       ON DUPLICATE KEY UPDATE user_name=%s""", (gid, uid, uname, uname))
    conn.commit(); conn.close()
    return jsonify(ok=True)


def _jotoe_hours(etc_val):
    """'조퇴 16:00' → 조퇴시간 N (17:30 - 조퇴시각, 점심 12:30~13:30 겹치면 제외). 시각 없으면 None."""
    m = re.search(r'(\d{1,2}):(\d{2})', etc_val or "")
    if not m:
        return None
    leave = int(m.group(1)) * 60 + int(m.group(2))
    end = 17 * 60 + 30          # 17:30
    if leave >= end:
        return 0
    diff = end - leave
    ls, le = 12 * 60 + 30, 13 * 60 + 30   # 점심 12:30~13:30
    os_, oe = max(leave, ls), min(end, le)
    if oe > os_:
        diff -= (oe - os_)
    return diff / 60.0


def _outing_hours(etc_val):
    """'외출 10:00~12:00' → 외출시간 N (복귀-외출, 점심 12:30~13:30 겹치면 제외). 시각 2개 없으면 None."""
    times = re.findall(r'(\d{1,2}):(\d{2})', etc_val or "")
    if len(times) < 2:
        return None
    t1 = int(times[0][0]) * 60 + int(times[0][1])   # 외출
    t2 = int(times[1][0]) * 60 + int(times[1][1])   # 복귀
    if t2 <= t1:
        return 0
    diff = t2 - t1
    ls, le = 12 * 60 + 30, 13 * 60 + 30
    os_, oe = max(t1, ls), min(t2, le)
    if oe > os_:
        diff -= (oe - os_)
    return diff / 60.0


def _fmt_hours(n):
    """4.0 → '4', 1.5 → '1.5'"""
    if n is None:
        return ""
    return str(int(n)) if float(n) == int(n) else str(n)


@app.route("/api/wr_report", methods=["POST"])
@_login_required
def save_wr_report():
    """보고서 생성/수정 (관리자가 그룹 멤버를 카테고리별로 배치)"""
    try:
        data = request.get_json(force=True)
        group_id = int(data["group_id"])
        report_date = str(data["report_date"]).replace("-", "")
        memo = str(data.get("memo", "")).strip()[:1000]
        entries = data.get("entries", [])
        edit_id = data.get("report_id")
        skip_leave_check = data.get("skip_leave_check", False)
        # 기타 항목 사유 미선택 차단
        etc_missing = [str(e.get("user_name", "")) for e in entries
                       if str(e.get("category", "")) == "기타"
                       and not int(e.get("skipped", 0))
                       and not str(e.get("etc_value", "")).strip()]
        if etc_missing:
            return jsonify(ok=False, msg="기타 항목의 사유가 선택되지 않았습니다: " + ", ".join(etc_missing))
        # 연차/무급 + 근무 동시배정 차단
        # 적재 시 휴가가 다른 값을 모두 덮어쓰므로(아래 a["leave"] 보정) 한 사람이
        # 주간기본과 기타/무급에 함께 들어가면 근무시간이 통째로 0 이 된다.
        # (2026-07-30 윤경진 8/2/0 → 0/0/0 으로 사라졌고 화면에도 표시되지 않았다)
        _work_cats = ("주간기본", "주간연장", "야간기본", "야간연장")
        _by_user = {}
        for e in entries:
            if int(e.get("skipped", 0)):
                continue
            u = _by_user.setdefault(str(e.get("user_name", "")), {"leave": "", "work": []})
            _cat = str(e.get("category", ""))
            _ev = str(e.get("etc_value", "")).strip()
            if _cat == "기타" and _ev in ("연차", "무급"):
                u["leave"] = _ev
            elif _cat in _work_cats:
                u["work"].append(_cat)
        both = [f"{nm}({u['leave']}+{'·'.join(u['work'])})"
                for nm, u in _by_user.items() if u["leave"] and u["work"]]
        if both:
            return jsonify(ok=False, msg="연차/무급과 근무가 함께 배정된 인원이 있습니다. "
                                         "한쪽을 빼주세요: " + ", ".join(both))
        conn = _conn(); cur = conn.cursor()
        # 연차 체크
        if not skip_leave_check:
            leave_conflicts = []
            for e in entries:
                uid = int(e["user_id"])
                skipped = int(e.get("skipped", 0))
                if skipped:
                    continue
                cur.execute("SELECT leave_type FROM leave_records WHERE e_id=%s AND leave_date=%s AND status='승인'",
                            (uid, report_date))
                lr = cur.fetchone()
                if lr:
                    leave_conflicts.append({"user_name": e.get("user_name", ""), "leave_type": lr[0]})
            if leave_conflicts:
                conn.close()
                return jsonify(ok=False, leave_conflicts=leave_conflicts,
                               msg="연차 사용 인원이 근무에 배정되어 있습니다.")
        user_name = session.get("user_name", "")
        if edit_id:
            edit_id = int(edit_id)
            cur.execute("SELECT status FROM wr_reports WHERE id=%s", (edit_id,))
            ex = cur.fetchone()
            if ex and ex[0] == '결재':
                conn.close()
                return jsonify(ok=False, msg="이미 최종결재된 보고서는 수정할 수 없습니다.")
            report_id = edit_id
            cur.execute("UPDATE wr_reports SET status='대기', created_by=%s, created_by_name=%s, memo=%s, reject_reason='', updated_at=NOW() WHERE id=%s",
                        (session.get("user_id"), user_name, memo, report_id))
            cur.execute("DELETE FROM wr_entries WHERE report_id=%s", (report_id,))
            cur.execute("DELETE FROM wr_approval_log WHERE report_id=%s", (report_id,))
        else:
            # 새 보고서 - "대기" 상태로 저장 (결재는 별도로 처리)
            cur.execute("""INSERT INTO wr_reports (group_id, report_date, status, created_by, created_by_name, memo)
                           VALUES (%s,%s,%s,%s,%s,%s)""", (group_id, report_date, "대기", session.get("user_id"), user_name, memo))
            report_id = cur.lastrowid
        for e in entries:
            uid = int(e["user_id"])
            uname = str(e.get("user_name", ""))
            cat = str(e.get("category", "주간기본"))
            ded = int(e.get("deduction", 0))
            etc_val = str(e.get("etc_value", ""))
            skipped = int(e.get("skipped", 0))
            cur.execute("""INSERT INTO wr_entries (report_id, user_id, user_name, category, deduction, etc_value, skipped)
                           VALUES (%s,%s,%s,%s,%s,%s,%s)
                           ON DUPLICATE KEY UPDATE deduction=%s, etc_value=%s, skipped=%s""",
                        (report_id, uid, uname, cat, ded, etc_val, skipped, ded, etc_val, skipped))
        conn.commit(); conn.close()
        return jsonify(ok=True, report_id=report_id)
    except Exception as e:
        return jsonify(ok=False, msg=f"저장 오류: {e}"), 500


@app.route("/api/wr_report/<int:report_id>")
@_login_required
def get_wr_report(report_id):
    conn = _conn(); cur = conn.cursor()
    cur.execute("""SELECT r.id, r.group_id, r.report_date, r.status, r.created_by_name,
                          r.created_at, g.group_name, r.reject_reason, IFNULL(r.memo,'')
                   FROM wr_reports r JOIN wr_groups g ON r.group_id = g.id
                   WHERE r.id=%s""", (report_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return jsonify(ok=False, msg="보고서를 찾을 수 없습니다.")
    rd = str(row[2])
    disp = f"{rd[:4]}-{rd[4:6]}-{rd[6:8]}" if len(rd) == 8 else rd
    # 요일 및 공휴일/주말 플래그
    disp_weekday = ""
    is_holiday = False
    is_weekend = False
    if len(rd) == 8:
        dt_obj = datetime.strptime(rd, "%Y%m%d").date()
        wd_idx = dt_obj.weekday()
        disp_weekday = ["월","화","수","목","금","토","일"][wd_idx]
        is_weekend = wd_idx >= 5
        is_holiday = dt_obj in KR_HOLIDAYS
    # 같은 날짜+그룹 회차 라벨
    gid = row[1]
    cur.execute("SELECT id FROM wr_reports WHERE group_id=%s AND report_date=%s ORDER BY id", (gid, rd))
    same_rows = [r2[0] for r2 in cur.fetchall()]
    orig_report_id = None
    if len(same_rows) > 1:
        seq = same_rows.index(report_id) + 1
        disp += f" (수정{seq-1})" if seq > 1 else " (원본)"
        if seq > 1:
            orig_report_id = same_rows[0]
    cur.execute("SELECT final_step FROM wr_groups WHERE id=%s", (gid,))
    fs_row = cur.fetchone()
    final_step = fs_row[0] if fs_row else 1
    report = {"id": row[0], "group_id": gid, "report_date": rd,
              "disp_date": disp,
              "disp_weekday": disp_weekday,
              "is_holiday": is_holiday,
              "is_weekend": is_weekend,
              "status": row[3], "created_by_name": row[4],
              "created_at": str(row[5]), "group_name": row[6],
              "reject_reason": row[7] or "", "memo": row[8] or "",
              "final_step": final_step,
              "orig_report_id": orig_report_id}
    cur.execute("SELECT id, user_id, user_name, category, deduction, etc_value, IFNULL(skipped,0) FROM wr_entries WHERE report_id=%s ORDER BY category, user_name", (report_id,))
    entries = [{"id": r[0], "user_id": r[1], "user_name": r[2], "category": r[3],
                "deduction": r[4], "etc_value": r[5], "skipped": r[6]} for r in cur.fetchall()]
    # 결재이력
    cur.execute("SELECT step, action, user_name, created_at FROM wr_approval_log WHERE report_id=%s ORDER BY id", (report_id,))
    approval_log = [{"step": r[0], "action": r[1], "user_name": r[2], "created_at": str(r[3])} for r in cur.fetchall()]
    conn.close()
    return jsonify(ok=True, report=report, entries=entries, approval_log=approval_log)


@app.route("/api/wr_report/delete", methods=["POST"])
@_login_required
def delete_wr_report():
    try:
        data = request.get_json(force=True)
        report_id = int(data["id"])
        conn = _conn(); cur = conn.cursor()
        cur.execute("SELECT status FROM wr_reports WHERE id=%s", (report_id,))
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify(ok=False, msg="보고서를 찾을 수 없습니다.")
        is_admin = session.get("role") == "admin"
        if row[0] == "결재" and not is_admin:
            conn.close()
            return jsonify(ok=False, msg="최종결재된 보고서는 삭제할 수 없습니다.")
        # 결재된 보고서 삭제 시 schedule_record도 정리
        if row[0] == "결재" and is_admin:
            cur.execute("SELECT group_id, report_date FROM wr_reports WHERE id=%s", (report_id,))
            rinfo = cur.fetchone()
            if rinfo:
                cur.execute("SELECT group_name FROM wr_groups WHERE id=%s", (rinfo[0],))
                gn = cur.fetchone()
                sheet_name = gn[0] if gn else "기타"
                cur.execute("SELECT user_name FROM wr_entries WHERE report_id=%s", (report_id,))
                for (uname,) in cur.fetchall():
                    cur.execute("DELETE FROM schedule_record WHERE emp_name=%s AND work_date=%s AND sheet_name=%s AND source_type IN ('보고서','수정근무보고서')",
                                (uname, rinfo[1], sheet_name))
        cur.execute("DELETE FROM wr_approval_log WHERE report_id=%s", (report_id,))
        cur.execute("DELETE FROM wr_entries WHERE report_id=%s", (report_id,))
        cur.execute("DELETE FROM wr_reports WHERE id=%s", (report_id,))
        conn.commit(); conn.close()
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, msg=str(e))


@app.route("/process_wr_request", methods=["POST"])
@_login_required
def process_wr_request():
    """근무보고서 결재 처리 (1차→2차→3차 또는 반려)"""
    data = request.get_json()
    report_id = int(data["id"])
    new_status = str(data["status"])
    reject_reason = str(data.get("reject_reason", "")).strip()
    if new_status not in ("결재", "반려"):
        return jsonify(ok=False, error="잘못된 상태"), 400
    if new_status == "반려" and not reject_reason:
        return jsonify(ok=False, error="반려 사유를 입력해 주세요"), 400
    conn = _conn(); cur = conn.cursor()
    cur.execute("SELECT group_id, status, report_date FROM wr_reports WHERE id=%s", (report_id,))
    rec = cur.fetchone()
    if not rec:
        conn.close()
        return jsonify(ok=False, error="보고서를 찾을 수 없습니다"), 404
    group_id, cur_status, report_date = rec
    cur.execute("SELECT final_step FROM wr_groups WHERE id=%s", (group_id,))
    grow = cur.fetchone()
    final_step = grow[0] if grow else 1
    login_uid = session.get("user_id")
    login_eid = session.get("e_id")  # tuser.id
    is_admin = session.get("role") == "admin"
    # reviewer 권한 (e_id로 조회)
    cur.execute("SELECT step FROM wr_group_reviewers WHERE group_id=%s AND user_id=%s", (group_id, login_eid))
    user_steps = [r[0] for r in cur.fetchall()]
    # member면 step=1 권한 추가
    cur.execute("SELECT COUNT(*) FROM wr_group_members WHERE group_id=%s AND user_id=%s", (group_id, login_eid))
    if cur.fetchone()[0] > 0 and 1 not in user_steps:
        user_steps.append(1)
    # writer(작성자 전용)도 step=1 권한 추가
    cur.execute("SELECT COUNT(*) FROM wr_group_writers WHERE group_id=%s AND user_id=%s", (group_id, login_eid))
    if cur.fetchone()[0] > 0 and 1 not in user_steps:
        user_steps.append(1)
    if new_status == "반려":
        if not is_admin and not user_steps:
            conn.close()
            return jsonify(ok=False, error="검토자 권한이 없습니다"), 403
        actual_status = "반려"
    else:
        if cur_status == "대기":
            if not is_admin and 1 not in user_steps:
                conn.close()
                return jsonify(ok=False, error="작성자만 처리 가능합니다"), 403
            actual_status = "결재" if final_step == 1 else "작성완료"
        elif cur_status == "작성완료":
            if not is_admin and 2 not in user_steps:
                conn.close()
                return jsonify(ok=False, error="검토자만 처리 가능합니다"), 403
            actual_status = "결재" if final_step == 2 else "검토"
        elif cur_status == "검토":
            if not is_admin and 3 not in user_steps:
                conn.close()
                return jsonify(ok=False, error="2차검토자(최종)만 처리 가능합니다"), 403
            actual_status = "결재"
        else:
            conn.close()
            return jsonify(ok=False, error=f"현재 상태({cur_status})에서 결재할 수 없습니다"), 400
    cur.execute("UPDATE wr_reports SET status=%s WHERE id=%s", (actual_status, report_id))
    if actual_status == "반려" and reject_reason:
        cur.execute("UPDATE wr_reports SET reject_reason=%s WHERE id=%s", (reject_reason, report_id))
    # 결재이력 저장
    approver_name = session.get("user_name", "")
    if actual_status == "작성완료":
        log_step = 1
    elif actual_status == "검토":
        log_step = 2
    elif actual_status == "결재":
        log_step = final_step
    elif actual_status == "반려":
        log_step = max(user_steps) if user_steps else 1
    else:
        log_step = 1
    cur.execute("INSERT INTO wr_approval_log (report_id, step, action, user_name) VALUES (%s,%s,%s,%s)",
                (report_id, log_step, actual_status, approver_name))
    # 최종 결재 시 schedule_record에 자동 반영
    if actual_status == "결재":
        # 수정근무보고서 여부 판정 — '작성 시각' 기준.
        # (예전에는 '이미 결재된 보고서가 있으면 수정본'으로 봤는데, 결재자가 원본보다
        #  수정본을 먼저 결재하면 라벨이 뒤바뀌었다. 2026-07-03 SMT / 07-17 QA 실제 사고)
        # 같은 그룹+날짜에서 나보다 먼저 작성된 보고서가 있으면 내가 수정본이다.
        cur.execute("""SELECT COUNT(*) FROM wr_reports
                       WHERE group_id=%s AND report_date=%s AND id<>%s
                         AND (created_at < (SELECT created_at FROM wr_reports WHERE id=%s)
                              OR (created_at = (SELECT created_at FROM wr_reports WHERE id=%s)
                                  AND id < %s))""",
                    (group_id, report_date, report_id, report_id, report_id, report_id))
        is_revision = cur.fetchone()[0] > 0
        src_type = '수정근무보고서' if is_revision else '보고서'
        # 휴일(토/일/공휴일) 여부 — 휴일이면 조퇴/외출 N을 기본근무에서 차감(8-N)
        try:
            _rd = datetime(int(report_date[:4]), int(report_date[4:6]), int(report_date[6:8]))
            is_hol_date = _rd.weekday() >= 5 or bool(KR_HOLIDAYS.get(_rd.date()))
        except Exception:
            is_hol_date = False
        # 그룹명으로 sheet_name 결정
        cur.execute("SELECT group_name FROM wr_groups WHERE id=%s", (group_id,))
        gn = cur.fetchone()
        sheet_name = gn[0] if gn else "기타"
        cur.execute("SELECT user_id, user_name, category, deduction, etc_value, IFNULL(skipped,0) FROM wr_entries WHERE report_id=%s", (report_id,))
        # 사람별 병합: 한 사람이 여러 카테고리(예: 주간기본 + 기타 조퇴)에 있어도 근무표 한 칸에 합산
        agg = OrderedDict()
        for uid, uname, cat, ded, etc_val, skipped in cur.fetchall():
            a = agg.setdefault(uname, {"uid": uid, "basic": 0, "ot": 0, "night": 0,
                                       "etc": None, "skipped": 0, "leave": None, "deduct": 0.0})
            if skipped:
                a["skipped"] = 1
                continue
            if cat == "주간기본":
                a["basic"] = max(a["basic"], 8)
                if ded and ded > 0:
                    a["etc"] = str(ded)
            elif cat == "주간연장":
                a["basic"] = max(a["basic"], 8); a["ot"] = max(a["ot"], 2)
            elif cat == "야간기본":
                a["basic"] = max(a["basic"], 8); a["night"] = max(a["night"], 6)
            elif cat == "야간연장":
                a["basic"] = max(a["basic"], 8); a["ot"] = max(a["ot"], 2); a["night"] = max(a["night"], 7)
            elif cat == "기타" and etc_val:
                if etc_val in ("연차", "무급"):
                    a["leave"] = etc_val
                elif etc_val.startswith("조퇴"):
                    # 조퇴: N(=17:30-조퇴시각) 누적 → 평일은 기타칸, 휴일은 기본근무 차감
                    _n = _jotoe_hours(etc_val)
                    if _n is not None:
                        a["deduct"] += _n
                    else:
                        a["etc"] = etc_val
                elif etc_val.startswith("외출"):
                    # 외출: N(=복귀-외출) 누적 → 평일은 기타칸, 휴일은 기본근무 차감
                    _n = _outing_hours(etc_val)
                    if _n is not None:
                        a["deduct"] += _n
                    else:
                        a["etc"] = etc_val
                else:
                    a["etc"] = etc_val
        # ── 수정보고서는 그날의 '확정본' ──
        # 수정본에서 제외된 인원 = 근무 취소. 예전에는 수정본에 있는 사람만 갱신해서
        # 취소된 인원의 원본 값이 그대로 남았다(2026-07-19 신민규 등 344건).
        # → 수정본이면 그 그룹·날짜의 기존 수정 레코드를 모두 지우고 새로 쓴다.
        if is_revision:
            cur.execute("""DELETE FROM schedule_record
                           WHERE work_date=%s AND sheet_name=%s
                             AND source_type='수정근무보고서'""",
                        (report_date, sheet_name))
            # 원본에만 있고 수정본에서 빠진 인원 → 근무 없음(0/0/0)으로 확정
            cur.execute("""SELECT DISTINCT emp_name FROM schedule_record
                           WHERE work_date=%s AND sheet_name=%s AND source_type='보고서'""",
                        (report_date, sheet_name))
            for (drop_nm,) in cur.fetchall():
                if drop_nm in agg:
                    continue
                cur.execute("""INSERT INTO schedule_record
                        (emp_name, work_date, basic_h, ot_h, night_h, etc,
                         source_type, sheet_name, uploaded_by)
                    VALUES (%s,%s,0,0,0,NULL,'수정근무보고서',%s,%s)""",
                            (drop_nm, report_date, sheet_name,
                             session.get("user_name", "")))

        for uname, a in agg.items():
            # 원본이면 기존 레코드(원본+수정) 모두 삭제. 수정본은 위에서 이미 정리됨.
            if not is_revision:
                cur.execute("DELETE FROM schedule_record WHERE emp_name=%s AND work_date=%s AND sheet_name=%s AND source_type IN ('보고서','수정근무보고서')",
                            (uname, report_date, sheet_name))
            if a["skipped"]:
                cur.execute("""
                    INSERT INTO schedule_record (emp_name, work_date, basic_h, ot_h, night_h, etc,
                                                 source_type, sheet_name, uploaded_by)
                    VALUES (%s,%s,0,0,0,'스킵',%s,%s,%s)
                """, (uname, report_date, src_type, sheet_name, session.get("user_name", "")))
                continue
            basic, ot, night, etc = a["basic"], a["ot"], a["night"], a["etc"]
            # 조퇴/외출 차감: 휴일이면 기본근무에서 빼고(8-N, 소수 가능), 평일이면 기타칸에 N 기록
            if a["deduct"] > 0:
                if is_hol_date:
                    basic = max(0, basic - a["deduct"])
                else:
                    etc = _fmt_hours(a["deduct"])
            # 연차/무급이면 다른 값 0 처리 (최종 보정, 차감보다 우선)
            if a["leave"]:
                basic, ot, night, etc = 0, 0, 0, a["leave"]
            cur.execute("""
                INSERT INTO schedule_record (emp_name, work_date, basic_h, ot_h, night_h, etc,
                                             source_type, sheet_name, uploaded_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (uname, report_date, basic, ot, night, etc, src_type, sheet_name,
                  session.get("user_name", "")))
            # 연차/무급이면 leave_records + annual_leave 동기화
            if a["leave"]:
                _sync_leave_from_other(cur, a["uid"], report_date, a["leave"])
    # 알림 (작성자에게)
    rd = report_date
    disp_date = f"{rd[:4]}-{rd[4:6]}-{rd[6:8]}" if len(rd) == 8 else rd
    approver_name = session.get("user_name", "")
    cur.execute("SELECT created_by FROM wr_reports WHERE id=%s", (report_id,))
    cb = cur.fetchone()
    if cb:
        noti_uid = cb[0]
        if actual_status == "결재":
            noti_msg = f"✅ {disp_date} 근무보고서가 최종결재 되었습니다 (결재자: {approver_name})"
        elif actual_status in ("작성완료", "검토"):
            noti_msg = f"📋 {disp_date} 근무보고서가 {actual_status} 되었습니다 (처리자: {approver_name})"
        else:
            noti_msg = f"❌ {disp_date} 근무보고서가 반려되었습니다 (처리자: {approver_name})\n📝 사유: {reject_reason}"
        cur.execute("INSERT INTO notifications (user_id, message, link) VALUES (%s,%s,%s)",
                    (noti_uid, noti_msg, "/work_report"))
    # 이메일 알림 수신자 수집
    mail_to = []
    mail_subj = mail_body = ""
    cur.execute("SELECT group_name FROM wr_groups WHERE id=%s", (group_id,))
    gn_row = cur.fetchone()
    gn = gn_row[0] if gn_row else ""
    if actual_status in ("작성완료", "검토"):
        next_step = 2 if actual_status == "작성완료" else 3
        cur.execute("SELECT user_name FROM wr_group_reviewers WHERE group_id=%s AND step=%s", (group_id, next_step))
        rev_names = [r[0] for r in cur.fetchall()]
        if rev_names:
            ph = ",".join(["%s"] * len(rev_names))
            cur.execute(f"SELECT email FROM employee_roster WHERE name IN ({ph}) AND email IS NOT NULL AND email <> ''", rev_names)
            mail_to = [r[0] for r in cur.fetchall()]
            step_label = "검토" if actual_status == "작성완료" else "2차검토"
            mail_subj = f"[근태관리시스템] 근무보고서 {step_label} 요청 ({disp_date})"
            mail_body = (
                f"<p>{gn} <b>{disp_date}</b> 근무보고서의 {step_label}을 요청합니다.</p>"
                f"<p>처리자: <b>{approver_name}</b></p>"
                f"<p>근태관리시스템에 로그인하여 검토해 주세요.</p>"
            )
    elif actual_status in ("결재", "반려"):
        cur.execute("SELECT created_by_name FROM wr_reports WHERE id=%s", (report_id,))
        auth_row = cur.fetchone()
        author_name = auth_row[0] if auth_row else ""
        if author_name:
            cur.execute("SELECT email FROM employee_roster WHERE name=%s AND email IS NOT NULL AND email <> '' LIMIT 1", (author_name,))
            em_row = cur.fetchone()
            if em_row:
                mail_to = [em_row[0]]
        if actual_status == "결재":
            mail_subj = f"[근태관리시스템] 근무보고서 최종결재 완료 ({disp_date})"
            mail_body = (
                f"<p>{gn} <b>{disp_date}</b> 근무보고서가 <b>최종결재</b> 되었습니다.</p>"
                f"<p>결재자: <b>{approver_name}</b></p>"
            )
        else:
            mail_subj = f"[근태관리시스템] 근무보고서 반려 ({disp_date})"
            mail_body = (
                f"<p>{gn} <b>{disp_date}</b> 근무보고서가 <b>반려</b> 되었습니다.</p>"
                f"<p>처리자: <b>{approver_name}</b></p>"
                f"<p>반려 사유: {reject_reason}</p>"
            )
    conn.commit(); conn.close()
    if mail_to:
        _send_email(mail_to, mail_subj, mail_body)
    return jsonify(ok=True)


@app.route("/api/wr_group_members/<int:group_id>")
@_login_required
def get_wr_group_members(group_id):
    conn = _conn(); cur = conn.cursor()
    cur.execute("SELECT user_id, user_name FROM wr_group_members WHERE group_id=%s ORDER BY user_name", (group_id,))
    members = [{"user_id": r[0], "user_name": r[1]} for r in cur.fetchall()]
    conn.close()
    return jsonify(ok=True, members=members)


@app.route("/api/wr_prev_report/<int:group_id>")
@_login_required
def get_wr_prev_report(group_id):
    """지정 날짜 직전의 보고서 데이터 반환 (전일 복사용)"""
    base_date = request.args.get("date", "").replace("-", "")
    if not base_date:
        base_date = datetime.now().strftime("%Y%m%d")
    conn = _conn(); cur = conn.cursor()
    cur.execute("""SELECT id, report_date FROM wr_reports
                   WHERE group_id=%s AND report_date < %s
                   ORDER BY report_date DESC LIMIT 1""", (group_id, base_date))
    row = cur.fetchone()
    if not row:
        conn.close()
        return jsonify(ok=False, msg="이전 보고서가 없습니다.")
    prev_id, prev_date = row
    cur.execute("""SELECT user_id, user_name, category, deduction, etc_value
                   FROM wr_entries WHERE report_id=%s""", (prev_id,))
    entries = [{"user_id": r[0], "user_name": r[1], "category": r[2],
                "deduction": r[3], "etc_value": r[4]} for r in cur.fetchall()]
    disp = f"{prev_date[:4]}-{prev_date[4:6]}-{prev_date[6:8]}" if len(prev_date) == 8 else prev_date
    conn.close()
    return jsonify(ok=True, entries=entries, prev_date=disp)


@app.route("/api/wr_approved_leaves/<int:group_id>")
@_login_required
def get_wr_approved_leaves(group_id):
    """해당 날짜에 승인된 연차/반차가 있는 그룹 멤버 반환"""
    date = request.args.get("date", "").replace("-", "")
    if not date:
        return jsonify(ok=True, leaves=[])
    conn = _conn(); cur = conn.cursor()
    cur.execute("SELECT user_id, user_name FROM wr_group_members WHERE group_id=%s", (group_id,))
    members = {r[0]: r[1] for r in cur.fetchall()}
    if not members:
        conn.close()
        return jsonify(ok=True, leaves=[])
    ph = ','.join(['%s'] * len(members))
    cur.execute(f"""
        SELECT e_id, leave_type FROM leave_records
        WHERE e_id IN ({ph}) AND leave_date=%s AND status='승인'
    """, list(members.keys()) + [date])
    leaves = []
    for e_id, leave_type in cur.fetchall():
        if e_id in members:
            leaves.append({"user_id": e_id, "user_name": members[e_id], "leave_type": leave_type})
    conn.close()
    return jsonify(ok=True, leaves=leaves)


# ═══════ 연차사용계획서 ═══════
@app.route("/api/leave_plan", methods=["POST"])
@_login_required
def create_leave_plan():
    """연차사용계획서 생성"""
    try:
        data = request.get_json()
        e_id = session.get("e_id")
        if not e_id:
            return jsonify(ok=False, error="직원 정보가 없습니다"), 400
        e_id = int(e_id)
        dates = data.get("dates", [])
        if not dates:
            return jsonify(ok=False, error="날짜를 선택해 주세요"), 400
        leave_type = str(data.get("type", "연차"))
        reason = str(data.get("reason", "")).strip()
        conn = _conn(); cur = conn.cursor()
        cur.execute("SELECT name FROM tuser WHERE id=%s", (e_id,))
        row = cur.fetchone()
        e_name = row[0] if row else ""
        # 소속 그룹 찾기 (lp_groups)
        cur.execute("SELECT g.id FROM lp_groups g JOIN lp_group_members m ON g.id=m.group_id WHERE m.user_id=%s LIMIT 1", (e_id,))
        grow = cur.fetchone()
        if not grow:
            conn.close()
            return jsonify(ok=False, error="소속 그룹이 없습니다. 관리자에게 그룹 배정을 요청하세요."), 400
        group_id = grow[0]
        # 계획월 결정 (가장 빠른 날짜 기준)
        sorted_dates = sorted([str(d).replace("-", "") for d in dates])
        plan_month = sorted_dates[0][:6]
        # 수정보고서 여부
        amend_from = data.get("amend_from")
        parent_id = int(amend_from) if amend_from else None
        # 계획서 생성
        cur.execute("""INSERT INTO leave_plans (user_id, user_name, group_id, plan_month, status, reason, parent_id)
                       VALUES (%s,%s,%s,%s,'대기',%s,%s)""", (e_id, e_name, group_id, plan_month, reason, parent_id))
        plan_id = cur.lastrowid
        for d in sorted_dates:
            cur.execute("INSERT INTO leave_plan_dates (plan_id, leave_date, leave_type) VALUES (%s,%s,%s)",
                        (plan_id, d, leave_type))
        conn.commit(); conn.close()
        disp_dates = [f"{d[:4]}-{d[4:6]}-{d[6:8]}" for d in sorted_dates]
        label = "수정보고서" if parent_id else "연차계획서"
        _send_telegram(f"📋 [{label}] {e_name} {leave_type} {len(sorted_dates)}건 (승인대기)\n📅 {', '.join(disp_dates)}\n📝 사유: {reason}", "leave")
        return jsonify(ok=True, plan_id=plan_id, count=len(sorted_dates))
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/api/leave_plan/<int:plan_id>", methods=["DELETE"])
@_login_required
def delete_leave_plan(plan_id):
    """연차사용계획서 삭제 (관리자 전용)"""
    if session.get("role") != "admin":
        return jsonify(ok=False, error="관리자만 삭제할 수 있습니다"), 403
    try:
        conn = _conn(); cur = conn.cursor()
        # 계획서 존재 확인
        cur.execute("SELECT id FROM leave_plans WHERE id=%s", (plan_id,))
        if not cur.fetchone():
            conn.close()
            return jsonify(ok=False, error="계획서를 찾을 수 없습니다"), 404
        # 관련 데이터 삭제
        cur.execute("DELETE FROM leave_plan_approval_log WHERE plan_id=%s", (plan_id,))
        cur.execute("DELETE FROM leave_plan_dates WHERE plan_id=%s", (plan_id,))
        cur.execute("DELETE FROM leave_plans WHERE id=%s", (plan_id,))
        conn.commit()
        conn.close()
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500


@app.route("/api/leave_plan/<int:plan_id>")
@_login_required
def get_leave_plan(plan_id):
    """계획서 상세 조회"""
    conn = _conn(); cur = conn.cursor()
    cur.execute("""SELECT p.id, p.user_id, p.user_name, p.group_id, p.plan_month,
                          p.status, p.reason, p.reject_reason, p.created_at,
                          g.group_name
                   FROM leave_plans p
                   LEFT JOIN lp_groups g ON p.group_id = g.id
                   WHERE p.id=%s""", (plan_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return jsonify(ok=False, error="계획서를 찾을 수 없습니다"), 404
    plan = {"id": row[0], "user_id": row[1], "user_name": row[2], "group_id": row[3],
            "plan_month": row[4], "status": row[5], "reason": row[6],
            "reject_reason": row[7] or "", "created_at": str(row[8]),
            "group_name": row[9] or "", "final_step": 2}
    cur.execute("SELECT leave_date, leave_type FROM leave_plan_dates WHERE plan_id=%s ORDER BY leave_date", (plan_id,))
    plan["dates"] = [{"leave_date": r[0], "leave_type": r[1]} for r in cur.fetchall()]
    # 결재이력
    cur.execute("SELECT step, action, user_name, created_at FROM leave_plan_approval_log WHERE plan_id=%s ORDER BY id", (plan_id,))
    plan["approval_log"] = [{"step": r[0], "action": r[1], "user_name": r[2], "created_at": str(r[3])} for r in cur.fetchall()]
    # 결재라인
    step_reviewers = {}
    cur.execute("SELECT step, user_name FROM lp_group_reviewers WHERE group_id=%s ORDER BY step", (plan["group_id"],))
    for s, n in cur.fetchall():
        step_reviewers.setdefault(s, []).append(n)
    plan["step_reviewers"] = {str(k): v for k, v in step_reviewers.items()}
    conn.close()
    return jsonify(ok=True, plan=plan)


@app.route("/leave_plan_view/<int:plan_id>")
@_login_required
def leave_plan_view(plan_id):
    """계획서 상세보기 페이지 (달력 + 결재표)"""
    conn = _conn(); cur = conn.cursor()
    cur.execute("""SELECT p.id, p.user_id, p.user_name, p.group_id, p.plan_month,
                          p.status, p.reason, p.reject_reason, p.created_at,
                          g.group_name
                   FROM leave_plans p
                   LEFT JOIN lp_groups g ON p.group_id = g.id
                   WHERE p.id=%s""", (plan_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        flash("계획서를 찾을 수 없습니다.", "danger")
        return redirect(url_for("leave_approval"))
    plan = {"id": row[0], "user_id": row[1], "user_name": row[2], "group_id": row[3],
            "plan_month": row[4], "status": row[5], "reason": row[6],
            "reject_reason": row[7] or "", "created_at": str(row[8]),
            "group_name": row[9] or "", "final_step": 2}
    cur.execute("SELECT leave_date, leave_type FROM leave_plan_dates WHERE plan_id=%s ORDER BY leave_date", (plan_id,))
    plan_dates = {r[0]: r[1] for r in cur.fetchall()}
    plan["dates"] = plan_dates
    plan["date_count"] = len(plan_dates)
    total_days = sum(0.5 if v == "반차" else 1 for v in plan_dates.values())
    plan["total_days"] = total_days
    # 결재라인
    step_reviewers = {}
    cur.execute("SELECT step, user_id, user_name FROM lp_group_reviewers WHERE group_id=%s ORDER BY step, user_name", (plan["group_id"],))
    for s, uid, n in cur.fetchall():
        step_reviewers.setdefault(s, []).append({"user_id": uid, "user_name": n})
    # 결재이력
    cur.execute("SELECT step, action, user_name, created_at FROM leave_plan_approval_log WHERE plan_id=%s ORDER BY id", (plan_id,))
    approval_log = [{"step": r[0], "action": r[1], "user_name": r[2], "created_at": str(r[3])} for r in cur.fetchall()]
    done_steps = {}
    for a in approval_log:
        if a["action"] != "반려":
            done_steps[a["step"]] = a
    # 달력 정보
    year_int = int(plan["plan_month"][:4])
    month_int = int(plan["plan_month"][4:6])
    _, dim = calendar.monthrange(year_int, month_int)
    first_wd = calendar.weekday(year_int, month_int, 1)
    day_info = []
    for d in range(1, dim + 1):
        dt = datetime(year_int, month_int, d)
        ds = dt.strftime("%Y%m%d")
        wd = dt.weekday()
        hol_name = KR_HOLIDAYS.get(dt.date(), "")
        lt = plan_dates.get(ds, "")
        day_info.append({
            "day": d, "ds": ds, "wd": wd,
            "is_sat": wd == 5, "is_sun": wd == 6,
            "is_holiday": bool(hol_name), "hol_name": hol_name,
            "leave_type": lt,
        })
    # 로그인 사용자 결재 권한
    login_uid = session.get("user_id")
    login_eid = session.get("e_id")
    is_admin = session.get("role") == "admin"
    cur.execute("SELECT step FROM lp_group_reviewers WHERE group_id=%s AND user_id=%s", (plan["group_id"], login_eid))
    user_steps = [r[0] for r in cur.fetchall()]
    # 연차 현황
    cur.execute("SELECT total, used, m1,m2,m3,m4,m5,m6,m7,m8,m9,m10,m11,m12 FROM annual_leave WHERE e_id=%s AND year=%s", (plan["user_id"], year_int))
    al_row = cur.fetchone()
    if al_row:
        al = {"total": al_row[0], "used": al_row[1]}
        for mi in range(1, 13):
            al[f"m{mi}"] = al_row[mi + 1]
    else:
        al = {"total": 0, "used": 0}
        for mi in range(1, 13):
            al[f"m{mi}"] = 0
    conn.close()
    # 다음 결재 단계 계산 (검토=1, 결재=2)
    status_to_next = {"대기": 1, "1차승인": 2}
    next_step = status_to_next.get(plan["status"], 0)
    can_approve = is_admin or (next_step in user_steps)
    is_owner = (int(login_eid) == plan["user_id"]) if login_eid else False
    step_labels = {1: "검토", 2: "결재"}
    return render_template("leave_plan_view.html",
                           plan=plan, day_info=day_info,
                           first_wd=first_wd, dim=dim,
                           year=year_int, month=month_int,
                           step_reviewers=step_reviewers,
                           done_steps=done_steps,
                           approval_log=approval_log,
                           step_labels=step_labels,
                           user_steps=user_steps,
                           is_admin=is_admin,
                           can_approve=can_approve,
                           is_owner=is_owner,
                           next_step=next_step,
                           al=al)


@app.route("/process_leave_plan", methods=["POST"])
@_login_required
def process_leave_plan():
    """연차계획서 승인/반려"""
    try:
        data = request.get_json()
        plan_id = int(data["id"])
        new_status = str(data["status"])
        reject_reason = str(data.get("reject_reason", "")).strip()
        if new_status not in ("승인", "반려"):
            return jsonify(ok=False, error="잘못된 상태"), 400
        if new_status == "반려" and not reject_reason:
            return jsonify(ok=False, error="반려 사유를 입력해 주세요"), 400
        conn = _conn(); cur = conn.cursor()
        cur.execute("SELECT user_id, user_name, group_id, status, plan_month, parent_id FROM leave_plans WHERE id=%s", (plan_id,))
        rec = cur.fetchone()
        if not rec:
            conn.close()
            return jsonify(ok=False, error="계획서를 찾을 수 없습니다"), 404
        e_id, e_name, group_id, cur_status, plan_month, parent_id = rec
        final_step = 2  # lp_groups: 검토(1)+결재(2) 고정
        login_uid = session.get("user_id")
        login_eid = session.get("e_id")
        is_admin = session.get("role") == "admin"
        cur.execute("SELECT step FROM lp_group_reviewers WHERE group_id=%s AND user_id=%s", (group_id, login_eid))
        user_steps = [r[0] for r in cur.fetchall()]
        approver_name = session.get("user_name", "")
        if new_status == "반려":
            if not is_admin and not user_steps:
                conn.close()
                return jsonify(ok=False, error="결재 권한이 없습니다"), 403
            actual_status = "반려"
            # 반려 시 어떤 단계에서인지 로그
            log_step = 0
            if cur_status == "대기": log_step = 1
            elif cur_status == "1차승인": log_step = 2
            cur.execute("INSERT INTO leave_plan_approval_log (plan_id, step, action, user_id, user_name) VALUES (%s,%s,'반려',%s,%s)",
                        (plan_id, log_step, login_uid, approver_name))
        else:
            if cur_status == "대기":
                if not is_admin and 1 not in user_steps:
                    conn.close()
                    return jsonify(ok=False, error="검토자만 처리 가능합니다"), 403
                actual_status = "1차승인"
                cur.execute("INSERT INTO leave_plan_approval_log (plan_id, step, action, user_id, user_name) VALUES (%s,1,'승인',%s,%s)",
                            (plan_id, login_uid, approver_name))
            elif cur_status == "1차승인":
                if not is_admin and 2 not in user_steps:
                    conn.close()
                    return jsonify(ok=False, error="결재자만 처리 가능합니다"), 403
                actual_status = "승인"
                cur.execute("INSERT INTO leave_plan_approval_log (plan_id, step, action, user_id, user_name) VALUES (%s,2,'승인',%s,%s)",
                            (plan_id, login_uid, approver_name))
            else:
                conn.close()
                return jsonify(ok=False, error=f"현재 상태({cur_status})에서 처리 불가"), 400
        cur.execute("UPDATE leave_plans SET status=%s WHERE id=%s", (actual_status, plan_id))
        if actual_status == "반려":
            cur.execute("UPDATE leave_plans SET reject_reason=%s WHERE id=%s", (reject_reason, plan_id))
        # 최종승인 시 leave_records + annual_leave 반영
        if actual_status == "승인":
            # 수정보고서인 경우: 원본 계획서의 leave_records 삭제
            if parent_id:
                cur.execute("SELECT leave_date FROM leave_plan_dates WHERE plan_id=%s", (parent_id,))
                old_dates = [r[0] for r in cur.fetchall()]
                new_plan_dates_set = set()
                cur.execute("SELECT leave_date FROM leave_plan_dates WHERE plan_id=%s", (plan_id,))
                for r in cur.fetchall():
                    new_plan_dates_set.add(r[0])
                for od in old_dates:
                    if od not in new_plan_dates_set:
                        cur.execute("DELETE FROM leave_records WHERE e_id=%s AND leave_date=%s AND status='승인'", (e_id, od))
                        cur.execute("DELETE FROM work_override WHERE e_id=%s AND e_date=%s AND col_type='other'", (e_id, od))
            cur.execute("SELECT leave_date, leave_type FROM leave_plan_dates WHERE plan_id=%s", (plan_id,))
            plan_dates = cur.fetchall()
            for leave_date, leave_type in plan_dates:
                cur.execute("""INSERT INTO leave_records (e_id, leave_date, leave_type, status, reason)
                               VALUES (%s,%s,%s,'승인',%s)
                               ON DUPLICATE KEY UPDATE status='승인', leave_type=%s""",
                            (e_id, leave_date, leave_type, f"계획서#{plan_id}", leave_type))
                # work_override 연동
                cur.execute("SELECT leave_type FROM leave_records WHERE e_id=%s AND leave_date=%s AND status='승인'", (e_id, leave_date))
                day_leaves = [r[0] for r in cur.fetchall()]
                if day_leaves:
                    other_val = "/".join(day_leaves)
                    cur.execute("""INSERT INTO work_override (e_id, e_date, col_type, value)
                                   VALUES (%s,%s,'other',%s)
                                   ON DUPLICATE KEY UPDATE value=%s, updated_at=NOW()""",
                                (e_id, leave_date, other_val, other_val))
            # annual_leave used 갱신
            year_val = int(plan_month[:4])
            cur.execute("SELECT leave_type, leave_date FROM leave_records WHERE e_id=%s AND leave_date LIKE %s AND status='승인'",
                        (e_id, f"{year_val}%"))
            total_used = 0
            month_used = [0] * 12
            for lt, ld in cur.fetchall():
                v = 0.5 if lt == "반차" else 1
                total_used += v
                m_idx = int(ld[4:6]) - 1
                month_used[m_idx] += v
            cur.execute("""INSERT INTO annual_leave (e_id, e_name, year, used, m1,m2,m3,m4,m5,m6,m7,m8,m9,m10,m11,m12)
                           VALUES (%s,%s,%s,%s, %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                           ON DUPLICATE KEY UPDATE used=%s, e_name=%s,
                               m1=%s,m2=%s,m3=%s,m4=%s,m5=%s,m6=%s,m7=%s,m8=%s,m9=%s,m10=%s,m11=%s,m12=%s,
                               updated_at=NOW()""",
                        (e_id, e_name, year_val, total_used, *month_used,
                         total_used, e_name, *month_used))
            # 근무보고서 자동생성
            for leave_date, leave_type in plan_dates:
                cur.execute("SELECT r.id FROM wr_reports r WHERE r.group_id=%s AND r.report_date=%s LIMIT 1", (group_id, leave_date))
                if not cur.fetchone():
                    cur.execute("""INSERT INTO wr_reports (group_id, report_date, status, created_by, created_by_name, memo)
                                   VALUES (%s,%s,'대기',%s,%s,'계획서승인 자동생성')""",
                                (group_id, leave_date, login_uid, approver_name))
                    rpt_id = cur.lastrowid
                    cur.execute("SELECT m.user_id, t.name FROM wr_group_members m JOIN tuser t ON m.user_id=t.id WHERE m.group_id=%s", (group_id,))
                    for m_uid, m_name in cur.fetchall():
                        cur.execute("SELECT leave_type FROM leave_records WHERE e_id=%s AND leave_date=%s AND status='승인'", (m_uid, leave_date))
                        m_lr = cur.fetchone()
                        if m_lr:
                            m_lc = m_lr[0] if m_lr[0] in ('연차', '무급') else '연차'
                            cur.execute("INSERT INTO wr_entries (report_id, user_id, user_name, category, deduction, etc_value, skipped) VALUES (%s,%s,%s,'기타',0,%s,0)",
                                        (rpt_id, m_uid, m_name, m_lc))
                        else:
                            cur.execute("INSERT INTO wr_entries (report_id, user_id, user_name, category, deduction, etc_value, skipped) VALUES (%s,%s,%s,'주간기본',0,'',0)",
                                        (rpt_id, m_uid, m_name))
        conn.commit(); conn.close()
        pm = plan_month
        disp_month = f"{pm[:4]}-{pm[4:6]}"
        if actual_status == "승인":
            _send_telegram(f"✅ [계획서최종승인] {e_name} {disp_month} 연차계획서 최종승인 (승인자: {approver_name})", "leave")
        elif actual_status == "반려":
            _send_telegram(f"❌ [계획서반려] {e_name} {disp_month} 연차계획서 반려 (처리자: {approver_name})\n📝 사유: {reject_reason}", "leave")
        else:
            _send_telegram(f"📋 [{actual_status}] {e_name} {disp_month} 연차계획서 {actual_status} (승인자: {approver_name})", "leave")
        return jsonify(ok=True, status=actual_status)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


# 디버그용 세션 확인 API
@app.route("/debug/session")
def debug_session():
    """현재 세션 정보 확인 (디버그용)"""
    return jsonify({
        "user_id": session.get("user_id"),
        "login_id": session.get("login_id"),
        "user_name": session.get("user_name"),
        "e_id": session.get("e_id"),
        "role": session.get("role"),
        "permissions": session.get("permissions")
    })


# ══════════════════════════════════════════════════════════════
#  ⚡ 전력관리 (KEPCO EDS Open API)
# ══════════════════════════════════════════════════════════════

from kepco_analyzer import (
    get_power_summary, get_history_data, get_alert_list,
    make_demo_summary, make_demo_history, make_demo_alerts,
)

KEPCO_KEY  = os.environ.get("KEPCO_API_KEY",  "")
KEPCO_CUST = os.environ.get("KEPCO_CUST_NO",  "1316065626")


def _power_demo() -> bool:
    """DB에 실제 데이터가 있으면 demo 아님 (API 키 없어도 스크래퍼 데이터 표시)"""
    if not KEPCO_CUST:
        return True
    try:
        conn = _conn()
        cur  = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM power_usage_raw WHERE cust_no=%s", (KEPCO_CUST,))
        count = cur.fetchone()[0]
        conn.close()
        return count == 0   # 데이터 없을 때만 demo
    except Exception:
        return True


@app.route("/power_dashboard")
@_login_required
def power_dashboard():
    demo = _power_demo()
    summary = {}
    alerts  = []
    if demo:
        summary = make_demo_summary()
        alerts  = make_demo_alerts()["rows"][:5]
    else:
        try:
            conn = _conn(); cur = conn.cursor()
            summary = get_power_summary(cur, KEPCO_CUST)
            alert_d = get_alert_list(cur, KEPCO_CUST, page=1, per_page=5)
            alerts  = alert_d["rows"]
            conn.close()
        except Exception as e:
            flash(f"전력 데이터 조회 오류: {e}", "danger")
    from datetime import datetime as _dt
    return render_template(
        "power_dashboard.html",
        demo=demo, summary=summary, alerts=alerts,
        active_page="power_dashboard",
        now_month=_dt.now().month,
    )


# ══════════════════════════════════════════════════════════
#  통합관제 — 수도(일일검침) · 전력 · 식수 · 통근버스 한 화면
# ══════════════════════════════════════════════════════════
@app.route("/control_center")
@_login_required
def control_center():
    from datetime import date as _date
    return render_template("control_center.html",
                           active_page="control_center",
                           today=_date.today().strftime("%Y-%m-%d"),
                           is_admin=(session.get("role") == "admin"))


@app.route("/api/cc_overview")
@_login_required
def api_cc_overview():
    """통합관제 요약 — 수도/전력/식수/버스를 한 번에 반환"""
    from datetime import date as _date, timedelta as _td, datetime as _dt
    today = _date.today()
    monday = today - _td(days=today.weekday())
    out = {"ok": True, "date": today.strftime("%Y-%m-%d")}

    conn = _conn(); cur = conn.cursor()
    try:
        # ── 수도: 최근 14일 검침 + 일일 사용량(지침 차이) ──
        cur.execute("""SELECT read_date, reading FROM water_meter
                       WHERE read_date >= %s ORDER BY read_date""",
                    (today - _td(days=14),))
        wrows = cur.fetchall()
        water_days, prev = [], None
        for rd, rdg in wrows:
            val = float(rdg or 0)
            water_days.append({
                "date": rd.strftime("%m/%d"),
                "reading": val,
                "usage": round(val - prev, 2) if prev is not None else None,
            })
            prev = val
        cur.execute("SELECT reading FROM water_meter WHERE read_date=%s", (today,))
        r = cur.fetchone()
        today_reading = float(r[0]) if r else None

        # 표에 뿌릴 최근 검침 목록 (일자 · 지침 · 사용량) — 최신순
        cur.execute("""SELECT read_date, reading, memo FROM water_meter
                       ORDER BY read_date DESC LIMIT 31""")
        recs = cur.fetchall()
        prev_map = {}
        cur.execute("""SELECT a.read_date, a.reading - b.reading
                       FROM water_meter a
                       JOIN water_meter b ON b.read_date = a.read_date - INTERVAL 1 DAY
                       ORDER BY a.read_date DESC LIMIT 31""")
        for rd, diff in cur.fetchall():
            prev_map[rd] = float(diff or 0)
        DOWK = ["월","화","수","목","금","토","일"]
        table = [{"date": rd.strftime("%Y-%m-%d"),
                  "md": rd.strftime("%m/%d"),
                  "dow": DOWK[rd.weekday()],
                  "reading": float(rg or 0),
                  "usage": prev_map.get(rd),
                  "memo": mo or ""} for rd, rg, mo in recs]

        out["water"] = {
            "days": water_days[-7:],
            "table": table,
            "today_reading": today_reading,
            "today_usage": (water_days[-1]["usage"] if water_days and
                            water_days[-1]["date"] == today.strftime("%m/%d") else None),
            "has_today": today_reading is not None,
            "last_date": recs[0][0].strftime("%Y-%m-%d") if recs else "",
        }

        # ── 전력: 오늘 누적 kWh + 최근 24시간 추이 ──
        cur.execute("""SELECT COALESCE(SUM(kwh),0) FROM power_usage_raw
                       WHERE cust_no=%s AND DATE(measured_at)=%s""", (KEPCO_CUST, today))
        power_today = float(cur.fetchone()[0] or 0)
        cur.execute("""SELECT COALESCE(SUM(kwh),0) FROM power_usage_raw
                       WHERE cust_no=%s AND DATE(measured_at)=%s""",
                    (KEPCO_CUST, today - _td(days=1)))
        power_yday = float(cur.fetchone()[0] or 0)
        cur.execute("""SELECT measured_at, kwh FROM power_usage_raw
                       WHERE cust_no=%s AND measured_at >= %s
                       ORDER BY measured_at""", (KEPCO_CUST, _dt.now() - _td(hours=24)))
        power_series = [{"t": m.strftime("%H:%M"), "kwh": float(k or 0)} for m, k in cur.fetchall()]

        # 이번 달 누적 · 최대 수요전력
        cur.execute("""SELECT COALESCE(SUM(kwh),0), COALESCE(MAX(kw_demand),0)
                       FROM power_usage_raw
                       WHERE cust_no=%s AND measured_at >= %s""",
                    (KEPCO_CUST, today.replace(day=1)))
        mon_kwh, mon_peak = cur.fetchone()

        # 최근 14일 일별 사용량 (오늘은 진행 중이라 제외)
        cur.execute("""SELECT DATE(measured_at) d, ROUND(SUM(kwh),1)
                       FROM power_usage_raw
                       WHERE cust_no=%s AND measured_at >= %s AND measured_at < %s
                       GROUP BY d ORDER BY d""",
                    (KEPCO_CUST, today - _td(days=14), today))
        power_daily = [{"date": d.strftime("%m/%d"), "kwh": float(k or 0)}
                       for d, k in cur.fetchall()]

        # 전기요금 (월결산자료 일반관리비 '(전기)')
        cur.execute("""SELECT `year`, `month`, amount FROM mgmt_cost_monthly
                       WHERE item='power' ORDER BY `year`, `month`""")
        pw_cost = [(y, m, int(a or 0)) for y, m, a in cur.fetchall()]
        pw_year = [x for x in pw_cost if x[0] == today.year and x[2] > 0]
        bill = pw_year[-1] if pw_year else None

        out["power"] = {"today_kwh": round(power_today, 1),
                        "yday_kwh": round(power_yday, 1),
                        "series": power_series,
                        "month_kwh": round(float(mon_kwh or 0), 1),
                        "month_peak": round(float(mon_peak or 0), 1),
                        "daily": power_daily,
                        "bill_ym": f"{bill[0]}-{bill[1]:02d}" if bill else "",
                        "bill_amount": bill[2] if bill else 0,
                        "bill_year_sum": sum(x[2] for x in pw_year)}

        # ── 식수: 월별 금액(월결산자료) + 월별 발주식수(식수현황) ──
        cur.execute("""SELECT `year`, `month`, amount FROM mgmt_cost_monthly
                       WHERE item='meal' ORDER BY `year`, `month`""")
        cost_map = {(y, m): int(a or 0) for y, m, a in cur.fetchall()}
        cur.execute("""SELECT `year`, `month`, breakfast, lunch, dinner, night, days
                       FROM meal_order_monthly ORDER BY `year`, `month`""")
        qty_map = {(y, m): {"breakfast": int(b or 0), "lunch": int(l or 0),
                            "dinner": int(dn or 0), "night": int(n or 0),
                            "days": int(dy or 0)}
                   for y, m, b, l, dn, n, dy in cur.fetchall()}

        # 올해 1~12월 금액 추이 (미집계 달은 0)
        meal_months = []
        for m in range(1, 13):
            q = qty_map.get((today.year, m), {})
            qty_total = q.get("breakfast", 0) + q.get("lunch", 0) + \
                        q.get("dinner", 0) + q.get("night", 0)
            meal_months.append({
                "ym": f"{today.year}-{m:02d}", "label": f"{m}월",
                "amount": cost_map.get((today.year, m), 0),
                "lunch": q.get("lunch", 0), "night": q.get("night", 0),
                "breakfast": q.get("breakfast", 0), "dinner": q.get("dinner", 0),
                "qty_total": qty_total,
            })
        # 금액이 집계된 가장 최근 달을 대표값으로 (당월은 결산 전이라 0인 경우가 있음)
        priced = [x for x in meal_months if x["amount"] > 0]
        latest = priced[-1] if priced else None
        prev   = priced[-2] if len(priced) > 1 else None

        # 이번 달 일별 식수 현황
        cur.execute("""SELECT order_date, breakfast, lunch, dinner, night
                       FROM meal_order_daily
                       WHERE order_date >= %s AND order_date < %s
                       ORDER BY order_date""",
                    (today.replace(day=1),
                     (today.replace(day=28) + _td(days=4)).replace(day=1)))
        DOWK = ["월","화","수","목","금","토","일"]
        cur_days = []
        for od, b, l, dn, n in cur.fetchall():
            cur_days.append({"day": od.day, "dow": DOWK[od.weekday()],
                             "breakfast": int(b or 0), "lunch": int(l or 0),
                             "dinner": int(dn or 0), "night": int(n or 0),
                             "total": int(b or 0) + int(l or 0) + int(dn or 0) + int(n or 0),
                             "is_today": od == today})
        this_month = {
            "ym": f"{today.year}-{today.month:02d}",
            "days": cur_days,
            "lunch": sum(x["lunch"] for x in cur_days),
            "night": sum(x["night"] for x in cur_days),
            "breakfast": sum(x["breakfast"] for x in cur_days),
            "dinner": sum(x["dinner"] for x in cur_days),
            "total": sum(x["total"] for x in cur_days),
        }

        out["meal"] = {
            "months": meal_months,
            "latest": latest, "prev": prev,
            "this_month": this_month,
            "year_sum": sum(v for (y, m), v in cost_map.items() if y == today.year),
        }

        # ── 통근버스: 등록 인원(호차별) + 실제 탑승 통계(경비실 운행일지) ──
        cur.execute("""SELECT bus_no, COUNT(*) FROM bus_members
                       WHERE active=1 GROUP BY bus_no ORDER BY bus_no""")
        by_bus = [{"bus_no": int(b or 0), "cnt": int(c or 0)} for b, c in cur.fetchall()]

        # 올해 1월~전월 월별 탑승 인원 (자료 있는 달만)
        cur.execute("""SELECT MONTH(ride_date) mo,
                              COUNT(DISTINCT ride_date) days,
                              SUM(on_0830), SUM(off_1730), SUM(off_2000),
                              ROUND(AVG(on_0830),1)
                       FROM bus_ridership
                       WHERE YEAR(ride_date)=%s AND ride_date < %s
                       GROUP BY mo ORDER BY mo""",
                    (today.year, today.replace(day=1)))
        by_month = [{"label": f"{int(mo)}월", "days": int(dd or 0),
                     "on": int(a or 0), "off1": int(b2 or 0), "off2": int(c2 or 0),
                     "avg_on": float(av or 0)}
                    for mo, dd, a, b2, c2, av in cur.fetchall()]

        # 최근 12개월 월별 평균/합계
        # (인자 없는 execute는 %-치환을 하지 않으므로 '%Y-%m'을 그대로 쓴다.
        #  '%%Y-%%m'로 쓰면 MySQL이 리터럴 '%Y-%m'으로 해석해 전 행이 한 그룹이 됨)
        cur.execute("""SELECT DATE_FORMAT(ride_date,'%Y-%m') ym,
                              COUNT(DISTINCT ride_date) days,
                              ROUND(AVG(on_0830),1), ROUND(AVG(off_1730),1),
                              SUM(on_0830)
                       FROM bus_ridership
                       GROUP BY ym ORDER BY ym DESC LIMIT 12""")
        monthly = [{"ym": ym, "days": int(dd or 0), "avg_on": float(ao or 0),
                    "avg_off": float(af or 0), "sum_on": int(so or 0)}
                   for ym, dd, ao, af, so in cur.fetchall()][::-1]

        cur.execute("SELECT MAX(ride_date) FROM bus_ridership")
        last_d = cur.fetchone()[0]
        out["bus"] = {"by_bus": by_bus,
                      "total": sum(b["cnt"] for b in by_bus),
                      "by_month": by_month, "monthly": monthly,
                      "year": today.year,
                      "last_date": last_d.strftime("%Y-%m-%d") if last_d else ""}

        # ── 전산실 온·습도: 현재값 + 오늘 최저/최고 + 최근 3시간 추이 ──
        cur.execute("""SELECT device_id, temperature, humidity, recorded_at
                       FROM mes_env_log ORDER BY id DESC LIMIT 1""")
        er = cur.fetchone()
        env = {"has": False}
        if er:
            cur.execute("""SELECT MIN(temperature), MAX(temperature)
                           FROM mes_env_log WHERE recorded_at >= %s""", (today,))
            mn, mx = cur.fetchone()
            cur.execute("""SELECT recorded_at, temperature FROM mes_env_log
                           WHERE recorded_at >= %s ORDER BY recorded_at""",
                        (_dt.now() - _td(hours=3),))
            series = [{"t": t.strftime("%H:%M"), "temp": float(v or 0)} for t, v in cur.fetchall()]
            temp = float(er[1] or 0)
            # 상태 판정: 28℃ 이상 경고, 32℃ 이상 위험
            state = "danger" if temp >= 32 else ("warn" if temp >= 28 else "ok")
            env = {"has": True, "place": er[0], "temp": temp,
                   "humi": float(er[2] or 0),
                   "at": er[3].strftime("%H:%M") if er[3] else "",
                   "min": float(mn or 0), "max": float(mx or 0),
                   "state": state, "series": series[::max(1, len(series)//40)]}
        out["env"] = env

        # ── 화재센서: 장비별 상태 + 최근 경보 ──
        cur.execute("""SELECT device_name, location, status, battery, last_seen
                       FROM fire_devices ORDER BY id""")
        fdevs = []
        for nm, loc, st, bat, seen in cur.fetchall():
            # last_seen이 2시간 이상 지났으면 통신 두절로 간주
            stale = bool(seen and (_dt.now() - seen).total_seconds() > 7200)
            fdevs.append({"name": nm, "loc": loc, "status": st or "",
                          "battery": int(bat or 0),
                          "seen": seen.strftime("%m/%d %H:%M") if seen else "-",
                          "stale": stale,
                          "state": "danger" if (st or "") not in ("normal", "") else
                                   ("warn" if stale or (bat or 0) < 20 else "ok")})
        cur.execute("""SELECT COUNT(*) FROM fire_alert_logs
                       WHERE severity IN ('warning','critical') AND created_at >= %s""",
                    (today - _td(days=7),))
        out["fire"] = {"devices": fdevs, "warn_7d": int(cur.fetchone()[0] or 0)}

        # ── MES 생산: 라인별 오늘 실적/목표 ──
        cur.execute("""SELECT device_id, total_count, target FROM mes_daily_summary
                       WHERE work_date=%s ORDER BY device_id""", (today,))
        lines = []
        for dev, cnt, tgt in cur.fetchall():
            cnt = int(cnt or 0); tgt = int(tgt or 0)
            lines.append({"line": dev, "count": cnt, "target": tgt,
                          "rate": round(cnt / tgt * 100, 1) if tgt else 0})
        out["mes"] = {"lines": lines}

        # ── CAPS 싱크 상태: 오늘 레벨별 건수 + 마지막 정상 시각 ──
        cur.execute("""SELECT level, COUNT(*) FROM sync_log
                       WHERE log_time >= %s GROUP BY level""", (today,))
        lv = {k: int(v or 0) for k, v in cur.fetchall()}

        # 일시적 ODBC 오류(-1905 / SQLDriverConnect)는 실오류에서 제외.
        # Windows ODBC 임시 DSN 레지스트리 충돌로, 30초 뒤 재시도하면 정상 수집됨.
        # 텔레그램 알림(_sync_log_checker)도 동일 기준으로 무시하므로 화면도 맞춘다.
        cur.execute("""SELECT COUNT(*) FROM sync_log
                       WHERE level='ERROR' AND log_time >= %s
                         AND (message LIKE %s OR message LIKE %s)""",
                    (today, "%SQLDriverConnect%", "%-1905%"))
        transient = int(cur.fetchone()[0] or 0)
        lv["ERROR"] = max(0, lv.get("ERROR", 0) - transient)
        cur.execute("SELECT MAX(log_time) FROM sync_log WHERE level IN ('OK','SKIP')")
        last_ok = cur.fetchone()[0]
        mins = int((_dt.now() - last_ok).total_seconds() // 60) if last_ok else None
        # 30분 넘게 정상 싱크가 없으면 경고, 2시간이면 위험
        sstate = "ok"
        if mins is None or mins >= 120: sstate = "danger"
        elif mins >= 30:                sstate = "warn"
        out["sync"] = {"ok": lv.get("OK", 0), "skip": lv.get("SKIP", 0),
                       "error": lv.get("ERROR", 0), "transient": transient,
                       "last_ok": last_ok.strftime("%m/%d %H:%M") if last_ok else "-",
                       "mins_ago": mins, "state": sstate}
    finally:
        conn.close()
    return jsonify(out)


@app.route("/api/wr_missing_check")
@_login_required
def api_wr_missing_check():
    """근무보고서 미작성일 조회 (관리자).
    기본은 조회만 하고, ?send=1 이면 작성자에게 메일 발송."""
    if session.get("role") != "admin":
        return jsonify({"ok": False, "error": "관리자만 가능합니다."}), 403
    do_send = request.args.get("send") == "1"

    conn = _conn(); cur = conn.cursor()
    try:
        start, end, per_writer, per_group, total = _wr_missing_scan(cur)
        emails = _wr_missing_emails(cur, per_writer) if total else {}
        try:
            cur.execute("""SELECT group_name FROM wr_groups
                           WHERE status='완료' AND alert_missing=0 ORDER BY id""")
            excluded = [r[0] for r in cur.fetchall()]
        except Exception:
            excluded = []
    finally:
        conn.close()

    sent = _wr_missing_send(per_writer, emails, start, end) if (do_send and total) else []
    return jsonify({
        "ok": True,
        "period": {"start": start, "end": end},
        "total": total,
        "excluded_groups": excluded,
        "groups": per_group,
        "writers": [{"name": w, "count": len(items),
                     "email": emails.get(w, ""), "items": items}
                    for w, items in per_writer.items()],
        "sent": sent,
    })


@app.route("/api/schedule_diff")
@_login_required
def api_schedule_diff():
    """근무표 ↔ 근무보고서 대조 결과 조회 (schedule_compare.py가 매일 08시 적재).
    month=YYYY-MM, type=유형, dept=부서, name=이름, resolved=0|1|all"""
    if not (session.get("role") == "admin" or _has_perm("schedule")):
        return jsonify({"ok": False, "error": "권한이 없습니다."}), 403

    month = (request.args.get("month") or datetime.now().strftime("%Y-%m")).strip()
    ym = month.replace("-", "")[:6]
    dtype = (request.args.get("type") or "").strip()
    dept = (request.args.get("dept") or "").strip()
    name = (request.args.get("name") or "").strip()
    resolved = (request.args.get("resolved") or "0").strip()

    where = ["ym=%s"]
    args = [ym]
    if resolved in ("0", "1"):
        where.append("resolved=%s"); args.append(int(resolved))
    if dtype:
        where.append("diff_type=%s"); args.append(dtype)
    if dept:
        where.append("dept=%s"); args.append(dept)
    if name:
        where.append("emp_name LIKE %s"); args.append(f"%{name}%")
    wsql = " AND ".join(where)

    conn = _conn(); cur = conn.cursor()
    try:
        cur.execute("SHOW TABLES LIKE 'schedule_diff'")
        if not cur.fetchone():
            return jsonify({"ok": True, "rows": [], "summary": {}, "depts": [],
                            "msg": "대조 데이터가 아직 없습니다."})
        cur.execute(f"""SELECT work_date, emp_name, dept, diff_type,
                               xl_basic, xl_ot, xl_night,
                               sys_basic, sys_ot, sys_night,
                               source_type, first_seen, resolved
                        FROM schedule_diff WHERE {wsql}
                        ORDER BY work_date, dept, emp_name LIMIT 1000""", args)
        rows = [{
            "date": r[0].strftime("%Y-%m-%d"),
            "name": r[1], "dept": r[2] or "", "type": r[3],
            "xl": f"{float(r[4]):g}/{float(r[5]):g}/{float(r[6]):g}",
            "sys": f"{float(r[7]):g}/{float(r[8]):g}/{float(r[9]):g}",
            "note": r[10] or "",
            "first_seen": r[11].strftime("%m-%d %H:%M") if r[11] else "",
            "resolved": int(r[12]),
        } for r in cur.fetchall()]

        # 유형별 요약 (미해소 기준)
        cur.execute("""SELECT diff_type, COUNT(*) FROM schedule_diff
                       WHERE ym=%s AND resolved=0 GROUP BY diff_type""", (ym,))
        summary = {k: int(v) for k, v in cur.fetchall()}
        cur.execute("""SELECT DISTINCT dept FROM schedule_diff
                       WHERE ym=%s AND dept<>'' ORDER BY dept""", (ym,))
        depts = [r[0] for r in cur.fetchall()]
        cur.execute("""SELECT COUNT(*), MAX(last_seen) FROM schedule_diff WHERE ym=%s""", (ym,))
        tot, last = cur.fetchone()
    finally:
        conn.close()

    return jsonify({"ok": True, "month": month, "rows": rows,
                    "summary": summary, "depts": depts,
                    "total_all": int(tot or 0),
                    "last_run": last.strftime("%Y-%m-%d %H:%M") if last else ""})


@app.route("/source_verify")
@_login_required
def source_verify():
    """3소스 검증 — 근무표(정본) · 근무보고서 · 근태기록(출입)을 한 화면에서 대조"""
    if not (session.get("role") == "admin" or _has_perm("schedule")):
        flash("접근 권한이 없습니다.", "danger")
        return redirect(url_for("dashboard"))
    return render_template("source_verify.html",
                           active_page="source_verify",
                           sel_month=datetime.now().strftime("%Y-%m"))


@app.route("/api/wr_reports_of_day")
@_login_required
def api_wr_reports_of_day():
    """특정 인원·날짜의 근무보고서 목록 (3소스 검증에서 일자 클릭 시 사용).
    해당 인원이 속한 그룹의 그날 보고서를 작성순으로 반환하고,
    각 보고서에서 그 사람이 어느 칸에 배정됐는지도 함께 준다."""
    if not (session.get("role") == "admin" or _has_perm("schedule")):
        return jsonify({"ok": False, "error": "권한이 없습니다."}), 403

    name = (request.args.get("name") or "").strip()
    d = (request.args.get("date") or "").strip().replace("-", "")
    if not name or len(d) != 8:
        return jsonify({"ok": False, "error": "이름과 날짜가 필요합니다."})

    conn = _conn(); cur = conn.cursor()
    try:
        # 해당 인원의 그룹 (근무보고서 그룹 기준, 없으면 schedule_record의 sheet_name)
        cur.execute("""SELECT g.id, g.group_name FROM wr_group_members m
                       JOIN wr_groups g ON m.group_id=g.id
                       WHERE m.user_name=%s""", (name,))
        gs = cur.fetchall()
        if not gs:
            cur.execute("""SELECT DISTINCT sheet_name FROM schedule_record
                           WHERE emp_name=%s AND work_date=%s""", (name, d))
            r = cur.fetchone()
            if r:
                cur.execute("SELECT id, group_name FROM wr_groups WHERE group_name=%s", (r[0],))
                gs = cur.fetchall()
        if not gs:
            return jsonify({"ok": True, "reports": [], "msg": "소속 그룹을 찾을 수 없습니다."})

        gids = [g[0] for g in gs]
        ph = ",".join(["%s"] * len(gids))
        cur.execute(f"""SELECT r.id, r.report_date, r.status, r.created_by_name,
                               r.created_at, g.group_name, IFNULL(r.memo,'')
                        FROM wr_reports r JOIN wr_groups g ON r.group_id=g.id
                        WHERE r.group_id IN ({ph}) AND r.report_date=%s
                        ORDER BY r.created_at, r.id""", (*gids, d))
        reps = cur.fetchall()

        out = []
        for i, (rid, rd, st, by, ca, gname, memo) in enumerate(reps):
            cur.execute("""SELECT category, deduction, etc_value, IFNULL(skipped,0)
                           FROM wr_entries WHERE report_id=%s AND user_name=%s""", (rid, name))
            mine = [{"category": c, "deduction": int(dd or 0),
                     "etc": e or "", "skipped": int(sk)} for c, dd, e, sk in cur.fetchall()]
            # 보고서 전체 배정 (칸별 인원)
            cur.execute("""SELECT category, user_name FROM wr_entries
                           WHERE report_id=%s ORDER BY category, user_name""", (rid,))
            by_cat = {}
            for c, u in cur.fetchall():
                by_cat.setdefault(c, []).append(u)
            out.append({
                "id": rid, "label": "원본" if i == 0 else f"수정{i}",
                "group": gname, "status": st, "writer": by or "",
                "created": ca.strftime("%Y-%m-%d %H:%M") if ca else "",
                "memo": memo, "mine": mine, "by_cat": by_cat,
                "is_final": i == len(reps) - 1,
            })
    finally:
        conn.close()
    return jsonify({"ok": True, "name": name,
                    "date": f"{d[:4]}-{d[4:6]}-{d[6:8]}", "reports": out})


@app.route("/api/source_verify_all")
@_login_required
def api_source_verify_all():
    """3소스 검증 — 전원 요약. 인원별 기록일수·오류건수·일치율을 한 번에 반환."""
    if not (session.get("role") == "admin" or _has_perm("schedule")):
        return jsonify({"ok": False, "error": "권한이 없습니다."}), 403

    month = (request.args.get("month") or datetime.now().strftime("%Y-%m")).strip()
    ym = month.replace("-", "")[:6]
    grp_f = (request.args.get("group") or "").strip()

    conn = _conn(); cur = conn.cursor()
    try:
        # 당일분은 아직 작성 중이라 오류로 보지 않는다 (조회월이 당월일 때만 적용)
        today_str = datetime.now().strftime("%Y%m%d")
        cutoff = today_str if ym == today_str[:6] else ym + "99"

        # 인원별 기록 일수 + 그룹
        cur.execute("""SELECT emp_name, sheet_name, COUNT(DISTINCT work_date)
                       FROM schedule_record WHERE work_date LIKE %s AND work_date < %s
                       GROUP BY emp_name, sheet_name""", (ym + "%", cutoff))
        base = {}
        for nm, sh, cnt in cur.fetchall():
            b = base.setdefault(nm, {"group": sh or "", "days": 0})
            b["days"] += int(cnt or 0)
            if not b["group"]:
                b["group"] = sh or ""

        # 인원별 미해소 오류 (유형별) — 당일 제외
        cur.execute("""SELECT emp_name, diff_type, COUNT(*) FROM schedule_diff
                       WHERE ym=%s AND resolved=0
                         AND DATE_FORMAT(work_date,'%%Y%%m%%d') < %s
                       GROUP BY emp_name, diff_type""", (ym, cutoff))
        errs = defaultdict(dict)
        for nm, dt, c in cur.fetchall():
            errs[nm][dt] = int(c or 0)

        # 부서
        cur.execute("""SELECT name, dept FROM employee_roster
                       WHERE dept IS NOT NULL AND dept<>''""")
        dept_of = dict(cur.fetchall())

        # 출입기록 일수 — 당일 제외
        cur.execute("""SELECT e_name, COUNT(DISTINCT e_date) FROM tenter
                       WHERE e_date LIKE %s AND e_date < %s AND e_id>=0
                       GROUP BY e_name""", (ym + "%", cutoff))
        tag_days = {r[0]: int(r[1] or 0) for r in cur.fetchall()}
    finally:
        conn.close()

    rows = []
    for nm, b in base.items():
        if grp_f and b["group"] != grp_f:
            continue
        e = errs.get(nm, {})
        bad = sum(e.values())
        days = b["days"]
        # 미작성일은 schedule_record에 행이 없어 days에 안 잡히므로 분모에 더해준다.
        # (안 그러면 오류수 > 기록일수가 되어 일치율이 음수로 나옴 — 박승규 사례)
        denom = days + e.get("미작성", 0) + e.get("근무표없음", 0)
        rows.append({
            "name": nm, "dept": dept_of.get(nm, ""), "group": b["group"],
            "days": days, "tag_days": tag_days.get(nm, 0),
            "errors": bad, "by_type": e,
            "checked": denom,
            "rate": round(max(0, denom - bad) / denom * 100, 1) if denom else 0,
        })
    rows.sort(key=lambda x: (-x["errors"], x["group"], x["name"]))
    groups = sorted({r["group"] for r in rows if r["group"]})
    return jsonify({"ok": True, "month": month, "rows": rows, "groups": groups,
                    "total_people": len(rows),
                    "total_errors": sum(r["errors"] for r in rows),
                    "clean": sum(1 for r in rows if r["errors"] == 0)})


@app.route("/api/source_verify")
@_login_required
def api_source_verify():
    """인원별 3소스 일자 대조.
    근무표(정본)는 schedule_diff에 적재된 대조결과에서 역산하지 않고,
    보고서·출입기록은 DB에서 직접 읽는다. 근무표 값은 schedule_diff에 기록된
    차이분만 알 수 있으므로, 차이가 없는 날은 '보고서와 동일'로 간주한다."""
    if not (session.get("role") == "admin" or _has_perm("schedule")):
        return jsonify({"ok": False, "error": "권한이 없습니다."}), 403

    month = (request.args.get("month") or datetime.now().strftime("%Y-%m")).strip()
    ym = month.replace("-", "")[:6]
    name = (request.args.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "이름을 입력하세요."})

    conn = _conn(); cur = conn.cursor()
    try:
        cur.execute("SELECT name, dept, emp_no FROM employee_roster WHERE name=%s", (name,))
        r = cur.fetchone()
        info = {"name": name, "dept": r[1] if r else "", "emp_no": r[2] if r else ""}

        # 근무보고서 (원본/수정 모두)
        cur.execute("""SELECT work_date, basic_h, ot_h, night_h, etc, source_type, sheet_name
                       FROM schedule_record WHERE emp_name=%s AND work_date LIKE %s""",
                    (name, ym + "%"))
        rep = {}
        grp = ""
        for wd, b, o, n, etc, st, sh in cur.fetchall():
            day = int(wd[6:8])
            rep.setdefault(day, {})[st] = {
                "v": f"{float(b or 0):g}/{float(o or 0):g}/{float(n or 0):g}",
                "etc": (etc or "")}
            grp = grp or (sh or "")
        info["group"] = grp

        # 출입기록
        cur.execute("""SELECT e_date, MIN(e_time), MAX(e_time), COUNT(*) FROM tenter
                       WHERE e_name=%s AND e_date LIKE %s AND e_id>=0 GROUP BY e_date""",
                    (name, ym + "%"))
        tag = {int(x[0][6:8]): {"in": x[1], "out": x[2], "cnt": int(x[3])}
               for x in cur.fetchall()}

        # 대조결과(근무표 값 포함)
        cur.execute("""SELECT work_date, diff_type, xl_basic, xl_ot, xl_night,
                              sys_basic, sys_ot, sys_night, source_type, resolved
                       FROM schedule_diff WHERE emp_name=%s AND ym=%s""", (name, ym))
        dif = {}
        for wd, dt, xb, xo, xn, sb, so, sn, note, rs in cur.fetchall():
            dif[wd.day] = {"type": dt,
                           "xl": f"{float(xb):g}/{float(xo):g}/{float(xn):g}",
                           "sys": f"{float(sb):g}/{float(so):g}/{float(sn):g}",
                           "note": note or "", "resolved": int(rs)}
    finally:
        conn.close()

    year, mon = int(ym[:4]), int(ym[4:6])
    dim = calendar.monthrange(year, mon)[1]
    DOWK = ["월", "화", "수", "목", "금", "토", "일"]
    today = date.today()
    days = []
    for d in range(1, dim + 1):
        dt = date(year, mon, d)
        # 당일은 아직 취합 전이라 판정에서 제외 (표시는 하되 '작성중'으로)
        pending = (dt >= today)
        srcs = rep.get(d, {})
        cur_rep = srcs.get("수정근무보고서") or srcs.get("보고서")
        df = dif.get(d)
        # 근무표 값: 차이가 기록돼 있으면 그 값, 아니면 보고서와 동일
        if df:
            xl = df["xl"]
        else:
            xl = cur_rep["v"] if cur_rep else ""
        days.append({
            "day": d, "dow": DOWK[dt.weekday()],
            "holiday": bool(KR_HOLIDAYS.get(dt)) or (mon == 7 and d == 1),
            "weekend": dt.weekday() >= 5,
            "xl": xl,
            "rep": (srcs.get("보고서") or {}).get("v", ""),
            "rep_etc": (srcs.get("보고서") or {}).get("etc", ""),
            "mod": (srcs.get("수정근무보고서") or {}).get("v", ""),
            "mod_etc": (srcs.get("수정근무보고서") or {}).get("etc", ""),
            "tag": tag.get(d),
            "diff": None if pending else df,
            "pending": pending,
        })
    return jsonify({"ok": True, "month": month, "info": info, "days": days})


@app.route("/api/schedule_accuracy")
@_login_required
def api_schedule_accuracy():
    """근무보고서 정확도(근무표 재현율) — 그룹별 + 월별 추이.
    근무표가 정본이므로 이 수치가 '보고서 자동화 전환 준비도'가 된다."""
    if not (session.get("role") == "admin" or _has_perm("schedule")):
        return jsonify({"ok": False, "error": "권한이 없습니다."}), 403
    month = (request.args.get("month") or datetime.now().strftime("%Y-%m")).strip()
    ym = month.replace("-", "")[:6]

    conn = _conn(); cur = conn.cursor()
    try:
        cur.execute("SHOW TABLES LIKE 'schedule_accuracy'")
        if not cur.fetchone():
            return jsonify({"ok": True, "groups": [], "trend": [],
                            "msg": "정확도 데이터가 아직 없습니다."})
        cur.execute("""SELECT grp, writers, cmp_cnt, ok_cnt, mismatch_cnt, missing_cnt,
                              accuracy, calc_day, updated_at
                       FROM schedule_accuracy WHERE ym=%s
                       ORDER BY accuracy DESC, grp""", (ym,))
        groups = []
        for g, w, c, o, mm, ms, a, cd, up in cur.fetchall():
            a = float(a or 0)
            groups.append({
                "group": g, "writers": w or "", "cmp": int(c or 0), "ok": int(o or 0),
                "mismatch": int(mm or 0), "missing": int(ms or 0),
                "accuracy": a, "calc_day": int(cd or 0),
                "level": ("ready" if a >= 99 else "near" if a >= 95
                          else "improve" if a >= 85 else "low"),
                "updated": up.strftime("%m-%d %H:%M") if up else "",
            })
        # 전체
        cur.execute("""SELECT SUM(cmp_cnt), SUM(ok_cnt), SUM(mismatch_cnt), SUM(missing_cnt)
                       FROM schedule_accuracy WHERE ym=%s""", (ym,))
        tc, to, tm, tms = cur.fetchone()
        tc, to = int(tc or 0), int(to or 0)
        total = {"cmp": tc, "ok": to, "mismatch": int(tm or 0), "missing": int(tms or 0),
                 "accuracy": round(to / tc * 100, 1) if tc else 0}
        # 월별 추이 (최근 12개월)
        cur.execute("""SELECT ym, SUM(cmp_cnt), SUM(ok_cnt) FROM schedule_accuracy
                       GROUP BY ym ORDER BY ym DESC LIMIT 12""")
        trend = [{"ym": f"{r[0][:4]}-{r[0][4:6]}",
                  "accuracy": round(int(r[2] or 0) / int(r[1]) * 100, 1) if r[1] else 0}
                 for r in cur.fetchall()][::-1]
    finally:
        conn.close()
    return jsonify({"ok": True, "month": month, "groups": groups,
                    "total": total, "trend": trend})


@app.route("/api/wr_my_missing")
@_login_required
def api_wr_my_missing():
    """로그인 사용자가 작성 담당인 그룹의 미작성일 (근무보고서 화면 경고 배너용).
    관리자는 전체 알림대상 그룹을 본다."""
    me = session.get("user_name", "")
    is_admin = session.get("role") == "admin"
    conn = _conn(); cur = conn.cursor()
    try:
        _s, _e, _pw, per_group, _t = _wr_missing_scan(cur)
    finally:
        conn.close()

    out = []
    for g in per_group:
        if is_admin or me in (g.get("writers") or []):
            out.append({"group": g["group"], "dates": g["dates"]})
    return jsonify({"ok": True,
                    "total": sum(len(g["dates"]) for g in out),
                    "groups": out})


@app.route("/api/cc_water_save", methods=["POST"])
@_login_required
def api_cc_water_save():
    """수도 일일 검침 저장 (같은 날짜는 덮어쓰기)"""
    d = request.get_json(silent=True) or {}
    read_date = (d.get("read_date") or "").strip()
    reading   = d.get("reading")
    memo      = (d.get("memo") or "").strip()[:200]
    if not read_date:
        return jsonify({"ok": False, "msg": "날짜를 입력하세요."})
    try:
        reading = float(reading)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "msg": "검침 지침을 숫자로 입력하세요."})
    if reading < 0:
        return jsonify({"ok": False, "msg": "검침 지침은 0 이상이어야 합니다."})

    conn = _conn(); cur = conn.cursor()
    try:
        # 직전 검침보다 작으면 오입력 가능성 — 경고만 반환하고 저장은 진행
        cur.execute("""SELECT reading FROM water_meter
                       WHERE read_date < %s ORDER BY read_date DESC LIMIT 1""", (read_date,))
        prev = cur.fetchone()
        warn = ""
        if prev and reading < float(prev[0]):
            warn = f"직전 검침({float(prev[0]):,.2f})보다 작습니다. 확인해 주세요."
        cur.execute("""INSERT INTO water_meter (read_date, reading, memo, created_by)
                       VALUES (%s,%s,%s,%s)
                       ON DUPLICATE KEY UPDATE reading=VALUES(reading), memo=VALUES(memo)""",
                    (read_date, reading, memo, session.get("user_name", "")))
        conn.commit()
    finally:
        conn.close()
    return jsonify({"ok": True, "warn": warn})


@app.route("/power_history")
@_login_required
def power_history():
    demo       = _power_demo()
    today      = datetime.now()
    start_date = request.args.get("start", (today - timedelta(days=6)).strftime("%Y-%m-%d"))
    end_date   = request.args.get("end",   today.strftime("%Y-%m-%d"))
    hist       = {}
    if demo:
        hist = make_demo_history(start_date, end_date)
    else:
        try:
            conn = _conn(); cur = conn.cursor()
            hist = get_history_data(cur, KEPCO_CUST, start_date, end_date)
            conn.close()
        except Exception as e:
            flash(f"조회 오류: {e}", "danger")
    return render_template(
        "power_history.html",
        demo=demo, hist=hist,
        start_date=start_date, end_date=end_date,
        active_page="power_history",
    )


@app.route("/power_alerts")
@_login_required
def power_alerts():
    demo = _power_demo()
    page = int(request.args.get("page", 1))
    data = {}
    if demo:
        data = make_demo_alerts()
    else:
        try:
            conn = _conn(); cur = conn.cursor()
            data = get_alert_list(cur, KEPCO_CUST, page=page, per_page=20)
            conn.close()
        except Exception as e:
            flash(f"조회 오류: {e}", "danger")
    return render_template(
        "power_alerts.html",
        demo=demo, data=data,
        active_page="power_alerts",
    )


@app.route("/power_settings", methods=["GET", "POST"])
@_login_required
@_admin_required
def power_settings():
    msg = None
    cfg = {"cust_no": KEPCO_CUST, "contract_kw": 500, "warn_pct": 90.0, "danger_pct": 95.0}
    try:
        conn = _conn(); cur = conn.cursor()
        if request.method == "POST":
            contract_kw = float(request.form.get("contract_kw", 500))
            warn_pct    = float(request.form.get("warn_pct",    90.0))
            danger_pct  = float(request.form.get("danger_pct",  95.0))
            cur.execute("""
                UPDATE power_settings
                SET contract_kw=%s, warn_pct=%s, danger_pct=%s
                WHERE id=1
            """, (contract_kw, warn_pct, danger_pct))
            conn.commit()
            msg = "저장됐습니다."
        cur.execute("""
            SELECT cust_no, contract_kw, warn_pct, danger_pct
            FROM power_settings LIMIT 1
        """)
        row = cur.fetchone()
        if row:
            cfg = {"cust_no": row[0], "contract_kw": float(row[1]),
                   "warn_pct": float(row[2]), "danger_pct": float(row[3])}
        conn.close()
    except Exception as e:
        flash(f"오류: {e}", "danger")
    return render_template(
        "power_settings.html",
        cfg=cfg, msg=msg,
        kepco_key_set=bool(KEPCO_KEY),
        kepco_cust_set=bool(KEPCO_CUST),
        active_page="power_settings",
    )


# ── 전력 데이터 업로드 API (파워플래너 스크래퍼용) ──────────────────
@app.route("/api/power_upload", methods=["POST"])
def api_power_upload():
    """
    파워플래너 스크래퍼가 수집한 데이터를 직접 POST로 수신해 DB에 저장.
    인증: X-Power-Token 헤더 (환경변수 POWER_UPLOAD_TOKEN과 비교)
    Body: {"cust_no":"...", "records":[{"measured_at":"YYYY-MM-DD HH:MM:SS","kwh":1.23,"kw_demand":4.56},...]}
    """
    token = request.headers.get("X-Power-Token", "")
    expected = os.environ.get("POWER_UPLOAD_TOKEN", "")
    if expected and token != expected:
        return jsonify({"ok": False, "error": "인증 실패"}), 401

    data = request.get_json(silent=True) or {}
    cust_no = data.get("cust_no", KEPCO_CUST)
    records = data.get("records", [])
    if not records:
        return jsonify({"ok": False, "error": "records 없음"}), 400

    try:
        conn = _conn()
        cur  = conn.cursor()
        # 테이블 없으면 생성
        cur.execute("""
            CREATE TABLE IF NOT EXISTS power_usage_raw (
                id          BIGINT AUTO_INCREMENT PRIMARY KEY,
                cust_no     VARCHAR(20)   NOT NULL,
                measured_at DATETIME      NOT NULL,
                kwh         DECIMAL(10,3) NOT NULL DEFAULT 0,
                kw_demand   DECIMAL(10,3) NOT NULL DEFAULT 0,
                created_at  DATETIME      NOT NULL DEFAULT NOW(),
                UNIQUE KEY uq_cust_time (cust_no, measured_at),
                INDEX idx_measured (measured_at)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        saved = 0
        for r in records:
            try:
                cur.execute("""
                    INSERT INTO power_usage_raw (cust_no, measured_at, kwh, kw_demand)
                    VALUES (%s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE
                        kwh       = IF(VALUES(kwh) > 0, VALUES(kwh), kwh),
                        kw_demand = IF(VALUES(kw_demand) > 0, VALUES(kw_demand), kw_demand)
                """, (cust_no, r["measured_at"], float(r.get("kwh", 0)), float(r.get("kw_demand", 0))))
                saved += cur.rowcount
            except Exception:
                continue
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "saved": saved, "total": len(records)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── 전력 청구요금 업로드 API (파워플래너 스크래퍼용) ───────────────
@app.route("/api/power_bill_upload", methods=["POST"])
def api_power_bill_upload():
    """
    파워플래너에서 수집한 월별 청구요금을 수신해 power_usage_monthly에 저장.
    Body: {"cust_no":"...", "bills":[{"ym":"YYYYMM","bill_amount":123456,"total_kwh":4567.0},...]}
    """
    token    = request.headers.get("X-Power-Token", "")
    expected = os.environ.get("POWER_UPLOAD_TOKEN", "")
    if expected and token != expected:
        return jsonify({"ok": False, "error": "인증 실패"}), 401

    data    = request.get_json(silent=True) or {}
    cust_no = data.get("cust_no", KEPCO_CUST)
    bills   = data.get("bills", [])
    if not bills:
        return jsonify({"ok": False, "error": "bills 없음"}), 400

    try:
        conn = _conn()
        cur  = conn.cursor()
        # 테이블 없으면 자동 생성
        cur.execute("""
            CREATE TABLE IF NOT EXISTS power_usage_monthly (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                cust_no     VARCHAR(20)   NOT NULL,
                usage_ym    CHAR(6)       NOT NULL,
                total_kwh   DECIMAL(14,3) NOT NULL DEFAULT 0,
                bill_amount INT           NOT NULL DEFAULT 0,
                peak_kw     DECIMAL(10,3) NOT NULL DEFAULT 0,
                updated_at  DATETIME      NOT NULL DEFAULT NOW() ON UPDATE NOW(),
                UNIQUE KEY uq_cust_ym (cust_no, usage_ym)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        saved = 0
        for b in bills:
            try:
                cur.execute("""
                    INSERT INTO power_usage_monthly (cust_no, usage_ym, bill_amount, total_kwh)
                    VALUES (%s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE
                        bill_amount = VALUES(bill_amount),
                        total_kwh   = CASE WHEN VALUES(total_kwh) > 0
                                          THEN VALUES(total_kwh) ELSE total_kwh END,
                        updated_at  = NOW()
                """, (cust_no, str(b.get("ym", "")),
                      int(b.get("bill_amount", 0)),
                      float(b.get("total_kwh", 0))))
                saved += 1
            except Exception:
                continue
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "saved": saved})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── 스마트뷰 실시간 요금 업로드 API ───────────────────────────
@app.route("/api/power_smartview_upload", methods=["POST"])
def api_power_smartview_upload():
    """
    Body: {cust_no, rt_kwh, rt_bill, basic_rate, basic_bill,
           demand_kw, demand_ym, max_demand_kw, max_demand_date,
           bill_start, bill_end, tariff, rate_table:[...]}
    """
    token    = request.headers.get("X-Power-Token", "")
    expected = os.environ.get("POWER_UPLOAD_TOKEN", "")
    if expected and token != expected:
        return jsonify({"ok": False, "error": "인증 실패"}), 401

    d = request.get_json(silent=True) or {}
    cust_no = d.get("cust_no", KEPCO_CUST)
    try:
        conn = _conn()
        cur  = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS power_smartview (
                id              INT AUTO_INCREMENT PRIMARY KEY,
                cust_no         VARCHAR(20)   NOT NULL,
                rt_kwh          DECIMAL(12,1) NOT NULL DEFAULT 0,
                rt_bill         INT           NOT NULL DEFAULT 0,
                tariff          VARCHAR(50)   NOT NULL DEFAULT '',
                basic_rate      INT           NOT NULL DEFAULT 0,
                basic_bill      INT           NOT NULL DEFAULT 0,
                demand_kw       DECIMAL(10,1) NOT NULL DEFAULT 0,
                demand_ym       VARCHAR(6)    NOT NULL DEFAULT '',
                max_demand_kw   DECIMAL(10,1) NOT NULL DEFAULT 0,
                max_demand_date VARCHAR(20)   NOT NULL DEFAULT '',
                bill_start      VARCHAR(12)   NOT NULL DEFAULT '',
                bill_end        VARCHAR(12)   NOT NULL DEFAULT '',
                rate_table_json TEXT,
                updated_at      DATETIME NOT NULL DEFAULT NOW() ON UPDATE NOW(),
                UNIQUE KEY uq_cust (cust_no)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        import json as _j
        cur.execute("""
            INSERT INTO power_smartview
              (cust_no, rt_kwh, rt_bill, tariff, basic_rate, basic_bill,
               demand_kw, demand_ym, max_demand_kw, max_demand_date,
               bill_start, bill_end, rate_table_json)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE
              rt_kwh=VALUES(rt_kwh), rt_bill=VALUES(rt_bill),
              tariff=VALUES(tariff), basic_rate=VALUES(basic_rate),
              basic_bill=VALUES(basic_bill), demand_kw=VALUES(demand_kw),
              demand_ym=VALUES(demand_ym), max_demand_kw=VALUES(max_demand_kw),
              max_demand_date=VALUES(max_demand_date),
              bill_start=VALUES(bill_start), bill_end=VALUES(bill_end),
              rate_table_json=VALUES(rate_table_json),
              updated_at=NOW()
        """, (cust_no,
              float(d.get("rt_kwh", 0)), int(d.get("rt_bill", 0)),
              str(d.get("tariff", "")),
              int(d.get("basic_rate", 0)), int(d.get("basic_bill", 0)),
              float(d.get("demand_kw", 0)), str(d.get("demand_ym", "")),
              float(d.get("max_demand_kw", 0)), str(d.get("max_demand_date", "")),
              str(d.get("bill_start", "")), str(d.get("bill_end", "")),
              _j.dumps(d.get("rate_table", []), ensure_ascii=False)))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── 전력 일별 데이터 업로드 API ────────────────────────────────
@app.route("/api/power_daily_upload", methods=["POST"])
def api_power_daily_upload():
    """
    일별 실측값을 power_usage_daily에 저장.
    Body: {"cust_no":"...", "days":[{"date":"YYYY-MM-DD","total_kwh":1234.56},...]}
    """
    token    = request.headers.get("X-Power-Token", "")
    expected = os.environ.get("POWER_UPLOAD_TOKEN", "")
    if expected and token != expected:
        return jsonify({"ok": False, "error": "인증 실패"}), 401

    data    = request.get_json(silent=True) or {}
    cust_no = data.get("cust_no", KEPCO_CUST)
    days    = data.get("days", [])
    if not days:
        return jsonify({"ok": False, "error": "days 없음"}), 400

    try:
        conn = _conn()
        cur  = conn.cursor()
        saved = 0
        for d in days:
            try:
                cur.execute("""
                    INSERT INTO power_usage_daily (cust_no, usage_date, total_kwh)
                    VALUES (%s, %s, %s)
                    ON DUPLICATE KEY UPDATE
                        total_kwh  = VALUES(total_kwh),
                        updated_at = NOW()
                """, (cust_no, str(d.get("date", "")), float(d.get("total_kwh", 0))))
                saved += 1
            except Exception:
                continue
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "saved": saved})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ══════════════════════════════════════════════════════════════
#  버스배차관리
# ══════════════════════════════════════════════════════════════

# 전기사업부 회사코드 (주말/휴일 주간 배차 제외 대상)
ELEC_DEPT_CODES = {"0002000000000000", "0004000000000000"}  # 전기사무, 전기생산

@app.route("/bus_members")
@_login_required
def bus_members():
    if not _has_perm("bus"):
        flash("권한이 없습니다.", "danger")
        return redirect("/")
    conn = _conn(); cur = conn.cursor()
    cur.execute("SELECT id, user_id, user_name, dept, bus_no, active FROM bus_members ORDER BY bus_no, user_name")
    members = [{"id": r[0], "user_id": r[1], "user_name": r[2], "dept": r[3],
                "bus_no": r[4], "active": r[5]} for r in cur.fetchall()]
    # tuser 전체 목록 (추가용) — employee_roster 조인으로 세부부서 포함
    cur.execute("""SELECT t.id, t.name, t.company, IFNULL(er.dept,'') as sub_dept
                   FROM tuser t LEFT JOIN employee_roster er ON t.name = er.name
                   WHERE t.name IS NOT NULL AND t.name <> '' ORDER BY t.name""")
    all_users = []
    for r in cur.fetchall():
        main_dept = DEPT_MAP.get((r[2] or "").strip(), "")
        sub_dept = (r[3] or "").strip()
        dept_display = sub_dept if sub_dept else main_dept
        all_users.append({"id": r[0], "name": r[1], "dept": dept_display, "main_dept": main_dept, "sub_dept": sub_dept})
    # 세부부서 목록 (필터용)
    cur.execute("SELECT DISTINCT dept FROM employee_roster WHERE dept IS NOT NULL AND dept <> '' ORDER BY dept")
    dept_list = [r[0] for r in cur.fetchall()]
    conn.close()
    return render_template("bus_members.html", active_page="bus_members",
                           members=members, all_users=all_users, dept_list=dept_list)


@app.route("/api/bus_member_save", methods=["POST"])
@_login_required
def api_bus_member_save():
    if not _has_perm("bus"):
        return jsonify({"ok": False, "error": "권한 없음"}), 403
    d = request.get_json(force=True)
    user_name = (d.get("user_name") or "").strip()
    bus_no = int(d.get("bus_no", 1))
    dept = (d.get("dept") or "").strip()
    user_id = d.get("user_id")
    if not user_name:
        return jsonify({"ok": False, "error": "이름 필수"})
    conn = _conn(); cur = conn.cursor()
    try:
        cur.execute("""INSERT INTO bus_members (user_id, user_name, dept, bus_no)
                       VALUES (%s,%s,%s,%s)
                       ON DUPLICATE KEY UPDATE bus_no=%s, dept=%s, active=1""",
                    (user_id, user_name, dept, bus_no, bus_no, dept))
        conn.commit()
    finally:
        conn.close()
    return jsonify({"ok": True})


@app.route("/api/bus_member_delete", methods=["POST"])
@_login_required
def api_bus_member_delete():
    if not _has_perm("bus"):
        return jsonify({"ok": False, "error": "권한 없음"}), 403
    d = request.get_json(force=True)
    mid = d.get("id")
    conn = _conn(); cur = conn.cursor()
    cur.execute("DELETE FROM bus_members WHERE id=%s", (mid,))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/bus_member_upload", methods=["POST"])
@_login_required
def api_bus_member_upload():
    """엑셀 업로드 — B열=이름, C열=호차(1/2)"""
    if not _has_perm("bus"):
        return jsonify({"ok": False, "error": "권한 없음"}), 403
    f = request.files.get("file")
    if not f:
        return jsonify({"ok": False, "error": "파일 없음"})
    import openpyxl, io
    wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
    ws = wb.active
    if len(wb.sheetnames) > 1:
        for sn in wb.sheetnames:
            if "첨부" in sn or "명단" in sn:
                ws = wb[sn]; break
    conn = _conn(); cur = conn.cursor()
    # name→세부부서 매핑 (employee_roster 우선, tuser fallback)
    cur.execute("""SELECT t.name, t.company, IFNULL(er.dept,'')
                   FROM tuser t LEFT JOIN employee_roster er ON t.name = er.name
                   WHERE t.name IS NOT NULL AND t.name <> ''""")
    name_dept = {}
    for nm, comp, sub in cur.fetchall():
        sub = (sub or "").strip()
        name_dept[nm.strip()] = sub if sub else DEPT_MAP.get((comp or "").strip(), "")
    added = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        # B열(idx1)=이름, C열(idx2)=호차 패턴
        name_val = None; bus_val = 1
        for i, cell in enumerate(row):
            if cell is None:
                continue
            cv = str(cell).strip()
            if cv in ("이름", "버스", "호차", ""):
                continue
            if cv in ("1", "2") or (isinstance(cell, (int, float)) and int(cell) in (1, 2)):
                bus_val = int(cell)
            elif cv and cv != "제외" and len(cv) <= 10 and not cv.startswith(" "):
                if name_val is None:
                    name_val = cv
        if not name_val:
            continue
        dept = name_dept.get(name_val, "")
        # tuser id 찾기
        cur.execute("SELECT id FROM tuser WHERE name=%s LIMIT 1", (name_val,))
        trow = cur.fetchone()
        uid = trow[0] if trow else None
        cur.execute("""INSERT INTO bus_members (user_id, user_name, dept, bus_no)
                       VALUES (%s,%s,%s,%s)
                       ON DUPLICATE KEY UPDATE bus_no=%s, dept=%s, active=1""",
                    (uid, name_val, dept, bus_val, bus_val, dept))
        added += 1
    conn.commit(); conn.close()
    return jsonify({"ok": True, "added": added})


@app.route("/bus_dispatch")
@_login_required
def bus_dispatch():
    if not _has_perm("bus"):
        flash("권한이 없습니다.", "danger")
        return redirect("/")
    return render_template("bus_dispatch.html", active_page="bus_dispatch",
                           is_admin=(session.get("role") == "admin"))


@app.route("/api/bus_dispatch_query")
@_login_required
def api_bus_dispatch_query():
    """배차 조회 API — date=YYYY-MM-DD, shift=주간|야간|전체"""
    if not _has_perm("bus"):
        return jsonify({"ok": False, "error": "권한 없음"}), 403
    sel_date = request.args.get("date", "")
    shift = request.args.get("shift", "주간")
    if not sel_date:
        return jsonify({"ok": False, "error": "날짜 필수"})

    dt = datetime.strptime(sel_date, "%Y-%m-%d")
    date_str = dt.strftime("%Y%m%d")
    is_weekend = dt.weekday() >= 5
    hol_name = KR_HOLIDAYS.get(dt.date(), "")
    is_holiday = is_weekend or bool(hol_name)
    display_date = f"{dt.month}월{dt.day}일"

    conn = _conn(); cur = conn.cursor()

    # 휴무자 조회 (공통)
    cur.execute("""
        SELECT DISTINCT e.user_name
        FROM wr_entries e JOIN wr_reports r ON e.report_id = r.id
        WHERE r.report_date = %s AND e.category = '기타'
    """, (date_str,))
    off_workers = set(r[0] for r in cur.fetchall())

    # 전기사업부 명단 (공통)
    elec_names = set()
    if is_holiday:
        cur.execute("SELECT name FROM tuser WHERE company IN (%s,%s) AND name IS NOT NULL",
                    tuple(ELEC_DEPT_CODES))
        elec_names = set(r[0] for r in cur.fetchall())

    # bus_members 마스터 (공통)
    cur.execute("SELECT user_name, bus_no, dept FROM bus_members WHERE active=1")
    bus_map = {}
    for nm, bno, dept in cur.fetchall():
        bus_map[nm] = {"bus_no": bno, "dept": dept}
    all_members = [{"name": nm, "bus_no": d["bus_no"], "dept": d["dept"]}
                   for nm, d in sorted(bus_map.items())]

    month_str = dt.strftime("%Y-%m")
    day_num = dt.day

    def _query_shift(s):
        """단일 시프트 배차 조회"""
        cats = ("주간기본", "주간연장") if s == "주간" else ("야간기본", "야간연장")
        cur.execute("""
            SELECT DISTINCT e.user_name
            FROM wr_entries e JOIN wr_reports r ON e.report_id = r.id
            WHERE r.report_date = %s AND e.category IN (%s, %s)
              AND IFNULL(e.skipped, 0) = 0
              AND (e.etc_value IS NULL OR e.etc_value NOT LIKE '%%연차%%')
              AND (e.etc_value IS NULL OR e.etc_value NOT LIKE '%%반차%%')
        """, (date_str, cats[0], cats[1]))
        workers = set(r[0] for r in cur.fetchall()) - off_workers
        s_elec_excluded = []
        if is_holiday and s == "주간":
            s_elec_excluded = sorted(elec_names & workers & set(bus_map.keys()))
            workers -= elec_names
        # 식수 연동
        mt = "석식" if s == "주간" else "조식"
        cur.execute("""SELECT `dept`, `count` FROM `meal_count`
                       WHERE `year_month`=%s AND `day`=%s AND `meal_type`=%s""",
                    (month_str, day_num, mt))
        mi = {}; ho = False; skip_depts = set()
        for dept, cnt in cur.fetchall():
            if str(cnt).strip() == '스킵':
                skip_depts.add(dept)
                mi[dept] = '스킵'
                continue
            try: c = int(cnt)
            except (ValueError, TypeError): c = 0
            if c > 0: ho = True
            mi[dept] = c
        # 스킵 부서 직원은 버스 제외
        b1 = []; b2 = []; un = []
        for nm in sorted(workers):
            if nm in bus_map:
                if bus_map[nm]["dept"] in skip_depts:
                    continue
                (b1 if bus_map[nm]["bus_no"] == 1 else b2).append(nm)
            else:
                un.append(nm)
        # 담당자 확인 플래그: 당일 조회 + 15:30 이후 + 식수 미입력(스킵/인원 모두 없음)
        now = datetime.now()
        needs_check = (
            dt.date() == now.date()
            and (now.hour * 60 + now.minute) >= (15 * 60 + 30)
            and len(mi) == 0
        )
        return {"shift": s, "bus1": b1, "bus2": b2, "unassigned": un,
                "has_overtime": ho, "meal_info": mi, "meal_type": mt,
                "elec_excluded": s_elec_excluded, "skip_depts": sorted(skip_depts),
                "needs_check": needs_check}

    # 금요일/휴일전일 판단
    is_pre_holiday = False
    extra_dispatches = []
    tomorrow = dt + timedelta(days=1)
    tom_is_holiday = (tomorrow.weekday() >= 5) or bool(KR_HOLIDAYS.get(tomorrow.date(), ""))
    if dt.weekday() == 4 or (tom_is_holiday and not is_holiday):
        is_pre_holiday = True
        check_dt = tomorrow
        for _ in range(3):
            c_hol = (check_dt.weekday() >= 5) or bool(KR_HOLIDAYS.get(check_dt.date(), ""))
            if not c_hol:
                break
            c_date_str = check_dt.strftime("%Y%m%d")
            for s in ["주간", "야간"]:
                ecats = ("주간기본", "주간연장") if s == "주간" else ("야간기본", "야간연장")
                cur.execute("""
                    SELECT DISTINCT e.user_name
                    FROM wr_entries e JOIN wr_reports r ON e.report_id=r.id
                    WHERE r.report_date=%s AND e.category IN (%s,%s) AND IFNULL(e.skipped,0)=0
                """, (c_date_str, ecats[0], ecats[1]))
                e_workers = set(r[0] for r in cur.fetchall())
                if s == "주간":
                    e_workers -= elec_names
                eb1 = []; eb2 = []; eun = []
                for nm in sorted(e_workers):
                    if nm in bus_map:
                        (eb1 if bus_map[nm]["bus_no"] == 1 else eb2).append(nm)
                    else:
                        eun.append(nm)
                extra_dispatches.append({
                    "date": check_dt.strftime("%m월%d일"),
                    "date_full": check_dt.strftime("%Y-%m-%d"),
                    "shift": s, "bus1": eb1, "bus2": eb2, "unassigned": eun
                })
            check_dt += timedelta(days=1)

    # 전체 조회: 주간+야간 둘 다
    if shift == "전체":
        shifts = [_query_shift("주간"), _query_shift("야간")]
        conn.close()
        return jsonify({
            "ok": True, "date": sel_date, "display_date": display_date, "shift": "전체",
            "is_holiday": is_holiday, "hol_name": hol_name,
            "shifts": shifts,
            "is_pre_holiday": is_pre_holiday, "extra_dispatches": extra_dispatches,
            "all_members": all_members
        })

    # 단일 시프트
    result = _query_shift(shift)
    conn.close()
    return jsonify({
        "ok": True, "date": sel_date, "display_date": display_date, "shift": shift,
        "is_holiday": is_holiday, "hol_name": hol_name,
        "bus1": result["bus1"], "bus2": result["bus2"], "unassigned": result["unassigned"],
        "has_overtime": result["has_overtime"], "meal_info": result["meal_info"],
        "meal_type": result["meal_type"],
        "is_pre_holiday": is_pre_holiday, "extra_dispatches": extra_dispatches,
        "elec_excluded": result["elec_excluded"],
        "skip_depts": result["skip_depts"],
        "needs_check": result["needs_check"],
        "all_members": all_members
    })


@app.route("/api/bus_dispatch_save", methods=["POST"])
@_login_required
def api_bus_dispatch_save():
    """배차 문자 내용 저장 (이력)"""
    if not _has_perm("bus"):
        return jsonify({"ok": False, "error": "권한 없음"}), 403
    d = request.get_json(force=True)
    conn = _conn(); cur = conn.cursor()
    cur.execute("""INSERT INTO bus_dispatch_log (dispatch_date, shift_type, content, bus1_names, bus2_names, created_by)
                   VALUES (%s,%s,%s,%s,%s,%s)""",
                (d.get("dispatch_date", ""), d.get("shift_type", ""),
                 d.get("content", ""), d.get("bus1_names", ""), d.get("bus2_names", ""),
                 session.get("user_name", "")))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/bus_sms_recipients")
@_login_required
def api_bus_sms_recipients():
    """총원명부 전화번호 목록 (배차 문자 수신자 선택용)"""
    if not _has_perm("bus"):
        return jsonify({"ok": False, "error": "권한 없음"}), 403
    conn = _conn(); cur = conn.cursor()
    cur.execute("""SELECT name, dept, phone FROM employee_roster
                   WHERE phone IS NOT NULL AND phone <> ''
                   ORDER BY dept, name""")
    recipients = [{"name": r[0], "dept": r[1], "phone": r[2]} for r in cur.fetchall()]
    conn.close()
    return jsonify({"ok": True, "recipients": recipients})


@app.route("/api/bus_sms_send", methods=["POST"])
@_login_required
def api_bus_sms_send():
    """솔라피 SMS 전송"""
    if not _has_perm("bus"):
        return jsonify({"ok": False, "error": "권한 없음"}), 403
    if not SOLAPI_KEY or not SOLAPI_SECRET or not SOLAPI_SENDER:
        return jsonify({"ok": False, "error": "솔라피 API 설정이 없습니다. .env를 확인하세요."})

    d = request.get_json(force=True)
    phones = d.get("receivers", [])
    msg   = (d.get("msg", "") or "").strip()
    if not phones:
        return jsonify({"ok": False, "error": "수신번호가 없습니다."})
    if not msg:
        return jsonify({"ok": False, "error": "메시지 내용이 없습니다."})

    # 전화번호 정제 (숫자만)
    cleaned = []
    for p in phones:
        p = re.sub(r"[^0-9]", "", str(p))
        if p and p not in cleaned:
            cleaned.append(p)
    if not cleaned:
        return jsonify({"ok": False, "error": "유효한 수신번호가 없습니다."})

    # 솔라피 HMAC 인증 헤더 생성
    dt = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.000Z")
    salt = os.urandom(16).hex()
    sig = _hmac.new(SOLAPI_SECRET.encode("utf-8"),
                    (dt + salt).encode("utf-8"), hashlib.md5).hexdigest()
    auth_header = f"HMAC-MD5 apiKey={SOLAPI_KEY}, date={dt}, salt={salt}, signature={sig}"

    # 솔라피 API 호출 — 단일: /send, 다중: /send-many
    from urllib.error import HTTPError
    if len(cleaned) == 1:
        body = json.dumps({
            "message": {"to": cleaned[0], "from": SOLAPI_SENDER, "text": msg}
        }).encode("utf-8")
        url = "https://api.solapi.com/messages/v4/send"
    else:
        messages = [{"to": p, "from": SOLAPI_SENDER, "text": msg} for p in cleaned]
        body = json.dumps({"messages": messages}).encode("utf-8")
        url = "https://api.solapi.com/messages/v4/send-many"
    try:
        req = Request(
            url,
            data=body,
            headers={"Content-Type": "application/json", "Authorization": auth_header},
            method="POST"
        )
        with urlopen(req, timeout=15) as resp:
            result = json.loads(resp.read().decode("utf-8"))
    except HTTPError as e:
        err_body = e.read().decode("utf-8")
        try:
            err_json = json.loads(err_body)
            return jsonify({"ok": False, "error": err_json.get("errorMessage", err_body)})
        except Exception:
            return jsonify({"ok": False, "error": f"HTTP {e.code}: {err_body}"})
    except Exception as e:
        return jsonify({"ok": False, "error": f"API 호출 오류: {e}"})

    # 솔라피 성공 응답: groupId 포함
    if "groupId" in result:
        sent = result.get("count", {}).get("total", len(cleaned))
        return jsonify({"ok": True, "sent": sent, "msg_id": result.get("groupId", "")})
    else:
        return jsonify({"ok": False, "error": result.get("errorMessage", "전송 실패"), "detail": result})


@app.route("/bus_sms_log")
@_login_required
def bus_sms_log():
    if not _has_perm("bus"):
        flash("권한이 없습니다.", "danger")
        return redirect("/")
    conn = _conn(); cur = conn.cursor()
    cur.execute("""SELECT id, dispatch_date, shift_type, content, bus1_names, bus2_names,
                          created_by, created_at
                   FROM bus_dispatch_log ORDER BY created_at DESC LIMIT 200""")
    logs = [{"id": r[0], "dispatch_date": r[1], "shift_type": r[2], "content": r[3],
             "bus1_names": r[4], "bus2_names": r[5], "created_by": r[6],
             "created_at": r[7].strftime("%Y-%m-%d %H:%M") if r[7] else ""} for r in cur.fetchall()]
    conn.close()
    return render_template("bus_sms_log.html", active_page="bus_sms_log", logs=logs)


# ── MES Blueprint 등록 ────────────────────────────────────
from mes_bp import mes_bp, init_mes_db
init_mes_db(app)
app.register_blueprint(mes_bp)

# ── 교육/훈련 Blueprint 등록 ──────────────────────────────
from edu_bp import edu_bp, init_edu_db
init_edu_db(app)
app.register_blueprint(edu_bp)

# ── 위험물·안전관리 Blueprint 등록 ────────────────────────
from hazmat_bp import hazmat_bp, init_hazmat_db, start_hazmat_monitor
init_hazmat_db(app)
app.register_blueprint(hazmat_bp)
start_hazmat_monitor()

# ── AI 안전진단 Blueprint 등록 ────────────────────────────
from safety_ai_bp import safety_ai_bp, init_safety_ai_db
init_safety_ai_db(app)
app.register_blueprint(safety_ai_bp)

# ── 경비일지 Blueprint 등록 ───────────────────────────────
from guard_bp import guard_bp, init_guard_db
init_guard_db(app)
app.register_blueprint(guard_bp)


# ── 페이스메이커 대시보드 ─────────────────────────────────
@app.route("/pacemaker")
@_login_required
def pacemaker():
    return render_template("pacemaker_dashboard.html", active_page="pacemaker")


# ── ESG 경영 (관리자 전용, 현재 준비중 플레이스홀더) ────────
def _require_admin():
    if "user_id" not in session:
        return redirect(url_for("login"))
    if session.get("role") != "admin":
        flash("관리자만 접근 가능합니다.", "danger")
        return redirect(url_for("dashboard"))
    return None


# ── ESG 환경 대시보드 ─────────────────────────────────────
ENV_CATEGORIES = [
    {"key": "waste_general",    "name": "일반 폐기물",     "icon": "🗑",  "unit": "kg"},
    {"key": "waste_recycle",    "name": "재활용 폐기물",   "icon": "♻️",  "unit": "kg"},
    {"key": "waste_designated", "name": "지정 폐기물",     "icon": "☣️",  "unit": "kg"},
    {"key": "water",            "name": "상수도",          "icon": "💧",  "unit": "㎥"},
    {"key": "gas_lng",          "name": "도시가스 (LNG)",  "icon": "🔥",  "unit": "㎥"},
    {"key": "gas_lpg",          "name": "LPG",            "icon": "🔥",  "unit": "kg"},
    {"key": "etc",              "name": "기타",            "icon": "📌",  "unit": "kg"},
]


@app.route("/esg/environment")
def esg_environment():
    guard = _require_admin()
    if guard: return guard

    from datetime import datetime as _dt, date as _date
    today = _date.today()
    ym = today.strftime("%Y%m")

    # 이번 달 전력 사용량 (power_usage_raw에서 집계)
    power_this_month = 0.0
    try:
        conn = _conn(); cur = conn.cursor()
        cur.execute("""
            SELECT COALESCE(SUM(CAST(usage_kwh AS DECIMAL(14,3))), 0)
            FROM power_usage_raw
            WHERE cust_no=%s AND DATE_FORMAT(usage_date,'%%Y%%m')=%s
        """, (KEPCO_CUST, ym))
        row = cur.fetchone()
        power_this_month = float(row[0] or 0)
        conn.close()
    except Exception:
        pass

    # 이번 달 환경 기록 합계 (카테고리별)
    env_totals = {c["key"]: 0.0 for c in ENV_CATEGORIES}
    try:
        conn = _conn(); cur = conn.cursor()
        cur.execute("""
            SELECT category, COALESCE(SUM(amount), 0)
            FROM esg_env_records
            WHERE DATE_FORMAT(record_date,'%%Y%%m')=%s
            GROUP BY category
        """, (ym,))
        for cat, total in cur.fetchall():
            env_totals[cat] = float(total or 0)
        conn.close()
    except Exception:
        pass

    # 최근 30건 입력 목록
    records = []
    try:
        conn = _conn(); cur = conn.cursor()
        cur.execute("""
            SELECT r.id, r.record_date, r.category, r.amount, r.unit, r.memo, r.created_at,
                   COALESCE(u.name, '')
            FROM esg_env_records r
            LEFT JOIN app_users u ON u.id = r.created_by
            ORDER BY r.record_date DESC, r.id DESC LIMIT 30
        """)
        cat_map = {c["key"]: c for c in ENV_CATEGORIES}
        for rid, rdate, cat, amt, unit, memo, ts, uname in cur.fetchall():
            cat_info = cat_map.get(cat, {"name": cat, "icon": "📌"})
            records.append({
                "id": rid,
                "date": rdate.strftime("%Y-%m-%d"),
                "category_key": cat,
                "category_name": cat_info["name"],
                "category_icon": cat_info["icon"],
                "amount": float(amt or 0),
                "unit": unit or "",
                "memo": memo or "",
                "created_at": ts.strftime("%m-%d %H:%M"),
                "created_by": uname,
            })
        conn.close()
    except Exception as e:
        flash(f"조회 오류: {e}", "danger")

    # 폐기물 합계 (3종 합산)
    waste_total = (env_totals["waste_general"]
                   + env_totals["waste_recycle"]
                   + env_totals["waste_designated"])

    return render_template(
        "env_dashboard.html",
        active_page="esg_environment",
        categories=ENV_CATEGORIES,
        env_totals=env_totals,
        waste_total=waste_total,
        power_this_month=power_this_month,
        records=records,
        today=today.isoformat(),
        ym=ym,
    )


@app.route("/esg/api/env_record", methods=["POST"])
def esg_api_add_env_record():
    guard = _require_admin()
    if guard: return guard
    try:
        data = request.get_json(force=True, silent=True) or {}
        record_date = str(data.get("record_date", "")).strip()
        category    = str(data.get("category", "")).strip()
        try:    amount = float(data.get("amount", 0) or 0)
        except (ValueError, TypeError):
            return jsonify(ok=False, error="invalid amount"), 400
        unit = str(data.get("unit", "")).strip()
        memo = str(data.get("memo", "")).strip()[:200]

        if not record_date or not category:
            return jsonify(ok=False, error="date/category required"), 400
        valid_cats = {c["key"] for c in ENV_CATEGORIES}
        if category not in valid_cats:
            return jsonify(ok=False, error="invalid category"), 400
        if amount < 0:
            return jsonify(ok=False, error="amount must be >= 0"), 400

        uid = session.get("user_id")
        conn = _conn(); cur = conn.cursor()
        cur.execute("""
            INSERT INTO esg_env_records (record_date, category, amount, unit, memo, created_by)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (record_date, category, amount, unit, memo, uid))
        new_id = cur.lastrowid
        conn.commit()
        conn.close()
        return jsonify(ok=True, id=new_id)
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500


@app.route("/esg/api/env_record/<int:rid>", methods=["DELETE"])
def esg_api_del_env_record(rid):
    guard = _require_admin()
    if guard: return guard
    try:
        conn = _conn(); cur = conn.cursor()
        cur.execute("DELETE FROM esg_env_records WHERE id=%s", (rid,))
        deleted = cur.rowcount
        conn.commit()
        conn.close()
        if deleted == 0:
            return jsonify(ok=False, error="not found"), 404
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500


# ── 전산관리 (IT 물품 요청) ─────────────────────────────────
IT_CATEGORIES = [
    {"key": "hardware",   "name": "하드웨어",    "icon": "💻"},
    {"key": "software",   "name": "소프트웨어",  "icon": "💿"},
    {"key": "network",    "name": "네트워크",    "icon": "🌐"},
    {"key": "peripheral", "name": "주변기기",    "icon": "🖱"},
    {"key": "repair",     "name": "수리/점검",   "icon": "🔧"},
    {"key": "other",      "name": "기타",        "icon": "📦"},
]
IT_STATUSES = [
    {"key": "pending",    "name": "대기",    "color": "secondary"},
    {"key": "accepted",   "name": "접수",    "color": "info"},
    {"key": "processing", "name": "처리중",  "color": "warning"},
    {"key": "completed",  "name": "완료",    "color": "success"},
    {"key": "rejected",   "name": "반려",    "color": "danger"},
]
IT_URGENCIES = [
    {"key": "low",    "name": "낮음",    "color": "#94a3b8"},
    {"key": "normal", "name": "보통",    "color": "#3b82f6"},
    {"key": "high",   "name": "높음",    "color": "#f59e0b"},
    {"key": "urgent", "name": "긴급",    "color": "#ef4444"},
]


def _it_enrich_requests(rows):
    """requests 목록을 보기 좋게 가공"""
    cat_map = {c["key"]: c for c in IT_CATEGORIES}
    st_map  = {s["key"]: s for s in IT_STATUSES}
    ug_map  = {u["key"]: u for u in IT_URGENCIES}
    result = []
    for r in rows:
        result.append({
            "id":           r[0],
            "requester_id": r[1],
            "requester":    r[2] or "",
            "category_key": r[3],
            "category":     cat_map.get(r[3], {}).get("name", r[3]),
            "category_icon":cat_map.get(r[3], {}).get("icon", "📦"),
            "location":     r[4] or "",
            "item_name":    r[5],
            "quantity":     r[6],
            "reason":       r[7] or "",
            "status_key":   r[8],
            "status":       st_map.get(r[8], {}).get("name", r[8]),
            "status_color": st_map.get(r[8], {}).get("color", "secondary"),
            "admin_memo":   r[9] or "",
            "processor":    r[10] or "",
            "processed_at": r[11].strftime("%Y-%m-%d %H:%M") if r[11] else "",
            "created_at":   r[12].strftime("%Y-%m-%d %H:%M"),
        })
    return result


@app.route("/it/request")
@_login_required
def it_request_page():
    """사용자: 요청하기 + 내 요청 이력"""
    uid = session.get("user_id")
    my_requests = []
    try:
        conn = _conn(); cur = conn.cursor()
        cur.execute("""
            SELECT r.id, r.requester_id, u.name,
                   r.category, r.location, r.item_name, r.quantity, r.reason,
                   r.status, r.admin_memo,
                   p.name, r.processed_at, r.created_at
            FROM it_requests r
            LEFT JOIN app_users u ON u.id = r.requester_id
            LEFT JOIN app_users p ON p.id = r.processed_by
            WHERE r.requester_id = %s
            ORDER BY r.created_at DESC
            LIMIT 100
        """, (uid,))
        my_requests = _it_enrich_requests(cur.fetchall())
        conn.close()
    except Exception as e:
        flash(f"조회 오류: {e}", "danger")

    return render_template(
        "it_request.html",
        active_page="it_request",
        categories=IT_CATEGORIES,
        urgencies=IT_URGENCIES,
        statuses=IT_STATUSES,
        my_requests=my_requests,
    )


@app.route("/it/manage")
def it_manage_page():
    """관리자: 전체 요청 관리"""
    guard = _require_admin()
    if guard: return guard

    status_filter = request.args.get("status", "all")
    try:
        conn = _conn(); cur = conn.cursor()
        where  = ""; params = []
        if status_filter != "all":
            where = "WHERE r.status = %s"
            params.append(status_filter)
        cur.execute(f"""
            SELECT r.id, r.requester_id, u.name,
                   r.category, r.location, r.item_name, r.quantity, r.reason,
                   r.status, r.admin_memo,
                   p.name, r.processed_at, r.created_at
            FROM it_requests r
            LEFT JOIN app_users u ON u.id = r.requester_id
            LEFT JOIN app_users p ON p.id = r.processed_by
            {where}
            ORDER BY
              CASE r.status
                WHEN 'pending'    THEN 1
                WHEN 'accepted'   THEN 2
                WHEN 'processing' THEN 3
                WHEN 'completed'  THEN 4
                WHEN 'rejected'   THEN 5
              END,
              r.created_at DESC
            LIMIT 500
        """, params)
        all_requests = _it_enrich_requests(cur.fetchall())

        # 상태별 카운트
        cur.execute("SELECT status, COUNT(*) FROM it_requests GROUP BY status")
        status_counts = {s: 0 for s in ["pending","accepted","processing","completed","rejected"]}
        for s, c in cur.fetchall():
            status_counts[s] = c
        conn.close()
    except Exception as e:
        all_requests = []
        status_counts = {}
        flash(f"조회 오류: {e}", "danger")

    return render_template(
        "it_manage.html",
        active_page="it_manage",
        categories=IT_CATEGORIES,
        urgencies=IT_URGENCIES,
        statuses=IT_STATUSES,
        requests=all_requests,
        status_counts=status_counts,
        status_filter=status_filter,
    )


@app.route("/it/api/request", methods=["POST"])
@_login_required
def it_api_submit_request():
    """신규 요청 제출"""
    try:
        data = request.get_json(force=True, silent=True) or {}
        category = str(data.get("category", "")).strip()
        location = str(data.get("location", "")).strip()[:100]
        item     = str(data.get("item_name", "")).strip()[:150]
        try:    qty = int(data.get("quantity", 1) or 1)
        except (ValueError, TypeError): qty = 1
        reason   = str(data.get("reason", "")).strip()[:1000]

        valid_cats = {c["key"] for c in IT_CATEGORIES}
        if category not in valid_cats:
            return jsonify(ok=False, error="invalid category"), 400
        if not item:
            return jsonify(ok=False, error="item required"), 400
        if qty < 1: qty = 1

        uid = session.get("user_id")
        conn = _conn(); cur = conn.cursor()
        cur.execute("""
            INSERT INTO it_requests (requester_id, category, location, item_name, quantity, reason)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (uid, category, location, item, qty, reason))
        new_id = cur.lastrowid
        conn.commit()
        conn.close()
        return jsonify(ok=True, id=new_id)
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500


@app.route("/it/api/request/<int:rid>/status", methods=["POST"])
def it_api_update_status(rid):
    """관리자: 요청 상태 변경 / 메모 추가"""
    guard = _require_admin()
    if guard: return guard
    try:
        data = request.get_json(force=True, silent=True) or {}
        status = str(data.get("status", "")).strip()
        memo   = str(data.get("admin_memo", "")).strip()

        valid = {s["key"] for s in IT_STATUSES}
        if status and status not in valid:
            return jsonify(ok=False, error="invalid status"), 400

        uid = session.get("user_id")
        updates, params = [], []
        if status:
            updates.append("status=%s"); params.append(status)
            # 완료/반려 시 처리자 기록
            if status in ("completed", "rejected", "accepted", "processing"):
                updates.append("processed_by=%s"); params.append(uid)
                updates.append("processed_at=NOW()")
        if "admin_memo" in data:
            updates.append("admin_memo=%s"); params.append(memo)

        if not updates:
            return jsonify(ok=False, error="nothing to update"), 400

        params.append(rid)
        conn = _conn(); cur = conn.cursor()
        cur.execute(f"UPDATE it_requests SET {', '.join(updates)} WHERE id=%s", params)
        updated = cur.rowcount
        conn.commit()
        conn.close()
        if updated == 0:
            return jsonify(ok=False, error="not found"), 404
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500


@app.route("/it/api/request/<int:rid>", methods=["DELETE"])
@_login_required
def it_api_delete_request(rid):
    """요청 삭제: 본인(pending 상태만) 또는 관리자"""
    try:
        uid = session.get("user_id")
        is_admin = session.get("role") == "admin"

        conn = _conn(); cur = conn.cursor()
        cur.execute("SELECT requester_id, status FROM it_requests WHERE id=%s", (rid,))
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify(ok=False, error="not found"), 404
        req_id, status = row

        # 관리자가 아니면 본인 + pending 상태만 삭제 가능
        if not is_admin:
            if req_id != uid:
                conn.close()
                return jsonify(ok=False, error="permission denied"), 403
            if status != "pending":
                conn.close()
                return jsonify(ok=False, error="only pending requests can be cancelled"), 400

        cur.execute("DELETE FROM it_requests WHERE id=%s", (rid,))
        conn.commit()
        conn.close()
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500


@app.route("/esg/social")
def esg_social():
    guard = _require_admin()
    if guard: return guard
    return render_template(
        "esg_placeholder.html",
        active_page="esg_social",
        page_icon="🤝",
        page_title="사회 (Social)",
        page_sub="안전보건·복리후생·교육훈련·사회공헌 등 사회 책임 관리",
        scope_items=[
            "안전보건 지표 (재해율·위험성평가)",
            "임직원 복리후생 및 만족도",
            "교육훈련 시간 및 이수율",
            "다양성·포용 지표 (성별·연령 등)",
            "협력사 상생·사회공헌 활동",
        ],
    )


@app.route("/esg/governance")
def esg_governance():
    guard = _require_admin()
    if guard: return guard
    return render_template(
        "esg_placeholder.html",
        active_page="esg_governance",
        page_icon="🏛",
        page_title="지배구조 (Governance)",
        page_sub="윤리경영·컴플라이언스·리스크 관리 등 거버넌스 관리",
        scope_items=[
            "윤리경영 교육 및 준법 감시",
            "컴플라이언스 점검 및 이슈 관리",
            "리스크 평가 및 내부통제",
            "이사회 운영 및 의사결정 공시",
            "ESG 정보 공개(TCFD·GRI) 및 보고서 발간",
        ],
    )


# ═══ TBM Blueprint 등록 (통합) ═══
try:
    from tbm_bp import tbm_bp as _tbm_bp
    app.register_blueprint(_tbm_bp)
except Exception as _e:
    print(f"[TBM Blueprint] 등록 실패: {_e}")


# ═══ 디지털 사이니지 Blueprint 등록 ═══
try:
    from signage_bp import signage_bp as _signage_bp
    app.register_blueprint(_signage_bp)
except Exception as _e:
    print(f"[Signage Blueprint] 등록 실패: {_e}")


if __name__ == "__main__":
    start_tuya_poller()
    t = threading.Thread(target=_sync_log_checker, daemon=True)
    t.start()
    t2 = threading.Thread(target=_telegram_bot_polling, daemon=True)
    t2.start()
    t3 = threading.Thread(target=_docker_sync_checker, daemon=True)
    t3.start()
    t4 = threading.Thread(target=_wr_overdue_checker, daemon=True)
    t4.start()
    t5 = threading.Thread(target=_wr_missing_checker, daemon=True)
    t5.start()
    app.config["TEMPLATES_AUTO_RELOAD"] = True
    app.run(debug=False, host="0.0.0.0", port=5050)
