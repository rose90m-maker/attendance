#!/usr/bin/env python3
"""
식수현황 엑셀 → attendance.meal_order_monthly 취합 (월별 발주식수)

원본: smb://192.168.100.3/taein_hq/이민영/식수현황/식수현황_YYYY_M.xlsx
시트 '식수관리'
  A열 = 일(숫자, '주간누계' 행은 문자라 자동 제외), B=요일
  E=조식, F=중식, G=석식, H=야식  (발주식수)

금액은 월결산자료의 일반관리비 '(식당)' 행에 있으며 mgmt_cost_import.py가 담당한다.

사용:
  python3 meal_order_import.py              # 폴더 전체
  python3 meal_order_import.py 2026_6 2026_7
"""
import os
import re
import sys
import unicodedata

import pymysql
from dotenv import load_dotenv

load_dotenv()

SRC_DIR = os.environ.get("MEAL_DIR", "/Volumes/taein_hq/이민영/식수현황")
SHEET = "식수관리"
DATA_START_ROW = 5
# 아워홈식수(19~22열) = 실제 배식된 '실적'. 발주식수(5~8열)는 고정 계약값이라 안 씀.
# 컬럼 배치는 2025~2026 전 파일에서 동일(아워홈식수@19, 조/중/석/야). 2026-07 재확인.
COLS = {"breakfast": 19, "lunch": 20, "dinner": 21, "night": 22}   # 아워홈 조/중/석/야


def _nfc(s):
    return unicodedata.normalize("NFC", s or "")


def _conn():
    return pymysql.connect(
        host=os.environ.get("DB_HOST", "192.168.100.11"),
        port=int(os.environ.get("DB_PORT", 3307)),
        user=os.environ.get("DB_USER", "root"),
        password=os.environ["DB_PASSWORD"],
        database="attendance",
        charset="utf8mb4",
    )


def list_files():
    try:
        names = os.listdir(SRC_DIR)
    except OSError:
        return []
    out = []
    for n in names:
        nn = _nfc(n)
        if nn.startswith("~$") or not nn.lower().endswith(".xlsx"):
            continue
        if nn.startswith("식수현황_"):
            out.append(os.path.join(SRC_DIR, n))
    return sorted(out, key=lambda p: _nfc(os.path.basename(p)))


def parse(path):
    """→ (월합계튜플, [일별튜플...]) 또는 (None, [])
       월합계: (year, month, breakfast, lunch, dinner, night, days)
       일별  : (date, breakfast, lunch, dinner, night)"""
    import openpyxl
    from datetime import datetime as _dt
    base = _nfc(os.path.basename(path))
    m = re.match(r"식수현황_(\d{4})_(\d{1,2})", base)
    if not m:
        return None, []
    year, month = int(m.group(1)), int(m.group(2))

    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"  ⚠️  '{SHEET}' 시트 없음 — 건너뜀: {base}")
        return None, []
    ws = wb[SHEET]

    tot = {k: 0 for k in COLS}
    days = 0
    daily = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        day = ws.cell(r, 1).value
        # '주간누계' 등 소계 행은 문자열이라 자동 제외
        if not isinstance(day, (int, float)) or not (1 <= day <= 31):
            continue
        days += 1
        vals = {}
        for k, c in COLS.items():
            v = ws.cell(r, c).value
            vals[k] = int(v) if isinstance(v, (int, float)) else 0
            tot[k] += vals[k]
        try:
            d = _dt(year, month, int(day)).date()
        except ValueError:
            continue
        daily.append((d, vals["breakfast"], vals["lunch"],
                      vals["dinner"], vals["night"]))
    if not days:
        return None, []
    return (year, month, tot["breakfast"], tot["lunch"],
            tot["dinner"], tot["night"], days), daily


def main():
    wanted = set(sys.argv[1:])          # 예: 2026_7
    files = list_files()
    if wanted:
        files = [f for f in files
                 if any(w in _nfc(os.path.basename(f)) for w in wanted)]
    if not files:
        print(f"❌ 대상 엑셀이 없습니다: {SRC_DIR}")
        return 1

    rows, daily_rows = [], []
    for f in files:
        rec, daily = parse(f)
        if rec:
            rows.append(rec)
            daily_rows.extend(daily)
            print(f"  📄 {_nfc(os.path.basename(f))} — "
                  f"{rec[6]}일 · 중식 {rec[3]:,} · 야식 {rec[5]:,}")
    if not rows:
        print("❌ 취합할 데이터가 없습니다.")
        return 1

    conn = _conn(); cur = conn.cursor()
    cur.executemany("""
        INSERT INTO meal_order_monthly
            (`year`, `month`, breakfast, lunch, dinner, night, days)
        VALUES (%s,%s,%s,%s,%s,%s,%s)
        ON DUPLICATE KEY UPDATE
            breakfast=VALUES(breakfast), lunch=VALUES(lunch),
            dinner=VALUES(dinner), night=VALUES(night), days=VALUES(days)
    """, rows)
    if daily_rows:
        cur.executemany("""
            INSERT INTO meal_order_daily
                (order_date, breakfast, lunch, dinner, night)
            VALUES (%s,%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE
                breakfast=VALUES(breakfast), lunch=VALUES(lunch),
                dinner=VALUES(dinner), night=VALUES(night)
        """, daily_rows)
    conn.commit()
    cur.execute("SELECT COUNT(*) FROM meal_order_monthly")
    n_m = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM meal_order_daily")
    n_d = cur.fetchone()[0]
    print(f"\n✅ 저장 완료 — 월별 {len(rows)}개월(누적 {n_m}) · "
          f"일별 {len(daily_rows)}건(누적 {n_d})")
    conn.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
