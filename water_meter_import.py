#!/usr/bin/env python3
"""
상수도 계량기 지침 일일관리대장(엑셀) → attendance.water_meter 취합

원본 엑셀 '일일입력' 시트 구조
  A열=날짜, B=요일, C=계량기 지침값(㎥), D=일일 사용량,
  E=판정 기준 사용량, F=평균 대비 증감량, G=증감률, H=상태, I=특이사항, J=입력자
  데이터는 5행부터

사용:
  python3 water_meter_import.py <엑셀경로>
  python3 water_meter_import.py            # 기본 경로 사용
"""
import os
import sys
from datetime import datetime

import pymysql
from dotenv import load_dotenv

load_dotenv()

DEFAULT_XLSX = os.environ.get(
    "WATER_XLSX",
    "/Users/changkooji/Downloads/상수도_계량기지침_일일관리대장_위험신호개선_최종.xlsx",
)
SHEET = "일일입력"
DATA_START_ROW = 5
COL_DATE, COL_READING, COL_STATE, COL_NOTE, COL_WRITER = 1, 3, 8, 9, 10


def _conn():
    return pymysql.connect(
        host=os.environ.get("DB_HOST", "192.168.100.11"),
        port=int(os.environ.get("DB_PORT", 3307)),
        user=os.environ.get("DB_USER", "root"),
        password=os.environ["DB_PASSWORD"],
        database="attendance",
        charset="utf8mb4",
    )


def parse(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"❌ '{SHEET}' 시트가 없습니다. (시트: {wb.sheetnames})")
        return []
    ws = wb[SHEET]
    rows = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        d = ws.cell(r, COL_DATE).value
        rdg = ws.cell(r, COL_READING).value
        if d is None or rdg in (None, ""):
            continue
        if isinstance(d, datetime):
            d = d.date()
        elif not hasattr(d, "year"):
            continue
        try:
            rdg = float(str(rdg).replace(",", ""))
        except ValueError:
            continue

        # 메모: 엑셀의 상태/특이사항을 합쳐 보관 (일일 사용량은 화면에서 지침 차이로 계산)
        parts = []
        for col in (COL_STATE, COL_NOTE):
            v = ws.cell(r, col).value
            s = str(v).strip() if v is not None else ""
            if s and s not in ("0", "None"):
                parts.append(s)
        memo = " / ".join(parts)[:200]

        writer = ws.cell(r, COL_WRITER).value
        writer = str(writer).strip()[:30] if writer else "경비실"
        rows.append((d, rdg, memo, writer))
    return rows


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    if not os.path.exists(path):
        print(f"❌ 파일을 찾을 수 없습니다: {path}")
        return 1

    rows = parse(path)
    if not rows:
        print("❌ 취합할 검침 데이터가 없습니다.")
        return 1
    print(f"  📄 {os.path.basename(path)} — {len(rows)}건 "
          f"({rows[0][0]} ~ {rows[-1][0]})")

    conn = _conn(); cur = conn.cursor()
    cur.executemany("""
        INSERT INTO water_meter (read_date, reading, memo, created_by)
        VALUES (%s,%s,%s,%s)
        ON DUPLICATE KEY UPDATE
            reading=VALUES(reading), memo=VALUES(memo)
    """, rows)
    conn.commit()
    cur.execute("SELECT COUNT(*), MIN(read_date), MAX(read_date) FROM water_meter")
    n, mn, mx = cur.fetchone()
    conn.close()
    print(f"\n✅ 저장 완료 — 이번 처리 {len(rows)}건 / 누적 {n}건 ({mn} ~ {mx})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
