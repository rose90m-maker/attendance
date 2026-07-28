#!/usr/bin/env python3
"""
월결산자료(일반관리비) → attendance.mgmt_cost_monthly 취합

원본: smb://192.168.100.3/taein_hq/이민영/YYYY 결산자료/YYYY.MM 결산자료.xlsx|xlsm
시트 '일반관리비'
  A열 = 항목명 ( (식당) / (상하수도) / (전기) / (폐기물) … )
  5행 = 월 헤더 (C5='1월' … N5='12월'), B열 = 연간 누계
  → 항목 × 월 금액(단위: 천원)을 뽑아 저장

가장 최근 파일 하나에 그 해의 1월~당월 금액이 모두 들어 있으므로
연도별로 최신 파일만 읽으면 된다.

사용:
  python3 mgmt_cost_import.py            # 기본 경로(2026 결산자료)에서 최신 파일
  python3 mgmt_cost_import.py <엑셀경로>
"""
import os
import re
import sys
import unicodedata

import pymysql
from dotenv import load_dotenv

load_dotenv()

SRC_DIR = os.environ.get("SETTLE_DIR", "/Volumes/taein_hq/이민영/2026 결산자료")
SHEET = "일반관리비"
MONTH_HEADER_ROW = 5
FIRST_MONTH_COL = 3      # C열 = 1월

# 저장할 항목: 엑셀 라벨(공백 제거) → 저장 키
ITEMS = {
    "(식당)":     "meal",
    "(상하수도)": "water",
    "(전기)":     "power",
    "(폐기물)":   "waste",
    "(업무용주유)": "fuel",
    "(기타경비)": "etc",
}


def _nfc(s):
    return unicodedata.normalize("NFC", s or "")


def _conn():
    return pymysql.connect(
        host=os.environ.get("DB_HOST", "192.168.100.11"),
        port=int(os.environ.get("DB_PORT", 3307)),
        user=os.environ.get("DB_USER", "root"),
        password=os.environ["DB_PASSWORD"],
        database="attendance",
        charset="utf8mb4",
    )


def latest_file(dirpath):
    """폴더에서 가장 최근 월의 결산자료 (YYYY.MM 결산자료.xlsx|xlsm)"""
    try:
        names = os.listdir(dirpath)
    except OSError:
        return None
    cands = []
    for n in names:
        nn = _nfc(n)
        if nn.startswith("~$") or not nn.lower().endswith((".xlsx", ".xlsm")):
            continue
        m = re.match(r"(\d{4})\.(\d{1,2})", nn)
        if m:
            cands.append(((int(m.group(1)), int(m.group(2))), os.path.join(dirpath, n)))
    if not cands:
        return None
    return max(cands)[1]


def parse(path):
    """→ [(year, month, item_key, amount_krw), ...]  금액 단위: 원"""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"❌ '{SHEET}' 시트 없음 (시트: {wb.sheetnames})")
        return []
    ws = wb[SHEET]

    m = re.match(r"(\d{4})", _nfc(os.path.basename(path)))
    year = int(m.group(1)) if m else None
    if not year:
        print("❌ 파일명에서 연도를 읽지 못했습니다.")
        return []

    # 월 헤더 → 열 매핑
    month_cols = {}
    for c in range(FIRST_MONTH_COL, ws.max_column + 1):
        v = ws.cell(MONTH_HEADER_ROW, c).value
        mm = re.match(r"(\d{1,2})\s*월", str(v or ""))
        if mm:
            month_cols[int(mm.group(1))] = c
    if not month_cols:
        print("❌ 월 헤더를 찾지 못했습니다.")
        return []

    rows = []
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(r, 1).value or "").replace(" ", "").strip()
        key = ITEMS.get(label)
        if not key:
            continue
        for month, c in sorted(month_cols.items()):
            v = ws.cell(r, c).value
            if v in (None, "") or isinstance(v, str):
                continue
            try:
                amt = float(v)
            except (TypeError, ValueError):
                continue
            if amt <= 0:
                continue    # 아직 미집계된 달
            rows.append((year, month, key, int(round(amt * 1000))))  # 천원 → 원
    return rows


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else latest_file(SRC_DIR)
    if not path or not os.path.exists(path):
        print(f"❌ 결산자료를 찾을 수 없습니다: {path or SRC_DIR}")
        print("   NAS(192.168.100.3) taein_hq 공유 마운트를 확인하세요.")
        return 1

    rows = parse(path)
    if not rows:
        print("❌ 취합할 데이터가 없습니다.")
        return 1
    print(f"  📄 {_nfc(os.path.basename(path))} — {len(rows)}건")

    conn = _conn(); cur = conn.cursor()
    cur.executemany("""
        INSERT INTO mgmt_cost_monthly (`year`, `month`, item, amount)
        VALUES (%s,%s,%s,%s)
        ON DUPLICATE KEY UPDATE amount=VALUES(amount)
    """, rows)
    conn.commit()

    cur.execute("""SELECT `year`, `month`, amount FROM mgmt_cost_monthly
                   WHERE item='meal' ORDER BY `year`, `month`""")
    print("\n  [식당(식수) 월별 금액]")
    for y, mo, amt in cur.fetchall():
        print(f"    {y}-{mo:02d} : {amt:,}원")
    conn.close()
    print(f"\n✅ 저장 완료 — {len(rows)}건")
    return 0


if __name__ == "__main__":
    sys.exit(main())
