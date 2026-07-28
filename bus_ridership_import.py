#!/usr/bin/env python3
"""
통근버스 운행일지(경비실 작성 엑셀) → attendance.bus_ridership 취합

원본: smb://192.168.100.3/taein_hq/이민영/매월말일경비실_통근버스운행일지_기사님작성양식/
      YYYYMM_평균탑승인원&집계표.xlsx  (월 1개)

엑셀 '집계표' 시트는 좌우 2블록(호차별)이며, 연도별로 열 위치가 다르다.
  2025년: 날짜가 '일(日)' 숫자, 2번째 블록이 3호차
  2026년: 날짜가 실제 날짜값, 2번째 블록이 2호차
→ 열을 하드코딩하지 않고 '날짜 / 출근인원 / 퇴근인원' 헤더를 찾아 자동 인식한다.

사용:
  python3 bus_ridership_import.py                 # 폴더 전체 취합
  python3 bus_ridership_import.py 202606 202607   # 특정 월만
"""
import os
import re
import sys
import glob
import unicodedata
from datetime import datetime

import pymysql
from dotenv import load_dotenv

load_dotenv()

SRC_DIR = os.environ.get(
    "BUS_LOG_DIR",
    "/Volumes/taein_hq/이민영/매월말일경비실_통근버스운행일지_기사님작성양식",
)

HEADER_SEARCH_ROWS = 12   # 헤더('날짜')를 찾을 최대 행


def _nfc(s):
    """macOS SMB는 한글 파일명을 NFD(자모 분리)로 돌려주므로 NFC로 정규화해 비교"""
    return unicodedata.normalize("NFC", s or "")


def list_source_files():
    """원본 폴더의 월별 집계표 목록 (한글 정규화 안전)"""
    try:
        names = os.listdir(SRC_DIR)
    except OSError:
        return []
    out = []
    for n in names:
        nn = _nfc(n)
        if nn.startswith("~$") or not nn.lower().endswith(".xlsx"):
            continue
        if "평균탑승인원" in nn:
            out.append(os.path.join(SRC_DIR, n))
    return sorted(out, key=lambda p: _nfc(os.path.basename(p)))


def _conn():
    return pymysql.connect(
        host=os.environ.get("DB_HOST", "192.168.100.11"),
        port=int(os.environ.get("DB_PORT", 3307)),
        user=os.environ.get("DB_USER", "root"),
        password=os.environ["DB_PASSWORD"],
        database="attendance",
        charset="utf8mb4",
    )


def _to_int(v):
    """셀 값 → 정수 (빈칸/문자 → None)"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip().replace(",", "")
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _to_date(v):
    """셀 값 → date (2026-6-1 / datetime / 문자열 혼재)"""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if hasattr(v, "year") and hasattr(v, "month"):   # date
        return v
    s = str(v).strip()
    m = re.match(r"(\d{4})[-./]\s*(\d{1,2})[-./]\s*(\d{1,2})", s)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3))).date()
        except ValueError:
            return None
    return None


def _cell_txt(ws, r, c):
    v = ws.cell(r, c).value
    return re.sub(r"\s+", "", str(v)) if v is not None else ""


def _sheet_year_month(ws, fallback_name=""):
    """시트 상단의 날짜값에서 연·월 추출 (없으면 파일명 YYYYMM)"""
    for r in range(1, 6):
        for c in range(1, 20):
            v = ws.cell(r, c).value
            if isinstance(v, datetime):
                return v.year, v.month
            d = _to_date(v)
            if d:
                return d.year, d.month
    m = re.match(r"(\d{4})(\d{2})", _nfc(fallback_name))
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None


def _find_blocks(ws):
    """'날짜' 헤더를 찾아 블록별 열 위치를 자동 인식.
    반환: [(header_row, c_date, c_on, c_off1730, c_off2000, bus_no, driver), ...]"""
    blocks = []
    for r in range(1, min(ws.max_row, HEADER_SEARCH_ROWS) + 1):
        for c in range(1, ws.max_column + 1):
            if _cell_txt(ws, r, c) != "날짜":
                continue
            c_on = c_off1 = c_off2 = None
            for cc in range(c + 1, min(c + 6, ws.max_column) + 1):
                t = _cell_txt(ws, r, cc)
                if "출근인원" in t and c_on is None:
                    c_on = cc
                elif "퇴근인원" in t:
                    if "17:30" in t and c_off1 is None:
                        c_off1 = cc
                    elif "20:00" in t and c_off2 is None:
                        c_off2 = cc
                    elif c_off1 is None:
                        c_off1 = cc
                    elif c_off2 is None:
                        c_off2 = cc
            if c_on is None:
                continue
            # 같은 블록의 '차량'/'기사' 라벨은 헤더 위쪽에 있음
            bus_no, driver = None, ""
            for rr in range(max(1, r - 6), r):
                for cc in range(max(1, c - 3), min(c + 8, ws.max_column) + 1):
                    t = _cell_txt(ws, rr, cc)
                    if t == "차량" and bus_no is None:
                        mm = re.search(r"(\d+)\s*호차", str(ws.cell(rr, cc + 1).value or ""))
                        if mm:
                            bus_no = int(mm.group(1))
                    elif t == "기사" and not driver:
                        driver = str(ws.cell(rr, cc + 1).value or "").strip()
            blocks.append([r, c, c_on, c_off1, c_off2, bus_no, driver])

    # 호차를 못 읽은 블록은 좌→우 순서로 1,2… 부여
    for i, b in enumerate(sorted(blocks, key=lambda x: x[1])):
        if b[5] is None:
            b[5] = i + 1
    return blocks


def parse_file(path):
    """엑셀 1개 → [(date, bus_no, driver, on, off1, off2, src), ...]"""
    import openpyxl
    base = os.path.basename(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    if "집계표" not in wb.sheetnames:
        print(f"  ⚠️  '집계표' 시트 없음 — 건너뜀: {_nfc(base)}")
        return []
    ws = wb["집계표"]
    year, month = _sheet_year_month(ws, base)
    if not year:
        print(f"  ⚠️  연·월 판별 실패 — 건너뜀: {_nfc(base)}")
        return []

    rows = []
    for hrow, c_date, c_on, c_off1, c_off2, bus_no, driver in _find_blocks(ws):
        for r in range(hrow + 1, ws.max_row + 1):
            raw = ws.cell(r, c_date).value
            d = _to_date(raw)
            if d is None:
                # 2025 양식: '1'(일 숫자) / 2026 양식: '1(월)'(일+요일)
                mm = re.match(r"\s*(\d{1,2})", str(raw or ""))
                day = int(mm.group(1)) if mm else None
                if day and 1 <= day <= 31:
                    try:
                        d = datetime(year, month, day).date()
                    except ValueError:
                        continue
            if d is None and c_date > 1:
                # 날짜 헤더 왼쪽 칸에 실제 날짜값이 있는 양식 대응
                d = _to_date(ws.cell(r, c_date - 1).value)
            if d is None or d.year != year or d.month != month:
                continue
            on   = _to_int(ws.cell(r, c_on).value)
            off1 = _to_int(ws.cell(r, c_off1).value) if c_off1 else None
            off2 = _to_int(ws.cell(r, c_off2).value) if c_off2 else None
            if on is None and off1 is None and off2 is None:
                continue   # 미작성일 — 0으로 저장하면 평균이 왜곡되므로 제외
            rows.append((d, bus_no, driver, on or 0, off1 or 0, off2 or 0, base))
    return rows


def main():
    if not os.path.isdir(SRC_DIR):
        print(f"❌ 원본 폴더에 접근할 수 없습니다: {SRC_DIR}")
        print("   NAS(192.168.100.3) taein_hq 공유가 마운트되어 있는지 확인하세요.")
        return 1

    wanted = set(sys.argv[1:])          # 예: 202606
    files = list_source_files()
    if wanted:
        files = [f for f in files
                 if any(_nfc(os.path.basename(f)).startswith(w) for w in wanted)]
    if not files:
        print("❌ 대상 엑셀이 없습니다.")
        return 1

    all_rows = []
    for f in files:
        rows = parse_file(f)
        print(f"  📄 {_nfc(os.path.basename(f))} — {len(rows)}건")
        all_rows.extend(rows)

    # (날짜,호차) 중복 시 인원 합이 큰 쪽을 채택
    # (원본 엑셀에 같은 날짜가 두 줄 들어간 달이 있음 — 한쪽은 전부 0)
    best = {}
    for row in all_rows:
        key = (row[0], row[1])
        if key not in best or sum(row[3:6]) > sum(best[key][3:6]):
            best[key] = row
    dropped = len(all_rows) - len(best)
    if dropped:
        print(f"  ℹ️  중복 {dropped}건 정리 (인원 큰 쪽 채택)")
    all_rows = sorted(best.values(), key=lambda r: (r[0], r[1]))

    if not all_rows:
        print("❌ 취합할 데이터가 없습니다.")
        return 1

    conn = _conn(); cur = conn.cursor()
    cur.executemany("""
        INSERT INTO bus_ridership
            (ride_date, bus_no, driver, on_0830, off_1730, off_2000, src_file)
        VALUES (%s,%s,%s,%s,%s,%s,%s)
        ON DUPLICATE KEY UPDATE
            driver=VALUES(driver), on_0830=VALUES(on_0830),
            off_1730=VALUES(off_1730), off_2000=VALUES(off_2000),
            src_file=VALUES(src_file)
    """, all_rows)
    conn.commit()

    cur.execute("SELECT COUNT(*), MIN(ride_date), MAX(ride_date) FROM bus_ridership")
    n, mn, mx = cur.fetchone()
    conn.close()
    print(f"\n✅ 저장 완료 — 이번 처리 {len(all_rows)}건 / 누적 {n}건 ({mn} ~ {mx})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
