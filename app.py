"""Flask attendance viewer - MariaDB version (for NAS deployment)"""
import os
import re
import ssl
import calendar
import threading
from collections import OrderedDict, defaultdict
from datetime import datetime, timedelta
from urllib.request import Request, urlopen
from urllib.parse import urlencode

import hashlib
from functools import wraps

import holidays
import pymysql
from flask import Flask, render_template, request, flash, jsonify, redirect, url_for, session, send_file

KR_HOLIDAYS = holidays.KR(years=range(2000, 2030), language="ko")

app = Flask(__name__)
app.secret_key = "attendance-viewer"

TELEGRAM_TOKEN = "8657100765:AAFKzTX4HWPXOnNZQW1EqbKb8p8y4CxYYmc"
TELEGRAM_CHAT_ID = "-5273610470"

MARIA = {
    "host": "127.0.0.1",
    "port": 3307,
    "user": "root",
    "password": "7602mr",
    "db": "attendance",
    "charset": "utf8mb4",
}

MODE_IN = "1"
MODE_OUT = "2"

DEPT_MAP = {
    "0001000000000000": "관리부",
    "0002000000000000": "전기사무",
    "0003000000000000": "전자사무",
    "0004000000000000": "전기생산",
    "0005000000000000": "전자생산",
    "0006000000000000": "임원",
    "0007000000000000": "경영기획팀",
}


def _conn():
    return pymysql.connect(**MARIA)


def _hash_pw(pw):
    return hashlib.sha256(pw.encode("utf-8")).hexdigest()


def _sync_leave_from_other(cur, e_id, e_date, value):
    """기타 而щ읆일연차/반차 ?낅젰 일leave_records ?먮룞 ?곕룞"""
    # 湲곗〈 ?대떦일leave_records 삭제 (기타 而щ읆 湲곕컲 ?곕룞遺?
    cur.execute("DELETE FROM leave_records WHERE e_id=%s AND leave_date=%s", (e_id, e_date))
    # 媛믪뿉 연차/반차 ?ы븿 일?덈줈 ?깅줉
    v = str(value).strip() if value else ""
    if "반차" in v:
        cur.execute("INSERT IGNORE INTO leave_records (e_id, leave_date, leave_type) VALUES (%s,%s,%s)",
                    (e_id, e_date, "반차"))
    elif "연차" in v:
        cur.execute("INSERT IGNORE INTO leave_records (e_id, leave_date, leave_type) VALUES (%s,%s,%s)",
                    (e_id, e_date, "연차"))
    # annual_leave.used ?ш퀎일
    year_val = int(e_date[:4])
    cur.execute("SELECT leave_type FROM leave_records WHERE e_id=%s AND leave_date LIKE %s",
                (e_id, f"{year_val}%"))
    total_used = 0
    for (lt,) in cur.fetchall():
        total_used += 0.5 if lt == "반차" else 1
    cur.execute("SELECT name FROM tuser WHERE id=%s", (e_id,))
    row = cur.fetchone()
    e_name = row[0] if row else ""
    cur.execute("""
        INSERT INTO annual_leave (e_id, e_name, year, used)
        VALUES (%s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE used=%s, e_name=%s, updated_at=NOW()
    """, (e_id, e_name, year_val, total_used, total_used, e_name))


def _send_telegram(msg, category=None):
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"

    def _post(chat_id):
        try:
            data = urlencode({"chat_id": chat_id, "text": msg}).encode()
            urlopen(Request(url, data), context=ctx, timeout=10)
        except Exception:
            pass

    try:
        conn2 = _conn(); cur2 = conn2.cursor()
        # ?꾩껜 ?뚮┝
        global_on = True
        if category:
            cur2.execute("SELECT `enabled` FROM telegram_settings WHERE `key`=%s", (category,))
            row = cur2.fetchone()
            if row and not row[0]:
                global_on = False
        if global_on:
            _post(TELEGRAM_CHAT_ID)
        # 媛쒖씤蹂일뚮┝
        if category:
            cur2.execute("SELECT tg_chat_id, tg_alerts FROM app_users WHERE tg_chat_id != '' AND tg_chat_id IS NOT NULL")
            for chat_id, tg_alerts in cur2.fetchall():
                alerts = [a.strip() for a in (tg_alerts or "").split(",") if a.strip()]
                if category in alerts:
                    _post(chat_id)
        conn2.close()
    except Exception:
        pass


_last_checked_sync_id = 0

def _sync_log_checker():
    """백그라운드: sync_log 테이블을 주기적으로 확인하여 텔레그램 알림 발송"""
    global _last_checked_sync_id
    while True:
        try:
            import time; time.sleep(60)
            conn = _conn(); cur = conn.cursor()
            # 理쒖큹 ?ㅽ뻾 일留덉?留?ID ?명똿
            if _last_checked_sync_id == 0:
                cur.execute("SELECT IFNULL(MAX(id),0) FROM sync_log")
                _last_checked_sync_id = cur.fetchone()[0]
                conn.close()
                continue
            # 일濡쒓렇 議고쉶
            cur.execute("SELECT id, level, message FROM sync_log WHERE id > %s ORDER BY id",
                        (_last_checked_sync_id,))
            rows = cur.fetchall()
            conn.close()
            for sid, level, message in rows:
                _last_checked_sync_id = sid
                msg = message or ""
                if level == "ERROR":
                    _send_telegram(f"⚠️ 싱크 에러 발생!\n❌{msg[:200]}", "sync_error")
        except Exception:
            pass


def _init_db():
    try:
        conn = _conn()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `work_override` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `e_id` INT NOT NULL,
                `e_date` VARCHAR(8) NOT NULL,
                `col_type` VARCHAR(10) NOT NULL,
                `value` VARCHAR(20) NOT NULL,
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_override` (`e_id`, `e_date`, `col_type`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # 湲곗〈 DECIMAL 而щ읆?대㈃ VARCHAR濡?蹂寃?
        cur.execute("""
            SELECT DATA_TYPE FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA=DATABASE() AND TABLE_NAME='work_override' AND COLUMN_NAME='value'
        """)
        row = cur.fetchone()
        if row and row[0].upper() != 'VARCHAR':
            cur.execute("ALTER TABLE work_override MODIFY `value` VARCHAR(20) NOT NULL")
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `annual_leave` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `e_id` INT NOT NULL,
                `e_name` VARCHAR(30) NOT NULL DEFAULT '',
                `year` SMALLINT NOT NULL,
                `total` FLOAT NOT NULL DEFAULT 15,
                `used` FLOAT NOT NULL DEFAULT 0,
                `memo` VARCHAR(200) NOT NULL DEFAULT '',
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_annual` (`e_id`, `year`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `app_users` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `login_id` VARCHAR(30) NOT NULL,
                `password` VARCHAR(64) NOT NULL,
                `name` VARCHAR(30) NOT NULL DEFAULT '',
                `e_id` INT DEFAULT NULL,
                `role` VARCHAR(10) NOT NULL DEFAULT 'user',
                `permissions` VARCHAR(100) NOT NULL DEFAULT '',
                `must_change_pw` TINYINT NOT NULL DEFAULT 1,
                `tg_chat_id` VARCHAR(50) NOT NULL DEFAULT '',
                `tg_alerts` VARCHAR(100) NOT NULL DEFAULT '',
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_login` (`login_id`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `leave_records` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `e_id` INT NOT NULL,
                `leave_date` VARCHAR(8) NOT NULL,
                `leave_type` VARCHAR(10) NOT NULL DEFAULT '연차',
                `memo` VARCHAR(100) NOT NULL DEFAULT '',
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_leave` (`e_id`, `leave_date`, `leave_type`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `meal_count` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `year_month` VARCHAR(7) NOT NULL,
                `dept` VARCHAR(10) NOT NULL,
                `meal_type` VARCHAR(10) NOT NULL,
                `day` INT NOT NULL,
                `count` VARCHAR(20) NOT NULL DEFAULT '0',
                `memo` VARCHAR(200) NOT NULL DEFAULT '',
                `writer` VARCHAR(20) NOT NULL DEFAULT '',
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_meal` (`year_month`, `dept`, `meal_type`, `day`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # memo 而щ읆 ?놁쑝硫?異붽?
        try:
            cur.execute("ALTER TABLE meal_count ADD COLUMN `memo` VARCHAR(200) NOT NULL DEFAULT '' AFTER `count`")
        except Exception:
            pass
        try:
            cur.execute("ALTER TABLE meal_count ADD COLUMN `writer` VARCHAR(20) NOT NULL DEFAULT '' AFTER `memo`")
        except Exception:
            pass
        try:
            cur.execute("ALTER TABLE meal_notice ADD COLUMN `writer` VARCHAR(20) NOT NULL DEFAULT '' AFTER `content`")
        except Exception:
            pass
        try:
            cur.execute("ALTER TABLE app_users ADD COLUMN `permissions` VARCHAR(100) NOT NULL DEFAULT '' AFTER `role`")
        except Exception:
            pass
        try:
            cur.execute("ALTER TABLE app_users ADD COLUMN `tg_chat_id` VARCHAR(50) NOT NULL DEFAULT '' AFTER `must_change_pw`")
        except Exception:
            pass
        try:
            cur.execute("ALTER TABLE app_users ADD COLUMN `tg_alerts` VARCHAR(100) NOT NULL DEFAULT '' AFTER `tg_chat_id`")
        except Exception:
            pass
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `meal_notice` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `year_month` VARCHAR(7) NOT NULL,
                `content` TEXT NOT NULL,
                `writer` VARCHAR(20) NOT NULL DEFAULT '',
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # ?붾젅洹몃옩 ?뚮┝ ?ㅼ젙 ?뚯씠釉?
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `telegram_settings` (
                `key` VARCHAR(30) PRIMARY KEY,
                `enabled` TINYINT NOT NULL DEFAULT 1,
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        for k in ('sync_error', 'leave', 'meal', 'notice'):
            cur.execute("INSERT IGNORE INTO telegram_settings (`key`, `enabled`) VALUES (%s, 1)", (k,))
        # 사원명부 ?뚯씠釉?
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `employee_roster` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `name` VARCHAR(30) NOT NULL,
                `dept` VARCHAR(30) NOT NULL DEFAULT '',
                `emp_no` VARCHAR(20) NOT NULL DEFAULT '',
                `gender` VARCHAR(10) NOT NULL DEFAULT '',
                `birth_date` VARCHAR(20) NOT NULL DEFAULT '',
                `calendar_type` VARCHAR(10) NOT NULL DEFAULT '',
                `age` INT NOT NULL DEFAULT 0,
                `appoint_date` VARCHAR(20) NOT NULL DEFAULT '',
                `position` VARCHAR(20) NOT NULL DEFAULT '',
                `phone` VARCHAR(30) NOT NULL DEFAULT '',
                `email` VARCHAR(100) NOT NULL DEFAULT '',
                `address` VARCHAR(300) NOT NULL DEFAULT '',
                `hire_date` VARCHAR(20) NOT NULL DEFAULT '',
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY `uq_empno` (`emp_no`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # 洹쇰Т?쒓린濡일뚯씠釉?
        cur.execute("""
            CREATE TABLE IF NOT EXISTS `schedule_record` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `emp_name` VARCHAR(50) NOT NULL,
                `work_date` VARCHAR(8) NOT NULL,
                `basic_h` INT NOT NULL DEFAULT 0,
                `ot_h` INT NOT NULL DEFAULT 0,
                `night_h` INT NOT NULL DEFAULT 0,
                `etc` VARCHAR(20) DEFAULT NULL,
                `source_type` VARCHAR(10) NOT NULL DEFAULT '원본',
                `sheet_name` VARCHAR(30) NOT NULL DEFAULT '전자',
                `uploaded_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                `uploaded_by` VARCHAR(50) DEFAULT NULL,
                UNIQUE KEY `uq_sr` (`emp_name`, `work_date`, `source_type`, `sheet_name`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # admin 기본 怨꾩젙
        cur.execute("""
            INSERT IGNORE INTO app_users (login_id, password, name, role, must_change_pw)
            VALUES ('admin', %s, '관리자', 'admin', 0)
        """, (_hash_pw("admin"),))
        # 湲곗〈 work_override 기타 而щ읆일연차/반차 일leave_records ?쇨큵 ?숆린일
        cur.execute("SELECT e_id, e_date, value FROM work_override WHERE col_type='other' AND (value LIKE '%%연차%%' OR value LIKE '%%반차%%')")
        for eid, edate, val in cur.fetchall():
            vstr = str(val).strip()
            if "반차" in vstr:
                cur.execute("INSERT IGNORE INTO leave_records (e_id, leave_date, leave_type) VALUES (%s,%s,'반차')", (eid, edate))
            if "연차" in vstr and "반차" not in vstr:
                cur.execute("INSERT IGNORE INTO leave_records (e_id, leave_date, leave_type) VALUES (%s,%s,'연차')", (eid, edate))
        conn.commit()
        conn.close()
    except Exception:
        pass


def _login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def _admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        if session.get("role") != "admin":
            flash("관리자 권한이 필요합니다.", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated


def _has_perm(perm):
    """admin은 모든 권한, 일반 사용자는 permissions 필드 확인"""
    if session.get("role") == "admin":
        return True
    perms = session.get("permissions", "")
    return perm in [p.strip() for p in perms.split(",") if p.strip()]


_init_db()


def _time_to_sec(t):
    if not t or len(t) < 6:
        return None
    return int(t[:2]) * 3600 + int(t[2:4]) * 60 + int(t[4:6])


def _sec_to_fmt(s):
    if s is None:
        return "-"
    h = int(s) // 3600
    m = (int(s) % 3600) // 60
    return f"{h:02d}:{m:02d}"


def _fmt_time(t):
    if not t or len(t) < 6:
        return ""
    return f"{t[:2]}:{t[2:4]}:{t[4:6]}"


def _fmt_date(d):
    if d and len(d) == 8:
        return f"{d[:4]}-{d[4:6]}-{d[6:8]}"
    return d or ""


def _merge_night_shifts(day_emp):
    """야간근무 대각선 패턴 병합
    패턴: 3/23 19:23출근 → 3/24 05:00퇴근 + 19:30출근 → 3/25 05:00퇴근 ...
    같은 날에 새벽퇴근(전날분) + 저녁출근(오늘분)이 섞인 경우를 분리 후 병합
    """
    emp_dates = defaultdict(list)
    for (e_date, eid) in list(day_emp.keys()):
        emp_dates[eid].append(e_date)
    remove_keys = []
    lost_morning_outs = {}
    for eid, dates in emp_dates.items():
        dates.sort()
        # 1?④퀎: ?媛곸꽑 遺꾨━ 일媛숈? 일?덈꼍?닿렐(?꾨궇遺? + ??곸텧洹일ㅻ뒛遺?
        for date in dates:
            key = (date, eid)
            rec = day_emp.get(key)
            if not rec:
                continue
            in_s = _time_to_sec(rec["in_time"])
            out_s = _time_to_sec(rec["out_time"])
            if (in_s is not None and in_s >= 64800 and
                    out_s is not None and out_s >= 43200 and out_s < in_s):
                # 야간 異쒓렐 吏곸쟾 ?닿렐湲곕줉: ?댁쟾 洹쇰Т일?닿렐?대?濡일쒓굅
                rec["out_time"] = None
                if "out_gate" in rec:
                    rec["out_gate"] = ""
            elif (in_s is not None and in_s >= 64800 and
                    out_s is not None and out_s < 43200):
                # ?덈꼍 ?닿렐일?꾨궇 ?덉퐫?쒕줈 ?대룞
                morning_out = rec["out_time"]
                morning_gate = rec.get("out_gate", "")
                rec["out_time"] = None
                if "out_gate" in rec:
                    rec["out_gate"] = ""
                try:
                    prev_date = (datetime.strptime(date, "%Y%m%d") - timedelta(days=1)).strftime("%Y%m%d")
                except ValueError:
                    continue
                prev_key = (prev_date, eid)
                prev = day_emp.get(prev_key)
                merged = False
                if prev:
                    prev_in_s = _time_to_sec(prev["in_time"])
                    if prev_in_s is not None and prev_in_s >= 64800 and prev["out_time"] is None:
                        prev["out_time"] = morning_out
                        if "out_gate" in prev:
                            prev["out_gate"] = morning_gate
                        prev["night_next_day"] = True
                        merged = True
                if not merged:
                    lost_morning_outs[prev_key] = {
                        "date": prev_date, "id": eid,
                        "name": rec["name"], "dept": rec.get("dept", ""),
                        "out_time": morning_out, "out_gate": morning_gate,
                    }
        # 2?④퀎: ?⑥? 誘몃ℓ移일일꾨궇 異쒓렐留일닿렐?놁쓬 일?ㅼ쓬일異쒓렐?놁쓬+?닿렐留?
        for i in range(len(dates) - 1):
            curr_key = (dates[i], eid)
            next_key = (dates[i + 1], eid)
            curr = day_emp.get(curr_key)
            nxt = day_emp.get(next_key)
            if not curr or not nxt:
                continue
            try:
                curr_dt = datetime.strptime(dates[i], "%Y%m%d")
                next_dt = datetime.strptime(dates[i + 1], "%Y%m%d")
            except ValueError:
                continue
            if (next_dt - curr_dt).days != 1:
                continue
            curr_in_s = _time_to_sec(curr["in_time"])
            if curr_in_s is not None and curr_in_s >= 64800 and curr["out_time"] is None:
                nxt_out_s = _time_to_sec(nxt["out_time"])
                if nxt["in_time"] is None and nxt_out_s is not None and nxt_out_s < 43200:
                    curr["out_time"] = nxt["out_time"]
                    if "out_gate" in curr:
                        curr["out_gate"] = nxt.get("out_gate", "")
                    curr["night_next_day"] = True
                    remove_keys.append(next_key)
    for k in remove_keys:
        if k in day_emp:
            del day_emp[k]
    # 3?④퀎: ?먯떎일?덈꼍 ?닿렐 蹂듦뎄 (異쒓렐?놁쓬+?닿렐留일덉퐫일?앹꽦)
    for key, info in lost_morning_outs.items():
        if key not in day_emp:
            day_emp[key] = {
                "date": info["date"], "id": info["id"],
                "name": info["name"], "dept": info.get("dept", ""),
                "in_time": None, "out_time": info["out_time"],
                "in_gate": "", "out_gate": info.get("out_gate", ""),
                "night_next_day": True,
            }
        elif day_emp[key]["out_time"] is None:
            day_emp[key]["out_time"] = info["out_time"]
            day_emp[key]["night_next_day"] = True
            if "out_gate" in day_emp[key]:
                day_emp[key]["out_gate"] = info.get("out_gate", "")
    # 4?④퀎: 怨좎븘 ?덈꼍?닿렐 일?꾨궇 야간洹쇰Т濡일대룞
    orphan_moves = {}
    for key in list(day_emp.keys()):
        rec = day_emp[key]
        if rec["in_time"] is not None:
            continue
        out_s = _time_to_sec(rec["out_time"])
        if out_s is None or out_s >= 28800:
            continue
        if rec.get("night_next_day"):
            continue
        e_date, eid = key
        try:
            prev_date = (datetime.strptime(e_date, "%Y%m%d") - timedelta(days=1)).strftime("%Y%m%d")
        except ValueError:
            continue
        prev_key = (prev_date, eid)
        if prev_key not in day_emp:
            orphan_moves[key] = prev_key
    for old_key, new_key in orphan_moves.items():
        rec = day_emp.pop(old_key)
        rec["date"] = new_key[0]
        rec["night_next_day"] = True
        day_emp[new_key] = rec


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        login_id = request.form.get("login_id", "").strip()
        password = request.form.get("password", "").strip()
        if not login_id or not password:
            flash("아이디와 비밀번호를 입력하세요.", "danger")
            return render_template("login.html")
        conn = _conn(); cur = conn.cursor()
        cur.execute("SELECT id, login_id, password, name, e_id, role, must_change_pw, `permissions` FROM app_users WHERE login_id=%s", (login_id,))
        row = cur.fetchone()
        if not row:
            # idno(?붾㈃ ?щ쾲) ?먮뒗 id(?대?PK)濡?tuser 議고쉶 일?먮룞 怨꾩젙 ?앹꽦
            cur.execute("SELECT id, name, idno FROM tuser WHERE (idno=%s OR CAST(id AS CHAR)=%s) AND name IS NOT NULL AND name <> ''", (login_id, login_id))
            tu = cur.fetchone()
            if tu:
                use_id = tu[2] if tu[2] else str(tu[0])  # idno ?곗꽑
                if password == login_id:
                    cur.execute("""
                        INSERT IGNORE INTO app_users (login_id, password, name, e_id, role, must_change_pw)
                        VALUES (%s, %s, %s, %s, 'user', 1)
                    """, (use_id, _hash_pw(use_id), tu[1], tu[0]))
                    conn.commit()
                    cur.execute("SELECT id, login_id, password, name, e_id, role, must_change_pw, `permissions` FROM app_users WHERE login_id=%s", (use_id,))
                    row = cur.fetchone()
        conn.close()
        if not row or row[2] != _hash_pw(password):
            flash("아이디 또는 비밀번호가 올바르지 않습니다.", "danger")
            return render_template("login.html")
        session["user_id"] = row[0]
        session["login_id"] = row[1]
        session["user_name"] = row[3]
        session["e_id"] = row[4]
        session["role"] = row[5]
        session["permissions"] = row[7] if len(row) > 7 else ""
        if row[6]:
            flash("초기 비밀번호입니다. 비밀번호를 변경해 주세요.", "warning")
            return redirect(url_for("change_password"))
        return redirect(url_for("dashboard"))
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/change_password", methods=["GET", "POST"])
@_login_required
def change_password():
    if request.method == "POST":
        cur_pw = request.form.get("current_password", "")
        new_pw = request.form.get("new_password", "")
        confirm_pw = request.form.get("confirm_password", "")
        if not new_pw or len(new_pw) < 4:
            flash("새 비밀번호는 4자리 이상이어야 합니다.", "danger")
            return render_template("change_password.html")
        if new_pw != confirm_pw:
            flash("새 비밀번호가 일치하지 않습니다.", "danger")
            return render_template("change_password.html")
        conn = _conn(); cur = conn.cursor()
        cur.execute("SELECT password FROM app_users WHERE id=%s", (session["user_id"],))
        row = cur.fetchone()
        if not row or row[0] != _hash_pw(cur_pw):
            flash("현재 비밀번호가 올바르지 않습니다.", "danger")
            conn.close()
            return render_template("change_password.html")
        cur.execute("UPDATE app_users SET password=%s, must_change_pw=0 WHERE id=%s",
                    (_hash_pw(new_pw), session["user_id"]))
        conn.commit(); conn.close()
        flash("비밀번호가 변경되었습니다.", "success")
        return redirect(url_for("dashboard"))
    return render_template("change_password.html")


@app.route("/user_management")
@_admin_required
def user_management():
    conn = _conn(); cur = conn.cursor()
    cur.execute("SELECT id, login_id, name, role, `permissions`, must_change_pw, created_at, tg_chat_id, tg_alerts FROM app_users ORDER BY role DESC, name")
    users = []
    for uid, lid, name, role, perms, mcp, cat, tg_cid, tg_al in cur.fetchall():
        users.append({"id": uid, "login_id": lid, "name": name, "role": role,
                       "permissions": perms or "", "must_change_pw": mcp,
                       "created_at": cat.strftime("%Y-%m-%d") if cat else "",
                       "tg_chat_id": tg_cid or "", "tg_alerts": tg_al or ""})
    conn.close()
    return render_template("user_management.html", users=users)

@app.route("/sync_users", methods=["POST"])
@_admin_required
def sync_users():
    """연차관리 재직자를 app_users에 일괄 등록 (없는 사람만)"""
    year_int = datetime.now().year
    conn = _conn(); cur = conn.cursor()
    cur.execute("""SELECT t.id, t.name, t.idno FROM tuser t
                   INNER JOIN annual_leave a ON t.id = a.e_id AND a.year = %s
                   WHERE t.name IS NOT NULL AND t.name <> ''""", (year_int,))
    employees = cur.fetchall()
    cur.execute("SELECT e_id FROM app_users WHERE e_id IS NOT NULL")
    existing = {r[0] for r in cur.fetchall()}
    added = 0
    for eid, name, idno in employees:
        if eid not in existing:
            login_id = idno if idno else str(eid)
            pw_hash = hashlib.sha256(login_id.encode()).hexdigest()
            cur.execute("""INSERT INTO app_users (login_id, password, name, e_id, role, must_change_pw, created_at)
                           VALUES (%s, %s, %s, %s, 'user', 1, NOW())""",
                        (login_id, pw_hash, name, eid))
            added += 1
    conn.commit(); conn.close()
    return jsonify(ok=True, added=added)


@app.route("/get_telegram_settings")
@_admin_required
def get_telegram_settings():
    conn = _conn(); cur = conn.cursor()
    cur.execute("SELECT `key`, `enabled` FROM telegram_settings")
    settings = {r[0]: bool(r[1]) for r in cur.fetchall()}
    conn.close()
    return jsonify(ok=True, settings=settings)


@app.route("/save_telegram_settings", methods=["POST"])
@_admin_required
def save_telegram_settings():
    data = request.get_json()
    conn = _conn(); cur = conn.cursor()
    for key in ("sync_error", "leave", "meal", "notice"):
        enabled = 1 if data.get(key) else 0
        cur.execute("UPDATE telegram_settings SET `enabled`=%s WHERE `key`=%s", (enabled, key))
    conn.commit(); conn.close()
    return jsonify(ok=True)


@app.route("/save_user_telegram", methods=["POST"])
@_admin_required
def save_user_telegram():
    try:
        data = request.get_json()
        uid = int(data["id"])
        tg_chat_id = str(data.get("tg_chat_id", "")).strip()
        tg_alerts = str(data.get("tg_alerts", "")).strip()
        conn = _conn(); cur = conn.cursor()
        cur.execute("UPDATE app_users SET tg_chat_id=%s, tg_alerts=%s WHERE id=%s", (tg_chat_id, tg_alerts, uid))
        conn.commit(); conn.close()
        return jsonify(ok=True)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex)), 500


@app.route("/save_user_perm", methods=["POST"])
@_admin_required
def save_user_perm():
    try:
        data = request.get_json()
        uid = int(data["id"])
        perms = str(data.get("permissions", "")).strip()
        conn = _conn(); cur = conn.cursor()
        cur.execute("UPDATE app_users SET `permissions`=%s WHERE id=%s", (perms, uid))
        conn.commit(); conn.close()
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/reset_user_pw", methods=["POST"])
@_admin_required
def reset_user_pw():
    try:
        data = request.get_json()
        uid = int(data["id"])
        conn = _conn(); cur = conn.cursor()
        cur.execute("SELECT login_id FROM app_users WHERE id=%s", (uid,))
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({"ok": False, "error": "사용자 없음"}), 404
        cur.execute("UPDATE app_users SET password=%s, must_change_pw=1 WHERE id=%s",
                    (_hash_pw(row[0]), uid))
        conn.commit(); conn.close()
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/delete_user", methods=["POST"])
@_admin_required
def delete_user():
    try:
        data = request.get_json()
        uid = int(data["id"])
        if uid == session.get("user_id"):
            return jsonify({"ok": False, "error": "본인 계정은 삭제할 수 없습니다"}), 400
        conn = _conn(); cur = conn.cursor()
        cur.execute("DELETE FROM app_users WHERE id=%s AND role<>'admin'", (uid,))
        conn.commit(); conn.close()
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/")
@_login_required
def dashboard():
    """dashboard"""
    from datetime import date as _date
    today = _date.today()
    today_str = today.strftime("%Y%m%d")
    today_month = today.month
    today_day = today.day
    is_admin = session.get("role") == "admin"

    notices = []
    att_summary = {"total": 0, "checked_in": 0, "not_in": 0, "depts": {}}
    leave_today = []
    birthdays = []

    try:
        conn = _conn(); cur = conn.cursor()
        # 怨듭일ы빆 (怨좎젙 ?곗꽑, 理쒓렐 10嫄?
        cur.execute("SELECT id, title, content, is_pinned, created_by, created_at FROM board_notices ORDER BY is_pinned DESC, created_at DESC LIMIT 10")
        notices = [{"id": r[0], "title": r[1], "content": r[2], "is_pinned": r[3], "created_by": r[4], "created_at": r[5]} for r in cur.fetchall()]

        # 금일 출근현황 (근태기록기 출근 타각자만)
        cur.execute("""
            SELECT u.name, u.company,
                   MIN(t.e_time) in_time
            FROM tenter t
            JOIN tuser u ON t.e_id=u.id
            WHERE t.e_date=%s AND t.e_mode=1
            GROUP BY t.e_id, u.name, u.company
        """, (today_str,))
        att_rows = cur.fetchall()
        dept_map = {}
        for r in att_rows:
            dept = DEPT_MAP.get((r[1] or '').strip(), '')
            if not dept:
                dept = '기타'
            if dept not in dept_map:
                dept_map[dept] = {"in": 0}
            dept_map[dept]["in"] += 1
        att_summary["checked_in"] = len(att_rows)
        att_summary["depts"] = dept_map

        # 湲덉씪 ?닿일?
        cur.execute("""
            SELECT u.name, u.company, lr.leave_type
            FROM leave_records lr
            JOIN tuser u ON lr.e_id=u.id
            WHERE lr.leave_date=%s
            ORDER BY u.name
        """, (today.strftime("%Y-%m-%d"),))
        leave_today = [{"name": r[0], "dept": DEPT_MAP.get((r[1] or '').strip(), '기타'), "type": r[2]} for r in cur.fetchall()]

        # 이번 달 생일자
        cur.execute("""
            SELECT er.name, er.dept, er.birth_date
            FROM employee_roster er
            WHERE er.birth_date IS NOT NULL AND er.birth_date <> ''
              AND er.name IS NOT NULL AND er.name <> ''
        """)
        for r in cur.fetchall():
            bday = str(r[2]).replace('-', '')
            if len(bday) >= 4:
                try:
                    mm = int(bday[4:6]); dd = int(bday[6:8])
                    if mm == today_month:
                        birthdays.append({"name": r[0], "dept": r[1] or '', "day": dd, "is_today": dd == today_day})
                except (ValueError, IndexError):
                    pass
        birthdays.sort(key=lambda x: x["day"])
        conn.close()
    except Exception as e:
        flash(f"대시보드 오류: {e}", "danger")

    can_notice = is_admin or _has_perm("notice")
    return render_template("dashboard.html", notices=notices,
                           att_summary=att_summary, leave_today=leave_today,
                           birthdays=birthdays, today=today, can_notice=can_notice)


@app.route("/save_notice", methods=["POST"])
@_login_required
def save_notice():
    if session.get("role") != "admin" and not _has_perm("notice"):
        return jsonify(ok=False, msg="\uad8c\ud55c\uc774 \uc5c6\uc2b5\ub2c8\ub2e4")
    nid = request.form.get("id", "").strip()
    title = request.form.get("title", "").strip()
    content = request.form.get("content", "").strip()
    is_pinned = 1 if request.form.get("is_pinned") else 0
    if not title:
        return jsonify(ok=False, msg="제목을 입력하세요")
    conn = _conn(); cur = conn.cursor()
    if nid:
        cur.execute("UPDATE board_notices SET title=%s, content=%s, is_pinned=%s WHERE id=%s",
                     (title, content, is_pinned, nid))
    else:
        cur.execute("INSERT INTO board_notices (title, content, is_pinned, created_by) VALUES (%s,%s,%s,%s)",
                     (title, content, is_pinned, session.get("user_name", "")))
    conn.commit(); conn.close()
    return jsonify(ok=True)


@app.route("/delete_notice", methods=["POST"])
@_login_required
def delete_notice():
    if session.get("role") != "admin" and not _has_perm("notice"):
        return jsonify(ok=False, msg="\uad8c\ud55c\uc774 \uc5c6\uc2b5\ub2c8\ub2e4")

    nid = request.form.get("id")
    conn = _conn(); cur = conn.cursor()
    cur.execute("DELETE FROM board_notices WHERE id=%s", (nid,))
    conn.commit(); conn.close()
    return jsonify(ok=True)


@app.route("/attendance")
@_login_required
def attendance():
    search = request.args.get("search", "").strip()
    sel_date = request.args.get("date", "").strip()
    sel_dept = request.args.get("dept", "").strip()
    sel_month = request.args.get("month", "").strip()
    is_admin = session.get("role") == "admin"
    if not is_admin:
        search = session.get("user_name", "")
    sel_date_raw = sel_date.replace("-", "") if sel_date else ""
    records = []
    monthly = {}
    stats = {"total": 0, "on_time": 0, "late": 0, "no_in": 0, "no_out": 0}
    mode = ""
    ranking = OrderedDict()
    dept_list = sorted(set(DEPT_MAP.values()))
    try:
        conn = _conn()
        cur = conn.cursor()
        if search:
            mode = "search"
            if not sel_month:
                sel_month = datetime.now().strftime("%Y-%m")
            # ?붾퀎?붿빟: 理쒓렐 1일議고쉶, 由ъ뒪일 ?좏깮일
            now = datetime.now()
            m_from = (now - timedelta(days=365)).strftime("%Y%m") + "01"
            m_to = now.strftime("%Y%m") + "31"
            # 1?④퀎: 而ㅻ쾭留일몃뜳?ㅻ줈 e_id 癒쇱? 議고쉶 (??ㅼ틪 諛⑹?)
            cur.execute("""
                SELECT DISTINCT e_id FROM tenter
                WHERE e_name LIKE %s OR e_idno LIKE %s
                LIMIT 50
            """, (f"%{search}%", f"%{search}%"))
            eids = [r[0] for r in cur.fetchall()]
            if eids:
                ph = ",".join(["%s"] * len(eids))
                # 2?④퀎: idx_eid ?몃뜳?ㅻ줈 鍮좊Ⅴ寃?議고쉶
                cur.execute(f"""
                    SELECT t.e_date, t.e_time, t.e_id, t.e_idno, t.e_name,
                           t.e_mode, g.name AS gate_name, u.company
                    FROM tenter t
                         LEFT JOIN tgate g ON t.g_id = g.id
                         LEFT JOIN tuser u ON t.e_id = u.id
                    WHERE t.e_id IN ({ph}) AND t.e_date >= %s AND t.e_date <= %s
                    ORDER BY t.e_date, t.e_time
                """, eids + [m_from, m_to])
            else:
                cur.execute("SELECT 1 WHERE 1=0")  # 결과없음
        elif sel_date_raw or sel_dept or sel_month:
            mode = "date"
            if sel_month and not sel_date_raw:
                # ?곗썡 ?좏깮: ?대떦 일?꾩껜 議고쉶
                ym = sel_month.replace("-", "")
                ym_year, ym_mon = int(ym[:4]), int(ym[4:6])
                ym_dim = calendar.monthrange(ym_year, ym_mon)[1]
                date_start = f"{ym}01"
                date_end = f"{ym}{ym_dim:02d}"
                # 야간洹쇰Т 蹂묓빀일?꾪빐 ?꾪썑 2일?ъ쑀
                q_start = (datetime(ym_year, ym_mon, 1) - timedelta(days=2)).strftime("%Y%m%d")
                q_end = (datetime(ym_year, ym_mon, ym_dim) + timedelta(days=2)).strftime("%Y%m%d")
            else:
                if not sel_date_raw:
                    sel_date_raw = datetime.now().strftime("%Y%m%d")
                    sel_date = datetime.now().strftime("%Y-%m-%d")
                date_start = sel_date_raw
                date_end = sel_date_raw
                q_start = sel_date_raw
                q_end = sel_date_raw
            sql = """
                SELECT t.e_date, t.e_time, t.e_id, t.e_idno, t.e_name,
                       t.e_mode, g.name AS gate_name, u.company
                FROM tenter t
                     LEFT JOIN tgate g ON t.g_id = g.id
                     LEFT JOIN tuser u ON t.e_id = u.id
                WHERE t.e_id >= 0 AND t.e_date >= %s AND t.e_date <= %s
            """
            params = [q_start, q_end]
            if sel_dept:
                dept_codes = [k for k, v in DEPT_MAP.items() if v == sel_dept]
                if dept_codes:
                    ph_d = ",".join(["%s"] * len(dept_codes))
                    sql += f" AND u.company IN ({ph_d})"
                    params += dept_codes
            sql += " ORDER BY t.e_name, t.e_time"
            cur.execute(sql, params)
        if mode:
            rows = cur.fetchall()
            day_emp = OrderedDict()
            for r in rows:
                e_date = r[0]
                e_time = r[1]
                eid = r[2]
                e_mode = (r[5] or "").strip()
                gate = (r[6] or "").strip()
                key = (e_date, eid)
                if key not in day_emp:
                    company = (r[7] or "").strip()
                    day_emp[key] = {
                        "date": e_date, "date_fmt": _fmt_date(e_date),
                        "id": eid,
                        "emp_no": (r[3] or "").strip() or str(eid),
                        "name": (r[4] or "").strip() or f"ID-{eid}",
                        "dept": DEPT_MAP.get(company, ""),
                        "in_time": None, "out_time": None,
                        "in_gate": "", "out_gate": "",
                    }
                e = day_emp[key]
                if e_mode == MODE_IN:
                    if e["in_time"] is None or e_time < e["in_time"]:
                        e["in_time"] = e_time
                        e["in_gate"] = gate
                elif e_mode == MODE_OUT:
                    if e["out_time"] is None or e_time > e["out_time"]:
                        e["out_time"] = e_time
                        e["out_gate"] = gate
            # 야간洹쇰Т 蹂묓빀 (?ㅼ쨷 ?좎쭨 ?곗씠일
            if mode == 'search' or (mode == 'date' and sel_month):
                _merge_night_shifts(day_emp)
            for e in day_emp.values():
                e["in_time_fmt"] = _fmt_time(e["in_time"])
                e["out_time_fmt"] = _fmt_time(e["out_time"]) + ("(+1일)" if e.get("night_next_day") else "")
                in_s = _time_to_sec(e["in_time"])
                out_s = _time_to_sec(e["out_time"])
                is_night = in_s is not None and in_s > 43140  # 11:59:00
                # ?좎쭨/휴일 ?먮퀎 (怨꾩궛 ?꾩뿉 寃곗젙)
                dt = datetime.strptime(e["date"], "%Y%m%d")
                e["weekday"] = dt.weekday()
                e["holiday"] = KR_HOLIDAYS.get(dt.date(), "")
                is_hol = e["weekday"] >= 5 or bool(e["holiday"])
                e["basic_fmt"] = "-"
                e["night_fmt"] = "-"
                e["other_fmt"] = "-"
                e["ot_hours"] = 0
                if in_s is not None and out_s is not None and out_s > in_s and not is_night:
                    wh = (out_s - in_s) / 3600
                    e["basic_fmt"] = f"{min(wh, 8):.0f}" if wh >= 8 else f"{wh:.1f}"
                    e["work_hours"] = wh
                    e["overtime"] = "2" if out_s >= 68400 else ""  # 19:00 ?댄썑留?연장
                    e["ot_hours"] = 2 if out_s >= 68400 else 0
                elif is_night and out_s is not None and out_s < in_s:
                    wh = (86400 - in_s + out_s) / 3600  # 異쒓렐~?먯젙 + ?먯젙~?닿렐
                    night_sec = max(0, 86400 - max(in_s, 79200)) + max(0, min(out_s, 21600))
                    ni_h = max(0, int(night_sec / 3600 - 1))  # 야간 22~06 - 1h ?닿쾶, 踰꾨┝
                    e["basic_fmt"] = f"{min(wh, 8):.0f}" if wh >= 8 else f"{wh:.1f}"
                    e["night_fmt"] = str(ni_h) if ni_h > 0 else "-"
                    e["work_hours"] = wh
                    e["overtime"] = "2" if out_s >= 28200 else ""  # 07:50 ?댄썑 ?닿렐 일연장
                    e["ot_hours"] = 2 if out_s >= 28200 else 0
                elif in_s is None and out_s is not None and out_s < 28800:  # 異쒓렐?놁쓬 + ?닿렐 08:00 ?댁쟾 일야간洹쇰Т
                    assumed_in = 68400  # ?꾨궇 19:00 異쒓렐 媛꾩＜
                    wh = (86400 - assumed_in + out_s) / 3600
                    night_sec = max(0, 86400 - max(assumed_in, 79200)) + max(0, min(out_s, 21600))
                    ni_h = max(0, int(night_sec / 3600 - 1))
                    e["basic_fmt"] = f"{min(wh, 8):.0f}" if wh >= 8 else f"{wh:.1f}"
                    e["night_fmt"] = str(ni_h) if ni_h > 0 else "-"
                    e["work_hours"] = wh
                    e["overtime"] = "2" if out_s >= 28200 else ""
                    e["ot_hours"] = 2 if out_s >= 28200 else 0
                elif in_s is None and out_s is not None and out_s < 71400:  # 異쒓렐?놁쓬 + ?닿렐 19:50 ?댁쟾 (二쇨컙)
                    e["basic_fmt"] = "8"
                    e["work_hours"] = 8
                    e["overtime"] = ""
                elif in_s is None and out_s is not None and out_s >= 71400:  # 異쒓렐?놁쓬 + ?닿렐 19:50 ?댄썑
                    e["basic_fmt"] = "8"
                    e["work_hours"] = 10
                    e["overtime"] = "2"
                    e["ot_hours"] = 2
                elif in_s is not None or out_s is not None:
                    e["basic_fmt"] = "8"
                    e["work_hours"] = 8
                    e["overtime"] = ""
                    if is_night and out_s is None:
                        assumed_out = 18000  # ?ㅼ쓬일05:00 ?닿렐 媛꾩＜
                        night_sec = max(0, 86400 - max(in_s, 79200)) + max(0, min(assumed_out, 21600))
                        ni_h = max(0, int(night_sec / 3600 - 1))
                        e["night_fmt"] = str(ni_h)
                        e["work_hours"] = 8
                else:
                    e["work_hours"] = 0
                    e["overtime"] = ""
                e["month"] = e["date"][:6]
                day_names = ["월", "화", "수", "목", "금", "토", "일"]
                e["date_fmt"] = f"{_fmt_date(e['date'])}({day_names[dt.weekday()]})"
                if sel_dept and e["dept"] != sel_dept:
                    continue
                if mode == 'date' and sel_month and e["date"] < date_start or \
                   mode == 'date' and sel_month and e["date"] > date_end:
                    continue
                records.append(e)
            # ?붾퀎 ?듦퀎 (理쒓렐 1일
            monthly = OrderedDict()
            all_in_secs_normal = []
            all_in_secs_late = []
            all_out_secs_normal = []
            all_out_secs_late = []
            month_cutoff = (datetime.now() - timedelta(days=365)).strftime("%Y%m")
            for r in records:
                m = r["month"]
                if m < month_cutoff:
                    continue
                if m not in monthly:
                    monthly[m] = {"label": f"{m[:4]}년 {int(m[4:6])}월",
                                  "total": 0, "on_time": 0, "late": 0, "no_in": 0, "no_out": 0,
                                  "total_work": 0, "total_overtime": 0,
                                  "sat_days": 0, "sat_work": 0, "sat_over": 0,
                                  "sun_days": 0, "sun_work": 0, "sun_over": 0,
                                  "hol_days": 0, "hol_work": 0, "hol_over": 0,
                                  "in_secs_normal": [], "in_secs_late": [],
                                  "out_secs_normal": [], "out_secs_late": []}
                ms = monthly[m]
                ms["total"] += 1
                ms["total_work"] += r["work_hours"]
                ms["total_overtime"] += r.get("ot_hours", 2 if r["overtime"] == "2" else 0)
                if r["weekday"] == 5:
                    ms["sat_days"] += 1
                    ms["sat_work"] += min(r["work_hours"], 8)
                    ms["sat_over"] += max(r["work_hours"] - 8, 0)
                elif r["weekday"] == 6:
                    ms["sun_days"] += 1
                    ms["sun_work"] += min(r["work_hours"], 8)
                    ms["sun_over"] += max(r["work_hours"] - 8, 0)
                if r.get("holiday"):
                    ms["hol_days"] += 1
                    ms["hol_work"] += min(r["work_hours"], 8)
                    ms["hol_over"] += max(r["work_hours"] - 8, 0)
                if r["in_time"] and r["in_time"] <= "115900":
                    ms["on_time"] += 1
                    stats["on_time"] += 1
                elif r["in_time"]:
                    ms["late"] += 1
                    stats["late"] += 1
                else:
                    ms["no_in"] += 1
                    stats["no_in"] += 1
                if not r["out_time"]:
                    ms["no_out"] += 1
                    stats["no_out"] += 1
                in_s = _time_to_sec(r["in_time"])
                out_s = _time_to_sec(r["out_time"])
                if r["in_time"] and r["in_time"] <= "115900":
                    if in_s is not None:
                        ms["in_secs_normal"].append(in_s)
                        all_in_secs_normal.append(in_s)
                    if out_s is not None:
                        ms["out_secs_normal"].append(out_s)
                        all_out_secs_normal.append(out_s)
                elif r["in_time"]:
                    if in_s is not None:
                        ms["in_secs_late"].append(in_s)
                        all_in_secs_late.append(in_s)
                    if out_s is not None:
                        ms["out_secs_late"].append(out_s)
                        all_out_secs_late.append(out_s)
            for ms in monthly.values():
                ms["avg_in_normal"] = _sec_to_fmt(sum(ms["in_secs_normal"]) / len(ms["in_secs_normal"])) if len(ms["in_secs_normal"]) >= 5 else "-"
                ms["avg_in_late"] = _sec_to_fmt(sum(ms["in_secs_late"]) / len(ms["in_secs_late"])) if len(ms["in_secs_late"]) >= 5 else "-"
                ms["avg_out_normal"] = _sec_to_fmt(sum(ms["out_secs_normal"]) / len(ms["out_secs_normal"])) if len(ms["out_secs_normal"]) >= 5 else "-"
                ms["avg_out_late"] = _sec_to_fmt(sum(ms["out_secs_late"]) / len(ms["out_secs_late"])) if len(ms["out_secs_late"]) >= 5 else "-"
            stats["total"] = len(records)
            stats["total_work"] = sum(ms["total_work"] for ms in monthly.values())
            stats["total_overtime"] = sum(ms["total_overtime"] for ms in monthly.values())
            stats["has_both"] = stats["on_time"] > 0 and stats["late"] > 0
            stats["avg_in_normal"] = _sec_to_fmt(sum(all_in_secs_normal) / len(all_in_secs_normal)) if len(all_in_secs_normal) >= 5 else "-"
            stats["avg_in_late"] = _sec_to_fmt(sum(all_in_secs_late) / len(all_in_secs_late)) if len(all_in_secs_late) >= 5 else "-"
            stats["avg_out_normal"] = _sec_to_fmt(sum(all_out_secs_normal) / len(all_out_secs_normal)) if len(all_out_secs_normal) >= 5 else "-"
            stats["avg_out_late"] = _sec_to_fmt(sum(all_out_secs_late) / len(all_out_secs_late)) if len(all_out_secs_late) >= 5 else "-"
        else:
            # ?쒕뵫 ?섏씠吏: TOP 10 일궧 (理쒓렐 1媛쒖썡)
            rank_start = (datetime.now() - timedelta(days=35)).strftime("%Y%m%d")
            cur.execute("""
                SELECT t.e_date, t.e_time, t.e_id, t.e_idno, t.e_name,
                       u.company
                FROM tenter t LEFT JOIN tuser u ON t.e_id = u.id
                WHERE t.e_id >= 0 AND t.e_mode = %s AND t.e_date >= %s
                ORDER BY t.e_date, t.e_time
            """, (MODE_IN, rank_start))
            all_rows = cur.fetchall()
            day_emp_in = {}
            emp_info = {}
            for r in all_rows:
                e_date, e_time, eid = r[0], r[1], r[2]
                key = (e_date, eid)
                if key not in day_emp_in or e_time < day_emp_in[key]:
                    day_emp_in[key] = e_time
                if eid not in emp_info:
                    company = (r[5] or "").strip()
                    emp_info[eid] = {
                        "name": (r[4] or "").strip() or f"ID-{eid}",
                        "emp_no": (r[3] or "").strip() or str(eid),
                        "dept": DEPT_MAP.get(company, "")
                    }
            emp_month = defaultdict(list)
            for (e_date, eid), in_time in day_emp_in.items():
                month = e_date[:6]
                in_s = _time_to_sec(in_time)
                if in_s is not None:
                    emp_month[(month, eid)].append(in_s)
            month_data = {}
            for (month, eid), secs in emp_month.items():
                if len(secs) < 5:
                    continue
                info = emp_info[eid]
                if sel_dept and info["dept"] != sel_dept:
                    continue
                avg = sum(secs) / len(secs)
                if month not in month_data:
                    month_data[month] = []
                month_data[month].append({
                    "name": info["name"],
                    "emp_no": info["emp_no"],
                    "dept": info["dept"],
                    "avg_in": _sec_to_fmt(avg),
                    "avg_in_sec": avg,
                    "count": len(secs)
                })
            recent_months = sorted(month_data.keys(), reverse=True)[:1]
            for month in sorted(recent_months):
                month_data[month].sort(key=lambda x: x["avg_in_sec"])
                label = f"{month[:4]}년 {int(month[4:6])}월"
                ranking[label] = month_data[month][:10]
        conn.close()
    except Exception as e:
        flash(f"error: {e}", "danger")
    sel_date_holiday = ""
    if sel_date_raw and len(sel_date_raw) == 8:
        try:
            sel_date_holiday = KR_HOLIDAYS.get(datetime.strptime(sel_date_raw, "%Y%m%d").date(), "")
        except ValueError:
            pass
    return render_template("attendance.html", records=records,
                           stats=stats, monthly=monthly,
                           search=search, sel_date=sel_date, mode=mode,
                           ranking=ranking, sel_date_holiday=sel_date_holiday,
                           sel_dept=sel_dept, dept_list=dept_list,
                           sel_month=sel_month)


@app.route("/weekly52")
@_login_required
def weekly52():
    search = request.args.get("search", "").strip()
    sel_month = request.args.get("month", "").strip()
    sel_dept = request.args.get("dept", "").strip()
    is_admin = session.get("role") == "admin"
    if not is_admin:
        search = session.get("user_name", "")
    weeks = []
    emp_name = ""
    dept_list = sorted(set(DEPT_MAP.values()))
    try:
        conn = _conn()
        cur = conn.cursor()
        if search:
            if not sel_month:
                sel_month = datetime.now().strftime("%Y-%m")
            # ?좏깮일짹7일(二?寃쎄퀎 ?ы븿)
            start_dt = datetime.strptime(sel_month + "-01", "%Y-%m-%d")
            q_start = (start_dt - timedelta(days=7)).strftime("%Y%m%d")
            dim52 = calendar.monthrange(start_dt.year, start_dt.month)[1]
            end_dt = datetime(start_dt.year, start_dt.month, dim52)
            q_end = (end_dt + timedelta(days=7)).strftime("%Y%m%d")
            # 1?④퀎: 而ㅻ쾭留일몃뜳?ㅻ줈 e_id 癒쇱? 議고쉶
            cur.execute("""
                SELECT DISTINCT e_id FROM tenter
                WHERE e_name LIKE %s OR e_idno LIKE %s LIMIT 50
            """, (f"%{search}%", f"%{search}%"))
            eids = [r[0] for r in cur.fetchall()]
            if eids:
                ph = ",".join(["%s"] * len(eids))
                # 2?④퀎: ?좎쭨 踰붿쐞 ?쒗븳 議고쉶
                cur.execute(f"""
                    SELECT t.e_date, t.e_time, t.e_id, t.e_idno, t.e_name, t.e_mode
                    FROM tenter t
                    WHERE t.e_id IN ({ph}) AND t.e_date >= %s AND t.e_date <= %s
                    ORDER BY t.e_date, t.e_time
                """, eids + [q_start, q_end])
            else:
                cur.execute("SELECT 1 WHERE 1=0")  # 결과없음
            rows = cur.fetchall()
            day_emp = OrderedDict()
            for r in rows:
                e_date, e_time, eid = r[0], r[1], r[2]
                e_mode = (r[5] or "").strip()
                key = (e_date, eid)
                if key not in day_emp:
                    day_emp[key] = {
                        "date": e_date, "id": eid,
                        "name": (r[4] or "").strip() or f"ID-{eid}",
                        "in_time": None, "out_time": None,
                    }
                e = day_emp[key]
                if e_mode == MODE_IN:
                    if e["in_time"] is None or e_time < e["in_time"]:
                        e["in_time"] = e_time
                elif e_mode == MODE_OUT:
                    if e["out_time"] is None or e_time > e["out_time"]:
                        e["out_time"] = e_time
            _merge_night_shifts(day_emp)
            records = []
            for e in day_emp.values():
                if not emp_name:
                    emp_name = e["name"]
                in_s = _time_to_sec(e["in_time"])
                out_s = _time_to_sec(e["out_time"])
                is_night = in_s is not None and in_s > 43140
                work_h = 0
                over_h = 0
                if in_s is not None and out_s is not None and out_s > in_s:
                    wh = (out_s - in_s) / 3600
                    work_h = min(wh, 8)
                    over_h = 2 if out_s >= 68400 else 0
                elif is_night and out_s is not None and out_s < in_s:
                    wh = (86400 - in_s + out_s) / 3600
                    work_h = min(wh, 8)
                    over_h = 2 if out_s >= 28200 else 0  # 07:50 ?댄썑
                elif in_s is None and out_s is not None and out_s < 71400:
                    work_h = 8
                elif in_s is None and out_s is not None and out_s >= 71400:
                    work_h, over_h = 8, 2
                elif in_s is not None or out_s is not None:
                    work_h = 8
                e["work_h"] = work_h
                e["over_h"] = over_h
                records.append(e)
            # 二쇱감蹂?吏묎퀎
            week_map = OrderedDict()
            for rec in records:
                dt = datetime.strptime(rec["date"], "%Y%m%d")
                if sel_month and dt.strftime("%Y-%m") != sel_month:
                    continue
                iso = dt.isocalendar()
                wk_key = f"{iso[0]}-W{iso[1]:02d}"
                if wk_key not in week_map:
                    mon = dt - timedelta(days=dt.weekday())
                    sun = mon + timedelta(days=6)
                    week_map[wk_key] = {
                        "week_no": iso[1],
                        "start": mon.strftime("%Y/%m/%d"),
                        "end": sun.strftime("%Y/%m/%d"),
                        "days": 0, "work": 0, "overtime": 0,
                    }
                w = week_map[wk_key]
                w["days"] += 1
                w["work"] += rec["work_h"]
                w["overtime"] += rec["over_h"]
            for w in week_map.values():
                w["work"] = int(w["work"])
                w["overtime"] = int(w["overtime"])
                w["total"] = w["work"] + w["overtime"]
            weeks = list(week_map.values())
        # 寃일?놁쓣 일 일吏곸썝 二?2?쒓컙 珥덇낵일
        over52_list = []
        over52_label = ""
        if not search:
            if sel_month:
                start_dt = datetime.strptime(sel_month + "-01", "%Y-%m-%d")
                dim_ov = calendar.monthrange(start_dt.year, start_dt.month)[1]
                d_from = (start_dt - timedelta(days=7)).strftime("%Y%m%d")
                d_to = (datetime(start_dt.year, start_dt.month, dim_ov) + timedelta(days=7)).strftime("%Y%m%d")
                over52_label = f"{start_dt.year}년 {start_dt.month}월 {start_dt.day}일주"
            else:
                today = datetime.now()
                four_weeks_ago = today - timedelta(weeks=4)
                d_from = four_weeks_ago.strftime("%Y%m%d")
                d_to = today.strftime("%Y%m%d")
                over52_label = "최근 4주"
            cur.execute("""
                SELECT t.e_date, t.e_time, t.e_id, t.e_idno, t.e_name,
                       t.e_mode, u.company
                FROM tenter t
                LEFT JOIN tgate g ON t.g_id = g.id
                LEFT JOIN tuser u ON t.e_id = u.id
                WHERE t.e_id >= 0 AND t.e_date >= %s AND t.e_date <= %s
                ORDER BY t.e_date, t.e_time
            """, (d_from, d_to))
            all52 = cur.fetchall()
            emp52 = {}
            for r in all52:
                e_date, e_time, eid = r[0], r[1], r[2]
                e_mode = (r[5] or "").strip()
                key = (e_date, eid)
                if key not in emp52:
                    company = (r[6] or "").strip()
                    emp52[key] = {
                        "date": e_date, "id": eid,
                        "name": (r[4] or "").strip() or f"ID-{eid}",
                        "emp_no": (r[3] or "").strip() or str(eid),
                        "dept": DEPT_MAP.get(company, ""),
                        "in_time": None, "out_time": None,
                    }
                e = emp52[key]
                if e_mode == MODE_IN:
                    if e["in_time"] is None or e_time < e["in_time"]:
                        e["in_time"] = e_time
                elif e_mode == MODE_OUT:
                    if e["out_time"] is None or e_time > e["out_time"]:
                        e["out_time"] = e_time
            _merge_night_shifts(emp52)
            for e in emp52.values():
                in_s = _time_to_sec(e["in_time"])
                out_s = _time_to_sec(e["out_time"])
                is_n = in_s is not None and in_s > 43140
                w, o = 0, 0
                if in_s is not None and out_s is not None and out_s > in_s:
                    wh = (out_s - in_s) / 3600
                    w = min(wh, 8)
                    o = 2 if out_s >= 68400 else 0
                elif is_n and out_s is not None and out_s < in_s:
                    wh = (86400 - in_s + out_s) / 3600
                    w, o = min(wh, 8), (2 if out_s >= 28200 else 0)  # 07:50 ?댄썑
                elif in_s is None and out_s is not None and out_s < 71400:
                    w = 8
                elif in_s is None and out_s is not None and out_s >= 71400:
                    w, o = 8, 2
                elif in_s is not None or out_s is not None:
                    w = 8
                e["work_h"] = w
                e["over_h"] = o
            week_emp = {}
            for e in emp52.values():
                dt = datetime.strptime(e["date"], "%Y%m%d")
                iso = dt.isocalendar()
                wk_key = f"{iso[0]}-W{iso[1]:02d}"
                eid = e["id"]
                k = (wk_key, eid)
                if k not in week_emp:
                    mon = dt - timedelta(days=dt.weekday())
                    sun = mon + timedelta(days=6)
                    week_emp[k] = {
                        "week_no": iso[1], "name": e["name"],
                        "emp_no": e["emp_no"], "dept": e["dept"],
                        "start": mon.strftime("%m/%d"),
                        "end": sun.strftime("%m/%d"),
                        "work": 0, "overtime": 0,
                    }
                we = week_emp[k]
                we["work"] += e["work_h"]
                we["overtime"] += e["over_h"]
            for we in week_emp.values():
                total = int(we["work"]) + int(we["overtime"])
                if total >= 52:
                    if sel_dept and we["dept"] != sel_dept:
                        continue
                    we["work"] = int(we["work"])
                    we["overtime"] = int(we["overtime"])
                    we["total"] = total
                    over52_list.append(we)
            over52_list.sort(key=lambda x: -x["total"])
        conn.close()
    except Exception as e:
        flash(f"error: {e}", "danger")
    return render_template("weekly52.html", weeks=weeks,
                           search=search, sel_month=sel_month,
                           emp_name=emp_name,
                           over52_list=over52_list if 'over52_list' in dir() else [],
                           over52_label=over52_label if 'over52_label' in dir() else "",
                           sel_dept=sel_dept, dept_list=dept_list)


@app.route("/work_schedule")
@_login_required
def work_schedule():
    search = request.args.get("search", "").strip()
    sel_month = request.args.get("month", "").strip()
    sel_dept = request.args.get("dept", "").strip()
    is_admin = session.get("role") == "admin"
    can_edit = is_admin or _has_perm("schedule")
    own_dept = ""
    initial_dept_view = (
        not is_admin
        and "dept" not in request.args
        and "search" not in request.args
        and "page" not in request.args
    )
    if not is_admin and (initial_dept_view or not sel_dept):
        try:
            conn_user = _conn()
            cur_user = conn_user.cursor()
            cur_user.execute(
                "SELECT dept FROM employee_roster WHERE name = %s AND dept IS NOT NULL AND dept <> '' LIMIT 1",
                (session.get("user_name", ""),),
            )
            row_user = cur_user.fetchone()
            if row_user and row_user[0]:
                own_dept = (row_user[0] or "").strip()
            conn_user.close()
        except Exception:
            own_dept = ""
    if initial_dept_view and own_dept:
        sel_dept = own_dept
    elif not can_edit and not sel_dept:
        search = session.get("user_name", "")
    page = max(1, int(request.args.get("page", 1) or 1))
    export_salary = request.args.get("export") == "salary"
    PER_PAGE = 20
    if not sel_month:
        sel_month = datetime.now().strftime("%Y-%m")
    year = int(sel_month[:4])
    mon = int(sel_month[5:7])
    dim = calendar.monthrange(year, mon)[1]
    month_label = f"{year}년 {mon}월"
    day_names = ["월", "화", "수", "목", "금", "토", "일"]
    day_info = []
    for d in range(1, dim + 1):
        dt = datetime(year, mon, d)
        wd = dt.weekday()
        hol = KR_HOLIDAYS.get(dt.date(), "")
        day_info.append({"day": d, "weekday": wd, "wd_name": day_names[wd],
                         "is_sat": wd == 5, "is_sun": wd == 6,
                         "is_holiday": bool(hol), "holiday_name": hol})
    employees = []
    total_pages = 1
    total_emp = 0
    override_log = []
    try:
        first_dt = datetime(year, mon, 1)
        last_dt = datetime(year, mon, dim)
        q_start = (first_dt - timedelta(days=2)).strftime("%Y%m%d")
        q_end = (last_dt + timedelta(days=2)).strftime("%Y%m%d")
        conn = _conn()
        cur = conn.cursor()
        # 1?④퀎: e_id 議고쉶 (遺일?꾪꽣瑜?SQL?먯꽌 泥섎━)
        params = []
        sql = "SELECT DISTINCT t.e_id FROM tenter t"
        if sel_dept:
            sql += " LEFT JOIN employee_roster er ON t.e_name = er.name"
        sql += " WHERE t.e_date >= %s AND t.e_date <= %s"
        params += [q_start, q_end]
        if search:
            sql += " AND (t.e_name LIKE %s OR t.e_idno LIKE %s)"
            params += [f"%{search}%", f"%{search}%"]
        if sel_dept:
            sql += " AND er.dept = %s"
            params += [sel_dept]
        cur.execute(sql, params)
        all_eids = [r[0] for r in cur.fetchall()]
        total_emp = len(all_eids)
        total_pages = max(1, (total_emp + PER_PAGE - 1) // PER_PAGE)
        page = min(page, total_pages)
        eids = all_eids if export_salary else all_eids[(page - 1) * PER_PAGE: page * PER_PAGE]
        if eids:
            ph = ",".join(["%s"] * len(eids))
            # 2?④퀎: idx_eid + ?좎쭨 踰붿쐞濡?鍮좊Ⅴ寃?議고쉶
            cur.execute(f"""
                SELECT t.e_date, t.e_time, t.e_id, t.e_idno, t.e_name,
                                             t.e_mode, u.company, er.dept
                FROM tenter t
                     LEFT JOIN tgate g ON t.g_id = g.id
                     LEFT JOIN tuser u ON t.e_id = u.id
                                         LEFT JOIN employee_roster er ON t.e_name = er.name
                WHERE t.e_id IN ({ph})
                  AND t.e_date >= %s AND t.e_date <= %s
                ORDER BY t.e_date, t.e_time
            """, eids + [q_start, q_end])
        else:
            cur.execute("SELECT 1 WHERE 1=0")  # 결과없음
        rows = cur.fetchall()
        day_emp = OrderedDict()
        emp_info = {}
        for r in rows:
            e_date, e_time, eid = r[0], r[1], r[2]
            e_mode = (r[5] or "").strip()
            key = (e_date, eid)
            if key not in day_emp:
                company = (r[6] or "").strip()
                roster_dept = (r[7] or "").strip()
                day_emp[key] = {
                    "date": e_date, "id": eid,
                    "name": (r[4] or "").strip() or f"ID-{eid}",
                    "emp_no": (r[3] or "").strip() or str(eid),
                    "dept": roster_dept or DEPT_MAP.get(company, ""),
                    "in_time": None, "out_time": None,
                }
                if eid not in emp_info:
                    emp_info[eid] = {
                        "name": day_emp[key]["name"],
                        "emp_no": day_emp[key]["emp_no"],
                        "dept": day_emp[key]["dept"],
                    }
            e = day_emp[key]
            if e_mode == MODE_IN:
                if e["in_time"] is None or e_time < e["in_time"]:
                    e["in_time"] = e_time
            elif e_mode == MODE_OUT:
                if e["out_time"] is None or e_time > e["out_time"]:
                    e["out_time"] = e_time
        _merge_night_shifts(day_emp)
        month_prefix = f"{year}{mon:02d}"
        emp_days = defaultdict(dict)
        for e in day_emp.values():
            if not e["date"].startswith(month_prefix):
                continue
            eid = e["id"]
            day_num = int(e["date"][6:8])
            in_s = _time_to_sec(e["in_time"])
            out_s = _time_to_sec(e["out_time"])
            is_night = in_s is not None and in_s > 43140
            dt2 = datetime.strptime(e["date"], "%Y%m%d")
            wd = dt2.weekday()
            is_hol = wd >= 5 or bool(KR_HOLIDAYS.get(dt2.date(), ""))
            basic, ot_h, ni_h, etc_h = 0, 0, 0, 0
            if is_night:
                if out_s is not None and out_s < in_s:
                    wh = (86400 - in_s + out_s) / 3600
                    basic = int(min(wh, 8))
                    ot_h = 2 if out_s >= 28200 else 0
                    night_sec = max(0, 86400 - max(in_s, 79200)) + max(0, min(out_s, 21600))
                    ni_h = max(0, round(night_sec / 3600 - 1))
                elif out_s is not None and out_s > in_s:
                    pass
                elif in_s is not None:
                    basic = 8
                    assumed_out = 18000
                    night_sec = max(0, 86400 - max(in_s, 79200)) + max(0, min(assumed_out, 21600))
                    ni_h = max(0, round(night_sec / 3600 - 1))
            else:
                if in_s is not None and out_s is not None and out_s > in_s:
                    wh = (out_s - in_s) / 3600
                    basic = int(min(wh, 8))
                    ot_h = 2 if out_s >= 68400 else 0
                elif in_s is None and out_s is not None and out_s < 28800:
                    assumed_in = 68400
                    wh = (86400 - assumed_in + out_s) / 3600
                    basic = int(min(wh, 8))
                    ot_h = 2 if out_s >= 28200 else 0
                    night_sec = max(0, 86400 - max(assumed_in, 79200)) + max(0, min(out_s, 21600))
                    ni_h = max(0, round(night_sec / 3600 - 1))
                elif in_s is None and out_s is not None and out_s >= 71400:
                    basic = 8
                    ot_h = 2
                elif in_s is None and out_s is not None:
                    basic = 8
                elif in_s is not None or out_s is not None:
                    basic = 8
                    if is_night and out_s is None:
                        assumed_out = 18000
                        night_sec = max(0, 86400 - max(in_s, 79200)) + max(0, min(assumed_out, 21600))
                        ni_h = max(0, round(night_sec / 3600 - 1))
            emp_days[eid][day_num] = {
                "basic": basic, "overtime": ot_h,
                "night": ni_h, "other": etc_h,
                "is_holiday": is_hol,
            }
        # ?ㅻ쾭?쇱씠일濡쒕뵫
        overrides = {}
        override_memos = {}
        override_updates = {}
        if eids:
            ph2 = ",".join(["%s"] * len(eids))
            month_str = f"{year}{mon:02d}"
            cur.execute(f"""
                SELECT e_id, e_date, col_type, value, memo, updated_at FROM work_override
                WHERE e_id IN ({ph2}) AND e_date LIKE %s
            """, eids + [f"{month_str}%"])
            for row in cur.fetchall():
                v = row[3]
                try:
                    v = float(v)
                except (ValueError, TypeError):
                    pass
                key3 = (row[0], row[1], row[2])
                overrides[key3] = v
                if row[4]:
                    override_memos[key3] = row[4]
                if row[5]:
                    override_updates[key3] = row[5].strftime("%m/%d %H:%M") if hasattr(row[5], 'strftime') else str(row[5])
        col_label = {"basic": "기본", "overtime": "연장", "night": "야간", "other": "기타"}
        override_log = []
        for eid, info in emp_info.items():
            dd = emp_days.get(eid, {})
            if not dd:
                continue
            b_r, o_r, n_r, e_r = [], [], [], []
            b_ov, o_ov, n_ov, e_ov = [], [], [], []
            b_memo, o_memo, n_memo, e_memo = [], [], [], []
            s_ot = s_ni = wd_ot = nw = hw = ho = 0
            unpaid = 0
            hol_basic = hol_ot_h = hol_night = 0
            hol_time = 0
            for d in range(1, dim + 1):
                rec = dd.get(d)
                date_str = f"{year}{mon:02d}{d:02d}"
                is_hol = rec["is_holiday"] if rec else day_info[d-1]["is_sat"] or day_info[d-1]["is_sun"] or day_info[d-1]["is_holiday"]
                bv = overrides.get((eid, date_str, "basic"),   rec["basic"]   if rec else None)
                ov = overrides.get((eid, date_str, "overtime"),rec["overtime"]if rec else None)
                nv = overrides.get((eid, date_str, "night"),   rec["night"]   if rec else None)
                ev = overrides.get((eid, date_str, "other"),   rec["other"]   if rec else None)
                b_r.append(bv); o_r.append(ov); n_r.append(nv); e_r.append(ev)
                b_ov.append((eid, date_str, "basic")    in overrides)
                o_ov.append((eid, date_str, "overtime") in overrides)
                n_ov.append((eid, date_str, "night")    in overrides)
                e_ov.append((eid, date_str, "other")    in overrides)
                b_memo.append(override_memos.get((eid, date_str, "basic"), ""))
                o_memo.append(override_memos.get((eid, date_str, "overtime"), ""))
                n_memo.append(override_memos.get((eid, date_str, "night"), ""))
                e_memo.append(override_memos.get((eid, date_str, "other"), ""))
                for ct in ["basic", "overtime", "night", "other"]:
                    k3 = (eid, date_str, ct)
                    if k3 in overrides and k3 in override_memos:
                        override_log.append({
                            "name": info["name"], "date": f"{d}일",
                            "col": col_label[ct], "value": overrides[k3],
                            "memo": override_memos[k3],
                            "updated": override_updates.get(k3, ""),
                        })
                if isinstance(ev, str) and "무급" in ev:
                    unpaid += 1
                    hol_time += 8
                elif isinstance(ev, str) and "연차" in ev:
                    pass
                elif isinstance(ev, (int, float)) and ev > 0:
                    hol_time += ev
                bv = bv or 0; ov = ov or 0; nv = nv or 0
                s_ot += ov; s_ni += nv; nw += nv
                # 휴일洹쇰줈: ?졖룹씪?붿씪+怨듯쑕?쇱쓽 기본?쒓컙留?
                is_hol_strict = day_info[d-1]["is_sat"] or day_info[d-1]["is_sun"] or day_info[d-1]["is_holiday"]
                if is_hol_strict:
                    hw += bv
                    ho += ov
                if is_hol:
                    hol_basic += bv; hol_ot_h += ov; hol_night += nv
                if not is_hol:
                    wd_ot += ov
            # 합산일override ?곸슜 (e_date = YYYYMM00)
            sum_date = f"{year}{mon:02d}00"
            final_sum_ot    = overrides.get((eid, sum_date, "sum_ot"),    0)
            final_sum_ni    = overrides.get((eid, sum_date, "sum_night"), 0)
            final_hol_work  = overrides.get((eid, sum_date, "hol_work"), 0)
            final_hol_ot    = overrides.get((eid, sum_date, "hol_ot"),   0)
            # ?レ옄 蹂일
            try: final_sum_ot   = float(final_sum_ot)
            except: final_sum_ot = 0
            try: final_sum_ni   = float(final_sum_ni)
            except: final_sum_ni = 0
            try: final_hol_work = float(final_hol_work)
            except: final_hol_work = 0
            try: final_hol_ot   = float(final_hol_ot)
            except: final_hol_ot = 0
            # 합산 override 硫붾え
            sum_ov = {}
            sum_memo_map = {}
            for sc in ["sum_ot", "sum_night", "hol_work", "hol_ot"]:
                sk = (eid, sum_date, sc)
                sum_ov[sc] = sk in overrides
                sum_memo_map[sc] = override_memos.get(sk, "")
                if sk in overrides and sk in override_memos:
                    col_kr = {"sum_ot":"연장","sum_night":"야간","hol_work":"휴일","hol_ot":"휴연"}
                    override_log.append({
                        "name": info["name"], "date": "합산",
                        "col": col_kr[sc], "value": overrides[sk],
                        "memo": override_memos[sk],
                        "updated": override_updates.get(sk, ""),
                    })
            employees.append({
                "id": eid,
                "emp_no": info["emp_no"], "dept": info["dept"], "name": info["name"],
                "basic": b_r, "overtime": o_r, "night": n_r, "other": e_r,
                "basic_ov": b_ov, "overtime_ov": o_ov, "night_ov": n_ov, "other_ov": e_ov,
                "basic_memo": b_memo, "overtime_memo": o_memo, "night_memo": n_memo, "other_memo": e_memo,
                "sum_ot": int(final_sum_ot), "sum_night": int(final_sum_ni),
                "hol_work": int(final_hol_work), "hol_ot": int(final_hol_ot),
                "wd_ot": int(wd_ot + final_sum_ot), "night_work": int(nw + final_sum_ni),
                "calc_hw": int(hw + final_hol_work), "calc_ho": int(ho + final_hol_ot),
                "unpaid": unpaid, "hol_time": int(hol_time),
                "hol_basic": int(hol_basic), "hol_ot_h": int(hol_ot_h),
                "hol_night": int(hol_night), "hol_total": int(hol_basic + hol_ot_h + hol_night),
                "sum_ov": sum_ov, "sum_memo": sum_memo_map,
            })
        conn.close()
    except Exception as ex:
        flash(f"error: {ex}", "danger")
    try:
        conn2 = _conn()
        cur2 = conn2.cursor()
        cur2.execute("SELECT DISTINCT dept FROM employee_roster WHERE dept IS NOT NULL AND dept <> '' ORDER BY dept")
        dept_list = [r[0] for r in cur2.fetchall()]
        conn2.close()
    except Exception:
        dept_list = sorted(set(DEPT_MAP.values()))
    # 급여출력폼Excel ?대낫?닿린
    if export_salary and employees:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        wbx = Workbook()
        wsx = wbx.active
        wsx.title = "급여출력폼"
        # R1: ?쒕ぉ (蹂묓빀)
        wsx.merge_cells("A1:F1")
        wsx["A1"].value = f"{month_label} 급여출력폼"
        wsx["A1"].font = Font(bold=True, size=14)
        wsx["A1"].alignment = Alignment(horizontal="center")
        # R2: ?ㅻ뜑
        headers = ["?ъ썝", "연장洹쇰줈?쒓컙", "야간洹쇰줈?쒓컙", "휴일洹쇰줈?쒓컙", "휴일연장洹쇰줈?쒓컙", "?몄텧議고눜?쒓컙"]
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for ci, h in enumerate(headers, 1):
            c = wsx.cell(2, ci, h)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")
            c.border = border
        # R3~: ?곗씠일
        for ri, emp in enumerate(employees, 3):
            vals = [emp["name"], emp["wd_ot"], emp["night_work"],
                    emp["calc_hw"], emp["calc_ho"], emp["hol_time"]]
            for ci, v in enumerate(vals, 1):
                c = wsx.cell(ri, ci, v)
                c.border = border
                if ci >= 2:
                    c.alignment = Alignment(horizontal="center")
        # 일?덈퉬
        wsx.column_dimensions["A"].width = 12
        for col in ["B", "C", "D", "E", "F"]:
            wsx.column_dimensions[col].width = 16
        buf = io.BytesIO()
        wbx.save(buf)
        buf.seek(0)
        fname = f"급여출력폼_{year}{mon:02d}.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return render_template("work_schedule.html",
                           search=search, sel_month=sel_month,
                           sel_dept=sel_dept, dept_list=dept_list,
                           month_label=month_label,
                           days_in_month=dim, day_info=day_info,
                           employees=employees,
                           page=page, total_pages=total_pages,
                           total_emp=total_emp,
                           override_log=override_log,
                           can_edit=can_edit)


@app.route("/save_override", methods=["POST"])
@_login_required
def save_override():
    if not _has_perm("schedule"):
        return jsonify({"ok": False, "error": "권한이 없습니다"}), 403
    try:
        data = request.get_json()
        e_id     = int(data["e_id"])
        e_date   = str(data["e_date"])
        col_type = str(data["col_type"])
        value    = data.get("value", "")
        memo     = data.get("memo", "").strip()
        conn = _conn(); cur = conn.cursor()
        if value == "" or value is None:
            cur.execute("DELETE FROM work_override WHERE e_id=%s AND e_date=%s AND col_type=%s",
                        (e_id, e_date, col_type))
        else:
            if not memo:
                return jsonify({"ok": False, "error": "메모를 입력하세요."}), 400
            v = value
            if col_type != 'other':
                v = str(float(value))
            cur.execute("""
                INSERT INTO work_override (e_id, e_date, col_type, value, memo)
                VALUES (%s,%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE value=%s, memo=%s, updated_at=NOW()
            """, (e_id, e_date, col_type, v, memo, v, memo))
        # 기타 而щ읆 연차/반차 일leave_records ?먮룞 ?곕룞
        if col_type == "other":
            _sync_leave_from_other(cur, e_id, e_date, value)
        conn.commit(); conn.close()
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/schedule_record")
@_login_required
def schedule_record():
    if session.get("role") != "admin" and not _has_perm("schedule"):
        flash("접근 권한이 없습니다.", "danger")
        return redirect(url_for("dashboard"))
    sel_month = request.args.get("month", "").strip()
    search = request.args.get("search", "").strip()
    if not sel_month:
        sel_month = datetime.now().strftime("%Y-%m")
    year = int(sel_month[:4])
    mon = int(sel_month[5:7])
    dim = calendar.monthrange(year, mon)[1]
    month_label = f"{year}년 {mon}월"
    day_names = ["월", "화", "수", "목", "금", "토", "일"]
    day_info = []
    for d in range(1, dim + 1):
        dt = datetime(year, mon, d)
        wd = dt.weekday()
        hol = KR_HOLIDAYS.get(dt.date(), "")
        day_info.append({"day": d, "weekday": wd, "wd_name": day_names[wd],
                         "is_sat": wd == 5, "is_sun": wd == 6,
                         "is_holiday": bool(hol), "holiday_name": hol})
    employees = []
    try:
        conn = _conn(); cur = conn.cursor()
        month_str = f"{year}{mon:02d}"
        # 수정 ?곗꽑, 원본 fallback 일전자 ?쒗듃留?
        cur.execute("""
            SELECT emp_name, work_date, basic_h, ot_h, night_h, etc, source_type
            FROM schedule_record
            WHERE work_date LIKE %s AND sheet_name='전자'
            ORDER BY emp_name, work_date
        """, (f"{month_str}%",))
        raw = cur.fetchall()
        # tuser 留ㅽ븨 (?대쫫?믪궗踰?遺일
        cur.execute("SELECT id, name, idno, company FROM tuser")
        tuser_map = {}
        for uid, uname, uidno, ucomp in cur.fetchall():
            n = (uname or "").strip()
            if not n:
                continue
            eno = (uidno or "").strip()
            if n not in tuser_map or (eno and not tuser_map[n]["emp_no_raw"]):
                tuser_map[n] = {"emp_no": eno or str(uid), "emp_no_raw": eno, "dept": DEPT_MAP.get((ucomp or "").strip(), "")}
        conn.close()
        # 수정 ?곗꽑 蹂묓빀
        merged = {}
        for name, wdate, bh, oh, nh, etc, src in raw:
            key = (name, wdate)
            if key not in merged or src == '수정':
                merged[key] = {"basic": bh, "ot": oh, "night": nh, "etc": etc, "src": src}
        # 吏곸썝蹂?洹몃９일
        emp_data = OrderedDict()
        for (name, wdate), vals in merged.items():
            if search and search not in name:
                continue
            if name not in emp_data:
                info = tuser_map.get(name, {"emp_no": "", "dept": ""})
                emp_data[name] = {"name": name, "emp_no": info["emp_no"], "dept": info["dept"], "days": {}}
            day_num = int(wdate[6:8])
            emp_data[name]["days"][day_num] = vals
        for idx, (name, ed) in enumerate(emp_data.items()):
            dd = ed["days"]
            b_r, o_r, n_r, e_r = [], [], [], []
            s_ot = s_ni = wd_ot = nw = hw = ho = 0
            unpaid = 0
            hol_basic = hol_ot_h = hol_night = hol_time = 0
            for d in range(1, dim + 1):
                rec = dd.get(d)
                is_hol = day_info[d-1]["is_sat"] or day_info[d-1]["is_sun"] or day_info[d-1]["is_holiday"]
                if rec:
                    bv, ov, nv, ev = rec["basic"], rec["ot"], rec["night"], rec["etc"]
                else:
                    bv, ov, nv, ev = None, None, None, None
                b_r.append(bv); o_r.append(ov); n_r.append(nv); e_r.append(ev)
                bval = bv or 0; oval = ov or 0; nval = nv or 0
                s_ot += oval; s_ni += nval; nw += nval
                is_hol_strict = day_info[d-1]["is_sat"] or day_info[d-1]["is_sun"] or day_info[d-1]["is_holiday"]
                if is_hol_strict:
                    hw += bval; ho += oval
                if is_hol:
                    hol_basic += bval; hol_ot_h += oval; hol_night += nval
                if not is_hol:
                    wd_ot += oval
                if isinstance(ev, str) and "무급" in ev:
                    unpaid += 1; hol_time += 8
                elif isinstance(ev, str) and "연차" in ev:
                    pass
                elif ev is not None:
                    try:
                        nv_etc = int(float(str(ev)))
                        if nv_etc > 0:
                            hol_time += nv_etc
                    except (ValueError, TypeError):
                        pass
            employees.append({
                "emp_no": ed["emp_no"], "dept": ed["dept"], "name": name,
                "basic": b_r, "overtime": o_r, "night": n_r, "other": e_r,
                "wd_ot": int(wd_ot), "night_work": int(nw),
                "calc_hw": int(hw), "calc_ho": int(ho),
                "unpaid": unpaid, "hol_time": int(hol_time),
                "hol_basic": int(hol_basic), "hol_ot_h": int(hol_ot_h),
                "hol_night": int(hol_night), "hol_total": int(hol_basic + hol_ot_h + hol_night),
            })
    except Exception as ex:
        flash(f"오류: {ex}", "danger")
    return render_template("schedule_record.html",
                           search=search, sel_month=sel_month,
                           month_label=month_label,
                           days_in_month=dim, day_info=day_info,
                           employees=employees)


@app.route("/upload_schedule_record", methods=["POST"])
@_login_required
def upload_schedule_record():
    if session.get("role") != "admin" and not _has_perm("schedule"):
        flash("접근 권한이 없습니다.", "danger")
        return redirect(url_for("dashboard"))
    file = request.files.get("file")
    year = request.form.get("year", "")
    if not file or not file.filename:
        flash("파일을 선택하세요.", "warning")
        return redirect(url_for("schedule_record"))
    if not year:
        year = datetime.now().year
    year = int(year)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file, data_only=True)
        # 전자 ?쒗듃 李얘린
        target_sheet = None
        for sn in wb.sheetnames:
            if "전자" in sn:
                target_sheet = sn
                break
        if not target_sheet:
            flash("'전자' 시트를 찾을 수 없습니다.", "danger")
            return redirect(url_for("schedule_record"))
        ws = wb[target_sheet]
        # R1: ?좎쭨 ?뚯떛 (醫뚯륫/?곗륫)
        blocks = []
        r1_left = ws.cell(1, 1).value or ""
        r1_right = ws.cell(1, 8).value or ""
        for title, cols in [(r1_left, "left"), (r1_right, "right")]:
            if not title:
                continue
            if "(수정)" not in str(title):
                continue
            m = re.search(r'(\d{2})/(\d{2})', str(title))
            if not m:
                continue
            mm, dd = int(m.group(1)), int(m.group(2))
            wdate = f"{year}{mm:02d}{dd:02d}"
            src = "수정" if "(수정)" in str(title) else "원본"
            blocks.append({"wdate": wdate, "src": src, "side": cols})
        if not blocks:
            flash("R1에서 날짜를 파싱할 수 없습니다.", "danger")
            return redirect(url_for("schedule_record"))
        # R4~ ?곗씠일?뚯떛
        records = []
        for blk in blocks:
            wdate, src, side = blk["wdate"], blk["src"], blk["side"]
            if side == "left":
                col_map = {1: "주간기본", 3: "주간연장", 4: "야간기본", 5: "야간연장", 6: "휴무"}
                ded_col = 2
                leave_col = 7
            else:
                col_map = {8: "주간기본", 9: "주간연장", 10: "야간기본", 11: "야간연장", 12: "휴무"}
                ded_col = None
                leave_col = None
            for r in range(4, ws.max_row + 1):
                for col_idx, cat in col_map.items():
                    name = ws.cell(r, col_idx).value
                    if not name or not isinstance(name, str):
                        continue
                    name = name.strip()
                    if not name:
                        continue
                    basic, ot, night, etc = 0, 0, 0, None
                    if cat == "주간기본":
                        basic = 8
                        if ded_col:
                            ded = ws.cell(r, ded_col).value
                            if ded is not None:
                                try:
                                    dv = int(float(str(ded)))
                                    if dv > 0:
                                        etc = str(dv)
                                except (ValueError, TypeError):
                                    pass
                    elif cat == "주간연장":
                        basic, ot = 8, 2
                    elif cat == "야간기본":
                        basic, night = 8, 6
                    elif cat == "야간연장":
                        basic, ot, night = 8, 2, 7
                    elif cat == "휴무":
                        if leave_col:
                            lt = ws.cell(r, leave_col).value
                            if lt and isinstance(lt, str) and lt.strip():
                                etc = lt.strip()
                    records.append((name, wdate, basic, ot, night, etc, src, "전자",
                                    session.get("user_name", "")))
        if not records:
            flash("파싱된 데이터가 없습니다.", "warning")
            return redirect(url_for("schedule_record"))
        conn = _conn(); cur = conn.cursor()
        inserted = 0
        for rec in records:
            cur.execute("""
                INSERT INTO schedule_record (emp_name, work_date, basic_h, ot_h, night_h, etc,
                                             source_type, sheet_name, uploaded_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE basic_h=%s, ot_h=%s, night_h=%s, etc=%s,
                                        uploaded_at=NOW(), uploaded_by=%s
            """, (rec[0], rec[1], rec[2], rec[3], rec[4], rec[5], rec[6], rec[7], rec[8],
                  rec[2], rec[3], rec[4], rec[5], rec[8]))
            inserted += 1
        conn.commit(); conn.close()
        dates_str = ", ".join(sorted(set(r[1] for r in records)))
        flash(f"일{inserted}건 업로드 완료 (날짜: {dates_str})", "success")
        # ?낅줈?쒗븳 ?щ줈 ?대룞
        first_date = records[0][1]
        redir_month = f"{first_date[:4]}-{first_date[4:6]}"
        return redirect(url_for("schedule_record", month=redir_month))
    except Exception as ex:
        flash(f"업로드 오류: {ex}", "danger")
        return redirect(url_for("schedule_record"))


@app.route("/annual_leave")
@_login_required
def annual_leave():
    search = request.args.get("search", "").strip()
    sel_month = request.args.get("month", "").strip()
    sel_dept = request.args.get("dept", "").strip()
    is_admin = session.get("role") == "admin"
    if not is_admin:
        search = session.get("user_name", "")
    if not sel_month:
        sel_month = datetime.now().strftime("%Y-%m")
    try:
        ym = sel_month.split("-")
        year_int, month_int = int(ym[0]), int(ym[1])
    except (ValueError, IndexError):
        year_int, month_int = datetime.now().year, datetime.now().month
        sel_month = f"{year_int:04d}-{month_int:02d}"
    dept_list = sorted(set(DEPT_MAP.values()))
    # ?щ젰 ?뺣낫
    _, dim = calendar.monthrange(year_int, month_int)
    first_wd = calendar.weekday(year_int, month_int, 1)  # 0=Mon
    day_info = []
    for d in range(1, dim + 1):
        dt = datetime(year_int, month_int, d)
        ds = dt.strftime("%Y%m%d")
        wd = dt.weekday()
        hol_name = KR_HOLIDAYS.get(dt.date(), "")
        day_info.append({
            "day": d, "ds": ds, "wd": wd,
            "is_sat": wd == 5, "is_sun": wd == 6,
            "is_holiday": bool(hol_name), "hol_name": hol_name,
        })
    conn = _conn(); cur = conn.cursor()
    # 吏곸썝 紐⑸줉 (annual_leave일?깅줉일吏곸썝留?
    sql_u = """SELECT t.id, t.name, t.company FROM tuser t
               INNER JOIN annual_leave a ON t.id = a.e_id AND a.year = %s
               WHERE t.name IS NOT NULL AND t.name <> ''"""
    params_u = [year_int]
    if sel_dept:
        dept_codes = [k for k, v in DEPT_MAP.items() if v == sel_dept]
        if dept_codes:
            ph = ",".join(["%s"] * len(dept_codes))
            sql_u += f" AND t.company IN ({ph})"
            params_u += dept_codes
    if search:
        sql_u += " AND (t.name LIKE %s OR CAST(t.id AS CHAR) LIKE %s OR t.idno LIKE %s)"
        params_u += [f"%{search}%", f"%{search}%", f"%{search}%"]
    sql_u += " ORDER BY t.company, t.name"
    cur.execute(sql_u, params_u)
    all_users = cur.fetchall()
    total_emp = len(all_users)
    users = all_users
    user_ids = [u[0] for u in users]
    # 연차 ?붿빟
    cur.execute("SELECT e_id, total, used, memo, deduct_prev, generated, m1,m2,m3,m4,m5,m6,m7,m8,m9,m10,m11,m12 FROM annual_leave WHERE year=%s", (year_int,))
    summary_map = {}
    for r in cur.fetchall():
        summary_map[r[0]] = {"total": r[1], "used": r[2], "memo": r[3],
                             "deduct_prev": r[4], "generated": r[5],
                             "months": list(r[6:18])}
    # ?대떦일媛쒕퀎 湲곕줉
    m_start = f"{year_int:04d}{month_int:02d}01"
    m_end = f"{year_int:04d}{month_int:02d}{dim:02d}"
    leave_map = {}  # e_id -> {date_str: [types]}
    year_leave_map = {}  # e_id -> total used (?곌컙)
    if user_ids:
        ph = ",".join(["%s"] * len(user_ids))
        cur.execute(f"SELECT e_id, leave_date, leave_type FROM leave_records WHERE e_id IN ({ph}) AND leave_date BETWEEN %s AND %s",
                    user_ids + [m_start, m_end])
        for eid, ld, lt in cur.fetchall():
            leave_map.setdefault(eid, {}).setdefault(ld, []).append(lt)
        # ?곌컙 ?꾩껜 ?ъ슜일怨꾩궛
        y_start = f"{year_int:04d}0101"
        y_end = f"{year_int:04d}1231"
        cur.execute(f"SELECT e_id, leave_type FROM leave_records WHERE e_id IN ({ph}) AND leave_date BETWEEN %s AND %s",
                    user_ids + [y_start, y_end])
        for eid, lt in cur.fetchall():
            year_leave_map[eid] = year_leave_map.get(eid, 0) + (0.5 if lt == "반차" else 1)
    conn.close()
    employees = []
    for uid, uname, company in users:
        if not uname:
            continue
        dept = DEPT_MAP.get((company or "").strip(), "")
        sm = summary_map.get(uid, {})
        total = sm.get("total", 15)
        used_calc = year_leave_map.get(uid, 0)
        used_db = sm.get("used", 0)
        used_final = max(used_calc, used_db)
        memo = sm.get("memo", "")
        deduct_prev = sm.get("deduct_prev", 0)
        generated = sm.get("generated", 0)
        remain = total - used_final
        emp_leaves = leave_map.get(uid, {})
        days = []
        for di in day_info:
            dl = emp_leaves.get(di["ds"], [])
            days.append(dl)
        employees.append({
            "id": uid, "name": uname, "dept": dept,
            "total": total, "used": used_final, "remain": remain, "memo": memo,
            "deduct_prev": deduct_prev, "generated": generated,
            "months": sm.get("months", [0]*12),
            "days": days,
        })
    return render_template("annual_leave.html",
                           employees=employees, sel_month=sel_month,
                           sel_dept=sel_dept, search=search,
                           dept_list=dept_list, day_info=day_info,
                           year=year_int, month=month_int, dim=dim,
                           first_wd=first_wd,
                           total_emp=total_emp,
                           can_edit=_has_perm("leave"))


@app.route("/save_leave_record", methods=["POST"])
@_login_required
def save_leave_record():
    if not _has_perm("leave"):
        return jsonify({"ok": False, "error": "권한이 없습니다"}), 403
    try:
        data = request.get_json()
        e_id = int(data["e_id"])
        leave_date = str(data["date"])
        leave_type = str(data.get("type", "연차"))
        action = str(data.get("action", "add"))
        conn = _conn(); cur = conn.cursor()
        if action == "delete":
            cur.execute("DELETE FROM leave_records WHERE e_id=%s AND leave_date=%s AND leave_type=%s",
                        (e_id, leave_date, leave_type))
        else:
            cur.execute("""
                INSERT IGNORE INTO leave_records (e_id, leave_date, leave_type)
                VALUES (%s, %s, %s)
            """, (e_id, leave_date, leave_type))
        # 洹쇰Т일기타 而щ읆?먮룄 ?먮룞 ?곕룞
        cur.execute("SELECT leave_type FROM leave_records WHERE e_id=%s AND leave_date=%s", (e_id, leave_date))
        day_leaves = [r[0] for r in cur.fetchall()]
        if day_leaves:
            other_val = "/".join(day_leaves)
            cur.execute("""
                INSERT INTO work_override (e_id, e_date, col_type, value)
                VALUES (%s,%s,'other',%s)
                ON DUPLICATE KEY UPDATE value=%s, updated_at=NOW()
            """, (e_id, leave_date, other_val, other_val))
        else:
            # 연차 湲곕줉일?놁쑝硫?기타 而щ읆?먯꽌 연차/반차 愿일override ?쒓굅
            cur.execute("SELECT value FROM work_override WHERE e_id=%s AND e_date=%s AND col_type='other'", (e_id, leave_date))
            ov_row = cur.fetchone()
            if ov_row and ov_row[0] in ("연차", "반차", "연차/반차", "반차/연차"):
                cur.execute("DELETE FROM work_override WHERE e_id=%s AND e_date=%s AND col_type='other'", (e_id, leave_date))
        # used ?먮룞 媛깆떊 + ?붾퀎(m1~m12) 媛깆떊
        year_val = int(leave_date[:4])
        cur.execute("SELECT leave_type, leave_date FROM leave_records WHERE e_id=%s AND leave_date LIKE %s",
                    (e_id, f"{year_val}%"))
        total_used = 0
        month_used = [0] * 12
        for lt, ld in cur.fetchall():
            v = 0.5 if lt == "반차" else 1
            total_used += v
            m_idx = int(ld[4:6]) - 1
            month_used[m_idx] += v
        cur.execute("SELECT name FROM tuser WHERE id=%s", (e_id,))
        row = cur.fetchone()
        e_name = row[0] if row else ""
        cur.execute("""
            INSERT INTO annual_leave (e_id, e_name, year, used, m1,m2,m3,m4,m5,m6,m7,m8,m9,m10,m11,m12)
            VALUES (%s, %s, %s, %s, %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE used=%s, e_name=%s,
                m1=%s,m2=%s,m3=%s,m4=%s,m5=%s,m6=%s,m7=%s,m8=%s,m9=%s,m10=%s,m11=%s,m12=%s,
                updated_at=NOW()
        """, (e_id, e_name, year_val, total_used,
              *month_used,
              total_used, e_name,
              *month_used))
        conn.commit(); conn.close()
        # ?붾젅洹몃옩 ?뚮┝
        d = leave_date
        disp_date = f"{d[:4]}-{d[4:6]}-{d[6:8]}" if len(d) == 8 else d
        if action == "delete":
            _send_telegram(f"🗑️ [연차삭제] {e_name} {disp_date} {leave_type} 삭제", "leave")
        else:
            _send_telegram(f"🌴 [연차등록] {e_name} {disp_date} {leave_type}", "leave")
        return jsonify({"ok": True, "used": total_used})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/save_annual_leave", methods=["POST"])
@_login_required
def save_annual_leave():
    if not _has_perm("leave"):
        return jsonify({"ok": False, "error": "권한이 없습니다"}), 403
    try:
        data = request.get_json()
        e_id = int(data["e_id"])
        year = int(data["year"])
        field = data["field"]
        value = data["value"]
        conn = _conn(); cur = conn.cursor()
        # ?대쫫 議고쉶
        cur.execute("SELECT name FROM tuser WHERE id=%s", (e_id,))
        row = cur.fetchone()
        e_name = row[0] if row else ""
        if field == "total":
            cur.execute("""
                INSERT INTO annual_leave (e_id, e_name, year, total)
                VALUES (%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE total=%s, e_name=%s, updated_at=NOW()
            """, (e_id, e_name, year, float(value), float(value), e_name))
        elif field == "used":
            cur.execute("""
                INSERT INTO annual_leave (e_id, e_name, year, used)
                VALUES (%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE used=%s, e_name=%s, updated_at=NOW()
            """, (e_id, e_name, year, float(value), float(value), e_name))
        elif field == "memo":
            cur.execute("""
                INSERT INTO annual_leave (e_id, e_name, year, memo)
                VALUES (%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE memo=%s, e_name=%s, updated_at=NOW()
            """, (e_id, e_name, year, str(value), str(value), e_name))
        else:
            return jsonify({"ok": False, "error": "invalid field"}), 400
        conn.commit(); conn.close()
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/meal_management")
@_login_required
def meal_management():
    if session.get("role") != "admin" and not _has_perm("meal"):
        flash("접근 권한이 없습니다.", "danger")
        return redirect(url_for("dashboard"))
    # 紐⑤컮일?먮룞 ?꾪솚 (view=pc ?뚮씪誘명꽣 ?놁쑝硫?
    if request.args.get("view") != "pc":
        ua = request.headers.get("User-Agent", "").lower()
        if any(k in ua for k in ["iphone", "android", "mobile", "ipod"]):
            return redirect(url_for("meal_mobile", month=request.args.get("month", datetime.now().strftime("%Y-%m"))))
    conn = _conn(); cur = conn.cursor()

    month_str = request.args.get("month", datetime.now().strftime("%Y-%m"))
    year, mon = int(month_str[:4]), int(month_str[5:7])
    days_in_month = calendar.monthrange(year, mon)[1]

    MEAL_DEPTS = ["전기", "전자", "관리"]
    MEAL_TYPES = ["조식", "석식", "야식"]

    # DB?먯꽌 ?대떦 일?곗씠일議고쉶
    cur.execute("SELECT dept, meal_type, `day`, `count`, `memo`, `writer`, updated_at FROM meal_count WHERE `year_month`=%s",
                (month_str,))
    meal_data = {}
    meal_memo_data = {}
    meal_time_data = {}
    meal_writer_data = {}
    meal_memos = []
    for dept, mtype, day, cnt, memo, writer, updated_at in cur.fetchall():
        meal_data[(dept, mtype, day)] = cnt
        time_str = updated_at.strftime("%m/%d %H:%M") if updated_at else ""
        if memo:
            meal_memo_data[(dept, mtype, day)] = memo
            meal_time_data[(dept, mtype, day)] = time_str
            meal_writer_data[(dept, mtype, day)] = writer or ""
            meal_memos.append({"dept": dept, "type": mtype, "day": day, "count": cnt, "memo": memo, "writer": writer or "", "time": time_str})
    # 怨듭? 議고쉶
    cur.execute("SELECT id, content, `writer`, created_at FROM meal_notice WHERE `year_month`=%s ORDER BY created_at DESC", (month_str,))
    meal_notices = []
    for nid, content, writer, created_at in cur.fetchall():
        t = created_at.strftime("%m/%d %H:%M") if created_at else ""
        meal_notices.append({"id": nid, "content": content, "writer": writer or "", "time": t})
    conn.close()

    # ?붿씪 / 怨듯쑕일?뺣낫
    DOW_KR = ["월", "화", "수", "목", "금", "토", "일"]
    days_info = []
    for d in range(1, days_in_month + 1):
        dt = datetime(year, mon, d)
        dow = DOW_KR[dt.weekday()]
        hol = KR_HOLIDAYS.get(dt.date(), "")
        is_holiday = (dt.weekday() >= 5) or bool(hol)
        days_info.append({"day": d, "dow": dow, "is_holiday": is_holiday, "hol": hol})

    return render_template("meal_management.html",
                           month=month_str, year=year, mon=mon,
                           days_in_month=days_in_month,
                           days_info=days_info,
                           meal_depts=MEAL_DEPTS,
                           meal_types=MEAL_TYPES,
                           meal_data=meal_data,
                           meal_memo_data=meal_memo_data,
                           meal_time_data=meal_time_data,
                           meal_writer_data=meal_writer_data,
                           meal_memos=meal_memos,
                           meal_notices=meal_notices,
                           can_edit=_has_perm("meal"))


@app.route("/meal_mobile")
@_login_required
def meal_mobile():
    if session.get("role") != "admin" and not _has_perm("meal"):
        flash("접근 권한이 없습니다.", "danger")
        return redirect(url_for("dashboard"))
    conn = _conn(); cur = conn.cursor()

    month_str = request.args.get("month", datetime.now().strftime("%Y-%m"))
    year, mon = int(month_str[:4]), int(month_str[5:7])
    days_in_month = calendar.monthrange(year, mon)[1]

    MEAL_DEPTS = ["전기", "전자", "관리"]
    MEAL_TYPES = ["조식", "석식", "야식"]

    cur.execute("SELECT dept, meal_type, `day`, `count`, `memo`, `writer`, updated_at FROM meal_count WHERE `year_month`=%s",
                (month_str,))
    meal_data = {}
    meal_memo_data = {}
    meal_time_data = {}
    meal_writer_data = {}
    for dept, mtype, day, cnt, memo, writer, updated_at in cur.fetchall():
        meal_data[(dept, mtype, day)] = cnt
        time_str = updated_at.strftime("%m/%d %H:%M") if updated_at else ""
        if memo:
            meal_memo_data[(dept, mtype, day)] = memo
            meal_time_data[(dept, mtype, day)] = time_str
            meal_writer_data[(dept, mtype, day)] = writer or ""

    cur.execute("SELECT id, content, `writer`, created_at FROM meal_notice WHERE `year_month`=%s ORDER BY created_at DESC", (month_str,))
    meal_notices = []
    for nid, content, writer, created_at in cur.fetchall():
        t = created_at.strftime("%m/%d %H:%M") if created_at else ""
        meal_notices.append({"id": nid, "content": content, "writer": writer or "", "time": t})
    conn.close()

    DOW_KR = ["월", "화", "수", "목", "금", "토", "일"]
    days_info = []
    for d in range(1, days_in_month + 1):
        dt = datetime(year, mon, d)
        dow = DOW_KR[dt.weekday()]
        hol = KR_HOLIDAYS.get(dt.date(), "")
        is_holiday = (dt.weekday() >= 5) or bool(hol)
        days_info.append({"day": d, "dow": dow, "is_holiday": is_holiday, "hol": hol})

    today_day = datetime.now().day if month_str == datetime.now().strftime("%Y-%m") else 1

    return render_template("meal_mobile.html",
                           month=month_str, year=year, mon=mon,
                           days_in_month=days_in_month,
                           days_info=days_info,
                           meal_depts=MEAL_DEPTS,
                           meal_types=MEAL_TYPES,
                           meal_data=meal_data,
                           meal_memo_data=meal_memo_data,
                           meal_time_data=meal_time_data,
                           meal_writer_data=meal_writer_data,
                           meal_notices=meal_notices,
                           today_day=today_day,
                           can_edit=_has_perm("meal"))


@app.route("/save_meal", methods=["POST"])
@_login_required
def save_meal():
    if not _has_perm("meal"):
        return jsonify({"ok": False, "error": "권한이 없습니다"}), 403
    try:
        data = request.get_json()
        year_month = str(data["year_month"])
        dept = str(data["dept"])
        meal_type = str(data["meal_type"])
        day = int(data["day"])
        count = str(data.get("count", "")).strip()
        memo = str(data.get("memo", "")).strip()
        conn = _conn(); cur = conn.cursor()
        if count == "" or count == "0":
            cur.execute("DELETE FROM meal_count WHERE `year_month`=%s AND dept=%s AND meal_type=%s AND `day`=%s",
                        (year_month, dept, meal_type, day))
            _send_telegram(f"🗑️ [식수삭제] {dept} {meal_type} {year_month}-{day:02d} (by {session.get('user_name','')})", "meal")
        else:
            writer = session.get("user_name", "")
            cur.execute("""
                INSERT INTO meal_count (`year_month`, dept, meal_type, `day`, `count`, `memo`, `writer`)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE `count`=%s, `memo`=%s, `writer`=%s, updated_at=NOW()
            """, (year_month, dept, meal_type, day, count, memo, writer, count, memo, writer))
        conn.commit(); conn.close()
        from datetime import datetime as _dt
        if count:
            memo_txt = f" 메모:{memo}" if memo else ""
            _send_telegram(f"🍽️ [식수등록] {dept} {meal_type} {year_month}-{day:02d} {count}명{memo_txt} (by {session.get('user_name','')})", "meal")
        return jsonify({"ok": True, "time": _dt.now().strftime("%m/%d %H:%M"), "writer": session.get("user_name", "")})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/save_meal_notice", methods=["POST"])
@_login_required
def save_meal_notice():
    if not _has_perm("meal"):
        return jsonify({"ok": False, "error": "권한이 없습니다"}), 403
    try:
        data = request.get_json()
        year_month = str(data["year_month"])
        content = str(data.get("content", "")).strip()
        if not content:
            return jsonify({"ok": False, "error": "내용을 입력하세요."}), 400
        writer = session.get("user_name", "")
        conn = _conn(); cur = conn.cursor()
        cur.execute("INSERT INTO meal_notice (`year_month`, content, `writer`) VALUES (%s, %s, %s)", (year_month, content, writer))
        nid = cur.lastrowid
        conn.commit(); conn.close()
        from datetime import datetime as _dt
        _send_telegram(f"📢 [식수공지] {content} (by {writer})", "notice")
        return jsonify({"ok": True, "id": nid, "time": _dt.now().strftime("%m/%d %H:%M"), "writer": writer})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/delete_meal_notice", methods=["POST"])
@_login_required
def delete_meal_notice():
    if not _has_perm("meal"):
        return jsonify({"ok": False, "error": "권한이 없습니다"}), 403
    try:
        data = request.get_json()
        nid = int(data["id"])
        conn = _conn(); cur = conn.cursor()
        cur.execute("SELECT content FROM meal_notice WHERE id=%s", (nid,))
        row = cur.fetchone()
        cur.execute("DELETE FROM meal_notice WHERE id=%s", (nid,))
        conn.commit(); conn.close()
        if row:
            _send_telegram(f"🗑️ [공지삭제] {row[0][:50]} (by {session.get('user_name','')})", "notice")
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500


@app.route("/log_management")
@_login_required
def log_management():
    if session.get("role") != "admin":
        flash("접근 권한이 없습니다.", "danger")
        return redirect(url_for("dashboard"))
    conn = _conn(); cur = conn.cursor()

    # --- ?꾪꽣 ?뚮씪誘명꽣 ---
    search   = request.args.get("search", "").strip()
    log_type = request.args.get("type", "all")          # all / override / leave / sync
    date_from = request.args.get("from", "")
    date_to   = request.args.get("to", "")

    # 湲곌컙 誘몄일일?理쒓렐 7일기본媛?
    if not date_from:
        date_from = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
    if not date_to:
        date_to = datetime.now().strftime("%Y-%m-%d")

    logs = []

    # --- work_override 濡쒓렇 ---
    if log_type in ("all", "override"):
        sql = """
            SELECT wo.e_id, u.name, wo.e_date, wo.col_type, wo.value,
                   wo.memo, wo.updated_at
            FROM work_override wo
            LEFT JOIN tuser u ON wo.e_id = u.id
            WHERE 1=1
        """
        params = []
        if search:
            sql += " AND (u.name LIKE %s OR CAST(wo.e_id AS CHAR) LIKE %s)"
            params += [f"%{search}%", f"%{search}%"]
        if date_from:
            sql += " AND wo.updated_at >= %s"
            params.append(date_from + " 00:00:00")
        if date_to:
            sql += " AND wo.updated_at <= %s"
            params.append(date_to + " 23:59:59")
        sql += " ORDER BY wo.updated_at DESC LIMIT 500"
        cur.execute(sql, params)
        col_label = {
            "basic": "기본", "overtime": "연장", "night": "야간", "other": "기타",
            "sum_ot": "전월연장", "sum_night": "전월야간",
            "hol_work": "전월휴일", "hol_ot": "전월휴연",
        }
        for eid, name, edate, ctype, val, memo, upd in cur.fetchall():
            # e_date ?щ㎎: YYYYMMDD or YYYYMM00
            dt = str(edate)
            if dt.endswith("00"):
                disp_date = f"{dt[:4]}-{dt[4:6]} 합산"
            else:
                disp_date = f"{dt[:4]}-{dt[4:6]}-{dt[6:8]}"
            logs.append({
                "type": "수정",
                "name": name or str(eid),
                "date": disp_date,
                "detail": col_label.get(ctype, ctype),
                "value": val,
                "memo": memo or "",
                "updated_at": upd,
            })

    # --- leave_records 濡쒓렇 ---
    if log_type in ("all", "leave"):
        sql = """
            SELECT lr.e_id, u.name, lr.leave_date, lr.leave_type,
                   lr.memo, lr.created_at
            FROM leave_records lr
            LEFT JOIN tuser u ON lr.e_id = u.id
            WHERE 1=1
        """
        params = []
        if search:
            sql += " AND (u.name LIKE %s OR CAST(lr.e_id AS CHAR) LIKE %s)"
            params += [f"%{search}%", f"%{search}%"]
        if date_from:
            sql += " AND lr.created_at >= %s"
            params.append(date_from + " 00:00:00")
        if date_to:
            sql += " AND lr.created_at <= %s"
            params.append(date_to + " 23:59:59")
        sql += " ORDER BY lr.created_at DESC LIMIT 500"
        cur.execute(sql, params)
        for eid, name, ldate, ltype, memo, crt in cur.fetchall():
            dt = str(ldate)
            disp_date = f"{dt[:4]}-{dt[4:6]}-{dt[6:8]}" if len(dt) == 8 else dt
            logs.append({
                "type": ltype,
                "name": name or str(eid),
                "date": disp_date,
                "detail": "연차관리",
                "value": ltype,
                "memo": memo or "",
                "updated_at": crt,
            })

    # --- sync_log 濡쒓렇 ---
    if log_type in ("all", "sync"):
        sql = "SELECT log_time, host, level, message FROM sync_log WHERE 1=1"
        params = []
        if date_from:
            sql += " AND log_time >= %s"
            params.append(date_from + " 00:00:00")
        if date_to:
            sql += " AND log_time <= %s"
            params.append(date_to + " 23:59:59")
        if search:
            sql += " AND message LIKE %s"
            params.append(f"%{search}%")
        sql += " ORDER BY id DESC LIMIT 500"
        cur.execute(sql, params)
        for lt, host, level, msg in cur.fetchall():
            logs.append({
                "type": "싱크",
                "name": host or "-",
                "date": "-",
                "detail": level,
                "value": (msg or "")[:120],
                "memo": "",
                "updated_at": lt,
            })

    # --- MDB 싱크 ?꾪솴 (일긽 議고쉶) ---
    cur.execute("SELECT COUNT(*) FROM tenter")
    sync_total = cur.fetchone()[0]
    cur.execute("SELECT MAX(e_date), MAX(e_uptime) FROM tenter")
    row = cur.fetchone()
    sync_max_date = row[0] if row else "-"
    sync_max_uptime = row[1] if row else "-"

    d7 = (datetime.now() - timedelta(days=7)).strftime("%Y%m%d")
    cur.execute("SELECT e_date, COUNT(*) FROM tenter WHERE e_date >= %s GROUP BY e_date ORDER BY e_date", (d7,))
    sync_daily = cur.fetchall()

    cur.execute("SELECT log_time, host, level, message FROM sync_log ORDER BY id DESC LIMIT 10")
    sync_recent = cur.fetchall()

    # 留덉?留일깃났 싱크
    cur.execute("SELECT log_time, message FROM sync_log WHERE level='OK' OR (level='INFO' AND message LIKE '%%sync done%%') ORDER BY id DESC LIMIT 1")
    last_ok = cur.fetchone()
    sync_last_ok_time = last_ok[0] if last_ok else None
    sync_last_ok_msg = last_ok[1] if last_ok else "-"

    # 留덉?留일먮윭
    cur.execute("SELECT log_time, message FROM sync_log WHERE level='ERROR' ORDER BY id DESC LIMIT 1")
    last_err = cur.fetchone()
    sync_last_err_time = last_err[0] if last_err else None
    sync_last_err_msg = last_err[1] if last_err else None

    conn.close()

    # ?듯빀 ?뺣젹 (理쒖떊일
    logs.sort(key=lambda x: x["updated_at"] or datetime.min, reverse=True)
    logs = logs[:500]

    return render_template("log_management.html",
                           logs=logs, search=search, log_type=log_type,
                           date_from=date_from, date_to=date_to,
                           sync_total=sync_total,
                           sync_max_date=sync_max_date,
                           sync_max_uptime=sync_max_uptime,
                           sync_daily=sync_daily,
                           sync_recent=sync_recent,
                           sync_last_ok_time=sync_last_ok_time,
                           sync_last_ok_msg=sync_last_ok_msg,
                           sync_last_err_time=sync_last_err_time,
                           sync_last_err_msg=sync_last_err_msg)


# ?? 사원명부 ??????????????????????????????????????
@app.route("/roster")
@_login_required
def roster():
    if session.get("role") != "admin":
        flash("접근 권한이 없습니다.", "danger")
        return redirect(url_for("dashboard"))
    search = request.args.get("search", "").strip()
    sel_dept = request.args.get("dept", "").strip()
    sel_pos = request.args.get("pos", "").strip()
    conn = _conn(); cur = conn.cursor()
    sql = "SELECT id, name, dept, emp_no, gender, birth_date, calendar_type, age, appoint_date, position, phone, email, address, hire_date FROM employee_roster WHERE 1=1"
    params = []
    if sel_dept:
        sql += " AND dept=%s"; params.append(sel_dept)
    if sel_pos:
        sql += " AND position=%s"; params.append(sel_pos)
    if search:
        sql += " AND (name LIKE %s OR emp_no LIKE %s OR phone LIKE %s OR position LIKE %s)"
        params += [f"%{search}%"] * 4
    sql += " ORDER BY dept, name"
    cur.execute(sql, params)
    rows = []
    for r in cur.fetchall():
        hire_date_str = r[13] or ""
        years_worked = ""
        if hire_date_str and len(hire_date_str) >= 10:
            try:
                hd = datetime.strptime(hire_date_str[:10], "%Y-%m-%d")
                diff = (datetime.now() - hd).days
                y = diff // 365
                years_worked = f"{y}년차"
            except:
                pass
        rows.append({"id": r[0], "name": r[1], "dept": r[2], "emp_no": r[3],
                      "gender": r[4], "birth_date": r[5], "calendar_type": r[6],
                      "age": r[7], "appoint_date": r[8], "position": r[9],
                      "phone": r[10], "email": r[11], "address": r[12], "hire_date": r[13],
                      "years_worked": years_worked})
    cur.execute("SELECT DISTINCT dept FROM employee_roster ORDER BY dept")
    dept_list = [r[0] for r in cur.fetchall()]
    cur.execute("SELECT DISTINCT position FROM employee_roster WHERE position!='' ORDER BY position")
    pos_list = [r[0] for r in cur.fetchall()]
    # ?대떖일?앹씪일
    from datetime import datetime as _dt
    cur_month = _dt.now().strftime("%m")
    cur.execute("""SELECT name, dept, position, birth_date, calendar_type, age, phone
                   FROM employee_roster
                   WHERE SUBSTRING(birth_date,6,2)=%s
                   ORDER BY SUBSTRING(birth_date,9,2)""", (cur_month,))
    birthday_list = [{"name":r[0],"dept":r[1],"position":r[2],"birth_date":r[3],
                      "calendar_type":r[4],"age":r[5],"phone":r[6],
                      "day":r[3][8:10] if r[3] and len(r[3])>=10 else ""} for r in cur.fetchall()]
    # ?꾧툑?쇳겕??곸옄 (留?6일?댁긽, ?댁궗/?뚯옣/?ъ옣 ?쒖쇅)
    cur_year = _dt.now().year
    cur.execute("""SELECT name, dept, position, birth_date, age, hire_date, phone
                   FROM employee_roster
                   WHERE age >= 56
                     AND position NOT LIKE '%%이사%%'
                     AND position NOT LIKE '%%회장%%'
                     AND position NOT LIKE '%%사장%%'
                   ORDER BY age DESC, name""")
    peak_list = [{"name":r[0],"dept":r[1],"position":r[2],"birth_date":r[3],
                  "age":r[4],"hire_date":r[5],"phone":r[6]} for r in cur.fetchall()]
    conn.close()
    is_admin = session.get("role") == "admin"
    return render_template("roster.html", rows=rows, dept_list=dept_list, pos_list=pos_list,
                           search=search, sel_dept=sel_dept, sel_pos=sel_pos, is_admin=is_admin,
                           birthday_list=birthday_list, peak_list=peak_list, cur_month=cur_month)


@app.route("/upload_roster", methods=["POST"])
@_admin_required
def upload_roster():
    f = request.files.get("file")
    if not f or not f.filename.endswith((".xlsx", ".xls")):
        return jsonify(ok=False, error="xlsx 파일을 선택하세요."), 400
    import openpyxl, io
    wb = openpyxl.load_workbook(io.BytesIO(f.read()))
    ws = wb.active
    # ?ㅻ뜑 일李얘린 일?ㅻ뜑 ?대쫫?쇰줈 而щ읆 ?몃뜳일?숈쟻 留ㅽ븨
    header_map = {"사원": "name", "부서": "dept", "사번": "emp_no",
                  "성별": "gender", "생년월일": "birth_date", "양력": "calendar_type",
                  "나이": "age", "발령일": "appoint_date", "직위": "position",
                  "휴대전화": "phone", "이메일": "email", "현주소": "address", "입사일": "hire_date"}
    col_map = {}
    header_row = None
    for r in range(1, min(10, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and str(v).strip() in header_map:
                col_map[header_map[str(v).strip()]] = c
        if "name" in col_map and "emp_no" in col_map:
            header_row = r
            break
        col_map.clear()
    if header_row is None or "emp_no" not in col_map:
        return jsonify(ok=False, error="헤더 행(사원,부서,사번,...)을 찾을 수 없습니다"), 400
    def _cell(r, key):
        ci = col_map.get(key)
        return ws.cell(r, ci).value if ci else None
    def _str(r, key):
        return str(_cell(r, key) or "").strip()
    def _date(r, key):
        v = _cell(r, key)
        return v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else str(v or "")
    conn = _conn(); cur = conn.cursor()
    cur.execute("DELETE FROM employee_roster")
    added = 0
    for r in range(header_row + 1, ws.max_row + 1):
        name = _str(r, "name")
        if not name:
            continue
        emp_no = _str(r, "emp_no")
        if not emp_no:
            continue
        try:
            age = int(_cell(r, "age"))
        except (ValueError, TypeError):
            age = 0
        cur.execute("""INSERT INTO employee_roster (name, dept, emp_no, gender, birth_date,
                       calendar_type, age, appoint_date, position, phone, email, address, hire_date)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (name, _str(r, "dept"), emp_no, _str(r, "gender"),
                     _date(r, "birth_date"), _str(r, "calendar_type"), age,
                     _date(r, "appoint_date"), _str(r, "position"), _str(r, "phone"),
                     _str(r, "email"), _str(r, "address"), _date(r, "hire_date")))
        added += 1
    conn.commit(); conn.close()
    return jsonify(ok=True, added=added, updated=0)


@app.route("/export_roster")
@_login_required
def export_roster():
    import openpyxl, io
    conn = _conn(); cur = conn.cursor()
    cur.execute("SELECT name, dept, emp_no, gender, birth_date, calendar_type, age, appoint_date, position, phone, email, address, hire_date FROM employee_roster ORDER BY dept, name")
    rows = cur.fetchall()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "사원명부"
    headers = ["사원", "부서", "사번", "성별", "생년월일", "양력", "나이", "발령일", "직위", "휴대전화", "이메일", "현주소", "입사일"]
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, download_name="사원명부.xlsx", as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    t = threading.Thread(target=_sync_log_checker, daemon=True)
    t.start()
    app.run(debug=False, host="0.0.0.0", port=5050)
