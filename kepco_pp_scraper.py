"""
한전 파워플래너 자동 데이터 수집 스크립트
==========================================
pp.kepco.co.kr 로그인 → 전기사용량 > 시간대별 → 엑셀 다운로드 → MariaDB 저장

사용법:
  python kepco_pp_scraper.py                      # 오늘 데이터
  python kepco_pp_scraper.py --date 20260415      # 특정 날짜
  python kepco_pp_scraper.py --month 202604       # 월 전체 (일별 반복)
  python kepco_pp_scraper.py --explore            # 로그인 후 화면 탐색 (디버그)

.env 필수 항목:
  PP_CUSTOMER_ID=고객번호(10자리)
  PP_PASSWORD=파워플래너_비밀번호
"""
import os, sys, csv, io, time, argparse
from datetime import datetime, timedelta
from pathlib import Path
from dotenv import load_dotenv
import pymysql
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

load_dotenv()

PP_URL = "https://pp.kepco.co.kr"
PP_ID  = os.environ.get("PP_CUSTOMER_ID", "")
PP_PW  = os.environ.get("PP_PASSWORD",    "")

MARIA = {
    "host":     os.environ.get("DB_HOST",     "127.0.0.1"),
    "port":     int(os.environ.get("DB_PORT", "3307")),
    "user":     os.environ.get("DB_USER",     "root"),
    "password": os.environ.get("DB_PASSWORD", "7602mr"),
    "db":       os.environ.get("DB_NAME",     "attendance"),
    "charset":  "utf8mb4",
}

# Flask API 업로드 설정 (로컬 실행 시 NAS MariaDB 대신 API POST 사용)
API_URL      = os.environ.get("POWER_API_URL",      "http://192.168.100.11:5050/api/power_upload")
BILL_API_URL = os.environ.get("POWER_BILL_API_URL", "http://192.168.100.11:5050/api/power_bill_upload")
API_TOKEN    = os.environ.get("POWER_UPLOAD_TOKEN", "")

SCREENSHOT_DIR  = Path(__file__).parent / "pp_screenshots"
DOWNLOAD_DIR    = Path(__file__).parent / "pp_downloads"
SCREENSHOT_DIR.mkdir(exist_ok=True)
DOWNLOAD_DIR.mkdir(exist_ok=True)

SMARTVIEW_API_URL = os.environ.get("POWER_SMARTVIEW_API_URL",
                                   "http://192.168.100.11:5050/api/power_smartview_upload")


# ── 로그인 ─────────────────────────────────────────────────────
def login(page):
    print("  🔐 로그인 중...")
    page.goto(PP_URL, wait_until="networkidle", timeout=30000)
    time.sleep(1.5)

    # 고객번호/ID 탭 클릭 시도
    for label in ["고객번호/ID 로그인", "고객번호", "ID 로그인"]:
        try:
            page.click(f"text={label}", timeout=2000)
            break
        except PWTimeout:
            continue

    time.sleep(0.5)

    # 고객번호 입력 (여러 셀렉터 시도)
    id_sel = [
        "input[placeholder*='고객번호']",
        "input[placeholder*='ID']",
        "input[name*='id']",
        "input[name*='custNo']",
        "input[type='text']:first-of-type",
    ]
    for sel in id_sel:
        try:
            inp = page.locator(sel).first
            inp.fill("")
            inp.fill(PP_ID)
            print(f"    고객번호 입력: {sel}")
            break
        except Exception:
            continue

    time.sleep(0.3)

    # 비밀번호 입력
    pw_sel = ["input[type='password']", "input[name*='pw']", "input[name*='pass']"]
    for sel in pw_sel:
        try:
            inp = page.locator(sel).first
            inp.fill(PP_PW)
            print(f"    비밀번호 입력: {sel}")
            break
        except Exception:
            continue

    time.sleep(0.3)

    # 로그인 버튼
    for btn in [
        "button:has-text('로그인')",
        "input[value='로그인']",
        "a:has-text('로그인')",
        ".btn-login",
        "#loginBtn",
    ]:
        try:
            page.click(btn, timeout=3000)
            break
        except PWTimeout:
            continue

    page.wait_for_load_state("networkidle", timeout=20000)
    time.sleep(2)

    # 팝업 닫기 시도
    for close in ["button:has-text('닫기')", "button:has-text('확인')", ".popup-close", ".modal-close"]:
        try:
            page.click(close, timeout=1500)
            time.sleep(0.5)
        except PWTimeout:
            pass

    if "login" in page.url.lower() and "logout" not in page.url.lower():
        shot = SCREENSHOT_DIR / f"login_fail_{datetime.now().strftime('%H%M%S')}.png"
        page.screenshot(path=str(shot))
        raise RuntimeError(f"로그인 실패 — 고객번호/비밀번호 확인\n스크린샷: {shot}")

    print(f"  ✅ 로그인 성공: {page.url}")
    return page


# 시간대별 페이지 직접 URL (탐색 모드로 확인된 실제 URL)
HOURLY_URL  = "https://pp.kepco.co.kr/rs/rs0101N.do?menu_id=O010201"
# 월별 청구요금 페이지
BILLING_URL = "https://pp.kepco.co.kr/cc/cc0102.do?menu_id=O010405"


# ── 시간대별 페이지로 이동 ──────────────────────────────────────
def go_to_hourly_page(page):
    """전기사용량 > 시간대별 페이지로 이동"""
    print("  📂 전기사용량 > 시간대별 이동...")
    try:
        page.goto(HOURLY_URL, wait_until="networkidle", timeout=15000)
        time.sleep(1.5)
        if _is_hourly_page(page):
            print(f"    ✅ 이동 성공: {page.url}")
            return True
    except PWTimeout:
        pass

    # fallback: 메뉴 클릭
    try:
        page.goto(PP_URL, wait_until="networkidle", timeout=10000)
        time.sleep(1)
        page.click("text=전기사용량", timeout=3000)
        time.sleep(0.5)
        page.click("text=시간대별", timeout=3000)
        page.wait_for_load_state("networkidle")
        time.sleep(1.5)
        if _is_hourly_page(page):
            print(f"    ✅ 메뉴 이동 성공: {page.url}")
            return True
    except PWTimeout:
        pass

    shot = SCREENSHOT_DIR / f"nav_fail_{datetime.now().strftime('%H%M%S')}.png"
    page.screenshot(path=str(shot))
    print(f"  ⚠️  시간대별 페이지 이동 실패. 스크린샷: {shot}")
    return False


def _is_hourly_page(page):
    """현재 페이지가 시간대별 페이지인지 확인"""
    try:
        content = page.content()
        return any(kw in content for kw in ["시간대별", "hourly", "시간대 별"])
    except Exception:
        return False


# ── 날짜 설정 + 조회 ───────────────────────────────────────────
def set_date_and_search(page, date_str: str):
    """
    date_str: 'YYYY-MM-DD' 형식
    날짜 입력(#SELECT_DT) → jQuery change 트리거 → getTotalData() 조회
    """
    print(f"  📅 날짜 설정: {date_str}")

    # jQuery로 날짜 설정 + change 이벤트 트리거
    page.evaluate(f"""
        if(typeof $ !== 'undefined') {{
            $('#SELECT_DT').val('{date_str}').trigger('change');
        }} else {{
            var inp = document.getElementById('SELECT_DT');
            if(inp) {{
                inp.value = '{date_str}';
                inp.dispatchEvent(new Event('change', {{bubbles:true}}));
            }}
        }}
    """)
    print(f"    날짜 설정: {date_str}")
    time.sleep(0.3)

    # 15분 라디오 확인 (이미 기본 선택됨 — 확인만)
    checked = page.evaluate("""
        () => {
            var r = document.querySelector("input[name='T_MODE'][value='15']");
            return r ? r.checked : null;
        }
    """)
    if checked is False:
        page.evaluate("document.querySelector(\"input[name='T_MODE'][value='15']\").click()")
        print("    15분 라디오 선택")
    else:
        print(f"    15분 라디오: {'이미 선택됨' if checked else '없음(기본 사용)'}")

    time.sleep(0.3)

    # 조회 버튼 클릭: a[href*="getTotalData"]
    try:
        page.click("a[href*='getTotalData']", timeout=5000)
        page.wait_for_load_state("networkidle", timeout=15000)
        time.sleep(2)
        print("    조회 완료")
        return True
    except PWTimeout:
        # fallback: JS로 직접 호출
        try:
            page.evaluate("getTotalData()")
            time.sleep(3)
            print("    조회 완료 (JS fallback)")
            return True
        except Exception as e:
            print(f"  ⚠️  조회 실패: {e}")
            return False


# ── 엑셀 다운로드 ──────────────────────────────────────────────
def download_excel(page, date_str: str):
    """엑셀 버튼 클릭 → 파일 다운로드"""
    print("  📥 엑셀 다운로드 중...")

    # 엑셀 버튼: a[href*="excelDown"] 또는 img[alt='엑셀'] 부모 클릭
    excel_selectors = [
        "a[href*='excelDown']",
        "img[alt='엑셀']",
        "a:has(img[alt='엑셀'])",
        "a.btn_blue_right:has(img[alt='엑셀'])",
    ]

    with page.expect_download(timeout=20000) as dl_info:
        downloaded = False
        for sel in excel_selectors:
            try:
                page.click(sel, timeout=3000)
                downloaded = True
                print(f"    클릭: {sel}")
                break
            except PWTimeout:
                continue

        if not downloaded:
            # JS fallback: excelDown() 직접 호출
            try:
                page.evaluate("excelDown()")
                downloaded = True
                print("    엑셀 다운로드 JS 호출")
            except Exception as e:
                shot = SCREENSHOT_DIR / f"excel_btn_{datetime.now().strftime('%H%M%S')}.png"
                page.screenshot(path=str(shot))
                raise RuntimeError(f"엑셀 버튼을 찾지 못했습니다. 스크린샷: {shot}")

    download = dl_info.value
    ext = Path(download.suggested_filename).suffix or ".xlsx"
    filename = f"pp_{date_str.replace('-','')}_{datetime.now().strftime('%H%M%S')}{ext}"
    save_path = DOWNLOAD_DIR / filename
    download.save_as(str(save_path))
    print(f"  💾 저장: {save_path.name} ({save_path.stat().st_size // 1024}KB)")
    return save_path


# ── Excel/CSV 파싱 → DB 저장 ───────────────────────────────────
def parse_and_save(file_path: Path, date_str: str) -> int:
    """
    파워플래너 엑셀 파일 파싱 → power_usage_raw INSERT IGNORE
    date_str: 'YYYY-MM-DD'
    """
    cust_no = PP_ID
    rows = []

    # Excel 파싱 (.xls → xlrd, .xlsx → openpyxl)
    if file_path.suffix.lower() == ".xls":
        try:
            import xlrd
            wb = xlrd.open_workbook(str(file_path))
            ws = wb.sheet_by_index(0)
            headers = None
            for i in range(ws.nrows):
                row = ws.row_values(i)
                if headers is None:
                    # 헤더 행: 첫 셀이 문자열이고 숫자가 아닌 행
                    if row and row[0] and isinstance(row[0], str) and not _is_number(row[0]):
                        headers = [str(c).strip() for c in row]
                    continue
                if not row or row[0] is None or row[0] == "":
                    continue
                if headers:
                    rows.append({h: v for h, v in zip(headers, row)})
        except ImportError:
            print("  ⚠️  xlrd 없음 — pip install xlrd")
        except Exception as e:
            print(f"  ⚠️  xls 파싱 오류: {e}")

    elif file_path.suffix.lower() == ".xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            headers = None
            for row in ws.iter_rows(values_only=True):
                if headers is None:
                    if row and row[0] and isinstance(row[0], str) and not _is_number(row[0]):
                        headers = [str(c).strip() if c else "" for c in row]
                    continue
                if not row or row[0] is None:
                    continue
                if headers:
                    rows.append({h: v for h, v in zip(headers, row)})
        except ImportError:
            print("  ⚠️  openpyxl 없음 — pip install openpyxl")
        except Exception as e:
            print(f"  ⚠️  xlsx 파싱 오류: {e}")

    # HTML 형식 파싱 (파워플래너 엑셀은 실제로 HTML 테이블)
    if not rows and file_path.suffix.lower() in (".xls", ".xlsx", ".html"):
        rows = _parse_html_table(file_path, date_str)

    # CSV 파싱 (fallback)
    if not rows:
        for enc in ["utf-8-sig", "euc-kr", "cp949"]:
            try:
                with open(file_path, encoding=enc) as f:
                    reader = csv.DictReader(f)
                    rows = list(reader)
                if rows:
                    break
            except Exception:
                continue

    if not rows:
        print(f"  ❌ 파싱 실패: {file_path.name}")
        return 0

    print(f"  📊 파싱 행수: {len(rows)}")

    # 레코드 직렬화 (미래 시간 제외)
    now = datetime.now()
    records = []
    for row in rows:
        dt  = row.get("measured_at")
        kwh = _to_float(row.get("kwh", 0))
        kw  = _to_float(row.get("kw_demand", 0))
        if not dt:
            continue
        if dt > now:
            continue
        records.append({
            "measured_at": dt.strftime("%Y-%m-%d %H:%M:%S"),
            "kwh":         kwh,
            "kw_demand":   kw,
        })

    if not records:
        print("  ⚠️  저장할 레코드 없음")
        return 0

    # 1순위: Flask API POST (로컬 실행 시)
    if API_URL:
        saved = _save_via_api(cust_no, records)
        if saved >= 0:
            return saved

    # 2순위: MariaDB 직접 (컨테이너 내부 실행 시)
    return _save_direct_db(cust_no, records)


def _save_via_api(cust_no: str, records: list) -> int:
    """Flask API POST로 저장 (192.168.100.11:5050/api/power_upload)"""
    import urllib.request, urllib.error, json as _json
    payload = _json.dumps({"cust_no": cust_no, "records": records}).encode()
    req = urllib.request.Request(
        API_URL,
        data=payload,
        method="POST",
        headers={
            "Content-Type": "application/json",
            "X-Power-Token": API_TOKEN,
        }
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            result = _json.loads(resp.read())
            saved = result.get("saved", 0)
            print(f"  ✅ API 저장: {saved}건 (전체 {result.get('total',0)}건)")
            return saved
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        print(f"  ⚠️  API 오류 {e.code}: {body}")
        return -1
    except Exception as e:
        print(f"  ⚠️  API 연결 실패: {e}")
        return -1


def _save_direct_db(cust_no: str, records: list) -> int:
    """MariaDB 직접 저장 (Docker 컨테이너 내부 실행 시)"""
    try:
        conn = pymysql.connect(**MARIA)
        cur  = conn.cursor()
        saved = 0
        for r in records:
            try:
                cur.execute("""
                    INSERT IGNORE INTO power_usage_raw (cust_no, measured_at, kwh, kw_demand)
                    VALUES (%s, %s, %s, %s)
                """, (cust_no, r["measured_at"], r["kwh"], r["kw_demand"]))
                saved += cur.rowcount
            except Exception:
                continue
        conn.commit()
        conn.close()
        print(f"  ✅ DB 저장: {saved}건")
        return saved
    except Exception as e:
        print(f"  ❌ DB 저장 실패: {e}")
        return 0


def _parse_html_table(file_path: Path, date_str: str) -> list:
    """
    파워플래너 엑셀(실제 HTML) 파싱
    tableListHour: 2열 구조 (시 01-12 / 13-24), 각 열에 사용량·최대수요·무효전력·CO2·역률
    date_str: 'YYYY-MM-DD'
    """
    import re
    rows = []
    try:
        for enc in ["utf-8-sig", "utf-8", "euc-kr", "cp949"]:
            try:
                content = file_path.read_bytes().decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            return rows

        # tableListHour 추출
        m = re.search(r'<table[^>]+id=["\']tableListHour["\'][^>]*>(.*?)</table>',
                      content, re.DOTALL | re.IGNORECASE)
        if not m:
            print("  ⚠️  tableListHour 없음")
            return rows

        table_html = m.group(1)
        tr_list    = re.findall(r'<tr[^>]*>(.*?)</tr>', table_html, re.DOTALL | re.IGNORECASE)

        base_date = datetime.strptime(date_str, "%Y-%m-%d")

        for tr in tr_list:
            cells = re.findall(r'<(?:td|th)[^>]*>(.*?)</(?:td|th)>',
                               tr, re.DOTALL | re.IGNORECASE)
            cells = [re.sub(r'<[^>]+>', '', c).replace('&nbsp;', '').strip() for c in cells]

            # 헤더 행 건너뛰기 (첫 셀이 숫자가 아닌 경우)
            if not cells or not _is_number(cells[0]):
                continue
            if len(cells) < 8:
                continue

            # 왼쪽 열: 시간 01-12
            # cells[0]=시, [1]=사용량kWh, [2]=최대수요kW, [3]=무효지상, [4]=무효진상,
            # [5]=CO2, [6]=역률지상, [7]=역률진상
            # 오른쪽 열: 시간 13-24
            # cells[8]=시, [9]=사용량kWh, [10]=최대수요kW, ...

            def _make_dt(hour_str):
                try:
                    h = int(float(hour_str))
                    if h == 24:  # 24시 = 다음날 00시
                        return base_date + timedelta(days=1)
                    return base_date.replace(hour=h % 24)
                except (ValueError, TypeError):
                    return None

            # 왼쪽 열
            dt_l = _make_dt(cells[0])
            if dt_l:
                rows.append({
                    "measured_at": dt_l,
                    "kwh":         cells[1],
                    "kw_demand":   cells[2],
                })

            # 오른쪽 열 (16열인 경우)
            if len(cells) >= 16:
                dt_r = _make_dt(cells[8])
                if dt_r:
                    rows.append({
                        "measured_at": dt_r,
                        "kwh":         cells[9],
                        "kw_demand":   cells[10],
                    })

        print(f"  📋 HTML 파싱: {len(rows)}건 (시간대별)")

    except Exception as e:
        print(f"  ⚠️  HTML 파싱 오류: {e}")

    return rows


def _to_float(v) -> float:
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0


def _is_number(s) -> bool:
    try:
        float(str(s).replace(",", ""))
        return True
    except (ValueError, TypeError):
        return False


def _find_col(row: dict, candidates: list) -> str:
    for c in candidates:
        for k in row.keys():
            if c.lower() in str(k).lower():
                return k
    return ""


# ── 하루치 수집 ────────────────────────────────────────────────
def collect_day(page, date_str: str) -> int:
    """
    date_str: 'YYYY-MM-DD'
    시간대별 페이지 → 날짜 설정 → 조회 → 엑셀 다운로드 → DB 저장
    """
    print(f"\n  📡 {date_str} 수집...")

    if not go_to_hourly_page(page):
        return 0

    page.screenshot(path=str(SCREENSHOT_DIR / f"hourly_{date_str}.png"))

    set_date_and_search(page, date_str)

    page.screenshot(path=str(SCREENSHOT_DIR / f"result_{date_str}.png"))

    try:
        file_path = download_excel(page, date_str)
    except Exception as e:
        print(f"  ❌ 다운로드 실패: {e}")
        return 0

    return parse_and_save(file_path, date_str)


# ── 월별 청구요금 스크래핑 ────────────────────────────────────
def scrape_billing(page, year: int = None) -> list:
    """
    파워플래너 월별 청구요금 페이지 스크래핑
    year: 특정 연도 선택 (None=기본/현재년도)
      2024 → KEPCO 2024년도 청구 (Apr 2024 – Mar 2025)
      2025 → KEPCO 2025년도 청구 (Apr 2025 – Mar 2026) = 기본값과 동일
    Returns: [{"ym":"YYYYMM","bill_amount":123456,"total_kwh":4567.0}, ...]
    """
    import time as _t
    print("  💳 월별 청구요금 페이지 이동...")
    page.goto(BILLING_URL, wait_until="networkidle", timeout=15000)
    _t.sleep(2)

    # 역률 팝업 등 모달 닫기
    _close_popups(page)

    # 연도 선택 (JavaScript로 직접 조작 — 팝업 클릭 충돌 회피)
    if year:
        print(f"    📅 연도 선택: {year}년")
        changed = page.evaluate(f"""() => {{
            var sels = document.querySelectorAll('select');
            for (var i = 0; i < sels.length; i++) {{
                var opts = sels[i].options;
                for (var j = 0; j < opts.length; j++) {{
                    if (opts[j].text.trim().startsWith('{year}') ||
                        opts[j].value.trim().startsWith('{year}')) {{
                        sels[i].selectedIndex = j;
                        sels[i].dispatchEvent(new Event('change', {{bubbles:true}}));
                        return sels[i].name || sels[i].id || 'ok';
                    }}
                }}
            }}
            return null;
        }}""")
        print(f"    select 변경(JS): {changed}")
        _t.sleep(0.5)

        # 조회 버튼 클릭 (KEPCO: a.btn_blue_right, 이미지 기반이라 text 없음)
        searched = False
        for btn in [
            "a.btn_blue_right",  # KEPCO 청구요금 페이지 조회 버튼 실제 클래스
            "input[value='조회']", "input[type='image'][alt='조회']",
            "button:text-is('조회')", "a:text-is('조회')",
        ]:
            try:
                page.click(btn, timeout=2000)
                searched = True
                print(f"    조회 클릭: {btn}")
                break
            except Exception:
                continue

        if not searched:
            # JS fallback: 알려진 함수명 (doSubmit = KEPCO 월별청구요금 조회 함수)
            result = page.evaluate("""() => {
                var fns = ['doSubmit','fnSearch','fn_search','doSearch',
                           'getMonthBillInfo','searchBill','goSearch'];
                for (var f of fns) {
                    if (typeof window[f] === 'function') { window[f](); return f; }
                }
                // href에 doSubmit 포함된 링크 클릭
                var links = document.querySelectorAll('a[href*=\"doSubmit\"]');
                if (links.length > 0) { links[0].click(); return 'a[href*=doSubmit]'; }
                return null;
            }""")
            print(f"    조회 JS: {result}")
            searched = bool(result)

        try:
            page.wait_for_load_state("networkidle", timeout=10000)
        except Exception:
            pass
        _t.sleep(2)

    shot = SCREENSHOT_DIR / f"billing_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
    page.screenshot(path=str(shot))
    print(f"    스크린샷: {shot.name}")

    # 1순위: HTML 직접 파싱 (팝업 간섭 없이)
    bills = _parse_billing_html(page.content())

    # 2순위: 엑셀 다운로드 시도 (HTML 파싱 실패 시)
    if not bills:
        try:
            with page.expect_download(timeout=10000) as dl_info:
                clicked = False
                for sel in ["a[href*='excelDown']", "a[href*='Excel']",
                            "img[alt='엑셀']", "a:has(img[alt='엑셀'])"]:
                    try:
                        page.click(sel, timeout=2000)
                        clicked = True
                        print(f"    엑셀 버튼 클릭: {sel}")
                        break
                    except PWTimeout:
                        continue
                if not clicked:
                    page.evaluate("if(typeof excelDown==='function') excelDown(); "
                                  "else if(typeof fnExcelDown==='function') fnExcelDown();")

            dl = dl_info.value
            ext = Path(dl.suggested_filename).suffix or ".xls"
            fp  = DOWNLOAD_DIR / f"billing_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
            dl.save_as(str(fp))
            print(f"    💾 {fp.name} ({fp.stat().st_size // 1024}KB)")
            bills = _parse_billing_file(fp)
        except Exception as e:
            print(f"    ⚠️  엑셀 다운로드 실패: {e}")

    return bills


def _close_popups(page):
    """역률 팝업, 모달 등 닫기"""
    import time as _t

    # ESC 키로 닫기 시도
    try:
        page.keyboard.press("Escape")
        _t.sleep(0.5)
    except Exception:
        pass

    # 팝업 닫기 버튼 클릭 시도
    for close_sel in [
        ".btn_close", "button.btn_close", "a.btn_close",
        ".popup_wrap .close", ".layer_popup .close",
        "button[title='닫기']", "a[title='닫기']",
        ".popup-close", ".modal-close",
        "a:has-text('×')", "button:has-text('×')",
        "a[onclick*='close']", "a[onclick*='Close']",
        ".pop_close", ".layer_close",
    ]:
        try:
            els = page.locator(close_sel).all()
            for el in els:
                if el.is_visible(timeout=300):
                    el.click()
                    _t.sleep(0.3)
                    print(f"    팝업 닫기: {close_sel}")
        except Exception:
            pass

    # JavaScript로 모든 팝업/레이어 숨기기 (최후 수단)
    try:
        closed = page.evaluate("""() => {
            var hidden = [];
            var sels = [
                '.layer_pop', '.popup_wrap', '.pop_wrap', '.modal-dialog',
                '.ui-dialog', '[class*="layer_"][class*="pop"]',
                '.dimm', '.dim_bg', '.overlay'
            ];
            sels.forEach(function(s) {
                document.querySelectorAll(s).forEach(function(el) {
                    if (el.offsetParent !== null) {  // visible
                        el.style.display = 'none';
                        hidden.push(s);
                    }
                });
            });
            return hidden;
        }""")
        if closed:
            print(f"    JS 팝업 숨김: {closed}")
    except Exception:
        pass

    _t.sleep(0.5)


def _parse_billing_file(fp: Path) -> list:
    """다운로드된 청구요금 파일 파싱 (HTML 위장 XLS 포함)"""
    # HTML 위장 파일 확인
    try:
        raw = fp.read_bytes()
        sample = raw[:2048].decode("utf-8-sig", errors="replace")
        if "<html" in sample.lower() or "<table" in sample.lower():
            return _parse_billing_html(raw.decode("utf-8-sig", errors="replace"))
    except Exception:
        pass

    # openpyxl 시도 (.xlsx)
    if fp.suffix.lower() == ".xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(fp, data_only=True)
            ws = wb.active
            rows = [[str(c) if c is not None else "" for c in row] for row in ws.iter_rows(values_only=True)]
            return _parse_billing_rows(rows)
        except Exception as e:
            print(f"    ⚠️  openpyxl 파싱 실패: {e}")

    return []


def _parse_billing_html(html: str) -> list:
    """
    파워플래너 청구요금 HTML 파싱.
    연월 + 청구금액(원) + 사용량(kWh) 추출.
    """
    import re
    bills = []
    seen_ym = set()

    # 연월 패턴: "2026년 03월" / "2026-03" / "202603"
    ym_pat  = re.compile(r'(\d{4})[년/\-]?\s*0?(\d{1,2})월?')
    # 금액: 5자리+ 숫자 (콤마 포함) — "1,234,567"
    amt_pat = re.compile(r'([\d,]{5,})')
    # kWh
    kwh_pat = re.compile(r'([\d,]+\.?\d*)\s*(?:kWh|KWH|kwh)')

    rows_html = re.findall(r'<tr[^>]*>(.*?)</tr>', html, re.DOTALL | re.IGNORECASE)
    for tr in rows_html:
        cells = re.findall(r'<(?:td|th)[^>]*>(.*?)</(?:td|th)>', tr, re.DOTALL | re.IGNORECASE)
        cells = [re.sub(r'<[^>]+>', '', c).replace('&nbsp;', ' ').strip() for c in cells]
        row_text = ' '.join(cells)

        ym_m = ym_pat.search(row_text)
        if not ym_m:
            continue

        year = int(ym_m.group(1))
        mon  = int(ym_m.group(2))
        if year < 2020 or year > 2030 or mon < 1 or mon > 12:
            continue
        ym = f"{year}{mon:02d}"
        if ym in seen_ym:
            continue

        # 금액 후보: 5자리+ 숫자 중 최대값 = 청구액
        amounts = []
        for a in amt_pat.findall(row_text):
            try:
                v = int(a.replace(',', ''))
                if v > 10000:   # 최소 1만원 이상
                    amounts.append(v)
            except ValueError:
                pass
        if not amounts:
            continue
        bill_amount = max(amounts)

        kwh_m = kwh_pat.search(row_text)
        total_kwh = float(kwh_m.group(1).replace(',', '')) if kwh_m else 0.0

        seen_ym.add(ym)
        bills.append({"ym": ym, "bill_amount": bill_amount, "total_kwh": total_kwh})
        print(f"    📅 {ym[:4]}-{ym[4:]} : {bill_amount:,}원  {total_kwh:,.1f}kWh")

    if not bills:
        print("    ⚠️  청구요금 파싱 결과 없음 (HTML 구조 확인 필요)")

    return bills


def _parse_billing_rows(rows: list) -> list:
    """엑셀 rows(list of list[str]) → 청구요금 dict 목록"""
    import re
    bills = []
    seen_ym = set()
    ym_pat  = re.compile(r'(\d{4})[년/\-]?\s*0?(\d{1,2})월?')
    for row in rows:
        row_text = ' '.join(str(c) for c in row)
        ym_m = ym_pat.search(row_text)
        if not ym_m:
            continue
        year, mon = int(ym_m.group(1)), int(ym_m.group(2))
        if year < 2020 or year > 2030 or mon < 1 or mon > 12:
            continue
        ym = f"{year}{mon:02d}"
        if ym in seen_ym:
            continue
        nums = []
        for c in row:
            try:
                v = int(str(c).replace(',', ''))
                if v > 10000:
                    nums.append(v)
            except (ValueError, TypeError):
                pass
        if not nums:
            continue
        seen_ym.add(ym)
        bill_amount = max(nums)
        bills.append({"ym": ym, "bill_amount": bill_amount, "total_kwh": 0.0})
        print(f"    📅 {ym[:4]}-{ym[4:]} : {bill_amount:,}원")
    return bills


def scrape_smartview(page) -> dict:
    """
    파워플래너 스마트뷰 실시간 요금 스크래핑 (로그인 직후 rm0101.do 페이지)
    반환: rt_kwh, rt_bill, basic_rate, basic_bill, demand_kw, demand_ym,
          max_demand_kw, max_demand_date, bill_start, bill_end, rate_table
    """
    import re, time as _t
    print("  📡 스마트뷰 실시간 요금 수집...")
    page.goto("https://pp.kepco.co.kr/rm/rm0101.do?menu_id=O010101",
              wait_until="networkidle", timeout=15000)
    _t.sleep(3)

    shot = SCREENSHOT_DIR / f"smartview_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
    page.screenshot(path=str(shot))
    print(f"    스크린샷: {shot.name}")

    raw = page.evaluate("""() => {
        function txt(id) {
            var el = document.getElementById(id);
            return el ? el.textContent.trim() : '';
        }
        function liVal(id) {
            var el = document.getElementById(id);
            if (!el) return '';
            var clone = el.cloneNode(true);
            clone.querySelectorAll('span').forEach(function(s){ s.remove(); });
            return clone.textContent.trim();
        }
        // 전력량요금 단가표 파싱
        var rateRows = [];
        var loads = ['경부하', '중간부하', '최대부하'];
        var tables = document.querySelectorAll('table');
        for (var t of tables) {
            var trs = t.querySelectorAll('tr');
            for (var tr of trs) {
                var tds = Array.from(tr.querySelectorAll('td'));
                var texts = tds.map(function(c){ return c.textContent.trim(); });
                var loadName = texts.find(function(t){ return loads.includes(t); });
                if (loadName) {
                    var nums = texts.filter(function(t){ return /^[0-9]+\\.?[0-9]*$/.test(t); });
                    if (nums.length >= 3) {
                        rateRows.push({
                            load: loadName,
                            summer: parseFloat(nums[0]),
                            spring_fall: parseFloat(nums[1]),
                            winter: parseFloat(nums[2])
                        });
                    }
                }
            }
            if (rateRows.length >= 3) break;
        }
        // 적용전기요금 (LI 첫번째, span 제거 후)
        var tariffEl = document.querySelector('li:first-child');
        var tariff = '';
        var lis = document.querySelectorAll('li');
        for (var li of lis) {
            var sp = li.querySelector('span');
            if (sp && sp.textContent.trim() === '적용전기요금') {
                var c = li.cloneNode(true);
                c.querySelectorAll('span').forEach(function(s){ s.remove(); });
                tariff = c.textContent.trim();
                break;
            }
        }
        return {
            rt_kwh:         parseFloat((txt('F_AP_QT') || '0').replace(/[^0-9.]/g,'')) || 0,
            rt_bill:        parseFloat((txt('TOTAL_CHARGE') || '0').replace(/[^0-9.]/g,'')) || 0,
            bill_start:     txt('START_DT'),
            bill_end:       txt('END_DT'),
            cur_date:       txt('SELECT_DT'),
            basic_rate_txt: liVal('showBb'),
            basic_bill_txt: liVal('showBb1'),
            demand_txt:     liVal('showBb2'),
            max_demand_txt: liVal('max_over_event'),
            tariff:         tariff,
            rate_table:     rateRows
        };
    }""")

    def _first_int(s):
        m = re.search(r'[\d,]+', str(s))
        return int(m.group().replace(',', '')) if m else 0

    def _first_float(s):
        m = re.search(r'[\d,]+\.?\d*', str(s))
        return float(m.group().replace(',', '')) if m else 0.0

    result = {
        "rt_kwh":          raw.get("rt_kwh", 0),
        "rt_bill":         int(raw.get("rt_bill", 0)),
        "bill_start":      raw.get("bill_start", ""),
        "bill_end":        raw.get("bill_end", ""),
        "cur_date":        raw.get("cur_date", ""),
        "tariff":          raw.get("tariff", ""),
        "basic_rate":      _first_int(raw.get("basic_rate_txt", "")),
        "basic_bill":      _first_int(raw.get("basic_bill_txt", "")),
        "demand_kw":       _first_float(raw.get("demand_txt", "")),
        "demand_ym":       "",
        "max_demand_kw":   _first_float(raw.get("max_demand_txt", "")),
        "max_demand_date": "",
        "rate_table":      raw.get("rate_table", []),
    }

    dm = re.search(r"'(\d{2})년\s*(\d{1,2})월", raw.get("demand_txt", ""))
    if dm:
        result["demand_ym"] = f"20{dm.group(1)}{int(dm.group(2)):02d}"

    dd = re.search(r'(\d{1,2}월\s*\d{1,2}일)', raw.get("max_demand_txt", ""))
    if dd:
        result["max_demand_date"] = dd.group(1)

    print(f"    실시간사용량: {result['rt_kwh']:,.0f} kWh | 실시간요금: {result['rt_bill']:,}원")
    print(f"    기본요금: {result['basic_bill']:,}원 | 최대수요: {result['max_demand_kw']:.0f} kW ({result['max_demand_date']})")
    return result


def _save_smartview_via_api(data: dict) -> bool:
    """Flask /api/power_smartview_upload POST로 저장"""
    import urllib.request, urllib.error, json as _json
    payload = _json.dumps({"cust_no": PP_ID, **data}).encode()
    req = urllib.request.Request(
        SMARTVIEW_API_URL,
        data=payload, method="POST",
        headers={"Content-Type": "application/json", "X-Power-Token": API_TOKEN}
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            result = _json.loads(resp.read())
            print(f"  ✅ 스마트뷰 API 저장: {result}")
            return True
    except Exception as e:
        print(f"  ⚠️  스마트뷰 API 저장 실패: {e}")
        return False


def _save_bills_via_api(bills: list) -> int:
    """Flask /api/power_bill_upload POST로 청구요금 저장"""
    import urllib.request, urllib.error, json as _json
    payload = _json.dumps({"cust_no": PP_ID, "bills": bills}).encode()
    req = urllib.request.Request(
        BILL_API_URL,
        data=payload,
        method="POST",
        headers={
            "Content-Type": "application/json",
            "X-Power-Token": API_TOKEN,
        }
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            result = _json.loads(resp.read())
            saved = result.get("saved", 0)
            print(f"  ✅ 청구요금 API 저장: {saved}건")
            return saved
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        print(f"  ⚠️  청구요금 API 오류 {e.code}: {body}")
        return 0
    except Exception as e:
        print(f"  ⚠️  청구요금 API 연결 실패: {e}")
        return 0


# ── 탐색 모드 ──────────────────────────────────────────────────
def explore(page):
    """로그인 후 화면 스크린샷 + 링크 목록 (main에서 login 후 호출)"""
    shot = SCREENSHOT_DIR / f"main_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
    page.screenshot(path=str(shot), full_page=True)
    print(f"  📸 메인 화면: {shot}")

    links = page.eval_on_selector_all(
        "a[href]",
        "els => els.map(e => ({text: e.innerText.trim(), href: e.href}))"
    )
    print("\n  📋 메뉴/링크:")
    for lnk in links:
        if lnk["text"] and len(lnk["text"]) < 30:
            print(f"    {lnk['text']:20s} → {lnk['href']}")

    btns = page.eval_on_selector_all(
        "button, a",
        "els => els.map(e => e.innerText.trim()).filter(t => t && "
        "(t.includes('엑셀') || t.includes('Excel') || t.includes('다운') || t.includes('CSV')))"
    )
    print(f"\n  📥 다운로드 버튼: {btns}")

    go_to_hourly_page(page)
    shot2 = SCREENSHOT_DIR / f"hourly_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
    page.screenshot(path=str(shot2), full_page=True)
    print(f"  📸 시간대별 화면: {shot2}")


# ── 메인 ───────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--explore",   action="store_true", help="로그인 후 화면 탐색")
    parser.add_argument("--smartview", action="store_true", help="스마트뷰 실시간 요금 수집")
    parser.add_argument("--billing",      action="store_true", help="월별 청구요금 수집")
    parser.add_argument("--billing-year", type=int, default=0, help="청구요금 연도 선택 (예: 2025)")
    parser.add_argument("--date",     default="",          help="수집 날짜 YYYYMMDD (기본: 오늘)")
    parser.add_argument("--month",    default="",          help="수집 월 YYYYMM (전체 일 반복)")
    parser.add_argument("--headless", action="store_true", help="헤드리스 모드")
    parser.add_argument("--backfill", type=int, default=0,
                        help="최근 N개월 데이터 백필 (예: --backfill 12)")
    args = parser.parse_args()

    if not PP_ID or not PP_PW:
        print("❌ .env에 PP_CUSTOMER_ID, PP_PASSWORD 를 설정하세요.")
        sys.exit(1)

    # 수집 날짜 목록 결정
    if args.month:
        ym = args.month  # YYYYMM
        year, month = int(ym[:4]), int(ym[4:6])
        # 해당 월의 모든 날짜
        from calendar import monthrange
        _, last_day = monthrange(year, month)
        dates = [
            f"{year}-{month:02d}-{d:02d}"
            for d in range(1, last_day + 1)
            if datetime(year, month, d) <= datetime.now()
        ]
    elif args.backfill:
        from calendar import monthrange
        dates = []
        base = datetime.now()
        for m_back in range(args.backfill, 0, -1):      # 오래된 달부터 순서대로
            total_months = base.year * 12 + base.month - 1 - m_back
            year  = total_months // 12
            month = total_months  % 12 + 1
            _, last_day = monthrange(year, month)
            for d in range(1, last_day + 1):
                dt = datetime(year, month, d)
                if dt <= datetime.now():
                    dates.append(f"{year}-{month:02d}-{d:02d}")
        print(f"  📅 백필 대상: {dates[0]} ~ {dates[-1]} ({len(dates)}일)")
    elif args.date:
        ds = args.date  # YYYYMMDD
        dates = [f"{ds[:4]}-{ds[4:6]}-{ds[6:8]}"]
    else:
        dates = [datetime.now().strftime("%Y-%m-%d")]

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=args.headless,
            downloads_path=str(DOWNLOAD_DIR),
        )
        ctx  = browser.new_context(accept_downloads=True)
        page = ctx.new_page()

        try:
            login(page)

            if args.explore:
                explore(page)
            elif args.smartview:
                print("\n  📡 스마트뷰 수집 모드")
                sv = scrape_smartview(page)
                if sv:
                    _save_smartview_via_api(sv)
                    print("\n✅ 스마트뷰 수집 완료")
                else:
                    print("\n⚠️  스마트뷰 데이터 없음")
            elif args.billing:
                print("\n  💳 청구요금 수집 모드")
                from datetime import datetime as _dt
                _now = _dt.now()
                yr = args.billing_year if args.billing_year else None
                bills = scrape_billing(page, year=yr)
                # KEPCO 회계연도는 4월 시작 — 4월 이후면 새 회계연도 탭도 추가 수집
                if not yr and _now.month >= 4:
                    next_fy = _now.year
                    print(f"\n  💳 {next_fy}년 탭 추가 수집 (이번달 청구 포함)...")
                    extra = scrape_billing(page, year=next_fy)
                    seen = {b["ym"] for b in bills}
                    bills += [b for b in extra if b["ym"] not in seen]
                if bills:
                    _save_bills_via_api(bills)
                    print(f"\n✅ 청구요금 수집 완료: 총 {len(bills)}개월")
                else:
                    print("\n⚠️  청구요금 데이터 없음 — pp_screenshots/billing_*.png 확인")
            else:
                total = 0
                for date_str in dates:
                    try:
                        n = collect_day(page, date_str)
                        total += n
                    except Exception as e:
                        shot = SCREENSHOT_DIR / f"error_{date_str}.png"
                        try:
                            page.screenshot(path=str(shot))
                        except Exception:
                            pass
                        print(f"  ❌ {date_str} 오류: {e}")
                        continue

                print(f"\n✅ 수집 완료: 총 {total}건 ({len(dates)}일)")

        except Exception as e:
            shot = SCREENSHOT_DIR / f"error_{datetime.now().strftime('%H%M%S')}.png"
            try:
                page.screenshot(path=str(shot))
                print(f"  📸 오류 스크린샷: {shot}")
            except Exception:
                pass
            print(f"❌ 오류: {e}")
            raise
        finally:
            browser.close()


if __name__ == "__main__":
    main()
